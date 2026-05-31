"""
SUNSHINE SHOPPING CENTER 订货系统 / Ordering System
====================================================
Single-file Streamlit app.

Run:
    streamlit run app.py --server.port 8502 --server.address 0.0.0.0
"""
from __future__ import annotations

import base64
import html
import io
import hashlib
import hmac
import json
import os
import re
import smtplib
import sqlite3
import ssl
import threading
import time
import traceback
import urllib.error
import urllib.request
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =========================================================================
# CONFIG / CONSTANTS
# =========================================================================
def _resolve_storage_paths() -> tuple[Path, Path, Path, Path, Path]:
    """Railway Volume: SUNSHINE_DATA_DIR=/data, DB_PATH=/data/orders.db (optional)."""
    data_env = os.getenv("SUNSHINE_DATA_DIR", "").strip()
    db_env = os.getenv("DB_PATH", "").strip()
    data_root = Path(data_env) if data_env else Path(".")
    if db_env:
        db_path = Path(db_env)
        if not data_env:
            data_root = db_path.parent
    else:
        db_path = data_root / "orders.db"
    for d in (data_root, db_path.parent):
        try:
            d.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
    return (
        db_path,
        data_root / "email_config.json",
        data_root / "email_log.json",
        data_root / "backups",
        data_root / "app_runtime.log",
    )


(
    DB_PATH,
    EMAIL_CONFIG_PATH,
    EMAIL_LOG_PATH,
    BACKUP_DIR,
    APP_LOG_PATH,
) = _resolve_storage_paths()


@dataclass
class ExcelSheetImportOutcome:
    """Excel 批量导入结果（价格 / 库存共用结构）。

    - n_written：实际写入数据库（或产生库存流水）的行数
    - n_skipped_benign：不构成错误、有意跳过的行（如数量为 0、价格与库内一致）
    - n_failed：无法处理或处理报错的行数
    """

    n_written: int = 0
    n_skipped_benign: int = 0
    n_failed: int = 0
    failure_messages: list[str] = field(default_factory=list)

    @property
    def n_ok(self) -> int:
        """界面「成功」：已生效的变更条数（与 n_written 一致）。"""
        return int(self.n_written)


def _import_row_ref(sheet_row: int, item_code: str, barcode: str, name: str) -> str:
    bits: list[str] = []
    nm = (name or "").strip()
    if nm:
        bits.append(nm[:48])
    ic = (item_code or "").strip()
    if ic:
        bits.append(f"编号:{ic}")
    bc = (barcode or "").strip()
    if bc:
        bits.append(f"条码:{bc}")
    label = " · ".join(bits) if bits else "（空行/无标识）"
    return f"表第 {sheet_row} 行 · {label}"


def _df_row_is_blank_catalog_row(row: pd.Series) -> bool:
    ic = str(row.get("ItemCode", "") or "").strip()
    bc = str(row.get("Barcode", "") or "").strip()
    nm = str(row.get("Name", "") or "").strip()
    return not ic and not bc and not nm
PRODUCTS_PATH = Path("products.xlsx")


def products_master_excel_path() -> Path:
    """商品主数据文件路径。

    优先使用根目录下「商品档案_*.xlsx」中**修改时间最新**的一份（例如
    商品档案_260506_080011.xlsx）；若不存在则回退到 products.xlsx。
    仓库保存/清空主数据时与读取使用同一路径，避免新旧两套表打架。"""
    root = Path(".")
    candidates = sorted(
        root.glob("商品档案_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]
    return PRODUCTS_PATH


def _normalize_product_sheet_columns(df: pd.DataFrame) -> pd.DataFrame:
    """将常见中文列名映射为 ItemCode / Barcode / Name，便于后续统一解析。"""
    if df is None or df.empty:
        return df
    col_by_lower = {
        str(c).replace("\ufeff", "").strip().lower(): c for c in df.columns
    }

    def _has(canonical: str) -> bool:
        return canonical in df.columns

    def _first_alias(cands: list[str]) -> str | None:
        for cand in cands:
            orig = col_by_lower.get(cand.strip().lower())
            if orig is not None:
                return orig
        return None

    renames: dict[str, str] = {}
    if not _has("ItemCode"):
        hit = _first_alias(
            ["商品编号", "货号", "编号", "内部编号", "sku", "itemcode", "item_code"]
        )
        if hit:
            renames[hit] = "ItemCode"
    if not _has("Barcode"):
        hit = _first_alias(
            ["条码", "商品条码", "条形码", "barcode", "ean", "upc"]
        )
        if hit:
            renames[hit] = "Barcode"
    if not _has("Name"):
        hit = _first_alias(["商品名称", "名称", "品名", "name", "description"])
        if hit:
            renames[hit] = "Name"
    if renames:
        df = df.rename(columns=renames)
    return df

# In-app page id in the address bar (?p=). Synced via st.query_params so the
# browser back/forward button matches Streamlit session (do not push a second
# history stack with window.history — that desyncs URL from page content).
URL_PAGE_QUERY_KEY = "p"

# Product image folder. File naming convention (in priority order):
#   images/<ItemCode>.<ext>   e.g. images/P001.jpg
#   images/<Barcode>.<ext>    e.g. images/8992730950194.png
# Supported extensions: jpg, jpeg, png, webp.
# Missing images are silently ignored — the UI shows a placeholder.
# Recommended: keep each image under 200KB and ≤ 800px on the longest edge
# so a 10-item search page stays snappy on slow phone connections.
IMAGES_DIR = Path("images")

# BACKUP_DIR resolved with SUNSHINE_DATA_DIR (see _resolve_storage_paths).
BACKUP_RETAIN = 30                    # keep at most this many auto-snapshots
BACKUP_MIN_INTERVAL_MINUTES = 60      # auto-snapshot at most once per hour

# Email paths resolved with DB_PATH / SUNSHINE_DATA_DIR above.
EMAIL_LOG_KEEP = 200                  # keep last N log entries


def _load_optional_env_file(path: Path) -> None:
    """Load KEY=VALUE lines into os.environ only when the key is unset.

    Used for local `gemini.env` next to app.py. Never logs file contents."""
    if not path.exists():
        return
    try:
        for raw in path.read_text(encoding="utf-8").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key and val and os.getenv(key) is None:
                os.environ[key] = val
    except Exception:
        pass


# Local dev: optional `gemini.env` (same folder as app.py). Env vars still win if set.
_load_optional_env_file(Path("gemini.env"))

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()


def _coerce_gemini_model_id(raw: str) -> str:
    """v1beta no longer serves many 1.5 model IDs for new keys — map to 2.0 Flash."""
    r = (raw or "").strip()
    if not r:
        return "gemini-2.0-flash"
    rl = r.lower()
    if rl.startswith("models/"):
        rl = rl[len("models/"):]
    if rl.startswith("gemini-1.5"):
        return "gemini-2.0-flash"
    return r


# Default to 2.0 Flash. System env may still set GEMINI_MODEL=gemini-1.5-flash — we coerce that.
GEMINI_MODEL = _coerce_gemini_model_id(
    os.getenv("GEMINI_MODEL", "gemini-2.0-flash").strip() or "gemini-2.0-flash"
)
# Comma-separated backup models when primary returns 429 (quota / rate limit).
GEMINI_FALLBACK_MODELS = os.getenv(
    "GEMINI_FALLBACK_MODELS",
    "gemini-2.5-flash,gemini-2.0-flash-lite",
).strip()
GEMINI_TIMEOUT_SEC = 20

# Email: Resend HTTPS API works on Railway Hobby (SMTP 587/465 is blocked there).
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "").strip()
RESEND_FROM = os.getenv("RESEND_FROM", "").strip()

# Product images live as plain files in this folder. Filename convention:
#   <ItemCode>.<ext>          e.g. P001.jpg
#   bc_<Barcode>.<ext>        e.g. bc_8801234567001.jpg  (when no ItemCode)
# This keeps images out of the SQLite DB (so backups stay small) and out of
# any added schema (so we don't need to migrate). Supported extensions: jpg
# jpeg png gif webp — whatever the phone camera or browser produces.
PRODUCT_IMAGES_DIR = Path("product_images")
IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".gif", ".webp")

BRANCHES: list[str] = [
    "NAMBER ONE STORE",
    "SUNSHINE MARKET",
    "SUNSHINE PS",
    "SUNSHINE FU-SANMA",
    "SUNSHINE FU-NEMO",
    "CHEN STORE-SARAKATA",
    "CHEN STORE-CHAPI",
]

WAREHOUSE_PASSWORD = os.getenv("SUNSHINE_WAREHOUSE_PASSWORD", "sunshine888")
ADMIN_PASSWORD = os.getenv("SUNSHINE_ADMIN_PASSWORD", "sunshine")
# 管理员手机号（可选）：设置后，管理员登录须同时校验手机号与密码。支持多个号码用英文逗号分隔。
# Comma-separated mobile numbers; when set, admin login requires matching phone + password.
ADMIN_PHONE_ENV = os.getenv("SUNSHINE_ADMIN_PHONE", "").strip()

# 临期预警天数：距离过期日 <= 该天数即触发预警。可用环境变量覆盖（默认 15 天）。
# Expiry warning window in days; SUNSHINE_EXPIRY_WARN_DAYS overrides the default.
EXPIRY_WARN_DAYS_DEFAULT = 15


def expiry_warn_days() -> int:
    """当前生效的预警天数（环境变量优先，非法值回退默认）。"""
    raw = os.getenv("SUNSHINE_EXPIRY_WARN_DAYS", "").strip()
    if raw.isdigit() and int(raw) > 0:
        return int(raw)
    return EXPIRY_WARN_DAYS_DEFAULT


def _normalize_phone_digits(raw: str) -> str:
    return re.sub(r"\D+", "", (raw or "").strip())


def _admin_phone_allowlist() -> set[str]:
    if not ADMIN_PHONE_ENV:
        return set()
    return {
        p for p in (_normalize_phone_digits(x) for x in ADMIN_PHONE_ENV.split(","))
        if p
    }

# Branch account permission codes (JSON stored per user; admin can grant subsets).
BRANCH_PERM_ORDER: list[tuple[str, str]] = [
    ("order", "perm_order"),
    ("my_orders", "perm_my_orders"),
    ("my_short", "perm_my_short"),
    ("messages", "perm_messages"),
    ("ai", "perm_ai"),
]
BRANCH_PERM_CODES: tuple[str, ...] = tuple(p for p, _ in BRANCH_PERM_ORDER)

ACCOUNT_STATUS_PENDING = "pending"
ACCOUNT_STATUS_APPROVED = "approved"
ACCOUNT_STATUS_REJECTED = "rejected"

AUDIT_EVENT_TYPES: tuple[str, ...] = (
    "login",
    "logout",
    "order_submit",
    "receive_confirm",
    "supplier_order",
    "catalog_reset",
    "catalog_save",
)


class Role:
    BRANCH = "branch"
    WAREHOUSE = "warehouse"
    ADMIN = "admin"


class OrderStatus:
    PENDING = "Pending"
    DISPATCHED = "Dispatched"
    RECEIVED = "Received"
    ALL = (PENDING, DISPATCHED, RECEIVED)


class ShortageStatus:
    OPEN = "Open"
    RESENDING = "Resending"
    OUT_OF_STOCK = "Out of Stock"
    RESOLVED = "Resolved"
    ALL = (OPEN, RESENDING, OUT_OF_STOCK, RESOLVED)


# =========================================================================
# I18N
# =========================================================================
T: dict[str, dict[str, str]] = {
    # Branding
    "app_brand":      {"en": "SUNSHINE SHOPPING CENTER", "zh": "阳光集团"},
    "app_subtitle":   {"en": "Ordering System",          "zh": "订货系统"},
    # Auth
    "select_role":    {"en": "Select your role",         "zh": "请选择您的角色"},
    "back_to_roles":  {"en": "◀ Back to role selection",  "zh": "◀ 返回选择角色"},
    "role_branch":    {"en": "🛒 Branch Staff",           "zh": "🛒 分店员工"},
    "role_warehouse": {"en": "📦 Warehouse Staff",        "zh": "📦 仓库员工"},
    "role_admin":     {"en": "🏭 Administrator",          "zh": "🏭 管理员"},
    "select_branch":  {"en": "Select Branch",            "zh": "选择分店"},
    "password":       {"en": "Password",                 "zh": "密码"},
    "login":          {"en": "Login",                    "zh": "登录"},
    "logout":         {"en": "Logout",                   "zh": "退出登录"},
    "wrong_pw":       {"en": "Wrong password",           "zh": "密码错误"},
    "admin_phone":    {"en": "Admin mobile number",      "zh": "管理员手机号"},
    "wrong_phone":    {"en": "Phone number not authorized",
                        "zh": "手机号未授权"},
    "admin_login_phone_hint": {
        "en": "Enter the mobile number configured for `SUNSHINE_ADMIN_PHONE` (server env).",
        "zh": "请输入服务器上 `SUNSHINE_ADMIN_PHONE` 环境变量中配置的管理员手机号。",
    },
    "nav_accounts":   {"en": "👤 Accounts & access",     "zh": "👤 账号与权限"},
    "nav_audit_log":  {"en": "📜 Audit log",            "zh": "📜 操作日志"},
    "audit_log_subtitle": {
        "en": "Login, orders, and receipt confirmations — filter by time, branch, or order.",
        "zh": "记录登录、下单、收货确认等操作，可按时间、分店、订单号筛选。",
    },
    "audit_filter_from": {"en": "From date",             "zh": "开始日期"},
    "audit_filter_to":   {"en": "To date",               "zh": "结束日期"},
    "audit_filter_events": {"en": "Event types",         "zh": "事件类型"},
    "audit_filter_branch": {"en": "Branch",              "zh": "分店"},
    "audit_filter_order": {"en": "Order ID contains",    "zh": "订单号包含"},
    "audit_filter_user": {"en": "Username contains",    "zh": "用户名包含"},
    "audit_export_csv": {"en": "Download CSV",         "zh": "导出 CSV"},
    "audit_no_rows":    {"en": "No matching records.",  "zh": "没有符合条件的记录。"},
    "audit_col_time":   {"en": "Time",                  "zh": "时间"},
    "audit_col_event":  {"en": "Event",                "zh": "事件"},
    "audit_col_role":   {"en": "Role",                 "zh": "角色"},
    "audit_col_account": {"en": "Account ID",          "zh": "账号ID"},
    "audit_col_user":   {"en": "Username",             "zh": "用户"},
    "audit_col_branch": {"en": "Branch",               "zh": "分店"},
    "audit_col_order":  {"en": "Order ID",             "zh": "订单号"},
    "audit_col_detail": {"en": "Detail",               "zh": "详情"},
    "audit_d_shared_login": {
        "en": "Shared password login ({role})",
        "zh": "共享密码登录（{role}）",
    },
    "audit_d_branch_login": {
        "en": "Branch account login · account ID {id}",
        "zh": "分店账号登录 · 账号 ID {id}",
    },
    "audit_d_logout_page": {
        "en": "Last screen before logout: {page}",
        "zh": "退出前所在页面：{page}",
    },
    "audit_d_line_count": {
        "en": "{n} item line(s)",
        "zh": "共 {n} 行商品明细",
    },
    "audit_d_has_short": {
        "en": "Receipt had shortage reported",
        "zh": "收货上报缺货",
    },
    "audit_d_no_short": {
        "en": "Receipt matched dispatch (no shortage)",
        "zh": "实收与发货一致（未报缺货）",
    },
    "audit_d_supplier_title": {
        "en": "Title: {title}",
        "zh": "标题：{title}",
    },
    "audit_d_supplier_lines": {
        "en": "{n} SKU line(s)",
        "zh": "订货 {n} 行",
    },
    "audit_ev_login":   {"en": "Login",                "zh": "登录"},
    "audit_ev_logout":  {"en": "Logout",               "zh": "登出"},
    "audit_ev_order":   {"en": "Place order",          "zh": "提交订单"},
    "audit_ev_receive": {"en": "Confirm receipt",      "zh": "确认收货"},
    "audit_ev_supplier": {"en": "Supplier order sent", "zh": "供货商下单已发送"},
    "audit_ev_catalog_reset": {"en": "Catalog wipe", "zh": "清空商品资料"},
    "audit_ev_catalog_save": {"en": "Shelf catalog save", "zh": "货架录入保存"},
    "audit_all_branches": {"en": "All branches",      "zh": "全部分店"},
    "acct_tab_login": {"en": "Account login",            "zh": "账号登录"},
    "acct_tab_apply": {"en": "Request an account",       "zh": "申请账号"},
    "acct_login_hint": {"en": "Log in with the username and password issued after admin approval.",
                        "zh": "请使用管理员审批通过后的用户名与密码登录。"},
    "acct_username":  {"en": "Username",                 "zh": "用户名"},
    "acct_display_name": {"en": "Display name",          "zh": "姓名/备注名"},
    "acct_phone":     {"en": "Phone (optional)",       "zh": "手机号（选填）"},
    "acct_password":  {"en": "Password",                 "zh": "密码"},
    "acct_password2": {"en": "Confirm password",         "zh": "确认密码"},
    "acct_apply_submit": {"en": "Submit application",   "zh": "提交申请"},
    "acct_apply_ok":  {"en": "Application received. An administrator must approve and grant access before you can use the system.",
                        "zh": "申请已提交。需由管理员审批并授权后，方可使用系统功能。"},
    "acct_apply_done_title": {"en": "Application submitted", "zh": "申请已提交"},
    "acct_back_branch_options": {"en": "Back to branch sign-in", "zh": "返回分店登录"},
    "acct_enter_store_title": {"en": "Enter your store", "zh": "进入门店"},
    "acct_enter_store_hint": {
        "en": "Select your store first. Login and account requests apply only to this store.",
        "zh": "请先选择您所在的门店。登录与申请账号仅针对当前门店，不可跨店使用。",
    },
    "acct_current_store": {"en": "Current store", "zh": "当前门店"},
    "acct_change_store": {"en": "Change store", "zh": "更换门店"},
    "acct_wrong_branch": {
        "en": "This account belongs to another store. Change store above or use the correct account.",
        "zh": "该账号不属于当前门店。请更换门店或使用对应门店的账号登录。",
    },
    "acct_apply_dup": {"en": "This username is already taken or you already have a pending/approved request.",
                        "zh": "该用户名已存在，或您已有待审/已通过的申请。"},
    "acct_user_invalid": {"en": "Use 3–32 characters: letters, numbers, underscore only.",
                          "zh": "用户名需 3–32 位，仅字母、数字、下划线。"},
    "acct_pw_short":  {"en": "Password must be at least 6 characters.",
                        "zh": "密码至少 6 位。"},
    "acct_pw_mismatch": {"en": "The two passwords do not match.",
                          "zh": "两次输入的密码不一致。"},
    "acct_branch_invalid": {"en": "Invalid branch.", "zh": "分店无效。"},
    "acct_pending_login": {"en": "This account is waiting for administrator approval.",
                          "zh": "该账号正在等待管理员审批。"},
    "acct_rejected_login": {"en": "This application was not approved. Contact your administrator.",
                            "zh": "该申请未通过，请联系管理员。"},
    "acct_no_access": {"en": "You do not have permission for this page.",
                       "zh": "您没有权限访问此页面。"},
    "acct_perm_need_admin": {"en": "No module permission assigned yet. Ask an administrator to grant access.",
                             "zh": "尚未分配任何功能权限，请联系管理员授权。"},
    "perm_order":     {"en": "Place orders",             "zh": "下单"},
    "perm_my_orders": {"en": "My orders",              "zh": "我的订单"},
    "perm_my_short":  {"en": "My shortages",           "zh": "我的缺货"},
    "perm_messages":  {"en": "Messages",               "zh": "消息中心"},
    "perm_ai":        {"en": "AI assistant",           "zh": "AI 助手"},
    "acct_pending_hdr": {"en": "Pending applications", "zh": "待审批申请"},
    "acct_approved_hdr": {"en": "Approved accounts", "zh": "已开通账号"},
    "acct_status":    {"en": "Status",                 "zh": "状态"},
    "acct_created":   {"en": "Submitted",              "zh": "提交时间"},
    "acct_actions":   {"en": "Actions",                "zh": "操作"},
    "acct_approve":   {"en": "Approve & set permissions", "zh": "批准并授权"},
    "acct_reject":    {"en": "Reject",                 "zh": "拒绝"},
    "acct_save_perms": {"en": "Save permissions",      "zh": "保存权限"},
    "acct_reset_pw":  {"en": "New password",           "zh": "新密码"},
    "acct_pw_apply":  {"en": "Update password",        "zh": "更新密码"},
    "acct_create_user": {"en": "Create account directly", "zh": "直接创建账号"},
    "acct_note_reject": {"en": "Reason (optional)",    "zh": "备注（选填）"},
    "acct_perms_pick":  {"en": "Allowed modules",      "zh": "可用功能模块"},
    "acct_updated":     {"en": "Saved.",               "zh": "已保存。"},
    "st_pending_acct": {"en": "Pending",               "zh": "待审批"},
    "st_approved_acct": {"en": "Approved",              "zh": "已通过"},
    "st_rejected_acct": {"en": "Rejected",              "zh": "已拒绝"},
    # Nav
    "nav_order":      {"en": "🛒 Place Order",           "zh": "🛒 下单"},
    "nav_stock":      {"en": "📅 Stock & Expiry",        "zh": "📅 库存/临期"},
    "stock_subtitle": {"en": "Branch batch inventory with expiry alerts",
                       "zh": "分店批次库存与临期/过期预警"},
    "stock_add_title": {"en": "Receive a batch (in-stock)", "zh": "批次入库"},
    "stock_pick_title": {"en": "Find from product library (or type manually below)",
                         "zh": "从商品库搜索（也可在下方手工输入）"},
    "stock_pick_ph":   {"en": "Search by name / barcode / item code",
                        "zh": "按名称 / 条码 / 编号搜索"},
    "stock_pick_select": {"en": "Select a product",       "zh": "选择商品"},
    "stock_pick_fill": {"en": "⬇️ Fill into form",         "zh": "⬇️ 填入下方表单"},
    "stock_pick_filled": {"en": "Filled. Adjust quantity & expiry below.",
                          "zh": "已填入，请在下方填写数量和过期日期。"},
    "stock_pick_clear": {"en": "🔄 Reset / type manually",
                         "zh": "🔄 重选 / 手工输入"},
    "stock_locked_hint": {"en": "Product filled from library (editable). "
                                "Enter cartons / pieces / expiry.",
                          "zh": "已从商品库带入（可微调）。请填写箱数 / 个数 / 过期日期。"},
    "stock_need_expire": {"en": "Please choose an expiry date.",
                          "zh": "请选择过期日期。"},
    "stock_name":     {"en": "Product name",             "zh": "商品名称"},
    "stock_code":     {"en": "Item code / barcode (optional)", "zh": "商品编号/条码（可选）"},
    "stock_unit":     {"en": "Unit (optional)",          "zh": "单位（可选）"},
    "stock_batch_no": {"en": "Batch no. (optional)",     "zh": "批次号（可选）"},
    "stock_qty_ct":   {"en": "Cartons",                  "zh": "箱数"},
    "stock_qty_pc":   {"en": "Pieces",                   "zh": "个数"},
    "stock_expire":   {"en": "Expiration date",          "zh": "过期日期"},
    "stock_add_btn":  {"en": "Add to branch stock",      "zh": "录入分店库存"},
    "stock_add_ok":   {"en": "Batch added.",             "zh": "批次已入库。"},
    "stock_list_title": {"en": "Current batches (this branch)", "zh": "本店当前批次"},
    "stock_warn_title": {"en": "Expiring / expired soon", "zh": "临期 / 已过期"},
    "stock_none":     {"en": "No batches yet.",          "zh": "暂无批次。"},
    "stock_expired":  {"en": "EXPIRED",                  "zh": "已过期"},
    "stock_days_left": {"en": "days left",               "zh": "天后过期"},
    "stock_need_qty": {"en": "Quantity must be greater than 0.",
                       "zh": "数量必须大于 0。"},
    "nav_expiry_dash": {"en": "📅 Expiry Dashboard",     "zh": "📅 临期看板"},
    "exp_dash_title": {"en": "Branch Stock & Expiry Dashboard",
                       "zh": "全局分店库存与临期统计看板"},
    "exp_dash_sub":   {"en": "Global expiry overview across all branches",
                       "zh": "全网各分店库存临期/过期统计总览"},
    "exp_filter_branch": {"en": "Branch filter",         "zh": "分店筛选"},
    "exp_filter_cat":  {"en": "Category filter",         "zh": "商品分类筛选"},
    "exp_kpi_branches": {"en": "Branches with stock",    "zh": "有库存分店数"},
    "exp_kpi_units":   {"en": "Total stock units",       "zh": "全网总库存量"},
    "exp_kpi_expired": {"en": "Expired items",           "zh": "已过期商品数"},
    "exp_kpi_expiring": {"en": "Expiring soon",          "zh": "即将过期商品数"},
    "exp_kpi_loss":    {"en": "Est. loss at risk",       "zh": "潜在损耗金额"},
    "exp_chart_title": {"en": "Expiring / expired by branch", "zh": "各分店临期/过期数量对比"},
    "exp_rank_title":  {"en": "Branch severity ranking",  "zh": "分店严重程度排行"},
    "exp_loss_title":  {"en": "Potential loss ranking",   "zh": "潜在资金损耗排行"},
    "exp_table_title": {"en": "Detail: batches across branches", "zh": "明细：各分店批次"},
    "exp_col_branch":  {"en": "Branch",                  "zh": "分店"},
    "exp_col_product": {"en": "Product",                 "zh": "商品名称"},
    "exp_col_stock":   {"en": "Remaining stock",         "zh": "剩余库存"},
    "exp_col_expire":  {"en": "Expire date",             "zh": "过期日期"},
    "exp_col_daysleft": {"en": "Days left",              "zh": "剩余天数"},
    "exp_col_manager": {"en": "Store manager",           "zh": "店长"},
    "exp_col_cat":     {"en": "Category",                "zh": "分类"},
    "exp_col_items":   {"en": "Items",                   "zh": "商品数"},
    "exp_col_ratio":   {"en": "At-risk ratio",           "zh": "临期占比"},
    "exp_only_risk":   {"en": "Show only expiring/expired",
                        "zh": "只看临期/已过期"},
    "exp_no_data":     {"en": "No batch data yet.",      "zh": "暂无批次数据。"},
    "nav_my_orders":  {"en": "📋 My Orders",             "zh": "📋 我的订单"},
    "nav_my_short":   {"en": "🔔 My Shortages",          "zh": "🔔 我的缺货"},
    "nav_pending":    {"en": "📦 Pending Dispatch",      "zh": "📦 待发货订单"},
    "nav_short_in":   {"en": "🔔 Shortage Notifications","zh": "🔔 缺货通知"},
    "nav_supplier_order": {"en": "🏭 Order from supplier", "zh": "🏭 向供货商下单"},
    "wh_supplier_subtitle": {
        "en": "Fill in what to order from the supplier. It will be sent to admin message center and configured email addresses.",
        "zh": "填写向供货商采购的内容。提交后将推送到管理员消息中心，并邮件通知在「邮件通知」中配置的收件人。",
    },
    "wh_supplier_subject": {"en": "Title / subject",     "zh": "标题/主题"},
    "wh_supplier_details": {
        "en": "Order lines (select products)",
        "zh": "订货明细（选择商品）",
    },
    "wh_supplier_pick": {"en": "Product", "zh": "选择商品"},
    "wh_supplier_add_line": {"en": "Add to order", "zh": "加入订货单"},
    "wh_supplier_search_hint": {
        "en": "Enter keywords to search the catalog.",
        "zh": "请输入关键词搜索商品。",
    },
    "wh_supplier_search_commit_hint": {
        "en": "Type a code or keyword, then click **Search** or press **Enter** in the field to load matches (Streamlit does not search while typing).",
        "zh": "输入编号或关键词后，请点击「搜索」，或在输入框内按 **Enter** 确认后再显示结果（仅输入不确认不会刷新列表）。",
    },
    "wh_supplier_cart": {"en": "Current order lines", "zh": "当前订货明细"},
    "wh_supplier_remove": {"en": "Remove", "zh": "移除"},
    "wh_supplier_send":   {"en": "Send",                 "zh": "发送"},
    "wh_supplier_need_lines": {"en": "Please add at least one product line.",
                                "zh": "请至少添加一行商品。"},
    "wh_supplier_default_subj": {"en": "Supplier restock",
                                  "zh": "供货商订货"},
    "wh_supplier_done_title": {"en": "Message sent",     "zh": "信息已发送"},
    "wh_supplier_done_msg": {
        "en": "Your supplier order request has been delivered to administrators and the configured email recipients.",
        "zh": "您的供货商下单信息已提交至管理员消息中心，并已按设置发送邮件。",
    },
    "wh_supplier_back": {"en": "New request",            "zh": "继续下单"},
    "wh_notif_inbox_title": {"en": "Supplier order",     "zh": "供货商下单"},
    "nav_dashboard":  {"en": "📊 Dashboard",              "zh": "📊 管理概览"},
    "nav_all_orders": {"en": "📋 All Orders",             "zh": "📋 所有订单"},
    "nav_dispatch":   {"en": "📦 Dispatch",               "zh": "📦 出库发货"},
    "nav_short_mgmt": {"en": "🔔 Shortages",              "zh": "🔔 缺货通知"},
    "nav_export":     {"en": "📥 Export Reports",         "zh": "📥 导出报表"},
    "nav_messages":   {"en": "🔔 Messages",               "zh": "🔔 消息中心"},
    "nav_arrivals":   {"en": "📦 Stock Arrivals",         "zh": "📦 进货通知"},
    "nav_inventory":  {"en": "📦 Inventory",              "zh": "📦 库存管理"},
    "nav_shelf_mobile": {"en": "📱 Shelf (mobile)",      "zh": "📱 货架录入（手机）"},
    "shelf_mobile_subtitle": {
        "en": "Enter products on the shelf: name, price, pieces per carton, and stock. "
              "For a full reset, use the danger zone on this page (clears database stock and the catalog file).",
        "zh": "在货架边用手机录入：商品名称、价格、每箱个数、库存等。需要彻底清空时，请使用本页「危险操作」"
              "（会删除数据库中的商品库存与价格覆盖，并重置商品表文件）。",
    },
    "shelf_pcs_per_carton": {"en": "Pcs / carton",        "zh": "每箱个数"},
    "shelf_wipe_title":   {"en": "Danger: clear all products & stock", "zh": "危险：清空全部商品与库存"},
    "shelf_wipe_blurb":   {
        "en": "Deletes `products.xlsx` contents, all inventory rows, price overrides, and inventory history. "
              "Does not delete past orders. Type the confirmation text to enable the button.",
        "zh": "将清空 `products.xlsx` 中的商品、数据库内全部库存、价格覆盖与库存流水。不会删除历史订单。输入确认文字后才会启用按钮。",
    },
    "shelf_wipe_confirm": {"en": "Type: CLEAR",            "zh": "输入确认：清空"},
    "shelf_wipe_ok":     {"en": "Wipe complete",         "zh": "已清空"},
    "shelf_save_ok":     {"en": "Catalog saved",         "zh": "商品资料已保存"},
    "shelf_need_name":   {"en": "Each row must have a product name.", "zh": "每一行必须填写商品名称。"},
    "shelf_save_btn":    {"en": "Save to catalog & sync stock", "zh": "保存到商品表并同步库存"},
    "shelf_editor_hint": {
        "en": "Add rows as you walk the aisle. Barcode and code are optional if unknown.",
        "zh": "沿货架逐行添加；不清楚条码或编号可先留空。",
    },
    # Statuses
    "st_pending":     {"en": "Pending",                  "zh": "待发货"},
    "st_dispatched":  {"en": "Dispatched",               "zh": "已发货"},
    "st_received":    {"en": "Received",                 "zh": "已收货"},
    "st_open":        {"en": "Open",                     "zh": "待处理"},
    "st_resending":   {"en": "Resending",                "zh": "补发中"},
    "st_oos":         {"en": "Out of Stock",             "zh": "缺货"},
    "st_resolved":    {"en": "Resolved",                 "zh": "已解决"},
    # Misc
    "current_branch": {"en": "Current branch",           "zh": "当前分店"},
    "search_product": {"en": "Search by name / barcode / code (scanner OK)",
                       "zh": "按名称/条码/编号搜索（支持扫码枪）"},
    "no_results":     {"en": "No products found",        "zh": "未找到商品"},
    "manual_add":     {"en": "+ Add product not in list","zh": "+ 添加列表外商品"},
    "manual":         {"en": "Manual",                   "zh": "手动"},
    "name":           {"en": "Name",                     "zh": "名称"},
    "unit":           {"en": "Unit",                     "zh": "单位"},
    "item_code":      {"en": "Item Code",                "zh": "商品编号"},
    "barcode":        {"en": "Barcode",                  "zh": "条码"},
    "cartons":        {"en": "Cartons",                  "zh": "箱数"},
    "each_pcs":       {"en": "Each pcs",                 "zh": "个数"},
    "order_qty_heading": {"en": "Order quantity",       "zh": "订货数量"},
    "qty_cartons":    {"en": "Cartons",                  "zh": "箱"},
    "qty_pcs":        {"en": "Pcs",                      "zh": "个"},
    "add_to_cart":    {"en": "Add to Cart",              "zh": "加入购物车"},
    "cart":           {"en": "Cart",                     "zh": "购物车"},
    "empty_cart":     {"en": "Cart is empty",            "zh": "购物车为空"},
    "branch_cart_sync_tip":
        {"en": "Your draft cart is saved to this account and comes back "
               "after a refresh or switching away from the browser.",
               "zh": "购物车草稿保存在本账号，刷新页面或从其他应用返回浏览器后会自动恢复。"},
    "submit_order":   {"en": "Submit Order",             "zh": "提交订单"},
    "order_submitted":{"en": "Order submitted",          "zh": "订单已提交"},
    "order_sent_done":{"en": "Order has been sent successfully",
                       "zh": "订单已发送完毕"},
    "order_sent_tip": {"en": "Warehouse will process your order soon.",
                       "zh": "仓库将尽快处理你的订单。"},
    "create_new_order":{"en": "Create another order",    "zh": "继续下新订单"},
    "view_my_orders_btn":{"en": "View my orders",        "zh": "查看我的订单"},
    "submit_busy":    {"en": "Submission in progress, please wait",
                       "zh": "正在提交，请稍候"},
    "dup_submit_block":{"en": "Duplicate submission blocked. Please refresh and check.",
                        "zh": "已拦截重复提交，请刷新后确认结果"},
    "order_id":       {"en": "Order ID",                 "zh": "订单号"},
    "order_date":     {"en": "Order Date",               "zh": "下单日期"},
    "remarks":        {"en": "Remarks",                  "zh": "备注"},
    "search":         {"en": "Search",                   "zh": "搜索"},
    "search_typing_hint": {"en": "Typing... searching shortly",
                           "zh": "输入中... 即将搜索"},
    "status":         {"en": "Status",                   "zh": "状态"},
    "branch":         {"en": "Branch",                   "zh": "分店"},
    "recent_login":   {"en": "Recent login",             "zh": "最近登录"},
    "use_recent_role":{"en": "Use recent role",          "zh": "使用最近角色"},
    "use_recent_branch":{"en": "Use recent branch",      "zh": "使用最近分店"},
    "filter_order_keyword":{"en": "Filter by order/item/barcode",
                            "zh": "按订单号/商品名/条码筛选"},
    "filter_branches":{"en": "Filter branches",          "zh": "筛选分店"},
    "filter_date_from":{"en": "Pending date from",       "zh": "待发货起始日期"},
    "filter_date_to":  {"en": "Pending date to",         "zh": "待发货结束日期"},
    "expand_match_only":{"en": "Expand matched orders only",
                         "zh": "仅展开命中订单"},
    "expand_all_orders":{"en": "Expand all orders",       "zh": "展开全部订单"},
    "collapse_all_orders":{"en": "Collapse all orders",   "zh": "折叠全部订单"},
    "recent_7_days":   {"en": "Recent 7 days",            "zh": "最近7天"},
    "all_dates":       {"en": "All dates",                "zh": "全部日期"},
    "all_branches":    {"en": "All branches",             "zh": "全部分店"},
    "filter_summary":  {"en": "Filter summary",           "zh": "筛选摘要"},
    "keyword_label":   {"en": "Keyword",                  "zh": "关键词"},
    "nav_ai":          {"en": "🤖 AI Assistant",          "zh": "🤖 AI助手"},
    "ai_title":        {"en": "AI Assistant",             "zh": "AI助手"},
    "ai_disclaimer":   {"en": "AI suggestions are for reference only. Confirm key actions against system data.",
                        "zh": "AI建议仅供参考，关键操作请以系统数据为准。"},
    "ai_missing_key":  {"en": "Gemini is not configured. Set GEMINI_API_KEY (env) or create gemini.env next to app.py.",
                        "zh": "未配置Gemini。请设置环境变量 GEMINI_API_KEY，或在 app.py 同目录创建 gemini.env。"},
    "ai_question":     {"en": "Ask about operation/process/issues",
                        "zh": "请输入你想咨询的操作/流程/问题"},
    "ai_ask_btn":      {"en": "Ask AI",                   "zh": "提问AI"},
    "ai_with_ctx":     {"en": "Include current system context",
                        "zh": "携带当前系统上下文"},
    "ai_clear_chat":   {"en": "Clear chat",               "zh": "清空对话"},
    "ai_quick_q":      {"en": "Quick questions",          "zh": "快捷问题"},
    "ai_q_shortage":   {"en": "How should I handle shortage workflow?",
                        "zh": "缺货流程该怎么处理？"},
    "ai_q_not_found":  {"en": "Why can't I find an order/product?",
                        "zh": "为什么找不到订单/商品？"},
    "ai_q_dispatch":   {"en": "How to dispatch orders faster and safely?",
                        "zh": "如何更快且安全地发货？"},
    "ai_rate_limit":   {"en": "Please wait a moment before next AI request.",
                        "zh": "请稍等片刻再发起下一次AI请求。"},
    "ai_working":      {"en": "AI is answering, please wait…",
                        "zh": "AI 正在回答，请稍候…"},
    "ai_empty_q":      {"en": "Question is empty.",       "zh": "问题不能为空。"},
    "ai_error":        {"en": "AI service unavailable, please try again later.",
                        "zh": "AI服务暂不可用，请稍后重试。"},
    "ai_quota_429":    {"en": "Gemini quota / rate limit (HTTP 429). Your key may have free-tier limit 0 for this model, or daily quota is exhausted.",
                        "zh": "Gemini 配额或频率限制（HTTP 429）。常见原因：免费层对该模型限额为 0，或当日/每分钟配额已用完。"},
    "ai_quota_links":  {"en": "Docs: https://ai.google.dev/gemini-api/docs/rate-limits · Usage: https://ai.dev/rate-limit",
                        "zh": "说明文档: https://ai.google.dev/gemini-api/docs/rate-limits · 用量: https://ai.dev/rate-limit"},
    "ai_quota_actions": {"en": "Actions: wait a few minutes and retry; enable billing in Google AI Studio; or set GEMINI_MODEL / GEMINI_FALLBACK_MODELS to a model your plan supports.",
                        "zh": "处理建议：等待几分钟后重试；在 Google AI Studio 为项目开通计费/提高配额；或在环境变量中设置 GEMINI_MODEL、GEMINI_FALLBACK_MODELS 为你账号可用的模型。"},
    "ai_key_leaked_403": {
        "en": "This Gemini API key was rejected (HTTP 403): Google reports it as leaked or revoked. It cannot be used anymore.",
        "zh": "当前 Gemini API 密钥已被拒绝（HTTP 403）：Google 判定该密钥已泄露或已被停用，无法继续使用。",
    },
    "ai_key_leaked_actions": {
        "en": "Create a **new** API key in Google AI Studio → delete/disable the old key → set `GEMINI_API_KEY` (environment) or edit `gemini.env` next to `app.py`, then restart the app. Never commit API keys to Git or share them in chat/screenshots.",
        "zh": "请在 [Google AI Studio](https://aistudio.google.com/apikey) **重新创建**密钥，并在控制台**删除/停用**旧密钥；在本机设置环境变量 `GEMINI_API_KEY`，或修改与 `app.py` 同目录的 `gemini.env`，然后**重启** Streamlit。切勿把密钥提交到 Git、聊天或截图外传。",
    },
    "ai_key_invalid_400": {
        "en": "The Gemini API key is not accepted (HTTP 400: API_KEY_INVALID). It may be wrong, copied with extra spaces, disabled, or for a different Google project.",
        "zh": "当前 Gemini API 密钥无效（HTTP 400）：可能被删除/复制错误、前后有空格、与 Google 项目不匹配，或环境变量里仍是旧值。",
    },
    "ai_key_invalid_actions": {
        "en": "1) In [AI Studio](https://aistudio.google.com/apikey) create a new key and enable **Generative Language API** for the project. 2) Set `GEMINI_API_KEY` in `gemini.env` (next to `app.py`) on **one line** with no spaces around `=`, or set the Windows user env var. 3) If both exist, **Windows env wins** — update or remove the old one. 4) **Fully restart** the Streamlit process (not just refresh the browser).",
        "zh": "1）打开 [Google AI Studio](https://aistudio.google.com/apikey) 新建 API 密钥，并确认已对该 Google 项目启用 **Generative Language API**。2）在 `app.py` 同目录的 `gemini.env` 中写 `GEMINI_API_KEY=新密钥`（等号两侧不要多余空格、不要加引号），或只改 Windows 用户环境变量。3）若同时存在**系统/用户环境变量**与 `gemini.env`，会**优先用环境变量** —— 请同步更新或删除旧的环境变量。4）改完后**完全退出并重新启动** Streamlit 进程，不要只刷新网页。",
    },
    "ai_model_used":   {"en": "(Answered with backup model: {model})",
                        "zh": "（已使用备用模型回答：{model}）"},
    "ai_ctx_role":     {"en": "Current role",             "zh": "当前角色"},
    "ai_ctx_branch":   {"en": "Current branch",           "zh": "当前分店"},
    "ai_ctx_open_short": {"en": "Open shortages",         "zh": "待处理缺货"},
    "ai_ctx_pending_orders": {"en": "Pending orders",     "zh": "待发货订单数"},
    "ai_ctx_latest_orders": {"en": "Latest orders summary","zh": "最新订单摘要"},
    "ai_answer":       {"en": "AI answer",                "zh": "AI回答"},
    "ai_role_only":    {"en": "Answer as a system operations assistant. Do not execute actions.",
                        "zh": "请作为系统操作助手回答，不执行任何实际操作。"},
    "no_data":        {"en": "No data",                  "zh": "暂无数据"},
    "no_pending":     {"en": "No pending orders",        "zh": "暂无待发货订单"},
    "no_shortages":   {"en": "No shortages",             "zh": "暂无缺货"},
    "receive":        {"en": "Receive Goods",            "zh": "收货确认"},
    "actual_cartons": {"en": "Actual Cartons",           "zh": "实收箱数"},
    "actual_pcs":     {"en": "Actual Pcs",               "zh": "实收个数"},
    "ordered":        {"en": "Ordered",                  "zh": "下单"},
    "dispatched":     {"en": "Dispatched",               "zh": "发货"},
    "confirm_receive":{"en": "Confirm Receipt",          "zh": "确认收货"},
    "receipt_done":   {"en": "Receipt confirmed",        "zh": "收货已确认"},
    "shortage_alert": {"en": "Shortage detected, warehouse notified",
                       "zh": "检测到缺货，已通知仓库"},
    "msg_center_title": {"en": "Message Center",          "zh": "消息中心"},
    "msg_empty":      {"en": "No messages yet",           "zh": "暂无消息"},
    "msg_mark_read":  {"en": "Mark as read",              "zh": "标记已读"},
    "msg_mark_all":   {"en": "Mark all as read",          "zh": "全部标记已读"},
    "msg_related_order": {"en": "Related order",          "zh": "关联订单"},
    "arrival_title":  {"en": "Arrival title",             "zh": "到货标题"},
    "arrival_notice": {"en": "Notice",                    "zh": "通知说明"},
    "arrival_items":  {"en": "Arrival items (one per line)", "zh": "到货商品清单（每行一个）"},
    "arrival_publish": {"en": "Publish arrival notice",   "zh": "发布到货通知"},
    "arrival_published": {"en": "Arrival notice published", "zh": "到货通知已发布"},
    "arrival_current": {"en": "Current active arrival notice", "zh": "当前生效到货通知"},
    "arrival_none":   {"en": "No active arrival notice",  "zh": "当前暂无生效到货通知"},
    "arrival_priority_tip": {"en": "New stock is available. Prioritize ordering these items.",
                             "zh": "有新货到达，请优先下单以下商品。"},
    "arrival_publish_hint_dialog": {
        "en": "Click the button below to open a window, search and tick products, then confirm — lines are appended to the list.",
        "zh": "点击下方按钮，在弹出窗口中搜索并勾选商品，确认后将自动写入下方「到货商品清单」。",
    },
    "arrival_pick_open_btn": {
        "en": "Pick products (popup)",
        "zh": "选择商品并追加到清单",
    },
    "arrival_pick_dialog_hint": {
        "en": "Search or scroll the list, tick products, then confirm.",
        "zh": "在窗口内搜索或浏览列表，勾选商品后点「确认追加」。",
    },
    "arrival_pick_dialog_select": {
        "en": "Products to add",
        "zh": "勾选要加入清单的商品",
    },
    "arrival_pick_confirm": {"en": "Confirm add", "zh": "确认追加"},
    "arrival_pick_cancel": {"en": "Close", "zh": "关闭"},
    "arrival_pick_appended": {"en": "Added {n} product(s) to the list.",
                               "zh": "已向清单追加 {n} 个商品。"},
    "arrival_pick_nothing_new": {
        "en": "No new lines (already in the list).",
        "zh": "没有新增行（可能已在清单中）。",
    },
    "arrival_pick_empty_sel": {
        "en": "Please tick at least one product.",
        "zh": "请先勾选至少一个商品。",
    },
    "inv_title":      {"en": "Inventory Management",      "zh": "库存管理"},
    "inv_import_append": {"en": "Import stock change from products.xlsx (+/- on current stock)",
                          "zh": "从 products.xlsx 导入库存变动（正数增加，负数减少）"},
    "inv_import_append_hint": {
        "en": "Only include SKUs that changed this time. Positive numbers add stock, negative numbers deduct stock (e.g. 1 / -1). Every row writes an inventory log; SKUs not listed stay unchanged.",
        "zh": "表中只填写本次有变动的 SKU：正数增加库存，负数减少库存（例如 1 / -1）。每一行都会写入库存流水日志；未填写的 SKU 保持不变。",
    },
    "inv_import_overwrite": {"en": "Import from products.xlsx — set inventory to spreadsheet values",
                             "zh": "从 products.xlsx 导入 — 按表内数量覆盖库存"},
    "inv_import_overwrite_hint": {
        "en": "**Caution**: set stock to carton/piece values from the sheet; SKUs not listed are left unchanged.",
        "zh": "**谨慎**：把库存设置为表中的箱数/个数（表中未出现的 SKU 不受影响）。用于盘点或纠错。",
    },
    "inv_import_overwrite_section": {
        "en": "Advanced — set stock to sheet values",
        "zh": "高级选项 — 按表格覆盖库存数量",
    },
    "inv_import_done": {"en": "Inventory import complete: {n} items updated",
                        "zh": "库存导入完成：更新 {n} 个商品"},
    "nav_price": {"en": "Price Management", "zh": "价格管理"},
    "price": {"en": "Price", "zh": "价格"},
    "updated_at": {"en": "Updated At", "zh": "修改日期"},
    "price_import_from_products": {
        "en": "Import prices from selected file",
        "zh": "导入选中文件中的价格",
    },
    "price_import_done": {
        "en": "Price import complete: {n} items updated",
        "zh": "价格导入完成：更新 {n} 个商品",
    },
    "price_apply": {"en": "Apply price", "zh": "应用价格"},
    "price_updated": {"en": "Price updated", "zh": "价格已更新"},
    "price_hint": {
        "en": "Upload an Excel file with columns ItemCode / Barcode / Name and Price (or 价格 / 单价). You can also edit below.",
        "zh": "请上传 Excel，需包含商品编号、条码、名称及价格列（支持列名 Price、价格、单价等）。也可在下方直接改价。",
    },
    "price_import_no_numeric_prices": {
        "en": "Parsed prices are all zero — check that the sheet has a Price / 价格 / 单价 column with numeric values.",
        "zh": "解析到的价格全部为 0：请确认表中有「价格」「单价」或 Price 列，且单元格为数字（勿用文本格式隐藏）。",
    },
    "select_import_file": {"en": "Select import file", "zh": "选择导入文件"},
    "download_inv_template": {
        "en": "📥 Download inventory import template (.xlsx)",
        "zh": "📥 下载库存导入模板（.xlsx）",
    },
    "download_price_template": {
        "en": "📥 Download price import template (.xlsx)",
        "zh": "📥 下载价格导入模板（.xlsx）",
    },
    "grid_keyword_search": {
        "en": "Keyword / code / barcode",
        "zh": "关键字 / 商品编号 / 条码",
    },
    "grid_search_apply": {"en": "Search", "zh": "筛选"},
    "grid_price_enter_keyword": {
        "en": "Enter a keyword or code and click **Search** to list products.",
        "zh": "请输入编号、条码或名称关键字，点击 **筛选** 显示商品列表。",
    },
    "grid_price_click_edit_hint": {
        "en": "Click the **Current price** cell to edit (web apps cannot use double‑click on rows).",
        "zh": "在「当前售价」列中**单击单元格**即可改价（网页端无法像桌面软件一样双击整行编辑）。",
    },
    "price_sheet_col": {
        "en": "List price (sheet)",
        "zh": "档案价（表格）",
    },
    "price_current_col": {
        "en": "Current price",
        "zh": "当前售价",
    },
    "grid_price_save": {"en": "Save price changes", "zh": "保存价格修改"},
    "grid_price_no_changes": {
        "en": "No price changes to save.",
        "zh": "没有检测到价格变动。",
    },
    "excel_batch_import": {"en": "📥 Excel batch import", "zh": "📥 Excel 批量导入"},
    "grid_inv_stock_ct": {"en": "Stock (cartons)", "zh": "库存箱数"},
    "grid_inv_stock_pc": {"en": "Stock (pcs)", "zh": "库存个数"},
    "grid_inv_delta_ct": {"en": "Δ cartons", "zh": "变动箱数"},
    "grid_inv_delta_pc": {"en": "Δ pcs", "zh": "变动个数"},
    "grid_inv_apply_rows": {
        "en": "Apply stock changes",
        "zh": "应用库存变动",
    },
    "grid_inv_no_delta": {
        "en": "No row has non‑zero stock change.",
        "zh": "没有填写非零的库存变动。",
    },
    "grid_inv_delta_hint": {
        "en": "Enter ± cartons / pcs in **Δ** columns, then click apply.",
        "zh": "在「变动箱数」「变动个数」列填写增减数量（可为负数），再点击「应用库存变动」。",
    },
    "price_full_list_expander": {
        "en": "📋 Full price list (optional)",
        "zh": "📋 查看全部商品价格（可选）",
    },
    "import_file_required": {"en": "Please choose an Excel file first.", "zh": "请先选择 Excel 文件。"},
    "import_file_invalid": {"en": "Unable to parse selected file. Check headers/format.", "zh": "无法解析所选文件，请检查表头和格式。"},
    "import_finished_toast": {
        "en": "Import finished — {n} row(s) written.",
        "zh": "导入完成：有 {n} 条已写入数据库（见下方「已更新」）。",
    },
    "import_finished_with_errors": {
        "en": "Import finished with some failed rows — see details below.",
        "zh": "导入已处理，但有部分行失败或未写入，请查看下方「错误明细」与失败条数。",
    },
    "import_finished_nothing_changed": {
        "en": "No rows were written: every row already matched the system, or had no effective change. This is normal if you re-import the same file.",
        "zh": "本次「已更新」为 0 是因为没有需要改写的数据：这些行与系统里当前价格/库存一致（例如售价与价格覆盖表相同），重新导入同一份表时常会出现。**若要改价，请先在 Excel 里改数字再导入。**",
    },
    "import_finished_no_effect_rows": {
        "en": "No countable rows were processed (all blank or skipped). Check the worksheet.",
        "zh": "未发现需要统计的有效数据行（可能全部为空或未匹配）。请检查工作表内容与列名。",
    },
    "import_metric_written": {
        "en": "Updated (written)",
        "zh": "已更新（写入）",
    },
    "import_metric_skipped": {
        "en": "Skipped (no change)",
        "zh": "跳过（无需变更）",
    },
    "import_metric_failed": {
        "en": "Failed",
        "zh": "失败",
    },
    "import_summary_skipped_hint": {
        "en": "Skipped rows are listed in the middle metric (not failures).",
        "zh": "中间数字为「与系统一致故未改写」的行数，不是错误。",
    },
    "import_error_detail": {
        "en": "Problem rows",
        "zh": "无法导入或有问题的明细",
    },
    "inv_adjust":     {"en": "Stock Adjustment",           "zh": "库存调整"},
    "inv_change_ct":  {"en": "Change cartons (+/-)",      "zh": "变动箱数（可正负）"},
    "inv_change_pc":  {"en": "Change pcs (+/-)",          "zh": "变动个数（可正负）"},
    "inv_apply":      {"en": "Apply adjustment",          "zh": "应用调整"},
    "inv_change_zero": {"en": "Change cannot be zero.", "zh": "变动不能为 0"},
    "inv_adjust_ok":   {"en": "Inventory updated.",      "zh": "库存已调整"},
    "inv_pick_product": {"en": "Select product",         "zh": "选择商品"},
    "inv_current":    {"en": "Current inventory",         "zh": "当前库存"},
    "admin_inventory_hub": {"en": "Inventory Workspace",  "zh": "库存工作台"},
    "admin_messages_hub":  {"en": "Message Workspace",    "zh": "消息工作台"},
    "choose_module":       {"en": "Choose module",        "zh": "选择功能模块"},
    "inv_low_stock":  {"en": "Insufficient inventory for dispatch",
                       "zh": "库存不足，无法完成发货"},
    "mark_dispatched":{"en": "Mark as Dispatched",       "zh": "标记已发货"},
    "shipment_marked":{"en": "Marked as dispatched",     "zh": "已标记发货"},
    "dispatch_busy":  {"en": "Dispatch is processing, please wait",
                       "zh": "正在处理发货，请稍候"},
    "dispatch_fill_ordered": {"en": "Fill with ordered qty",
                              "zh": "按下单数量填充"},
    "dispatch_clear_all": {"en": "Clear all dispatch qty",
                           "zh": "清空全部发货数"},
    "dispatch_changed_row": {"en": "Adjusted from ordered qty",
                             "zh": "已偏离下单数量"},
    "shortage_qty":   {"en": "Shortage Qty",             "zh": "缺货数量"},
    "warehouse_reply":{"en": "Warehouse Reply",          "zh": "仓库回复"},
    "resend":         {"en": "Resend",                   "zh": "补发"},
    "mark_oos":       {"en": "Mark Out of Stock",        "zh": "标记缺货"},
    "confirm_resend": {"en": "Confirm Resend Received",  "zh": "确认补发已收到"},
    "card_today":     {"en": "Orders Today",             "zh": "今日订单"},
    "card_pending":   {"en": "Pending Dispatch",         "zh": "待发货"},
    "card_dispatched":{"en": "Dispatched",               "zh": "已发货"},
    "card_short":     {"en": "Open Shortages",           "zh": "待处理缺货"},
    "branch_status":  {"en": "Branch Status",            "zh": "各分店状态"},
    "latest_orders":  {"en": "Latest Orders",            "zh": "最新订单"},
    "latest_short":   {"en": "Latest Shortages",         "zh": "最新缺货"},
    "from_date":      {"en": "From",                     "zh": "起始日期"},
    "to_date":        {"en": "To",                       "zh": "结束日期"},
    "exp_picking":    {"en": "Picking List (per branch)","zh": "捡货单（各店分列）"},
    "exp_recon":      {"en": "Reconciliation (Dispatched vs Received)",
                       "zh": "对账单（发货 vs 实收）"},
    "exp_short":      {"en": "Shortage Report",          "zh": "缺货报告"},
    "exp_inventory":  {"en": "Inventory Report (All Stock)", "zh": "库存报表（全部库存）"},
    "generate":       {"en": "Generate",                 "zh": "生成"},
    "download":       {"en": "Download",                 "zh": "下载"},
    "tbd":            {"en": "(to be implemented)",      "zh": "（待实现）"},
    # Pagination
    "page":           {"en": "Page",                     "zh": "页"},
    "of":             {"en": "of",                       "zh": "/"},
    "prev_page":      {"en": "◀ Previous",               "zh": "◀ 上一页"},
    "next_page":      {"en": "Next ▶",                   "zh": "下一页 ▶"},
    "showing":        {"en": "Showing",                  "zh": "显示"},
    "items":          {"en": "items",                    "zh": "条"},
    "results_count":  {"en": "results",                  "zh": "结果"},
    "per_page":       {"en": "per page",                 "zh": "每页"},
    # Dispatch history
    "nav_dispatch_history":  {"en": "📜 Dispatch History",
                              "zh": "📜 出库历史"},
    "dispatch_history":      {"en": "Dispatch History",  "zh": "出库历史"},
    "select_date":           {"en": "Select Date",       "zh": "选择日期"},
    "today_btn":             {"en": "Today",             "zh": "今天"},
    "all_dates":             {"en": "All Dates",         "zh": "所有日期"},
    "dispatch_date":         {"en": "Dispatch Date",     "zh": "出库日期"},
    "dispatch_time":         {"en": "Dispatch Time",     "zh": "出库时间"},
    "no_dispatch_history":   {"en": "No dispatch history",
                              "zh": "暂无出库记录"},
    "dispatched_orders":     {"en": "Dispatched Orders", "zh": "已发货订单"},
    "received_status":       {"en": "Receipt status",    "zh": "收货状态"},
    "fully_received":        {"en": "Fully received",    "zh": "已全部收货"},
    "partial_received":      {"en": "Partially received","zh": "部分收货"},
    "awaiting_receipt":      {"en": "Awaiting receipt",  "zh": "待收货"},
    # Picking-list downloads in warehouse view
    "dl_picking":     {"en": "🖨️ Picking List",          "zh": "🖨️ 捡货单"},
    "dl_this_order":  {"en": "🖨️ Download Picking Slip", "zh": "🖨️ 下载捡货单"},
    "dl_branch_all":  {"en": "🖨️ Download All for Branch", "zh": "🖨️ 下载本店全部"},
    "dl_all_pending": {"en": "🖨️ Download All Pending",  "zh": "🖨️ 下载全部待发货"},
    "picking_slip":   {"en": "Picking Slip",             "zh": "捡货单"},
    "lines":          {"en": "Lines",                    "zh": "行数"},
    "total_items":    {"en": "Total Lines",              "zh": "合计行数"},
    "orders":         {"en": "orders",                   "zh": "订单"},
    # Batch add / confirmation flow
    "add_selected":      {"en": "📥 Add Selected to Cart",
                          "zh": "📥 追加购物车（已选商品）"},
    "added_n_items":     {"en": "Added {n} items to cart",
                          "zh": "已追加 {n} 项到购物车"},
    "no_qty_selected":   {"en": "No items have a quantity. Enter cartons or pcs first.",
                          "zh": "没有商品填写数量。请先输入箱数或个数。"},
    "review_cart":       {"en": "📝 Review & Send Order",
                          "zh": "📝 确认并发送订单"},
    "cart_checkout":     {"en": "🧾 Checkout", "zh": "🧾 去结算"},
    "prod_date_opt":     {"en": "Production date (optional)",
                          "zh": "生产日期（可选）"},
    "review_title":      {"en": "Review Your Order",
                          "zh": "确认订单"},
    "review_subtitle":   {"en": "Check each line, edit quantities or remove items, then send.",
                          "zh": "请核对每行数量，可修改或删除，然后发送。"},
    "send_order":        {"en": "📤 Send Order",
                          "zh": "📤 发送订单"},
    "back_to_browse":    {"en": "◀ Back to browsing",
                          "zh": "◀ 返回继续选购"},
    "items_count":       {"en": "Items in this order",
                          "zh": "本次订单商品数"},
    "clear_qty":         {"en": "Clear all entered quantities",
                          "zh": "清空已输入数量"},
    "qty_cleared":       {"en": "Quantities cleared",
                          "zh": "数量已清空"},
    "saved_for_batch":   {"en": "Saved",
                          "zh": "已保存"},
    # Warehouse dispatch quantity input
    "dispatch_qty":      {"en": "Dispatch Qty",
                          "zh": "发货数量"},
    "dispatch_cartons":  {"en": "Dispatch Cartons",
                          "zh": "发货箱数"},
    "dispatch_pcs":      {"en": "Dispatch Pcs",
                          "zh": "发货个数"},
    "dispatch_hint":     {"en": "Enter actual quantities being shipped. "
                                "Defaults to ordered amount.",
                          "zh": "请输入实际发货数量，默认等于订单数量。"},
    "less_than_ordered": {"en": "Short of order",
                          "zh": "少于订单"},
    "ordered_qty":       {"en": "Ordered",
                          "zh": "订货"},
    # Database backup / restore
    "nav_backup":        {"en": "💾 Database Backup",
                          "zh": "💾 数据库备份"},
    "backup_title":      {"en": "Database Backup & Restore",
                          "zh": "数据库备份与恢复"},
    "backup_subtitle":   {"en": "Snapshots are stored locally in the "
                                "backups/ folder. Auto-snapshot runs on "
                                "startup and at most once per hour.",
                          "zh": "备份文件保存在 backups/ 目录。"
                                "启动时和每小时最多一次会自动备份。"},
    "backup_now":        {"en": "💾 Backup Now",
                          "zh": "💾 立即备份"},
    "backup_done":       {"en": "Backup created",
                          "zh": "备份已创建"},
    "backup_failed":     {"en": "Backup failed",
                          "zh": "备份失败"},
    "available_backups": {"en": "Available Backups",
                          "zh": "可用备份"},
    "no_backups":        {"en": "No backups yet",
                          "zh": "尚无备份"},
    "backup_file":       {"en": "File",
                          "zh": "文件"},
    "backup_size":       {"en": "Size (KB)",
                          "zh": "大小 (KB)"},
    "backup_created":    {"en": "Created",
                          "zh": "创建时间"},
    "backup_counts":     {"en": "Records (orders / shipments / shortages)",
                          "zh": "记录数 (订单 / 发货 / 缺货)"},
    "backup_actions":    {"en": "Actions",
                          "zh": "操作"},
    "download_backup":   {"en": "⬇️ Download",
                          "zh": "⬇️ 下载"},
    "restore_backup":    {"en": "♻️ Restore",
                          "zh": "♻️ 恢复"},
    "delete_backup":     {"en": "🗑️ Delete",
                          "zh": "🗑️ 删除"},
    "confirm_restore_q": {"en": "⚠️ Restore from this backup? "
                                "Current data will be replaced. "
                                "A safety snapshot of current data will "
                                "be created first.",
                          "zh": "⚠️ 确认从此备份恢复？"
                                "当前数据将被覆盖。"
                                "恢复前会自动备份当前数据库。"},
    "yes_restore":       {"en": "✅ Yes, restore now",
                          "zh": "✅ 是，立即恢复"},
    "cancel":            {"en": "Cancel",
                          "zh": "取消"},
    "restore_done":      {"en": "Restore complete. Safety snapshot saved as: ",
                          "zh": "恢复完成。安全备份保存为: "},
    "export_sql":        {"en": "📤 Export as SQL Dump",
                          "zh": "📤 导出 SQL 文件"},
    "import_sql":        {"en": "📥 Restore from SQL Dump",
                          "zh": "📥 从 SQL 文件恢复"},
    "upload_sql":        {"en": "Upload .sql file",
                          "zh": "上传 .sql 文件"},
    "import_done":       {"en": "Import complete. Safety snapshot: ",
                          "zh": "导入完成。安全备份: "},
    "current_db_info":   {"en": "Current Database",
                          "zh": "当前数据库"},
    # Product images
    "nav_images":        {"en": "🖼️ Product Images",
                          "zh": "🖼️ 商品图片"},
    "img_title":         {"en": "Product Image Management",
                          "zh": "商品图片管理"},
    "img_subtitle":      {"en": "Upload a photo for each product. "
                                "On phones, the upload button opens the camera.",
                          "zh": "为每个商品上传图片。"
                                "在手机上点击上传按钮会自动调用相机。"},
    "img_search":        {"en": "Find product to upload image",
                          "zh": "搜索要上传图片的商品"},
    "img_upload":        {"en": "📷 Take photo / Choose image",
                          "zh": "📷 拍照 / 选择图片"},
    "img_uploaded":      {"en": "Image saved",
                          "zh": "图片已保存"},
    "img_delete":        {"en": "🗑️ Remove image",
                          "zh": "🗑️ 删除图片"},
    "img_deleted":       {"en": "Image removed",
                          "zh": "图片已删除"},
    "img_has_image":     {"en": "Has image",
                          "zh": "已有图片"},
    "img_no_image":      {"en": "No image yet",
                          "zh": "暂无图片"},
    "img_show_in_search":{"en": "Show product images in search results",
                          "zh": "在搜索结果中显示商品图片"},
    # Batch image upload
    "img_batch_title":   {"en": "🚀 Batch upload by filename",
                          "zh": "🚀 批量上传（按文件名配对）"},
    "img_batch_hint":    {"en": "Name files <ItemCode>.jpg or <Barcode>.jpg "
                                "to auto-match. Multiple files allowed.",
                          "zh": "文件命名为 <商品编号>.jpg 或 <条码>.jpg 即可自动配对。"
                                "支持多文件同时上传。"},
    "img_batch_choose":  {"en": "Choose multiple image files",
                          "zh": "选择多个图片文件"},
    "img_batch_preview": {"en": "Pairing preview",
                          "zh": "配对预览"},
    "img_matched":       {"en": "Matched",
                          "zh": "已匹配"},
    "img_unmatched":     {"en": "Unmatched",
                          "zh": "未匹配"},
    "img_filename":      {"en": "File name",
                          "zh": "文件名"},
    "img_matched_to":    {"en": "Matched to",
                          "zh": "匹配到"},
    "img_match_reason":  {"en": "Reason",
                          "zh": "原因"},
    "img_batch_confirm": {"en": "✅ Confirm and upload {n} images",
                          "zh": "✅ 确认上传 {n} 张图片"},
    "img_batch_done":    {"en": "{n} images uploaded successfully",
                          "zh": "成功上传 {n} 张图片"},
    "img_batch_failed":  {"en": "{n} files failed to save",
                          "zh": "{n} 个文件保存失败"},
    "img_batch_cancel":  {"en": "Cancel",
                          "zh": "取消"},
    # Auto-scan root directory for images
    "img_scan_title":    {"en": "📂 Auto-import from app folder",
                          "zh": "📂 自动导入应用根目录"},
    "img_scan_hint":     {"en": "Drop image files (named <ItemCode>.jpg or "
                                "<Barcode>.jpg) into the same folder as "
                                "app.py. They will be moved into "
                                "product_images/ on next start, or click "
                                "below to do it now.",
                          "zh": "把图片文件（命名为 <商品编号>.jpg 或 "
                                "<条码>.jpg）拖到 app.py 所在的根目录。"
                                "下次启动时会自动归档到 product_images/，"
                                "或点击下方按钮立即归档。"},
    "img_scan_now":      {"en": "🔍 Scan & import now",
                          "zh": "🔍 立即扫描并归档"},
    "img_scan_moved":    {"en": "Files imported on startup",
                          "zh": "启动时归档"},
    "img_scan_left":     {"en": "Files left in root (no match)",
                          "zh": "根目录留下（无法匹配）"},
    "img_scan_none":     {"en": "Nothing to import — root folder is clean.",
                          "zh": "无需导入——根目录干净。"},
    "img_scan_errors":   {"en": "Errors",
                          "zh": "错误"},
    # Email notifications
    "nav_email":         {"en": "📧 Email Notifications",
                          "zh": "📧 邮件通知"},
    "email_title":       {"en": "Email Notification Settings",
                          "zh": "邮件通知设置"},
    "email_subtitle":    {"en": "Send automatic email alerts for new orders, "
                                "dispatches, shortages, and warehouse supplier "
                                "requests. Sending is asynchronous — UI is "
                                "never blocked.",
                          "zh": "为新订单、发货、缺货以及仓库向供货商下单等事件发送邮件。"
                                "邮件异步发送，不会阻塞页面操作。"},
    "email_enabled":     {"en": "Enable email notifications",
                          "zh": "启用邮件通知"},
    "smtp_settings":     {"en": "SMTP Settings",
                          "zh": "SMTP 设置"},
    "smtp_host":         {"en": "SMTP Host",
                          "zh": "SMTP 服务器"},
    "smtp_port":         {"en": "SMTP Port",
                          "zh": "SMTP 端口"},
    "smtp_user":         {"en": "SMTP Username",
                          "zh": "SMTP 用户名"},
    "smtp_password":     {"en": "SMTP Password",
                          "zh": "SMTP 密码"},
    "smtp_from":         {"en": "From Address",
                          "zh": "发件人地址"},
    "smtp_use_tls":      {"en": "Use STARTTLS",
                          "zh": "使用 STARTTLS"},
    "event_settings":    {"en": "Per-event Settings",
                          "zh": "事件设置"},
    "ev_new_order":      {"en": "📦 New Order",
                          "zh": "📦 新订单"},
    "ev_dispatched":     {"en": "🚚 Dispatched",
                          "zh": "🚚 已发货"},
    "ev_shortage":       {"en": "🔔 Shortage",
                          "zh": "🔔 缺货"},
    "ev_supplier_order": {"en": "🏭 Supplier purchase request",
                          "zh": "🏭 向供货商下单"},
    "recipients":        {"en": "Recipients (one per line)",
                          "zh": "收件人（每行一个邮箱）"},
    "branch_recipients": {"en": "Per-branch Email (optional)",
                          "zh": "分店专属邮箱（可选）"},
    "branch_email_hint": {"en": "If set, dispatched-event emails for that "
                                "branch will additionally CC this address.",
                          "zh": "若填写，发货邮件会额外抄送给该分店的此邮箱。"},
    "save_email_cfg":    {"en": "💾 Save Settings",
                          "zh": "💾 保存设置"},
    "email_cfg_saved":   {"en": "Settings saved",
                          "zh": "设置已保存"},
    "test_email":        {"en": "📧 Send Test Email",
                          "zh": "📧 发送测试邮件"},
    "test_to":           {"en": "Send test to",
                          "zh": "测试邮件发送至"},
    "test_sent_ok":      {"en": "✅ Test email sent successfully",
                          "zh": "✅ 测试邮件发送成功"},
    "test_sent_fail":    {"en": "❌ Test email failed",
                          "zh": "❌ 测试邮件发送失败"},
    "email_resend_active": {
        "en": "Sending via **Resend API** (HTTPS). Railway Hobby/Fire plans cannot use Gmail SMTP on ports 587/465.",
        "zh": "当前通过 **Resend API**（HTTPS）发信。Railway 的 Free/Hobby 套餐无法使用 Gmail 的 SMTP（587/465 端口被封锁）。",
    },
    "email_railway_smtp_hint": {
        "en": "On Railway Free/Hobby, outbound SMTP is blocked. Either upgrade to **Pro** and redeploy, or set `RESEND_API_KEY` + `RESEND_FROM` in Railway Variables (see README).",
        "zh": "Railway 的 Free/Hobby 会封锁出站 SMTP。请二选一：① 升级到 **Pro** 并 Redeploy；② 在 Railway Variables 设置 `RESEND_API_KEY` 与 `RESEND_FROM`（见 README）。",
    },
    "email_log_title":   {"en": "Recent Email Activity",
                          "zh": "近期邮件记录"},
    "email_log_empty":   {"en": "No email activity yet",
                          "zh": "暂无邮件记录"},
    "log_time":          {"en": "Time",
                          "zh": "时间"},
    "log_event":         {"en": "Event",
                          "zh": "事件"},
    "log_to":            {"en": "Recipients",
                          "zh": "收件人"},
    "log_status":        {"en": "Status",
                          "zh": "状态"},
    "log_ok":            {"en": "OK",
                          "zh": "成功"},
    "log_fail":          {"en": "Failed",
                          "zh": "失败"},
    "log_clear":         {"en": "🗑️ Clear log",
                          "zh": "🗑️ 清空记录"},
}

STATUS_LABEL: dict[str, str] = {
    OrderStatus.PENDING:    "st_pending",
    OrderStatus.DISPATCHED: "st_dispatched",
    OrderStatus.RECEIVED:   "st_received",
    ShortageStatus.OPEN:         "st_open",
    ShortageStatus.RESENDING:    "st_resending",
    ShortageStatus.OUT_OF_STOCK: "st_oos",
    ShortageStatus.RESOLVED:     "st_resolved",
}


def t(key: str) -> str:
    lang = st.session_state.get("lang", "en")
    entry = T.get(key)
    if not entry:
        return key
    return entry.get(lang) or entry.get("en") or key


# =========================================================================
# DATABASE
# =========================================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        TEXT    NOT NULL,
    branch          TEXT    NOT NULL,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    price           REAL    DEFAULT 0,
    qty_cartons     INTEGER NOT NULL DEFAULT 0,
    qty_pcs         INTEGER NOT NULL DEFAULT 0,
    is_manual       INTEGER NOT NULL DEFAULT 0,
    status          TEXT    NOT NULL DEFAULT 'Pending',
    order_date      TEXT    NOT NULL,
    dispatch_date   TEXT,
    dispatch_cartons INTEGER,                           -- what warehouse actually shipped
    dispatch_pcs    INTEGER,                            -- (may differ from qty_* if low stock)
    receive_date    TEXT,
    actual_cartons  INTEGER,
    actual_pcs      INTEGER,
    receive_remarks TEXT
);
CREATE INDEX IF NOT EXISTS idx_orders_order_id ON orders(order_id);
CREATE INDEX IF NOT EXISTS idx_orders_branch   ON orders(branch);
CREATE INDEX IF NOT EXISTS idx_orders_status   ON orders(status);
CREATE INDEX IF NOT EXISTS idx_orders_date     ON orders(order_date);

CREATE TABLE IF NOT EXISTS shipments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        TEXT    NOT NULL,
    branch          TEXT    NOT NULL,
    dispatch_date   TEXT    NOT NULL,
    dispatched_by   TEXT
);
CREATE INDEX IF NOT EXISTS idx_shipments_order_id ON shipments(order_id);

CREATE TABLE IF NOT EXISTS shortages (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        TEXT    NOT NULL,
    order_line_id   INTEGER NOT NULL,
    branch          TEXT    NOT NULL,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    short_cartons   INTEGER NOT NULL DEFAULT 0,
    short_pcs       INTEGER NOT NULL DEFAULT 0,
    status          TEXT    NOT NULL DEFAULT 'Open',
    warehouse_reply TEXT,
    branch_remarks  TEXT,
    reported_date   TEXT    NOT NULL,
    resolved_date   TEXT
);
CREATE INDEX IF NOT EXISTS idx_short_status ON shortages(status);
CREATE INDEX IF NOT EXISTS idx_short_branch ON shortages(branch);
CREATE INDEX IF NOT EXISTS idx_short_order  ON shortages(order_id);

CREATE TABLE IF NOT EXISTS notifications (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    event_type      TEXT    NOT NULL,
    title           TEXT    NOT NULL,
    message         TEXT    NOT NULL,
    order_id        TEXT,
    target_role     TEXT    NOT NULL,
    target_branch   TEXT,
    is_read         INTEGER NOT NULL DEFAULT 0,
    created_at      TEXT    NOT NULL,
    read_at         TEXT
);
CREATE INDEX IF NOT EXISTS idx_notify_target
ON notifications(target_role, target_branch, is_read, created_at DESC);

CREATE TABLE IF NOT EXISTS stock_arrivals (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    title           TEXT    NOT NULL,
    notice          TEXT,
    items_text      TEXT    NOT NULL,
    is_active       INTEGER NOT NULL DEFAULT 1,
    created_at      TEXT    NOT NULL,
    created_by      TEXT
);
CREATE INDEX IF NOT EXISTS idx_arrival_active
ON stock_arrivals(is_active, created_at DESC);

CREATE TABLE IF NOT EXISTS inventory (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    item_key        TEXT    NOT NULL UNIQUE,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    stock_cartons   INTEGER NOT NULL DEFAULT 0,
    stock_pcs       INTEGER NOT NULL DEFAULT 0,
    updated_at      TEXT    NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_inventory_item_code ON inventory(item_code);
CREATE INDEX IF NOT EXISTS idx_inventory_barcode ON inventory(barcode);

CREATE TABLE IF NOT EXISTS inventory_txn (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    txn_type        TEXT    NOT NULL, -- IN / OUT / ADJUST
    item_key        TEXT    NOT NULL,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    order_id        TEXT,
    change_cartons  INTEGER NOT NULL DEFAULT 0,
    change_pcs      INTEGER NOT NULL DEFAULT 0,
    before_cartons  INTEGER NOT NULL DEFAULT 0,
    before_pcs      INTEGER NOT NULL DEFAULT 0,
    after_cartons   INTEGER NOT NULL DEFAULT 0,
    after_pcs       INTEGER NOT NULL DEFAULT 0,
    operator        TEXT,
    created_at      TEXT    NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_inventory_txn_created ON inventory_txn(created_at DESC);

CREATE TABLE IF NOT EXISTS product_prices (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    item_key        TEXT    NOT NULL UNIQUE,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    price           REAL    NOT NULL DEFAULT 0,
    operator        TEXT,
    updated_at      TEXT    NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_product_prices_updated ON product_prices(updated_at DESC);
-- 临期看板按 item_code/barcode 关联取价，加索引避免全表扫描。
CREATE INDEX IF NOT EXISTS idx_product_prices_item_code ON product_prices(item_code);
CREATE INDEX IF NOT EXISTS idx_product_prices_barcode ON product_prices(barcode);

CREATE TABLE IF NOT EXISTS user_accounts (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    username        TEXT    NOT NULL COLLATE NOCASE,
    password_salt   TEXT    NOT NULL,
    password_hash   TEXT    NOT NULL,
    branch          TEXT    NOT NULL,
    display_name    TEXT,
    phone           TEXT,
    status          TEXT    NOT NULL DEFAULT 'pending',
    permissions     TEXT    NOT NULL DEFAULT '[]',
    created_at      TEXT    NOT NULL,
    reviewed_at     TEXT,
    review_note     TEXT,
    UNIQUE(username)
);
CREATE INDEX IF NOT EXISTS idx_user_accounts_status
ON user_accounts(status, created_at DESC);
CREATE INDEX IF NOT EXISTS idx_user_accounts_branch ON user_accounts(branch);

CREATE TABLE IF NOT EXISTS audit_log (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    created_at      TEXT    NOT NULL,
    event_type      TEXT    NOT NULL,
    role            TEXT    NOT NULL,
    account_id      INTEGER,
    username        TEXT,
    branch          TEXT,
    order_id        TEXT,
    detail          TEXT
);
CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_log(created_at DESC);
CREATE INDEX IF NOT EXISTS idx_audit_event ON audit_log(event_type);
CREATE INDEX IF NOT EXISTS idx_audit_branch ON audit_log(branch);
CREATE INDEX IF NOT EXISTS idx_audit_order ON audit_log(order_id);
CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_log(username);

CREATE TABLE IF NOT EXISTS branch_cart_draft (
    account_id  INTEGER NOT NULL,
    branch      TEXT NOT NULL,
    cart_json   TEXT NOT NULL DEFAULT '[]',
    updated_at  TEXT NOT NULL,
    PRIMARY KEY (account_id, branch)
);

CREATE TABLE IF NOT EXISTS product_catalog (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    price           REAL    NOT NULL DEFAULT 0,
    category        TEXT    DEFAULT 'General',
    pcs_per_carton  REAL    DEFAULT 0,
    created_at      TEXT    NOT NULL,
    updated_at      TEXT    NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_product_catalog_item_code ON product_catalog(item_code);
CREATE INDEX IF NOT EXISTS idx_product_catalog_barcode ON product_catalog(barcode);
CREATE INDEX IF NOT EXISTS idx_product_catalog_category ON product_catalog(category);

-- =========================================================
-- BRANCH_INVENTORY_BATCHES  (分店批次库存 + 过期预警)
-- 各分店独立持有；一行 = 某分店某商品的一个批次（带过期日期）。
-- 与中央 inventory 表互不影响：inventory 仍是仓库权威库存，
-- 本表面向"分店货架上的实际批次"做临期/过期管理。
-- =========================================================
CREATE TABLE IF NOT EXISTS branch_inventory_batches (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    branch          TEXT    NOT NULL,            -- 分店ID（须属于 BRANCHES）
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    batch_no        TEXT,                         -- 批次号（可选，便于追溯）
    qty_cartons     INTEGER NOT NULL DEFAULT 0,   -- 该批次剩余库存（箱）
    qty_pcs         INTEGER NOT NULL DEFAULT 0,   -- 该批次剩余库存（个）
    production_date TEXT,                          -- 生产/入库日期 YYYY-MM-DD（可选）
    shelf_life_days INTEGER,                       -- 保质期天数（可选）
    expire_date     TEXT    NOT NULL,             -- 过期日期 YYYY-MM-DD（核心，必填）
    status          TEXT    NOT NULL DEFAULT 'active',  -- active / depleted / discarded
    received_by     TEXT,
    note            TEXT,
    created_at      TEXT    NOT NULL,
    updated_at      TEXT    NOT NULL
);
-- 临期扫描的核心查询：按分店 + 状态 + 过期日期范围过滤。
CREATE INDEX IF NOT EXISTS idx_batch_branch_expire
ON branch_inventory_batches(branch, status, expire_date);
CREATE INDEX IF NOT EXISTS idx_batch_expire
ON branch_inventory_batches(expire_date);
CREATE INDEX IF NOT EXISTS idx_batch_item
ON branch_inventory_batches(item_code, barcode);
"""


@contextmanager
def db_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    with db_conn() as conn:
        conn.executescript(SCHEMA_SQL)
        # Migration: older orders.db files won't have dispatch_cartons/pcs.
        # SQLite's CREATE TABLE IF NOT EXISTS leaves an existing table alone,
        # so add the columns explicitly when missing. This is idempotent.
        existing_cols = {
            r[1] for r in conn.execute("PRAGMA table_info(orders)").fetchall()
        }
        if "dispatch_cartons" not in existing_cols:
            conn.execute("ALTER TABLE orders ADD COLUMN dispatch_cartons INTEGER")
        if "dispatch_pcs" not in existing_cols:
            conn.execute("ALTER TABLE orders ADD COLUMN dispatch_pcs INTEGER")


def _branch_cart_session_ctx() -> tuple[int, str] | None:
    """(account_id, branch) when a branch clerk session may own a server-side cart."""
    if st.session_state.get("role") != Role.BRANCH:
        return None
    aid = st.session_state.get("account_id")
    branch = st.session_state.get("branch") or ""
    if aid is None or not str(branch).strip():
        return None
    try:
        return (int(aid), str(branch).strip())
    except (TypeError, ValueError):
        return None


def _parse_branch_cart_payload(raw: str | None) -> list[dict]:
    """Load cart lines from JSON; skips invalid entries."""
    if not raw or not str(raw).strip():
        return []
    try:
        data = json.loads(raw)
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    out: list[dict] = []
    for it in data:
        if not isinstance(it, dict):
            continue
        name = str(it.get("name") or "").strip()
        if not name:
            continue
        try:
            q_ct = int(it.get("qty_cartons") or 0)
            q_pc = int(it.get("qty_pcs") or 0)
        except (TypeError, ValueError):
            continue
        if q_ct < 0:
            q_ct = 0
        if q_pc < 0:
            q_pc = 0
        try:
            price = float(it.get("price") or 0.0)
        except (TypeError, ValueError):
            price = 0.0
        try:
            im_raw = int(it.get("is_manual"))
        except (TypeError, ValueError):
            im_raw = 0
        out.append({
            "item_code": str(it.get("item_code") or "").strip(),
            "barcode": str(it.get("barcode") or "").strip(),
            "name": name,
            "unit": str(it.get("unit") or "").strip(),
            "price": price,
            "qty_cartons": q_ct,
            "qty_pcs": q_pc,
            "is_manual": 1 if im_raw else 0,
        })
    return out


def _persist_branch_cart_to_db(account_id: int, branch: str, cart: list[dict]) -> None:
    """Upsert serialized cart draft for branch staff (shared SQLite)."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        blob = json.dumps(cart or [], ensure_ascii=False)
    except (TypeError, ValueError):
        blob = "[]"
    with db_conn() as conn:
        conn.execute(
            """
            INSERT INTO branch_cart_draft (account_id, branch, cart_json, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(account_id, branch) DO UPDATE SET
                cart_json = excluded.cart_json,
                updated_at = excluded.updated_at
            """,
            (account_id, branch, blob, now),
        )


def _delete_branch_cart_draft(account_id: int, branch: str) -> None:
    with db_conn() as conn:
        conn.execute(
            "DELETE FROM branch_cart_draft WHERE account_id = ? AND branch = ?",
            (account_id, branch),
        )


def _hydrate_branch_cart_if_needed() -> None:
    """Restore cart from DB once per Streamlit session if memory is empty."""
    ctx = _branch_cart_session_ctx()
    if not ctx:
        return
    aid, branch = ctx
    if st.session_state.get("_branch_cart_hydrated"):
        return
    st.session_state["_branch_cart_hydrated"] = True
    if len(st.session_state.get("cart") or []) > 0:
        return
    row = None
    with db_conn() as conn:
        row = conn.execute(
            "SELECT cart_json FROM branch_cart_draft WHERE account_id = ? AND branch = ?",
            (aid, branch),
        ).fetchone()
    if not row:
        return
    loaded = _parse_branch_cart_payload(row["cart_json"])
    loaded = [
        x for x in loaded
        if x["qty_cartons"] > 0 or x["qty_pcs"] > 0
    ]
    if loaded:
        st.session_state.cart = loaded


def _persist_branch_cart() -> None:
    ctx = _branch_cart_session_ctx()
    if not ctx:
        return
    aid, branch = ctx
    _persist_branch_cart_to_db(aid, branch, list(st.session_state.get("cart") or []))


def _normalize_username(raw: str) -> str:
    return (raw or "").strip().lower()


def _validate_username(username: str) -> bool:
    u = _normalize_username(username)
    if len(u) < 3 or len(u) > 32:
        return False
    return bool(re.fullmatch(r"[a-z0-9_]+", u))


def _hash_new_password(password: str) -> tuple[str, str]:
    salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 120_000)
    return salt.hex(), dk.hex()


def _verify_account_password(password: str, salt_hex: str, hash_hex: str) -> bool:
    try:
        salt = bytes.fromhex(salt_hex)
        dk = hashlib.pbkdf2_hmac(
            "sha256", password.encode("utf-8"), salt, 120_000
        )
        ex = bytes.fromhex(hash_hex)
        return hmac.compare_digest(dk, ex)
    except Exception:
        return False


def _parse_permissions_json(raw: str | None) -> list[str]:
    if not raw or not str(raw).strip():
        return []
    try:
        data = json.loads(raw)
    except Exception:
        return []
    if not isinstance(data, list):
        return []
    out: list[str] = []
    for x in data:
        if isinstance(x, str) and x in BRANCH_PERM_CODES:
            out.append(x)
    return out


def _default_all_permissions() -> list[str]:
    return list(BRANCH_PERM_CODES)


def account_fetch_by_username(username: str) -> sqlite3.Row | None:
    u = _normalize_username(username)
    if not u:
        return None
    with db_conn() as conn:
        return conn.execute(
            "SELECT * FROM user_accounts WHERE lower(username) = ?",
            (u,),
        ).fetchone()


def account_insert_application(
    username: str,
    password: str,
    branch: str,
    display_name: str,
    phone: str,
) -> tuple[bool, str]:
    """Returns (ok, error_key_or_empty). error_key is a t() key."""
    if not _validate_username(username):
        return False, "acct_user_invalid"
    if branch not in BRANCHES:
        return False, "acct_branch_invalid"
    if len(password) < 6:
        return False, "acct_pw_short"
    u = _normalize_username(username)
    existing = account_fetch_by_username(u)
    salt, ph = _hash_new_password(password)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if existing is not None:
        st = existing["status"]
        if st in (ACCOUNT_STATUS_PENDING, ACCOUNT_STATUS_APPROVED):
            return False, "acct_apply_dup"
        if st == ACCOUNT_STATUS_REJECTED:
            with db_conn() as conn:
                conn.execute(
                    """
                    UPDATE user_accounts SET
                        password_salt = ?, password_hash = ?, branch = ?,
                        display_name = ?, phone = ?, status = ?,
                        permissions = ?, created_at = ?, reviewed_at = NULL,
                        review_note = NULL
                    WHERE id = ?
                    """,
                    (
                        salt,
                        ph,
                        branch,
                        (display_name or "").strip() or None,
                        (phone or "").strip() or None,
                        ACCOUNT_STATUS_PENDING,
                        json.dumps([], ensure_ascii=False),
                        now,
                        existing["id"],
                    ),
                )
            return True, ""
        return False, "acct_apply_dup"
    try:
        with db_conn() as conn:
            conn.execute(
                """
                INSERT INTO user_accounts (
                    username, password_salt, password_hash, branch,
                    display_name, phone, status, permissions, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    u,
                    salt,
                    ph,
                    branch,
                    (display_name or "").strip() or None,
                    (phone or "").strip() or None,
                    ACCOUNT_STATUS_PENDING,
                    json.dumps([], ensure_ascii=False),
                    now,
                ),
            )
    except sqlite3.IntegrityError:
        return False, "acct_apply_dup"
    return True, ""


def account_try_login(username: str, password: str) -> tuple[str, sqlite3.Row | None]:
    """Returns status: ok | pending | rejected | bad_password | not_found"""
    row = account_fetch_by_username(username)
    if row is None:
        return "not_found", None
    if not _verify_account_password(
        password, row["password_salt"], row["password_hash"]
    ):
        return "bad_password", row
    st = row["status"]
    if st == ACCOUNT_STATUS_PENDING:
        return "pending", row
    if st == ACCOUNT_STATUS_REJECTED:
        return "rejected", row
    if st != ACCOUNT_STATUS_APPROVED:
        return "rejected", row
    perms = _parse_permissions_json(row["permissions"])
    if not perms:
        return "no_permissions", row
    return "ok", row


def account_set_status(
    uid: int,
    status: str,
    note: str | None = None,
) -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with db_conn() as conn:
        conn.execute(
            """
            UPDATE user_accounts
            SET status = ?, reviewed_at = ?, review_note = ?
            WHERE id = ?
            """,
            (status, now, (note or "").strip() or None, uid),
        )


def account_approve(uid: int, permissions: list[str]) -> None:
    clean = [p for p in permissions if p in BRANCH_PERM_CODES]
    if not clean:
        clean = _default_all_permissions()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with db_conn() as conn:
        conn.execute(
            """
            UPDATE user_accounts
            SET status = ?, permissions = ?, reviewed_at = ?, review_note = NULL
            WHERE id = ?
            """,
            (ACCOUNT_STATUS_APPROVED, json.dumps(clean, ensure_ascii=False), now, uid),
        )


def account_update_permissions(uid: int, permissions: list[str]) -> None:
    clean = [p for p in permissions if p in BRANCH_PERM_CODES]
    with db_conn() as conn:
        conn.execute(
            "UPDATE user_accounts SET permissions = ? WHERE id = ?",
            (json.dumps(clean, ensure_ascii=False), uid),
        )


def account_set_password(uid: int, new_password: str) -> None:
    salt, ph = _hash_new_password(new_password)
    with db_conn() as conn:
        conn.execute(
            "UPDATE user_accounts SET password_salt = ?, password_hash = ? WHERE id = ?",
            (salt, ph, uid),
        )


def account_create_direct(
    username: str,
    password: str,
    branch: str,
    permissions: list[str] | None,
) -> tuple[bool, str]:
    if not _validate_username(username):
        return False, "acct_user_invalid"
    if branch not in BRANCHES:
        return False, "acct_branch_invalid"
    if len(password) < 6:
        return False, "acct_pw_short"
    u = _normalize_username(username)
    if account_fetch_by_username(u) is not None:
        return False, "acct_apply_dup"
    clean = (
        [p for p in (permissions or []) if p in BRANCH_PERM_CODES]
        or _default_all_permissions()
    )
    salt, ph = _hash_new_password(password)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with db_conn() as conn:
            conn.execute(
                """
                INSERT INTO user_accounts (
                    username, password_salt, password_hash, branch,
                    display_name, phone, status, permissions, created_at, reviewed_at
                ) VALUES (?, ?, ?, ?, NULL, NULL, ?, ?, ?, ?)
                """,
                (
                    u,
                    salt,
                    ph,
                    branch,
                    ACCOUNT_STATUS_APPROVED,
                    json.dumps(clean, ensure_ascii=False),
                    now,
                    now,
                ),
            )
    except sqlite3.IntegrityError:
        return False, "acct_apply_dup"
    return True, ""


def account_list_by_status(status: str | None = None) -> list[sqlite3.Row]:
    with db_conn() as conn:
        if status:
            return list(
                conn.execute(
                    """
                    SELECT * FROM user_accounts
                    WHERE status = ?
                    ORDER BY created_at DESC
                    """,
                    (status,),
                ).fetchall()
            )
        return list(
            conn.execute(
                "SELECT * FROM user_accounts ORDER BY created_at DESC"
            ).fetchall()
        )


def count_open_shortages() -> int:
    with db_conn() as conn:
        return conn.execute(
            "SELECT COUNT(*) FROM shortages WHERE status = ?",
            (ShortageStatus.OPEN,),
        ).fetchone()[0]


def _inventory_item_key(item_code: str, barcode: str, name: str) -> str:
    ic = (item_code or "").strip()
    bc = (barcode or "").strip()
    nm = (name or "").strip()
    if ic:
        return f"ic:{ic.lower()}"
    if bc:
        return f"bc:{bc.lower()}"
    return f"nm:{nm.lower()}"


def _inventory_version() -> str:
    try:
        with db_conn() as conn:
            row = conn.execute(
                "SELECT COALESCE(MAX(updated_at), '') AS v FROM inventory"
            ).fetchone()
        return str(row["v"] if row and "v" in row.keys() else "")
    except Exception:
        return ""


def _price_version() -> str:
    try:
        with db_conn() as conn:
            row = conn.execute(
                "SELECT COALESCE(MAX(updated_at), '') AS v FROM product_prices"
            ).fetchone()
        return str(row["v"] if row and "v" in row.keys() else "")
    except Exception:
        return ""


def _upsert_inventory_line(
    conn: sqlite3.Connection,
    item_code: str,
    barcode: str,
    name: str,
    unit: str,
    stock_ct: int,
    stock_pc: int,
) -> None:
    key = _inventory_item_key(item_code, barcode, name)
    conn.execute(
        """
        INSERT INTO inventory
        (item_key, item_code, barcode, name, unit, stock_cartons, stock_pcs, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(item_key) DO UPDATE SET
            item_code=excluded.item_code,
            barcode=excluded.barcode,
            name=excluded.name,
            unit=excluded.unit,
            stock_cartons=excluded.stock_cartons,
            stock_pcs=excluded.stock_pcs,
            updated_at=excluded.updated_at
        """,
        (
            key,
            (item_code or "").strip() or None,
            (barcode or "").strip() or None,
            (name or "").strip() or "-",
            (unit or "").strip() or None,
            int(stock_ct),
            int(stock_pc),
            now_str(),
        ),
    )


def import_inventory_from_products_detailed(
    overwrite: bool = False, df: pd.DataFrame | None = None
) -> ExcelSheetImportOutcome:
    """Import stock from spreadsheet; returns counts + per-row failure reasons."""
    out = ExcelSheetImportOutcome()
    if df is None:
        df = _load_products_sheet_for_inventory_import()
    if df.empty:
        return out

    with db_conn() as conn:
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            if _df_row_is_blank_catalog_row(row):
                continue
            item_code = str(row.get("ItemCode", "") or "").strip()
            barcode = str(row.get("Barcode", "") or "").strip()
            name = str(row.get("Name", "") or "").strip()
            unit = str(row.get("Unit", "") or "").strip()
            if not name:
                out.n_failed += 1
                out.failure_messages.append(
                    f"{_import_row_ref(i, item_code, barcode, name)} → 缺少商品名称，已跳过"
                )
                continue

            qty_ct = int(float(row.get("StockCartons", 0) or 0))
            qty_pc = int(float(row.get("StockPcs", 0) or 0))
            qty_total = int(float(row.get("StockTotal", 0) or 0))
            if qty_ct == 0 and qty_pc == 0 and qty_total != 0:
                qty_pc = qty_total

            if overwrite:
                inv_row = _get_inventory_row(conn, item_code, barcode, name)
                before_ct = int(inv_row["stock_cartons"]) if inv_row else 0
                before_pc = int(inv_row["stock_pcs"]) if inv_row else 0
                target_ct, target_pc = qty_ct, qty_pc
                d_ct = target_ct - before_ct
                d_pc = target_pc - before_pc
                if d_ct == 0 and d_pc == 0:
                    out.n_skipped_benign += 1
                    continue
                try:
                    _apply_inventory_change(
                        conn,
                        txn_type="ADJUST",
                        item_code=item_code,
                        barcode=barcode,
                        name=name,
                        unit=unit,
                        change_ct=d_ct,
                        change_pc=d_pc,
                        operator="import_products_xlsx_overwrite",
                    )
                except Exception as e:
                    out.n_failed += 1
                    out.failure_messages.append(
                        f"{_import_row_ref(i, item_code, barcode, name)} → {e}"
                    )
                    log_exception("import_inventory_row", e)
                    continue
                out.n_written += 1
            else:
                if qty_ct == 0 and qty_pc == 0:
                    # 数量为 0 的行视为「本条不导入」，不计入失败。
                    out.n_skipped_benign += 1
                    continue
                if qty_ct >= 0 and qty_pc >= 0:
                    txn_type = "IN"
                elif qty_ct <= 0 and qty_pc <= 0:
                    txn_type = "OUT"
                else:
                    txn_type = "ADJUST"
                try:
                    _apply_inventory_change(
                        conn,
                        txn_type=txn_type,
                        item_code=item_code,
                        barcode=barcode,
                        name=name,
                        unit=unit,
                        change_ct=qty_ct,
                        change_pc=qty_pc,
                        operator="import_products_xlsx",
                    )
                except Exception as e:
                    out.n_failed += 1
                    out.failure_messages.append(
                        f"{_import_row_ref(i, item_code, barcode, name)} → {e}"
                    )
                    log_exception("import_inventory_row", e)
                    continue
                out.n_written += 1
    return out


def import_inventory_from_products(
    overwrite: bool = False, df: pd.DataFrame | None = None
) -> int:
    """Import stock — backward-compatible: returns number of DB-updated rows."""
    return import_inventory_from_products_detailed(overwrite, df).n_written


def _upsert_product_price(
    conn: sqlite3.Connection,
    item_code: str,
    barcode: str,
    name: str,
    price: float,
    operator: str,
) -> None:
    key = _inventory_item_key(item_code, barcode, name)
    conn.execute(
        """
        INSERT INTO product_prices
        (item_key, item_code, barcode, name, price, operator, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(item_key) DO UPDATE SET
            item_code=excluded.item_code,
            barcode=excluded.barcode,
            name=excluded.name,
            price=excluded.price,
            operator=excluded.operator,
            updated_at=excluded.updated_at
        """,
        (
            key,
            (item_code or "").strip() or None,
            (barcode or "").strip() or None,
            (name or "").strip() or "-",
            float(price or 0),
            (operator or "").strip() or None,
            now_str(),
        ),
    )


def import_prices_from_products_detailed(
    df: pd.DataFrame | None = None,
) -> ExcelSheetImportOutcome:
    """Import prices into DB price overlay table; returns structured outcome."""
    out = ExcelSheetImportOutcome()
    if df is None:
        df = _load_products_sheet_for_inventory_import()
    if df.empty:
        return out
    with db_conn() as conn:
        for i, (_, row) in enumerate(df.iterrows(), start=2):
            if _df_row_is_blank_catalog_row(row):
                continue
            item_code = str(row.get("ItemCode", "") or "").strip()
            barcode = str(row.get("Barcode", "") or "").strip()
            name = str(row.get("Name", "") or "").strip()
            if not name:
                out.n_failed += 1
                out.failure_messages.append(
                    f"{_import_row_ref(i, item_code, barcode, name)} → 缺少商品名称，已跳过"
                )
                continue
            raw_p = row.get("Price")
            try:
                if raw_p is None or (
                    isinstance(raw_p, float) and pd.isna(raw_p)
                ):
                    price = float("nan")
                else:
                    price = float(raw_p)
            except (TypeError, ValueError):
                out.n_failed += 1
                out.failure_messages.append(
                    f"{_import_row_ref(i, item_code, barcode, name)} → 价格格式无效（{raw_p!r}）"
                )
                continue
            if price != price:  # NaN
                out.n_failed += 1
                out.failure_messages.append(
                    f"{_import_row_ref(i, item_code, barcode, name)} → 缺少有效价格"
                )
                continue
            if price < 0:
                out.n_failed += 1
                out.failure_messages.append(
                    f"{_import_row_ref(i, item_code, barcode, name)} → 价格不能为负数（{price}）"
                )
                continue
            key = _inventory_item_key(item_code, barcode, name)
            old = conn.execute(
                "SELECT price FROM product_prices WHERE item_key = ?",
                (key,),
            ).fetchone()
            old_price = float(old["price"]) if old else None
            if old_price is not None and abs(old_price - price) < 1e-9:
                out.n_skipped_benign += 1
                continue
            _upsert_product_price(
                conn,
                item_code=item_code,
                barcode=barcode,
                name=name,
                price=price,
                operator="import_products_xlsx_price",
            )
            out.n_written += 1
    load_products.clear()
    return out


def import_prices_from_products(df: pd.DataFrame | None = None) -> int:
    """Import prices — backward-compatible: returns number of upserted price rows."""
    return import_prices_from_products_detailed(df).n_written


def _get_inventory_row(
    conn: sqlite3.Connection,
    item_code: str,
    barcode: str,
    name: str,
) -> sqlite3.Row | None:
    key = _inventory_item_key(item_code, barcode, name)
    return conn.execute(
        "SELECT * FROM inventory WHERE item_key = ?",
        (key,),
    ).fetchone()


def _apply_inventory_change(
    conn: sqlite3.Connection,
    txn_type: str,
    item_code: str,
    barcode: str,
    name: str,
    unit: str,
    change_ct: int,
    change_pc: int,
    order_id: str = "",
    operator: str = "",
) -> tuple[int, int]:
    row = _get_inventory_row(conn, item_code, barcode, name)
    before_ct = int(row["stock_cartons"]) if row else 0
    before_pc = int(row["stock_pcs"]) if row else 0
    after_ct = before_ct + int(change_ct)
    after_pc = before_pc + int(change_pc)
    _upsert_inventory_line(
        conn,
        item_code=item_code,
        barcode=barcode,
        name=name,
        unit=unit,
        stock_ct=after_ct,
        stock_pc=after_pc,
    )
    conn.execute(
        """
        INSERT INTO inventory_txn
        (txn_type, item_key, item_code, barcode, name, order_id,
         change_cartons, change_pcs, before_cartons, before_pcs,
         after_cartons, after_pcs, operator, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            txn_type,
            _inventory_item_key(item_code, barcode, name),
            (item_code or "").strip() or None,
            (barcode or "").strip() or None,
            (name or "").strip() or "-",
            (order_id or "").strip() or None,
            int(change_ct),
            int(change_pc),
            before_ct,
            before_pc,
            after_ct,
            after_pc,
            (operator or "").strip() or None,
            now_str(),
        ),
    )
    return after_ct, after_pc


# =========================================================================
# 分店批次库存 & 临期/过期预警 (Branch batch inventory & expiry alerts)
# =========================================================================
def _parse_ymd(value: str) -> datetime | None:
    """把 'YYYY-MM-DD' 解析为 datetime；非法返回 None。"""
    s = (value or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d")
    except ValueError:
        return None


def compute_expire_date(
    production_date: str = "",
    shelf_life_days: int | None = None,
    explicit_expire: str = "",
) -> str:
    """计算过期日期，返回 'YYYY-MM-DD'。

    优先级：
      1. 直接给定过期日期 explicit_expire（最权威）。
      2. 否则用 生产/入库日期 + 保质期天数 推算。
    两者都缺则抛 ValueError —— 入库必须能确定过期日期。
    """
    exp = _parse_ymd(explicit_expire)
    if exp is not None:
        return exp.strftime("%Y-%m-%d")
    prod = _parse_ymd(production_date)
    if prod is not None and shelf_life_days is not None and int(shelf_life_days) >= 0:
        return (prod + timedelta(days=int(shelf_life_days))).strftime("%Y-%m-%d")
    raise ValueError("必须提供过期日期，或同时提供生产日期与保质期天数")


def add_branch_batch(
    *,
    branch: str,
    name: str,
    item_code: str = "",
    barcode: str = "",
    unit: str = "",
    qty_cartons: int = 0,
    qty_pcs: int = 0,
    production_date: str = "",
    shelf_life_days: int | None = None,
    expire_date: str = "",
    batch_no: str = "",
    received_by: str = "",
    note: str = "",
) -> int:
    """入库/调拨：把一个批次商品写入指定分店的批次库存。

    - 强制要求 branch 属于 BRANCHES、name 非空、库存为正、且能确定过期日期。
    - 返回新批次行 id。失败抛 ValueError。
    """
    branch = (branch or "").strip()
    if branch not in BRANCHES:
        raise ValueError(f"未知分店: {branch}")
    name = (name or "").strip()
    if not name:
        raise ValueError("商品名称不能为空")
    qc, qp = int(qty_cartons or 0), int(qty_pcs or 0)
    if qc < 0 or qp < 0 or (qc == 0 and qp == 0):
        raise ValueError("入库数量必须大于 0")
    # 入库接口强制要求过期日期（可直接给，或由生产日期+保质期推算）。
    expire = compute_expire_date(production_date, shelf_life_days, expire_date)
    ts = now_str()
    with db_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO branch_inventory_batches
            (branch, item_code, barcode, name, unit, batch_no,
             qty_cartons, qty_pcs, production_date, shelf_life_days,
             expire_date, status, received_by, note, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?, ?, ?)
            """,
            (
                branch,
                (item_code or "").strip() or None,
                (barcode or "").strip() or None,
                name,
                (unit or "").strip() or None,
                (batch_no or "").strip() or None,
                qc,
                qp,
                (production_date or "").strip() or None,
                int(shelf_life_days) if shelf_life_days is not None else None,
                expire,
                (received_by or "").strip() or None,
                (note or "").strip() or None,
                ts,
                ts,
            ),
        )
        return int(cur.lastrowid)


def scan_expiring_batches(
    warn_days: int | None = None,
    today: str = "",
) -> dict[str, list[sqlite3.Row]]:
    """扫描所有分店，找出临期/已过期且仍有库存的活动批次，按分店归类。

    判定条件（与需求一致）：
      当前日期 + 预警天数 >= 过期日期  ⇔  过期日期 <= 当前日期 + 预警天数
      且 status='active' 且 (qty_cartons > 0 或 qty_pcs > 0)

    返回: { 分店ID: [批次行, ...] }，每个分店内部按过期日期升序（最紧急在前）。
    """
    days = expiry_warn_days() if warn_days is None else int(warn_days)
    base = _parse_ymd(today) or datetime.now()
    # 阈值日期：过期日期 <= 该日期即纳入预警（含已过期）。
    threshold = (base + timedelta(days=days)).strftime("%Y-%m-%d")
    with db_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM branch_inventory_batches
            WHERE status = 'active'
              AND (qty_cartons > 0 OR qty_pcs > 0)
              AND expire_date <= ?
            ORDER BY branch ASC, expire_date ASC
            """,
            (threshold,),
        ).fetchall()
    grouped: dict[str, list[sqlite3.Row]] = {}
    for r in rows:
        grouped.setdefault(r["branch"], []).append(r)
    return grouped


def _format_expiry_lines(batches: list[sqlite3.Row], today_dt: datetime) -> str:
    """把某分店的临期批次拼成可读的多行文本（用于消息中心与邮件正文）。"""
    lines: list[str] = []
    for b in batches:
        exp = _parse_ymd(b["expire_date"])
        days_left = (exp - today_dt).days if exp else None
        if days_left is None:
            tag = ""
        elif days_left < 0:
            tag = f"（已过期 {abs(days_left)} 天）"
        elif days_left == 0:
            tag = "（今天到期）"
        else:
            tag = f"（剩 {days_left} 天）"
        qty_parts = []
        if int(b["qty_cartons"] or 0) > 0:
            qty_parts.append(f"{int(b['qty_cartons'])} 箱")
        if int(b["qty_pcs"] or 0) > 0:
            qty_parts.append(f"{int(b['qty_pcs'])} 个")
        qty_txt = " ".join(qty_parts) or "0"
        code = (b["item_code"] or b["barcode"] or "").strip()
        code_txt = f" [{code}]" if code else ""
        lines.append(
            f"- {b['name']}{code_txt} · 库存 {qty_txt} · 过期 {b['expire_date']} {tag}"
        )
    return "\n".join(lines)


def run_expiry_scan_and_notify(
    warn_days: int | None = None,
    today: str = "",
    send_email: bool = True,
) -> dict:
    """临期预警主流程（供定时任务调用）：

      1. 扫描临期/过期批次并按分店归类。
      2. 每个有临期商品的分店：写一条消息中心通知（target=该分店）。
      3. 如配置了店长邮箱（email_config.branch_emails[branch]），发邮件。
      4. 给管理员写一条汇总通知。

    返回执行汇总（便于日志/控制台输出）。本函数尽量不抛异常。
    """
    days = expiry_warn_days() if warn_days is None else int(warn_days)
    today_dt = _parse_ymd(today) or datetime.now()
    today_label = today_dt.strftime("%Y-%m-%d")
    grouped = scan_expiring_batches(days, today_label)

    summary = {
        "ran_at": now_str(),
        "warn_days": days,
        "branches_alerted": 0,
        "items_total": 0,
        "emails_sent": 0,
        "details": {},
    }
    if not grouped:
        return summary

    try:
        cfg = load_email_config()
    except Exception as e:
        log_exception("expiry_load_email_config", e)
        cfg = {}
    branch_emails = (cfg.get("branch_emails") or {}) if isinstance(cfg, dict) else {}

    admin_lines: list[str] = []
    for branch, batches in grouped.items():
        body = _format_expiry_lines(batches, today_dt)
        title = f"⚠️ 临期/过期预警：{len(batches)} 项商品（{branch}）"
        message = (
            f"以下商品距离过期 ≤ {days} 天（或已过期），请尽快处理：\n\n{body}"
        )
        # 2) 消息中心：发往该分店
        create_notification(
            event_type="expiry_alert",
            title=title,
            message=message,
            target_role=Role.BRANCH,
            target_branch=branch,
        )
        # 3) 邮件：发往店长邮箱（若已配置）
        addr = (branch_emails.get(branch) or "").strip()
        if send_email and addr:
            subject = f"[SUNSHINE 临期预警] {branch} · {len(batches)} 项商品"
            notify(
                "expiry_alert",
                subject,
                f"分店：{branch}\n预警天数：{days} 天\n日期：{today_label}\n\n{message}",
                extra_to=[addr],
            )
            summary["emails_sent"] += 1

        summary["branches_alerted"] += 1
        summary["items_total"] += len(batches)
        summary["details"][branch] = len(batches)
        admin_lines.append(f"{branch}: {len(batches)} 项")

    # 4) 管理员汇总
    create_notification(
        event_type="expiry_alert",
        title=f"⚠️ 临期预警汇总：{summary['items_total']} 项 / {summary['branches_alerted']} 家分店",
        message="各分店临期/过期商品数量：\n" + "\n".join(admin_lines),
        target_role=Role.ADMIN,
    )
    return summary


# =========================================================================
# 管理员临期看板 · 统计接口 (Admin expiry dashboard analytics)
# =========================================================================
# 设计要点：
#   - 全部用单条聚合 SQL（COUNT/SUM/CASE WHEN）在数据库内算好再返回，
#     避免把大量批次行拉到 Python 里循环。
#   - JOIN 商品主档 product_catalog（取分类、每箱个数），并用相关子查询
#     取生效价（product_prices 覆盖优先，回退 product_catalog 价）。
#   - catalog JOIN 用"单行匹配子查询(pc.id = (SELECT ... LIMIT 1))"，
#     防止 item_code/barcode 双键 OR 关联产生行膨胀（重复计数）。

# 单品有效数量（统一折算为"个"当量）：个数 + 箱数 × 每箱个数（未知按 1 估）。
_EXPIRY_UNITS_SQL = (
    "(b.qty_pcs + b.qty_cartons * COALESCE(NULLIF(pc.pcs_per_carton, 0), 1))"
)

# catalog 单行匹配（避免 OR 双键 JOIN 的行膨胀）。
_EXPIRY_CATALOG_JOIN = (
    "LEFT JOIN product_catalog pc ON pc.id = ("
    "  SELECT c.id FROM product_catalog c"
    "  WHERE (b.item_code IS NOT NULL AND b.item_code <> '' AND c.item_code = b.item_code)"
    "     OR (b.barcode  IS NOT NULL AND b.barcode  <> '' AND c.barcode  = b.barcode)"
    "  LIMIT 1)"
)

# 生效单价：价格覆盖表优先，其次主档价，最后 0。
_EXPIRY_PRICE_SQL = (
    "COALESCE("
    "  (SELECT pr.price FROM product_prices pr"
    "   WHERE (b.item_code IS NOT NULL AND b.item_code <> '' AND pr.item_code = b.item_code)"
    "      OR (b.barcode  IS NOT NULL AND b.barcode  <> '' AND pr.barcode  = b.barcode)"
    "   LIMIT 1),"
    "  pc.price, 0)"
)


def _expiry_filter_clause(
    branch_filter: list[str] | None,
    category_filter: str | None,
) -> tuple[str, dict]:
    """构造 WHERE 过滤（分店多选 + 分类），返回 (sql, params)。"""
    clauses = ["b.status = 'active'", "(b.qty_cartons > 0 OR b.qty_pcs > 0)"]
    params: dict = {}
    if branch_filter:
        placeholders = ",".join(f":br{i}" for i in range(len(branch_filter)))
        clauses.append(f"b.branch IN ({placeholders})")
        for i, br in enumerate(branch_filter):
            params[f"br{i}"] = br
    if category_filter:
        clauses.append("pc.category = :cat")
        params["cat"] = category_filter
    return " AND ".join(clauses), params


def admin_expiry_branch_stats(
    warn_days: int | None = None,
    branch_filter: list[str] | None = None,
    category_filter: str | None = None,
    today: str = "",
) -> list[dict]:
    """各分店临期统计（一次聚合查询，按分店一行）。

    返回每分店：批次数、总库存当量、已过期数、临期数、临期占比、潜在损耗金额。
    """
    days = expiry_warn_days() if warn_days is None else int(warn_days)
    base = _parse_ymd(today) or datetime.now()
    today_label = base.strftime("%Y-%m-%d")
    threshold = (base + timedelta(days=days)).strftime("%Y-%m-%d")
    where, params = _expiry_filter_clause(branch_filter, category_filter)
    params.update({"today": today_label, "threshold": threshold})
    sql = f"""
        SELECT
            b.branch AS branch,
            COUNT(*) AS batch_count,
            SUM({_EXPIRY_UNITS_SQL}) AS total_units,
            SUM(CASE WHEN b.expire_date <= :today THEN 1 ELSE 0 END) AS expired_n,
            SUM(CASE WHEN b.expire_date > :today AND b.expire_date <= :threshold
                     THEN 1 ELSE 0 END) AS expiring_n,
            SUM(CASE WHEN b.expire_date <= :threshold
                     THEN {_EXPIRY_UNITS_SQL} * {_EXPIRY_PRICE_SQL}
                     ELSE 0 END) AS loss_value
        FROM branch_inventory_batches b
        {_EXPIRY_CATALOG_JOIN}
        WHERE {where}
        GROUP BY b.branch
        ORDER BY expired_n DESC, expiring_n DESC
    """
    with db_conn() as conn:
        rows = conn.execute(sql, params).fetchall()
    out: list[dict] = []
    for r in rows:
        bc = int(r["batch_count"] or 0)
        risk = int(r["expired_n"] or 0) + int(r["expiring_n"] or 0)
        out.append({
            "branch": r["branch"],
            "batch_count": bc,
            "total_units": int(r["total_units"] or 0),
            "expired_n": int(r["expired_n"] or 0),
            "expiring_n": int(r["expiring_n"] or 0),
            "risk_n": risk,
            "risk_ratio": (risk / bc) if bc else 0.0,
            "loss_value": round(float(r["loss_value"] or 0), 2),
        })
    return out


def admin_expiry_overview(
    warn_days: int | None = None,
    branch_filter: list[str] | None = None,
    category_filter: str | None = None,
    today: str = "",
) -> dict:
    """全局总览 KPI（基于分店聚合结果二次汇总；分店数 ≤ 7，开销极小）。"""
    stats = admin_expiry_branch_stats(warn_days, branch_filter, category_filter, today)
    return {
        "total_branches": len(BRANCHES),
        "branches_with_stock": len(stats),
        "total_units": sum(s["total_units"] for s in stats),
        "total_batches": sum(s["batch_count"] for s in stats),
        "expired_total": sum(s["expired_n"] for s in stats),
        "expiring_total": sum(s["expiring_n"] for s in stats),
        "loss_total": round(sum(s["loss_value"] for s in stats), 2),
        "stats": stats,
    }


def admin_product_categories() -> list[str]:
    """商品主档里出现过的分类（用于看板筛选下拉）。"""
    try:
        with db_conn() as conn:
            rows = conn.execute(
                "SELECT DISTINCT category FROM product_catalog "
                "WHERE category IS NOT NULL AND category <> '' ORDER BY category"
            ).fetchall()
        return [r["category"] for r in rows]
    except Exception:
        return []


def _branch_manager_label(branch: str, branch_emails: dict | None = None) -> str:
    """店长信息：姓名取该分店一个已审核账号的显示名，联系方式取 branch_emails。

    系统未单独存"店长姓名"字段，这里用已审核分店账号的 display_name 近似，
    并附上邮箱联系方式。两者皆空则返回 '-'。
    """
    name = ""
    try:
        with db_conn() as conn:
            row = conn.execute(
                "SELECT display_name, username FROM user_accounts "
                "WHERE branch = ? AND status = 'approved' "
                "ORDER BY created_at ASC LIMIT 1",
                (branch,),
            ).fetchone()
        if row:
            name = (row["display_name"] or row["username"] or "").strip()
    except Exception:
        pass
    if branch_emails is None:
        try:
            branch_emails = (load_email_config().get("branch_emails") or {})
        except Exception:
            branch_emails = {}
    email = (branch_emails.get(branch) or "").strip()
    if name and email:
        return f"{name} · {email}"
    return name or email or "-"


def admin_expiry_detail_rows(
    warn_days: int | None = None,
    branch_filter: list[str] | None = None,
    category_filter: str | None = None,
    only_risk: bool = False,
    today: str = "",
    limit: int = 1000,
) -> list[sqlite3.Row]:
    """看板主数据表的明细行（带分类、生效单价；按过期日期升序）。"""
    days = expiry_warn_days() if warn_days is None else int(warn_days)
    base = _parse_ymd(today) or datetime.now()
    threshold = (base + timedelta(days=days)).strftime("%Y-%m-%d")
    where, params = _expiry_filter_clause(branch_filter, category_filter)
    params.update({"threshold": threshold, "limit": int(limit)})
    risk_clause = " AND b.expire_date <= :threshold" if only_risk else ""
    sql = f"""
        SELECT b.branch, b.name, b.item_code, b.barcode, b.unit,
               b.qty_cartons, b.qty_pcs, b.expire_date,
               pc.category AS category,
               {_EXPIRY_PRICE_SQL} AS unit_price
        FROM branch_inventory_batches b
        {_EXPIRY_CATALOG_JOIN}
        WHERE {where}{risk_clause}
        ORDER BY b.expire_date ASC
        LIMIT :limit
    """
    with db_conn() as conn:
        return conn.execute(sql, params).fetchall()


def page_admin_expiry_dashboard() -> None:
    """管理员：全局分店库存与临期统计看板。"""
    render_page_heading(t("exp_dash_title"), t("exp_dash_sub"))
    warn_days = expiry_warn_days()

    # ---- 筛选条 ----
    cats = admin_product_categories()
    fc1, fc2, fc3 = st.columns([2, 2, 1])
    sel_branches = fc1.multiselect(t("exp_filter_branch"), BRANCHES, default=[])
    sel_cat = fc2.selectbox(t("exp_filter_cat"), ["—"] + cats, index=0)
    only_risk = fc3.checkbox(t("exp_only_risk"), value=False)
    branch_filter = sel_branches or None
    category_filter = None if sel_cat in ("—", "") else sel_cat

    ov = admin_expiry_overview(warn_days, branch_filter, category_filter)
    stats = ov["stats"]

    # ---- KPI 卡片 ----
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric(t("exp_kpi_branches"), f"{ov['branches_with_stock']}/{ov['total_branches']}")
    k2.metric(t("exp_kpi_units"), f"{ov['total_units']:,}")
    k3.metric(t("exp_kpi_expired"), f"{ov['expired_total']:,}")
    k4.metric(t("exp_kpi_expiring"), f"{ov['expiring_total']:,}")
    k5.metric(t("exp_kpi_loss"), f"{ov['loss_total']:,.0f}")

    if not stats:
        st.info(t("exp_no_data"))
        return

    # ---- 各分店临期/过期对比图（Streamlit 原生柱状图）----
    st.markdown(f"#### {t('exp_chart_title')}")
    chart_df = pd.DataFrame(
        [
            {
                "branch": s["branch"],
                t("exp_kpi_expired"): s["expired_n"],
                t("exp_kpi_expiring"): s["expiring_n"],
            }
            for s in stats
        ]
    ).set_index("branch")
    st.bar_chart(chart_df)

    # ---- 分店严重程度排行 ----
    st.markdown(f"#### {t('exp_rank_title')}")
    rank = sorted(
        stats,
        key=lambda s: (s["expired_n"], s["risk_n"], s["loss_value"]),
        reverse=True,
    )
    st.dataframe(
        pd.DataFrame([
            {
                t("exp_col_branch"): s["branch"],
                t("exp_col_items"): s["batch_count"],
                t("exp_kpi_expired"): s["expired_n"],
                t("exp_kpi_expiring"): s["expiring_n"],
                t("exp_col_ratio"): f"{s['risk_ratio'] * 100:.0f}%",
                t("exp_kpi_loss"): f"{s['loss_value']:,.0f}",
            }
            for s in rank
        ]),
        use_container_width=True,
        hide_index=True,
    )

    # ---- 潜在损耗排行 ----
    st.markdown(f"#### {t('exp_loss_title')}")
    loss_rank = sorted(stats, key=lambda s: s["loss_value"], reverse=True)
    st.dataframe(
        pd.DataFrame([
            {
                t("exp_col_branch"): s["branch"],
                t("exp_kpi_loss"): f"{s['loss_value']:,.0f}",
                t("exp_kpi_expired"): s["expired_n"],
                t("exp_kpi_expiring"): s["expiring_n"],
            }
            for s in loss_rank
        ]),
        use_container_width=True,
        hide_index=True,
    )

    # ---- 全局明细表 ----
    st.markdown(f"#### {t('exp_table_title')}")
    rows = admin_expiry_detail_rows(
        warn_days, branch_filter, category_filter, only_risk=only_risk
    )
    if not rows:
        st.info(t("exp_no_data"))
        return
    try:
        bemails = (load_email_config().get("branch_emails") or {})
    except Exception:
        bemails = {}
    today = datetime.now().date()
    mgr_cache: dict[str, str] = {}
    table = []
    for r in rows:
        exp = _parse_ymd(r["expire_date"])
        dl = (exp.date() - today).days if exp else None
        if dl is None:
            days_txt = ""
        elif dl < 0:
            days_txt = t("stock_expired")
        else:
            days_txt = str(dl)
        stock_parts = []
        if int(r["qty_cartons"] or 0) > 0:
            stock_parts.append(f"{int(r['qty_cartons'])} {t('stock_qty_ct')}")
        if int(r["qty_pcs"] or 0) > 0:
            stock_parts.append(f"{int(r['qty_pcs'])} {t('stock_qty_pc')}")
        br = r["branch"]
        if br not in mgr_cache:
            mgr_cache[br] = _branch_manager_label(br, bemails)
        table.append({
            t("exp_col_branch"): br,
            t("exp_col_product"): r["name"],
            t("exp_col_cat"): r["category"] or "",
            t("exp_col_stock"): " / ".join(stock_parts) or "0",
            t("exp_col_expire"): r["expire_date"],
            t("exp_col_daysleft"): days_txt,
            t("exp_col_manager"): mgr_cache[br],
        })
    st.dataframe(pd.DataFrame(table), use_container_width=True, hide_index=True)


def _notification_target_filter_sql(role: str, branch: str | None) -> tuple[str, tuple]:
    if role == Role.BRANCH:
        return (
            "target_role = ? AND (target_branch IS NULL OR target_branch = ?)",
            (role, branch or ""),
        )
    return ("target_role = ?", (role,))


def count_unread_notifications() -> int:
    role = st.session_state.get("role")
    if role not in (Role.BRANCH, Role.WAREHOUSE, Role.ADMIN):
        return 0
    branch = st.session_state.get("branch")
    where_sql, params = _notification_target_filter_sql(role, branch)
    with db_conn() as conn:
        row = conn.execute(
            f"SELECT COUNT(*) FROM notifications WHERE {where_sql} AND is_read = 0",
            params,
        ).fetchone()
    return int(row[0] if row else 0)


def create_notification(
    event_type: str,
    title: str,
    message: str,
    target_role: str,
    target_branch: str | None = None,
    order_id: str = "",
) -> None:
    try:
        with db_conn() as conn:
            conn.execute(
                """
                INSERT INTO notifications
                (event_type, title, message, order_id, target_role, target_branch, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    event_type,
                    title,
                    message,
                    (order_id or "").strip() or None,
                    target_role,
                    (target_branch or "").strip() or None,
                    now_str(),
                ),
            )
    except Exception as e:
        log_exception("create_notification", e)


def load_notifications(limit: int = 100) -> list[sqlite3.Row]:
    role = st.session_state.get("role")
    if role not in (Role.BRANCH, Role.WAREHOUSE, Role.ADMIN):
        return []
    branch = st.session_state.get("branch")
    where_sql, params = _notification_target_filter_sql(role, branch)
    with db_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT * FROM notifications
            WHERE {where_sql}
            ORDER BY is_read ASC, created_at DESC
            LIMIT ?
            """,
            (*params, int(limit)),
        ).fetchall()
    return rows


def header_ticker_text() -> str:
    """Short rolling text shown in the top header."""
    role = st.session_state.get("role")
    if role not in (Role.BRANCH, Role.WAREHOUSE, Role.ADMIN):
        return ""
    rows = load_notifications(limit=5)
    unread_rows = [r for r in rows if int(r["is_read"] or 0) == 0]
    if unread_rows:
        parts = [f"{r['title']}" for r in unread_rows[:3]]
        return " | ".join(parts)
    arrival = get_active_stock_arrival()
    if arrival is not None:
        return f"📦 {(arrival['title'] or '').strip() or t('nav_arrivals')}"
    return ""


def mark_notification_read(notification_id: int) -> None:
    with db_conn() as conn:
        conn.execute(
            """
            UPDATE notifications
            SET is_read = 1, read_at = COALESCE(read_at, ?)
            WHERE id = ?
            """,
            (now_str(), int(notification_id)),
        )


def mark_all_notifications_read() -> None:
    role = st.session_state.get("role")
    if role not in (Role.BRANCH, Role.WAREHOUSE, Role.ADMIN):
        return
    branch = st.session_state.get("branch")
    where_sql, params = _notification_target_filter_sql(role, branch)
    with db_conn() as conn:
        conn.execute(
            f"""
            UPDATE notifications
            SET is_read = 1, read_at = COALESCE(read_at, ?)
            WHERE {where_sql} AND is_read = 0
            """,
            (now_str(), *params),
        )


def get_active_stock_arrival() -> sqlite3.Row | None:
    with db_conn() as conn:
        row = conn.execute(
            """
            SELECT *
            FROM stock_arrivals
            WHERE is_active = 1
            ORDER BY created_at DESC
            LIMIT 1
            """
        ).fetchone()
    return row


def publish_stock_arrival(title: str, notice: str, items_text: str) -> int:
    """Create a new active arrival bulletin and retire old ones."""
    with db_conn() as conn:
        conn.execute("UPDATE stock_arrivals SET is_active = 0 WHERE is_active = 1")
        conn.execute(
            """
            INSERT INTO stock_arrivals
            (title, notice, items_text, is_active, created_at, created_by)
            VALUES (?, ?, ?, 1, ?, ?)
            """,
            (
                (title or "").strip() or "New arrivals",
                (notice or "").strip(),
                (items_text or "").strip(),
                now_str(),
                st.session_state.get("role") or "admin",
            ),
        )
        new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    return int(new_id)


# =========================================================================
# BACKUP / RESTORE
# =========================================================================
# Strategy:
#   - On startup: take one snapshot (so even a "code change broke the app"
#     scenario leaves us with a fresh restore point).
#   - On meaningful writes: auto-snapshot, but throttled so we don't make
#     hundreds of files during a busy ordering session.
#   - Manual backup button in the admin panel for explicit checkpoints.
#   - Restore is admin-only and requires explicit confirmation. Before
#     overwriting the live DB, the *current* DB is itself backed up under
#     a "pre_restore" name so the restore is reversible.
#
# We use SQLite's native backup() API rather than file copy, so we get a
# transactionally consistent snapshot even if writes are happening
# concurrently. backup() pages the source into the destination atomically.
def _ensure_backup_dir() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def backup_database(label: str = "auto") -> Path | None:
    """Snapshot the live DB to backups/orders_<label>_<timestamp>.db.

    Uses sqlite3's native backup API for atomic consistency. Returns the
    backup file path, or None if the source DB doesn't exist yet."""
    if not DB_PATH.exists():
        return None
    _ensure_backup_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(c for c in label if c.isalnum() or c in "-_") or "auto"
    dest = BACKUP_DIR / f"orders_{safe_label}_{ts}.db"
    # If a backup with the exact same timestamp already exists (two backups
    # in the same second), append a counter so we never silently overwrite.
    if dest.exists():
        n = 1
        while True:
            cand = BACKUP_DIR / f"orders_{safe_label}_{ts}_{n}.db"
            if not cand.exists():
                dest = cand
                break
            n += 1

    src = sqlite3.connect(DB_PATH)
    try:
        dst = sqlite3.connect(dest)
        try:
            src.backup(dst)  # atomic page-level copy with consistent view
        finally:
            dst.close()
    finally:
        src.close()

    return dest


def _prune_old_auto_backups() -> int:
    """Keep only the most recent BACKUP_RETAIN auto-* snapshots.
    Manual / pre_restore backups are never pruned. Returns count deleted."""
    if not BACKUP_DIR.exists():
        return 0
    auto = sorted(
        BACKUP_DIR.glob("orders_auto_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    to_delete = auto[BACKUP_RETAIN:]
    for p in to_delete:
        try:
            p.unlink()
        except OSError:
            pass
    return len(to_delete)


def auto_snapshot_if_due() -> Path | None:
    """Throttled auto-snapshot. Skips if the most recent auto-snapshot was
    less than BACKUP_MIN_INTERVAL_MINUTES ago."""
    if not DB_PATH.exists():
        return None
    _ensure_backup_dir()
    auto = sorted(
        BACKUP_DIR.glob("orders_auto_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if auto:
        age_min = (datetime.now().timestamp() - auto[0].stat().st_mtime) / 60.0
        if age_min < BACKUP_MIN_INTERVAL_MINUTES:
            return None
    path = backup_database("auto")
    _prune_old_auto_backups()
    return path


def list_backups() -> list[dict]:
    """Return all backups with a quick row-count peek per table.
    Sorted newest first. Each entry: {path, name, size_kb, created, counts}."""
    if not BACKUP_DIR.exists():
        return []
    out = []
    files = sorted(
        BACKUP_DIR.glob("orders_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for p in files:
        stat = p.stat()
        # Peek table counts. Use a try/except so a corrupt backup just
        # shows as "?" rather than crashing the listing.
        counts = {"orders": "?", "shipments": "?", "shortages": "?"}
        try:
            c = sqlite3.connect(p)
            try:
                for table in counts:
                    try:
                        n = c.execute(
                            f"SELECT COUNT(*) FROM {table}"
                        ).fetchone()[0]
                        counts[table] = int(n)
                    except sqlite3.DatabaseError:
                        counts[table] = "?"
            finally:
                c.close()
        except sqlite3.DatabaseError:
            pass
        out.append({
            "path": p,
            "name": p.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "created": datetime.fromtimestamp(stat.st_mtime).strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            "counts": counts,
        })
    return out


def restore_from_backup(backup_path: Path) -> Path:
    """Restore the live DB from a backup file. Before overwriting, the
    current DB is itself snapshotted as `pre_restore_*` so the operation
    is reversible. Returns path to the safety snapshot."""
    if not backup_path.exists():
        raise FileNotFoundError(f"Backup not found: {backup_path}")

    # Sanity check: backup must look like one of our DB files
    try:
        c = sqlite3.connect(backup_path)
        try:
            tables = {
                r[0] for r in c.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
            }
        finally:
            c.close()
    except sqlite3.DatabaseError as e:
        raise ValueError(f"Backup file is not a valid SQLite DB: {e}")

    if not {"orders", "shipments", "shortages"}.issubset(tables):
        raise ValueError(
            f"Backup is missing required tables. Found: {tables}"
        )

    # Safety snapshot of whatever is currently live
    safety = backup_database("pre_restore")

    # Overwrite the live DB with the backup. Using sqlite's backup API into
    # the live file gives us an atomic, transactionally clean restore even
    # if the live DB is in use by another connection.
    src = sqlite3.connect(backup_path)
    try:
        dst = sqlite3.connect(DB_PATH)
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()

    return safety


def export_db_dump_sql() -> bytes:
    """Export the live DB as a portable SQL text dump. Useful for migrating
    across SQLite versions or inspecting the data in a text editor."""
    if not DB_PATH.exists():
        return b""
    conn = sqlite3.connect(DB_PATH)
    try:
        lines = list(conn.iterdump())
    finally:
        conn.close()
    return "\n".join(lines).encode("utf-8")


def restore_from_sql_dump(sql_bytes: bytes) -> Path:
    """Replace the live DB with one built from a SQL dump.
    Pre-restore safety backup is taken first."""
    text = sql_bytes.decode("utf-8") if isinstance(sql_bytes, (bytes, bytearray)) else sql_bytes
    safety = backup_database("pre_restore")
    # Replace the live DB. Build into a temp file then move into place so
    # we never leave a half-written orders.db.
    tmp_path = DB_PATH.with_suffix(".db.restoring")
    if tmp_path.exists():
        tmp_path.unlink()
    conn = sqlite3.connect(tmp_path)
    try:
        conn.executescript(text)
        conn.commit()
    finally:
        conn.close()
    # Atomic replace
    if DB_PATH.exists():
        DB_PATH.unlink()
    tmp_path.rename(DB_PATH)
    return safety


# =========================================================================
# EMAIL NOTIFICATIONS
# =========================================================================
# Three event types:
#   - new_order   (branch submitted an order  → notify warehouse + admin)
#   - dispatched  (warehouse marked dispatch  → notify branch + admin)
#   - shortage    (branch reported short qty  → notify warehouse + admin)
#
# Design constraints:
#   1. Sending must be ASYNCHRONOUS — never block a UI button click on SMTP.
#      Failures degrade silently (logged, never raised to the user).
#   2. Config is in a JSON file, NOT the SQLite DB, so backups of orders.db
#      don't carry SMTP creds.
#   3. Per-event "enabled" toggles + per-event recipient lists.
#   4. Every send attempt — success or failure — gets a row in email_log.json.
#      Admin page reads & displays this so the operator can see what's flowing.
#   5. Test-send button so the operator can verify SMTP without triggering a
#      real order.

EVENT_KEYS = ("new_order", "dispatched", "shortage", "supplier_order", "expiry_alert")


def _default_email_config() -> dict:
    return {
        "enabled": False,
        "smtp_host": "",
        "smtp_port": 587,
        "smtp_user": "",
        "smtp_password": "",
        "from_addr": "",
        "use_tls": True,
        "events": {
            "new_order":  {"enabled": True, "to": []},
            "dispatched": {"enabled": True, "to": []},
            "shortage":   {"enabled": True, "to": []},
            "supplier_order": {"enabled": True, "to": []},
            # 临期/过期预警：每天定时扫描后，按分店发往店长邮箱（branch_emails）。
            "expiry_alert": {"enabled": True, "to": []},
        },
        # Optional per-branch recipient — a "dispatched" event for a given
        # branch will additionally CC the address listed here.
        "branch_emails": {},
    }


def _env_bool(name: str) -> bool | None:
    v = os.getenv(name, "").strip().lower()
    if not v:
        return None
    return v in ("1", "true", "yes", "on")


def _apply_email_env_overrides(cfg: dict) -> dict:
    """Production (e.g. Railway): SMTP/recipients from env when JSON is absent.

    Env vars override file values when set (same precedence as GEMINI_*)."""
    en = _env_bool("SUNSHINE_EMAIL_ENABLED")
    if en is not None:
        cfg["enabled"] = en
    for key, env_name in (
        ("smtp_host", "SUNSHINE_SMTP_HOST"),
        ("smtp_user", "SUNSHINE_SMTP_USER"),
        ("smtp_password", "SUNSHINE_SMTP_PASSWORD"),
        ("from_addr", "SUNSHINE_SMTP_FROM"),
    ):
        v = os.getenv(env_name, "").strip()
        if v:
            cfg[key] = v
    pw_env = os.getenv("SUNSHINE_SMTP_PASSWORD", "").strip()
    if pw_env:
        cfg["smtp_password"] = _normalize_smtp_password(pw_env)
    port_s = os.getenv("SUNSHINE_SMTP_PORT", "").strip()
    if port_s.isdigit():
        cfg["smtp_port"] = int(port_s)
    tls = _env_bool("SUNSHINE_SMTP_USE_TLS")
    if tls is not None:
        cfg["use_tls"] = tls
    to_all = os.getenv("SUNSHINE_EMAIL_NOTIFY_TO", "").strip()
    if to_all:
        addrs = [a.strip() for a in to_all.split(",") if a.strip()]
        if addrs:
            for ev in EVENT_KEYS:
                cfg["events"][ev]["to"] = list(addrs)
                cfg["events"][ev]["enabled"] = True
    return cfg


def load_email_config() -> dict:
    """Load config from disk; merge with defaults so a missing field never
    crashes the page."""
    base = _default_email_config()
    if not EMAIL_CONFIG_PATH.exists():
        merged = dict(base)
        merged["events"] = {k: dict(v) for k, v in base["events"].items()}
        return _apply_email_env_overrides(merged)
    try:
        with EMAIL_CONFIG_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        merged = dict(base)
        merged["events"] = {k: dict(v) for k, v in base["events"].items()}
        return _apply_email_env_overrides(merged)
    # Shallow-merge top-level, deep-merge events
    merged = {**base, **data}
    merged_events = base["events"].copy()
    for k, v in (data.get("events") or {}).items():
        if k in merged_events:
            merged_events[k] = {**merged_events[k], **v}
    merged["events"] = merged_events
    merged["branch_emails"] = data.get("branch_emails", {}) or {}
    return _apply_email_env_overrides(merged)


def save_email_config(cfg: dict) -> None:
    with EMAIL_CONFIG_PATH.open("w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# --- Logging --------------------------------------------------------------
_log_lock = threading.Lock()


def _append_email_log(entry: dict) -> None:
    """Append one log entry, keeping only the most recent EMAIL_LOG_KEEP."""
    with _log_lock:
        log = []
        if EMAIL_LOG_PATH.exists():
            try:
                with EMAIL_LOG_PATH.open("r", encoding="utf-8") as f:
                    log = json.load(f)
                if not isinstance(log, list):
                    log = []
            except Exception:
                log = []
        log.append(entry)
        log = log[-EMAIL_LOG_KEEP:]
        try:
            with EMAIL_LOG_PATH.open("w", encoding="utf-8") as f:
                json.dump(log, f, indent=2, ensure_ascii=False)
        except Exception:
            pass


def load_email_log() -> list[dict]:
    if not EMAIL_LOG_PATH.exists():
        return []
    try:
        with EMAIL_LOG_PATH.open("r", encoding="utf-8") as f:
            log = json.load(f)
        if not isinstance(log, list):
            return []
        return log
    except Exception:
        return []


# --- Sending --------------------------------------------------------------
def _normalize_smtp_password(pw: str) -> str:
    """Gmail app passwords are often shown with spaces — strip them."""
    return re.sub(r"\s+", "", (pw or ""))


def _railway_smtp_blocked_hint(err: str) -> str:
    low = (err or "").lower()
    if any(
        x in low
        for x in (
            "timed out", "timeout", "unreachable", "network is unreachable",
            "errno 101", "[errno 101]",
        )
    ):
        return " " + t("email_railway_smtp_hint")
    return ""


def _send_resend_api(
    cfg: dict,
    to: list[str],
    subject: str,
    body: str,
    attachments: list[tuple[str, bytes, str]] | None = None,
) -> tuple[bool, str]:
    """HTTPS email — works on Railway when SMTP ports are blocked."""
    if not RESEND_API_KEY:
        return False, "RESEND_API_KEY not set"
    from_addr = (RESEND_FROM or (cfg.get("from_addr") or "")).strip()
    if not from_addr:
        return False, "Set RESEND_FROM (verified sender domain in Resend)"
    if not to:
        return False, "No recipients"
    payload: dict = {
        "from": from_addr,
        "to": to,
        "subject": subject,
        "text": body,
    }
    if attachments:
        payload["attachments"] = [
            {
                "filename": fname or "attachment.bin",
                "content": base64.b64encode(data).decode("ascii"),
            }
            for fname, data, _mime in attachments
        ]
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
        if resp.status not in (200, 201):
            return False, f"HTTP {resp.status}: {_safe_text(raw, 500)}"
        return True, ""
    except urllib.error.HTTPError as e:
        try:
            detail = e.read().decode("utf-8", errors="replace")
        except Exception:
            detail = ""
        return False, f"HTTP {e.code}: {_safe_text(detail, 800)}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _deliver_email(
    cfg: dict,
    to: list[str],
    subject: str,
    body: str,
    attachments: list[tuple[str, bytes, str]] | None = None,
) -> tuple[bool, str]:
    if RESEND_API_KEY:
        return _send_resend_api(cfg, to, subject, body, attachments=attachments)
    ok, err = _send_smtp(cfg, to, subject, body, attachments=attachments)
    if not ok:
        err = (err or "") + _railway_smtp_blocked_hint(err)
    return ok, err


def _send_smtp(
    cfg: dict,
    to: list[str],
    subject: str,
    body: str,
    attachments: list[tuple[str, bytes, str]] | None = None,
) -> tuple[bool, str]:
    """Synchronous send — DO NOT call from UI code. Returns (ok, error_msg)."""
    if not cfg.get("smtp_host") or not cfg.get("from_addr"):
        return False, "SMTP not fully configured"
    if not to:
        return False, "No recipients"

    msg = EmailMessage()
    msg["From"] = cfg["from_addr"]
    msg["To"] = ", ".join(to)
    msg["Subject"] = subject
    msg.set_content(body)
    for fname, data, mime in (attachments or []):
        mtype = (mime or "").strip().lower()
        if "/" in mtype:
            maintype, subtype = mtype.split("/", 1)
        else:
            maintype, subtype = "application", "octet-stream"
        msg.add_attachment(
            data,
            maintype=maintype,
            subtype=subtype,
            filename=fname or "attachment.bin",
        )

    host = cfg["smtp_host"]
    port = int(cfg.get("smtp_port") or 587)
    user = cfg.get("smtp_user") or ""
    pw = _normalize_smtp_password(cfg.get("smtp_password") or "")
    use_tls = bool(cfg.get("use_tls", True))

    try:
        # Port 465 = implicit SSL; everything else uses STARTTLS when enabled
        if port == 465:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=15) as s:
                if user:
                    s.login(user, pw)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.ehlo()
                if use_tls:
                    s.starttls(context=ssl.create_default_context())
                    s.ehlo()
                if user:
                    s.login(user, pw)
                s.send_message(msg)
        return True, ""
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _send_async(
    cfg: dict,
    to: list[str],
    subject: str,
    body: str,
    event: str,
    attachments: list[tuple[str, bytes, str]] | None = None,
) -> None:
    """Background-thread wrapper around _send_smtp. Logs every attempt."""
    def _worker():
        ok, err = _deliver_email(cfg, to, subject, body, attachments=attachments)
        _append_email_log({
            "ts":      now_str(),
            "event":   event,
            "to":      to,
            "subject": subject,
            "ok":      ok,
            "error":   err if not ok else "",
        })
    t = threading.Thread(target=_worker, daemon=True)
    t.start()


def notify(event: str, subject: str, body: str,
           extra_to: list[str] | None = None,
           attachments: list[tuple[str, bytes, str]] | None = None) -> None:
    """Public entry point used by business code. Non-blocking, never raises.

    `event` must be one of EVENT_KEYS. Recipients = config.events[event].to
    plus any `extra_to` (e.g. a specific branch's address)."""
    if event not in EVENT_KEYS:
        return
    try:
        cfg = load_email_config()
    except Exception as e:
        log_exception("notify_load_email_config", e)
        return
    if not cfg.get("enabled"):
        return
    ev = cfg.get("events", {}).get(event, {})
    if not ev.get("enabled", True):
        return
    to: list[str] = []
    for addr in (ev.get("to") or []) + (extra_to or []):
        addr = (addr or "").strip()
        if addr and addr not in to:
            to.append(addr)
    if not to:
        return
    _send_async(cfg, to, subject, body, event, attachments=attachments)


def send_test_email(cfg: dict, to_addr: str) -> tuple[bool, str]:
    """Synchronous test send used by the admin UI. Logs the attempt and
    returns (ok, error_msg) so the page can show the result immediately."""
    subject = "[SUNSHINE 阳光集团] Test email / 测试邮件"
    body = (
        "This is a test email from the SUNSHINE ordering system.\n\n"
        "If you can read this, your SMTP configuration is working.\n\n"
        "—— SUNSHINE 阳光集团 订货系统"
    )
    ok, err = _deliver_email(cfg, [to_addr], subject, body)
    _append_email_log({
        "ts":      now_str(),
        "event":   "test",
        "to":      [to_addr],
        "subject": subject,
        "ok":      ok,
        "error":   err if not ok else "",
    })
    return ok, err


# =========================================================================
# PRODUCTS
# =========================================================================
def _products_mtime() -> float:
    p = products_master_excel_path()
    return p.stat().st_mtime if p.exists() else -1.0


def _parse_products_excel_after_read(df: pd.DataFrame) -> tuple[pd.DataFrame, bool]:
    """Normalize dtypes / stock aliases. Does NOT overlay SQLite inventory."""
    for col in [
        "ItemCode", "Barcode", "Name", "Unit", "Price", "Category",
        "PcsPerCarton",
        "StockCartons", "StockPcs", "StockTotal",
    ]:
        if col not in df.columns:
            df[col] = "" if col not in ("PcsPerCarton",) else 0.0
    for c in ["ItemCode", "Barcode", "Name", "Unit", "Category"]:
        df[c] = df[c].astype(str).fillna("").replace({"nan": "", "None": ""})
    for c in ["ItemCode", "Barcode"]:
        df[c] = df[c].str.replace(r"\.0$", "", regex=True)
    col_by_lower = {str(c).strip().lower(): c for c in df.columns}

    def _find_col(candidates: list[str]) -> str | None:
        for cand in candidates:
            hit = col_by_lower.get(cand.strip().lower())
            if hit:
                return hit
        return None

    # Price: Excel often uses 价格/单价 instead of "Price"; map aliases first.
    price_candidates = [
        "Price", "价格", "单价", "售价", "零售价", "批发价", "进货价",
        "UnitPrice", "unit_price", "ListPrice", "MSRP",
        "单价(元)", "价格(元)", "售价(元)",
    ]
    price_src = _find_col(price_candidates)
    if price_src:
        df["Price"] = pd.to_numeric(df[price_src], errors="coerce").fillna(0.0)
    else:
        df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0.0)

    pcs_carton_candidates = [
        "PcsPerCarton", "每箱个数", "箱规", "箱装数", "PackSize",
        "QtyPerCarton", "PCS_PER_CARTON", "每箱包数",
    ]
    pcs_src = _find_col(pcs_carton_candidates)
    if pcs_src:
        df["PcsPerCarton"] = pd.to_numeric(df[pcs_src], errors="coerce").fillna(0.0)
    df["PcsPerCarton"] = pd.to_numeric(df["PcsPerCarton"], errors="coerce").fillna(0.0)

    stock_ct_candidates = [
        "StockCartons", "库存箱数", "库存(箱)", "库存箱", "CartonsStock", "Stock_Cartons",
    ]
    stock_pc_candidates = [
        "StockPcs", "库存个数", "库存(个)", "库存个", "PcsStock", "Stock_Pcs",
    ]
    stock_total_candidates = [
        "StockTotal", "Stock", "库存", "可用库存", "Available", "QtyAvailable",
    ]
    has_stock_info = False

    ct_col = _find_col(stock_ct_candidates)
    if ct_col:
        df["StockCartons"] = pd.to_numeric(df[ct_col], errors="coerce").fillna(0.0)
        has_stock_info = True
    if "StockCartons" not in df.columns:
        df["StockCartons"] = 0.0
    pc_col = _find_col(stock_pc_candidates)
    if pc_col:
        df["StockPcs"] = pd.to_numeric(df[pc_col], errors="coerce").fillna(0.0)
        has_stock_info = True
    if "StockPcs" not in df.columns:
        df["StockPcs"] = 0.0
    total_col = _find_col(stock_total_candidates)
    if total_col:
        df["StockTotal"] = pd.to_numeric(df[total_col], errors="coerce").fillna(0.0)
        has_stock_info = True
    if "StockTotal" not in df.columns:
        df["StockTotal"] = 0.0
    return df, has_stock_info


def _load_products_sheet_for_inventory_import() -> pd.DataFrame:
    """Rows and stock columns straight from spreadsheet (before DB overlay)."""
    master = products_master_excel_path()
    if not master.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(
            master,
            dtype={"ItemCode": str, "Barcode": str, "Category": str, "Unit": str},
        )
        df = df.rename(columns=lambda c: str(c).replace("\ufeff", "").strip())
        df = _normalize_product_sheet_columns(df)
        df, _ = _parse_products_excel_after_read(df)
        df.loc[df["Category"].str.strip() == "", "Category"] = "General"
        return df
    except Exception as e:
        log_exception("load_products_sheet_for_inventory_import", e)
        return pd.DataFrame()


def _load_products_sheet_from_upload(uploaded_file) -> pd.DataFrame:
    """Load products-like Excel from uploaded file for safe explicit import."""
    if uploaded_file is None:
        return pd.DataFrame()
    try:
        df = pd.read_excel(
            uploaded_file,
            dtype={"ItemCode": str, "Barcode": str, "Category": str, "Unit": str},
        )
        df = df.rename(columns=lambda c: str(c).replace("\ufeff", "").strip())
        df = _normalize_product_sheet_columns(df)
        df, _ = _parse_products_excel_after_read(df)
        df.loc[df["Category"].str.strip() == "", "Category"] = "General"
        return df
    except Exception as e:
        log_exception("load_products_sheet_from_upload", e)
        return pd.DataFrame()


def _build_inventory_import_template_bytes() -> bytes:
    """Excel template for inventory import; sheet1 matches parser column names.

    A second sheet contains brief Chinese instructions. Import only reads the
    first worksheet (default read_excel behavior)."""
    n = 8
    main = pd.DataFrame({
        "ItemCode":     [""] * n,
        "Barcode":      [""] * n,
        "Name":         [""] * n,
        "Unit":         [""] * n,
        "StockCartons": [""] * n,
        "StockPcs":     [""] * n,
    })
    lines = [
        "填写说明（导入时不会读取本页）",
        "",
        "① 请在「库存导入」工作表中填写。导入程序只读取第一张表。",
        "② 必填：商品名称(Name)。建议填写 商品编号(ItemCode) 或 条码(Barcode)，便于与商品档案对应。",
        "③ 库存变动导入：填写 StockCartons（库存箱数）、StockPcs（库存个数）；正数加仓、负数减仓；两列均为 0 的行会跳过。",
        "④ 列名也支持中文别名，例如：库存箱数、库存个数、库存（与系统导入说明一致）。",
        "⑤「覆盖库存」模式会把该行库存设为表中的箱数/个数，请谨慎使用。",
    ]
    instr = pd.DataFrame({"说明": lines})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        main.to_excel(writer, sheet_name="库存导入", index=False)
        instr.to_excel(writer, sheet_name="填写说明", index=False)
    buf.seek(0)
    return buf.getvalue()


def _build_price_import_template_bytes() -> bytes:
    """Excel template for price import; first sheet has ItemCode, Barcode, Name, Price."""
    n = 8
    main = pd.DataFrame({
        "ItemCode": [""] * n,
        "Barcode":  [""] * n,
        "Name":     [""] * n,
        "Price":    [""] * n,
    })
    lines = [
        "填写说明（导入时不会读取本页）",
        "",
        "① 请在「价格导入」工作表中填写。导入程序只读取第一张表。",
        "② 必填：商品名称(Name)。必须包含价格列：Price，或使用中文列名 价格、单价、售价 等。",
        "③ 建议同时填写 ItemCode 或 Barcode，与商品档案一致。价格为数字（元）。",
        "④ 未改价的商品可留空不填；名称为空的行会被跳过。",
    ]
    instr = pd.DataFrame({"说明": lines})
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        main.to_excel(writer, sheet_name="价格导入", index=False)
        instr.to_excel(writer, sheet_name="填写说明", index=False)
    buf.seek(0)
    return buf.getvalue()


@st.cache_data(
    ttl=300,
    # Replaces the default “Running load_products(...)” status with a branded line.
    show_spinner="☀️ 载入商品资料 · Loading catalog…",
)
def load_products(
    _mtime: float = -1.0,
    _inv_ver: str = "",
    _price_ver: str = "",
) -> pd.DataFrame:
    """Load product master from products.xlsx; auto-create demo if missing.

    Recognized columns: ItemCode, Barcode, Name, Unit, Price, Category.
    Category is optional. When present, it's used to group items in the
    warehouse picking-slip exports (Frozen / Rice / Beverage / etc.).
    Items without a category fall back to 'General'."""

    # === TRY DATABASE CATALOG FIRST ===
    try:
        with db_conn() as conn:
            rows = conn.execute(
                "SELECT item_code, barcode, name, unit, price, category, pcs_per_carton "
                "FROM product_catalog ORDER BY id"
            ).fetchall()
        if rows:
            df = pd.DataFrame([dict(r) for r in rows])
            df = df.rename(columns={
                "item_code": "ItemCode", "barcode": "Barcode", "name": "Name",
                "unit": "Unit", "price": "Price", "category": "Category",
                "pcs_per_carton": "PcsPerCarton",
            })
            for c in ["ItemCode", "Barcode", "Name", "Unit", "Category"]:
                df[c] = df[c].fillna("").astype(str).replace({"nan": "", "None": ""})
            df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0.0)
            df["PcsPerCarton"] = pd.to_numeric(df["PcsPerCarton"], errors="coerce").fillna(0.0)
            for sc in ["StockCartons", "StockPcs", "StockTotal"]:
                df[sc] = 0.0
            # Overlay inventory from inventory table
            try:
                inv = conn.execute(
                    "SELECT item_code, barcode, name, stock_cartons, stock_pcs FROM inventory"
                ).fetchall()
                inv_by_ic = {str(r["item_code"]).strip().lower(): r for r in inv if (r["item_code"] or "").strip()}
                inv_by_bc = {str(r["barcode"]).strip().lower(): r for r in inv if (r["barcode"] or "").strip()}
                inv_by_nm = {str(r["name"]).strip().lower(): r for r in inv if (r["name"] or "").strip()}
                for idx, row in df.iterrows():
                    ic = str(row.get("ItemCode", "") or "").strip().lower()
                    bc = str(row.get("Barcode", "") or "").strip().lower()
                    nm = str(row.get("Name", "") or "").strip().lower()
                    hit = (inv_by_ic.get(ic) if ic else None) or \
                          (inv_by_bc.get(bc) if bc else None) or \
                          (inv_by_nm.get(nm) if nm else None)
                    if hit:
                        df.at[idx, "StockCartons"] = int(hit["stock_cartons"] or 0)
                        df.at[idx, "StockPcs"] = int(hit["stock_pcs"] or 0)
                df["StockTotal"] = pd.to_numeric(df["StockCartons"], errors="coerce").fillna(0) + \
                                   pd.to_numeric(df["StockPcs"], errors="coerce").fillna(0)
            except Exception:
                pass
            # Overlay prices from product_prices table
            try:
                price_rows = conn.execute(
                    "SELECT item_code, barcode, name, price FROM product_prices"
                ).fetchall()
                by_ic = {str(r["item_code"]).strip().lower(): float(r["price"] or 0) for r in price_rows if (r["item_code"] or "").strip()}
                by_bc = {str(r["barcode"]).strip().lower(): float(r["price"] or 0) for r in price_rows if (r["barcode"] or "").strip()}
                by_nm = {str(r["name"]).strip().lower(): float(r["price"] or 0) for r in price_rows if (r["name"] or "").strip()}
                for idx, row in df.iterrows():
                    ic = str(row.get("ItemCode", "") or "").strip().lower()
                    bc = str(row.get("Barcode", "") or "").strip().lower()
                    nm = str(row.get("Name", "") or "").strip().lower()
                    hit = (by_ic.get(ic) if ic else None) or (by_bc.get(bc) if bc else None) or (by_nm.get(nm) if nm else None)
                    if hit is not None:
                        df.at[idx, "Price"] = float(hit)
            except Exception:
                pass
            df["_has_stock_info"] = True
            df.loc[df["Category"].str.strip() == "", "Category"] = "General"
            return df
    except Exception:
        pass  # Table might not exist yet, fall through to Excel
    # === END DATABASE CATALOG ===

    master = products_master_excel_path()
    if not master.exists():
        demo = pd.DataFrame({
            "ItemCode": ["P001", "P002", "P003", "P004", "P005"],
            "Barcode":  ["8801234567001", "8801234567002", "8801234567003",
                         "8801234567004", "8801234567005"],
            "Name":     ["Demo Rice 5kg", "Demo Sugar 1kg", "Demo Cooking Oil 1L",
                         "Demo Soap Bar", "Demo Soft Drink"],
            "Unit":     ["bag", "bag", "bottle", "pc", "can"],
            "Price":    [12.5, 3.2, 4.8, 1.5, 2.0],
            "Category": ["Rice", "General", "General", "General", "Beverage"],
            "PcsPerCarton": [10, 20, 12, 1, 24],
        })
        try:
            demo.to_excel(PRODUCTS_PATH, index=False)
        except Exception as e:
            log_exception("products_autocreate_demo", e)
        return demo
    try:
        # Read ItemCode and Barcode as strings from the get-go. Otherwise
        # pandas infers them as floats (especially large numeric barcodes
        # like 8801234567001), and "8801234567001.0" never matches anything.
        df = pd.read_excel(
            master,
            dtype={"ItemCode": str, "Barcode": str, "Category": str, "Unit": str},
        )
        # Normalize headers from Excel/WPS (trim spaces, remove BOM).
        df = df.rename(columns=lambda c: str(c).replace("\ufeff", "").strip())
        df = _normalize_product_sheet_columns(df)
        df, has_stock_info = _parse_products_excel_after_read(df)
        # Prefer inventory table values when available.
        try:
            with db_conn() as conn:
                inv = conn.execute(
                    "SELECT item_code, barcode, name, stock_cartons, stock_pcs FROM inventory"
                ).fetchall()
            inv_by_ic = {
                str(r["item_code"]).strip().lower(): r for r in inv
                if (r["item_code"] or "").strip()
            }
            inv_by_bc = {
                str(r["barcode"]).strip().lower(): r for r in inv
                if (r["barcode"] or "").strip()
            }
            inv_by_nm = {
                str(r["name"]).strip().lower(): r for r in inv
                if (r["name"] or "").strip()
            }
            matched = 0
            for idx, row in df.iterrows():
                ic = str(row.get("ItemCode", "") or "").strip().lower()
                bc = str(row.get("Barcode", "") or "").strip().lower()
                nm = str(row.get("Name", "") or "").strip().lower()
                hit = (inv_by_ic.get(ic) if ic else None) or \
                      (inv_by_bc.get(bc) if bc else None) or \
                      (inv_by_nm.get(nm) if nm else None)
                if hit:
                    df.at[idx, "StockCartons"] = int(hit["stock_cartons"] or 0)
                    df.at[idx, "StockPcs"] = int(hit["stock_pcs"] or 0)
                    matched += 1
            if inv:
                has_stock_info = True
            if matched > 0:
                df["StockTotal"] = (
                    pd.to_numeric(df["StockCartons"], errors="coerce").fillna(0)
                    + pd.to_numeric(df["StockPcs"], errors="coerce").fillna(0)
                )
        except Exception as e:
            log_exception("load_products_inventory_overlay", e)
        try:
            with db_conn() as conn:
                price_rows = conn.execute(
                    "SELECT item_code, barcode, name, price FROM product_prices"
                ).fetchall()
            by_ic = {
                str(r["item_code"]).strip().lower(): float(r["price"] or 0)
                for r in price_rows if (r["item_code"] or "").strip()
            }
            by_bc = {
                str(r["barcode"]).strip().lower(): float(r["price"] or 0)
                for r in price_rows if (r["barcode"] or "").strip()
            }
            by_nm = {
                str(r["name"]).strip().lower(): float(r["price"] or 0)
                for r in price_rows if (r["name"] or "").strip()
            }
            for idx, row in df.iterrows():
                ic = str(row.get("ItemCode", "") or "").strip().lower()
                bc = str(row.get("Barcode", "") or "").strip().lower()
                nm = str(row.get("Name", "") or "").strip().lower()
                hit = (by_ic.get(ic) if ic else None)
                if hit is None:
                    hit = (by_bc.get(bc) if bc else None)
                if hit is None:
                    hit = (by_nm.get(nm) if nm else None)
                if hit is not None:
                    df.at[idx, "Price"] = float(hit)
        except Exception as e:
            log_exception("load_products_price_overlay", e)
        df["_has_stock_info"] = bool(has_stock_info)
        # Items with no category fall back to "General" so grouping always works
        df.loc[df["Category"].str.strip() == "", "Category"] = "General"
        return df
    except Exception as e:
        st.error(f"Error loading catalog ({master.name}): {e}")
        return pd.DataFrame(columns=[
            "ItemCode", "Barcode", "Name", "Unit", "Price", "Category",
            "PcsPerCarton",
            "StockCartons", "StockPcs", "StockTotal", "_has_stock_info",
        ])


def search_products(query: str, limit: int = 80) -> pd.DataFrame:
    df = load_products(_products_mtime(), _inventory_version(), _price_version())
    if df.empty:
        return df
    q = (query or "").strip().lower()
    if not q:
        return df.head(limit)

    def _col_q(col: str) -> pd.Series:
        """Robust string match: handles numeric SKUs read as float (e.g. 11.0)."""
        s = df[col].fillna("").astype(str).str.strip().str.lower()
        s = s.str.replace(r"\.0$", "", regex=True)
        return s.str.contains(q, na=False, regex=False)

    mask = _col_q("Name") | _col_q("Barcode") | _col_q("ItemCode")
    return df[mask].head(limit)


def _product_master_excel_columns() -> list[str]:
    return [
        "ItemCode", "Barcode", "Name", "Unit", "Price", "Category",
        "PcsPerCarton", "StockCartons", "StockPcs", "StockTotal",
    ]


def _shelf_bootstrap_dataframe() -> pd.DataFrame:
    raw = _load_products_sheet_for_inventory_import()
    if raw is not None and not raw.empty:
        return raw
    return pd.DataFrame(
        [
            {
                "ItemCode": "", "Barcode": "", "Name": "", "Unit": "箱",
                "Price": 0.0, "Category": "General",
                "PcsPerCarton": 1.0,
                "StockCartons": 0.0, "StockPcs": 0.0,
            }
        ]
    )


def admin_wipe_all_products_and_stock() -> None:
    """Remove catalog file rows, inventory, price overrides, and inventory history."""
    with db_conn() as conn:
        conn.execute("DELETE FROM inventory")
        conn.execute("DELETE FROM inventory_txn")
        conn.execute("DELETE FROM product_prices")
    empty = pd.DataFrame(columns=_product_master_excel_columns())
    empty.to_excel(products_master_excel_path(), index=False)
    load_products.clear()


def admin_save_shelf_catalog(edited: pd.DataFrame) -> tuple[int, int]:
    """Persist shelf editor dataframe to Excel, clear DB prices, sync inventory.

    Returns (saved_product_rows, inventory_adjust_rows)."""
    cols = _product_master_excel_columns()
    df = edited.copy()
    for c in cols:
        if c not in df.columns:
            df[c] = 0.0 if c in (
                "Price", "PcsPerCarton", "StockCartons", "StockPcs", "StockTotal",
            ) else ""
    df = df[cols]
    for sc in ["ItemCode", "Barcode", "Name", "Unit", "Category"]:
        df[sc] = df[sc].fillna("").astype(str).str.strip()
    for nc in ["Price", "PcsPerCarton", "StockCartons", "StockPcs", "StockTotal"]:
        df[nc] = pd.to_numeric(df[nc], errors="coerce").fillna(0.0)
    df.loc[df["Category"].str.strip() == "", "Category"] = "General"
    df.loc[df["Unit"].str.strip() == "", "Unit"] = "箱"
    df["StockTotal"] = df["StockCartons"] * df["PcsPerCarton"] + df["StockPcs"]
    mask_name = df["Name"].str.strip() != ""
    df = df[mask_name].copy()
    if df.empty:
        return 0, 0
    with db_conn() as conn:
        conn.execute("DELETE FROM product_prices")
    df.to_excel(products_master_excel_path(), index=False)
    load_products.clear()
    inv_n = import_inventory_from_products(overwrite=True, df=df)
    return len(df), inv_n


def _sheet_price_lookup() -> dict[str, float]:
    """Item_key → Price as stored in products.xlsx (before DB price overlay)."""
    df = _load_products_sheet_for_inventory_import()
    if df.empty:
        return {}
    out: dict[str, float] = {}
    for _, row in df.iterrows():
        ic = str(row.get("ItemCode", "") or "").strip()
        bc = str(row.get("Barcode", "") or "").strip()
        nm = str(row.get("Name", "") or "").strip()
        if not nm:
            continue
        k = _inventory_item_key(ic, bc, nm)
        out[k] = float(row.get("Price", 0) or 0)
    return out


# =========================================================================
# UTILITIES
# =========================================================================
def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def unix_ts() -> float:
    return datetime.now().timestamp()


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _append_app_log(level: str, where: str, detail: str) -> None:
    """Best-effort local runtime log. Never raises."""
    line = f"{now_str()} [{level}] [{where}] {detail}\n"
    try:
        with APP_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass


def log_exception(where: str, exc: Exception) -> None:
    _append_app_log("ERROR", where, f"{type(exc).__name__}: {exc}")


def audit_write(
    event_type: str,
    *,
    order_id: str | None = None,
    branch: str | None = None,
    detail: str | None = None,
    extra: dict | None = None,
) -> None:
    """Append one row to audit_log. Best-effort; does not raise to callers."""
    try:
        created = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        role = str(st.session_state.get("role") or "")
        account_id = st.session_state.get("account_id")
        username = st.session_state.get("account_username")
        if role == Role.WAREHOUSE and not username:
            username = "warehouse"
        elif role == Role.ADMIN and not username:
            username = "admin"
        b = branch if branch is not None else st.session_state.get("branch")
        parts: dict = {}
        if extra:
            parts.update(extra)
        if detail:
            parts["message"] = detail
        detail_json = json.dumps(parts, ensure_ascii=False) if parts else None
        with db_conn() as conn:
            conn.execute(
                """
                INSERT INTO audit_log (
                    created_at, event_type, role, account_id, username,
                    branch, order_id, detail
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    created,
                    event_type,
                    role,
                    account_id,
                    username,
                    b,
                    order_id,
                    detail_json,
                ),
            )
    except Exception as e:
        log_exception("audit_write", e)


def _audit_event_label(code: str) -> str:
    key = {
        "login": "audit_ev_login",
        "logout": "audit_ev_logout",
        "order_submit": "audit_ev_order",
        "receive_confirm": "audit_ev_receive",
        "supplier_order": "audit_ev_supplier",
        "catalog_reset": "audit_ev_catalog_reset",
        "catalog_save": "audit_ev_catalog_save",
    }.get(code)
    return t(key) if key else code


def _format_audit_detail_display(detail_raw: str | None, event_type: str) -> str:
    """Turn stored JSON detail into human-readable text for the admin table."""
    s = (detail_raw or "").strip()
    if not s:
        return ""
    try:
        d: object = json.loads(s)
    except json.JSONDecodeError:
        return s
    if not isinstance(d, dict):
        return s

    parts: list[str] = []

    if d.get("method") == "shared_password":
        pr = str(d.get("portal_role") or "")
        role_lab = {
            Role.WAREHOUSE: t("role_warehouse"),
            Role.ADMIN: t("role_admin"),
            Role.BRANCH: t("role_branch"),
        }.get(pr, pr or "-")
        parts.append(t("audit_d_shared_login").format(role=role_lab))
    elif d.get("method") == "branch_account" and d.get("account_id") is not None:
        parts.append(
            t("audit_d_branch_login").format(id=int(d["account_id"]))
        )

    if event_type == "logout" and d.get("page") is not None:
        parts.append(t("audit_d_logout_page").format(page=str(d["page"])))

    if "line_count" in d and event_type in ("order_submit", "receive_confirm"):
        parts.append(t("audit_d_line_count").format(n=int(d["line_count"])))

    if event_type == "receive_confirm" and "any_shortage" in d:
        parts.append(
            t("audit_d_has_short")
            if d.get("any_shortage")
            else t("audit_d_no_short")
        )

    if event_type == "supplier_order":
        if (d.get("title") or "").strip():
            tit = str(d.get("title") or "").strip()
            parts.append(t("audit_d_supplier_title").format(title=tit[:240]))
        if d.get("lines") is not None:
            parts.append(t("audit_d_supplier_lines").format(n=int(d["lines"])))

    msg = (d.get("message") or "").strip()
    if msg and msg not in parts:
        parts.append(msg)

    out = " · ".join(p for p in parts if p)
    return out if out else s


def _order_fingerprint(branch: str, cart_snapshot: list[dict]) -> str:
    """Stable hash to prevent accidental double-submission by rapid clicks."""
    payload = json.dumps(
        {"branch": branch, "cart": cart_snapshot},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _safe_text(s: str, max_len: int = 1200) -> str:
    txt = (s or "").strip()
    if len(txt) <= max_len:
        return txt
    return txt[:max_len] + "..."


def _build_ai_context() -> str:
    role = st.session_state.get("role") or "-"
    branch = st.session_state.get("branch") or "-"
    lines = [
        f"{t('ai_ctx_role')}: {role}",
        f"{t('ai_ctx_branch')}: {branch}",
    ]
    try:
        with db_conn() as conn:
            open_short = conn.execute(
                "SELECT COUNT(*) FROM shortages WHERE status = 'Open'"
            ).fetchone()[0]
            pending_orders = conn.execute(
                "SELECT COUNT(DISTINCT order_id) FROM orders WHERE status = 'Pending'"
            ).fetchone()[0]
            latest = conn.execute(
                "SELECT order_id, branch, status, order_date "
                "FROM orders ORDER BY order_date DESC LIMIT 5"
            ).fetchall()
        lines.append(f"{t('ai_ctx_open_short')}: {open_short}")
        lines.append(f"{t('ai_ctx_pending_orders')}: {pending_orders}")
        if latest:
            latest_rows = [
                f"- {r['order_id']} | {r['branch']} | {r['status']} | {str(r['order_date'])[:16]}"
                for r in latest
            ]
            lines.append(f"{t('ai_ctx_latest_orders')}:\n" + "\n".join(latest_rows))
    except Exception as e:
        log_exception("build_ai_context", e)
    return _safe_text("\n".join(lines), max_len=1800)


def _normalize_gemini_model_id(model: str) -> str:
    m = (model or "").strip()
    if m.startswith("models/"):
        m = m[len("models/"):]
    return m or "gemini-2.0-flash"


def _build_gemini_model_chain() -> list[str]:
    """Primary model first, then comma-separated fallbacks (deduped)."""
    seen: set[str] = set()
    out: list[str] = []
    parts = [GEMINI_MODEL] + [
        p.strip() for p in (GEMINI_FALLBACK_MODELS or "").split(",") if p.strip()
    ]
    for raw in parts:
        mid = _normalize_gemini_model_id(_coerce_gemini_model_id(raw))
        if mid not in seen:
            seen.add(mid)
            out.append(mid)
    return out


def _extract_gemini_retry_seconds(err_text: str) -> str | None:
    m = re.search(r"Please retry in ([0-9.]+)\s*s", err_text or "", re.I)
    if not m:
        return None
    try:
        return str(int(float(m.group(1)) + 0.999))
    except ValueError:
        return None


def _gemini_app_knowledge_block() -> str:
    """Static UI / workflow facts for the assistant (not user data)."""
    return (
        "[AppKnowledge — SUNSHINE 阳光集团 订货系统 / Streamlit]\n"
        "Language: top-right EN / 中文 toggles UI strings.\n"
        "\n"
        "【分店 branch】Sidebar: 下单, 我的订单, 我的缺货, 🤖 AI助手.\n"
        "- 下单: search (scanner OK), 5 products per page, enter 箱数/个数, "
        "「追加购物车」→「确认并发送订单」→ success page 「订单已发送完毕」.\n"
        "- 我的订单: statuses Pending / Dispatched / Received. "
        "When Dispatched: enter 实收, submit 确认收货; shortages vs dispatch qty may create 缺货 records + email.\n"
        "- 我的缺货: view branch shortages + warehouse reply; Resending → button 确认补发已收到.\n"
        "\n"
        "【仓库 warehouse】Sidebar: 待发货订单, 缺货通知, 出库历史, 🤖 AI助手.\n"
        "- 待发货: filters (branch/date/keyword), per-order dispatch qty, "
        "buttons 按下单数量填充 / 清空, then 标记已发货.\n"
        "- 缺货通知: Open shortages → 补发 or 标记缺货 + reply text.\n"
        "\n"
        "【管理员 admin】Sidebar: 管理概览, 所有订单, 出库发货, 缺货通知, 导出报表, "
        "商品图片, 数据库备份, 邮件通知, 🤖 AI助手.\n"
        "\n"
        "[Order flow] Pending → Dispatched → Received.\n"
        "[Shortage flow] Open → Resending or Out of Stock → Resolved.\n"
    )


def _gemini_error_detail(data: dict) -> str:
    err = data.get("error") or {}
    if isinstance(err, dict):
        msg = (err.get("message") or err.get("status") or "").strip()
        if msg:
            return msg
    pf = data.get("promptFeedback") or {}
    if isinstance(pf, dict):
        br = pf.get("blockReason")
        if br:
            return f"promptFeedback.blockReason={br}"
    return ""


def _call_gemini_single_model(model_id: str, prompt: str) -> tuple[bool, str, int]:
    """One generateContent call. Returns (ok, answer_or_error_detail, http_code).

    http_code is 0 on success, HTTP status on HTTPError, -1 on other errors."""
    endpoint = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{model_id}:generateContent"
    )
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0.35, "maxOutputTokens": 2048},
    }
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        endpoint,
        data=body,
        headers={
            "Content-Type": "application/json",
            "x-goog-api-key": GEMINI_API_KEY,
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=GEMINI_TIMEOUT_SEC) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
        data = json.loads(raw)
        cands = data.get("candidates") or []
        if not cands:
            detail = _gemini_error_detail(data) or json.dumps(
                data, ensure_ascii=False)[:800]
            log_exception(
                "call_gemini_no_candidates",
                Exception(f"{model_id}: {detail}"),
            )
            return False, f"HTTP 200 (no candidates): {_safe_text(detail, 800)}", 200
        parts = (cands[0].get("content") or {}).get("parts") or []
        answer = "\n".join((p.get("text") or "").strip() for p in parts).strip()
        fin = (cands[0].get("finishReason") or "").strip()
        if fin == "MAX_TOKENS" and answer:
            answer += (
                "\n\n---\n⚠️ 输出因长度达到上限被截断；可把问题拆成更短的一句再问。"
                "\n(Output hit max length; try a shorter question.)"
            )
        if not answer:
            detail = _gemini_error_detail(data) or str(cands[0])[:600]
            log_exception(
                "call_gemini_empty_text",
                Exception(f"{model_id}: {detail}"),
            )
            return False, f"HTTP 200 (empty text): {_safe_text(detail, 800)}", 200
        return True, _safe_text(answer, max_len=12000), 0
    except urllib.error.HTTPError as e:
        try:
            detail = e.read().decode("utf-8", errors="replace")
        except Exception:
            detail = ""
        msg = f"HTTP {e.code}: {_safe_text(detail, 1500)}"
        log_exception("call_gemini_http", Exception(f"{model_id} {msg}"))
        return False, msg, int(e.code)
    except Exception as e:
        log_exception("call_gemini_exc", Exception(f"{model_id}: {e}"))
        return False, f"{type(e).__name__}: {e}", -1


def _call_gemini_api(question: str, context_text: str) -> tuple[bool, str]:
    if not GEMINI_API_KEY:
        return False, t("ai_missing_key")
    ui_lang = st.session_state.get("lang") or "zh"
    # Answer language follows the user's question, not the UI toggle (users may
    # keep UI in 中文 but ask in English).
    lang_rule = (
        "【语言 / Language】Answer in the same primary language as [UserQuestion]: "
        "if the user writes mainly in English, the entire reply must be in English; "
        "if mainly in Chinese (简体), the entire reply must be in 中文（简体）. "
        "For mixed questions, follow the dominant language. "
        "When naming sidebar entries, keep the real on-screen Chinese labels; in an English reply you may add "
        "a short English gloss in parentheses once (e.g. 待发货订单 (pending dispatch queue))."
    )
    weekly_data_rule = (
        "【经营数据】你现在拥有最近一周的经营摘要数据（见 [SystemContext]：近7日已发货出库量排名与当前 "
        "status=Open 的缺货列表）。若用户询问出库情况、热点商品或需要处理建议，请结合这些数据给出"
        "具体的统计回答与补货/处理建议。"
        if ui_lang == "zh"
        else "[Ops data] You have a 7-day dispatched outbound summary and open shortage rows in "
        "[SystemContext]. For outbound volume, hot SKUs, or replenishment/handling advice, cite "
        "those figures and give concrete guidance."
    )
    system_text = "\n".join([
        t("ai_role_only"),
        lang_rule,
        "",
        _gemini_app_knowledge_block().strip(),
        "",
        weekly_data_rule,
        "",
        "【回答规则 / Answer rules】",
        "1) 先直接回答用户问题：用编号步骤（1. 2. 3.），写清楚要点，不要只报数字。",
        "2) 步骤里必须指向本系统侧边栏真实入口：中文回答写界面中文全称；英文回答用英文叙述步骤，并写出中文入口名（可加括号英译，如 我的订单 (My orders)）。",
        "3) [SystemContext] 只是当前数据库快照。若其中待办数量为 0，只用一句话说明「当前无待办」，",
        "   然后仍然完整写出「有数据时」应如何操作的全程步骤，不得用 0 代替教程。",
        "4) 禁止用「当前没有…」作为唯一答案结束流程类问题；流程类问题至少 8 行有效说明（可含小标题）。",
        "5) 句子写完整，禁止写到一半断开；除导航标注规则外，不要中英无意义来回切换。",
        "6) 若问题超出本应用能力，明确说不知道，不要编造。",
        "7) 不要声称已在系统中替用户点击或保存任何操作。",
        "",
        "Never claim an action has been executed in the database.",
    ])
    prompt = (
        f"{system_text}\n\n"
        f"[SystemContext — 仅供参考]\n{_safe_text(context_text, 2800)}\n\n"
        f"[UserQuestion]\n{_safe_text(question, 1200)}"
    )
    chain = _build_gemini_model_chain()
    last_err = ""
    last_http = 0
    for idx, mid in enumerate(chain):
        ok, msg, http = _call_gemini_single_model(mid, prompt)
        if ok:
            if idx > 0:
                msg = t("ai_model_used").format(model=mid) + "\n\n" + msg
            return True, msg
        last_err = msg
        last_http = http
        if http in (429, 503):
            continue
        break

    if last_http == 429 or "429" in last_err or "RESOURCE_EXHAUSTED" in last_err:
        wait = _extract_gemini_retry_seconds(last_err)
        lines = [
            t("ai_error"),
            "",
            t("ai_quota_429"),
            t("ai_quota_actions"),
            t("ai_quota_links"),
        ]
        if wait:
            lines.insert(3, f"Google 建议约 {wait} 秒后再重试。 / Retry after ~{wait}s.")
        lines.append("")
        lines.append(f"({_safe_text(last_err, 2000)})")
        return False, "\n\n".join(lines)

    err_l = (last_err or "").lower()
    # Google may disable keys exposed in repos/chats; message contains "leaked".
    if last_http == 403 and "leaked" in err_l:
        return False, "\n\n".join([
            t("ai_key_leaked_403"),
            "",
            t("ai_key_leaked_actions"),
            "",
            f"({_safe_text(last_err, 2000)})",
        ])

    if last_http == 400 and (
        "api key not valid" in err_l
        or "api_key_invalid" in err_l
        or "invalid api key" in err_l
    ):
        return False, "\n\n".join([
            t("ai_key_invalid_400"),
            "",
            t("ai_key_invalid_actions"),
            "",
            f"({_safe_text(last_err, 2000)})",
        ])

    return False, f"{t('ai_error')}\n\n({_safe_text(last_err, 1500)})"


def gen_order_id(branch: str) -> str:
    safe = "".join(c for c in branch if c.isalnum()).upper()[:10]
    return f"{safe}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"


# --- Product images -------------------------------------------------------
# Simple file-on-disk approach. No DB schema, no compression, no resizing —
# just save the raw upload under a predictable filename. The browser scales
# the image at render time, so this is "fast enough" for a few hundred
# products. If we ever need to optimize, we can add Pillow compression
# without changing the public helpers below.
def _safe_filename_part(s: str) -> str:
    """Strip path separators / weird characters from a string before
    using it as a filename. Defends against weird ItemCode values like
    'A/B' or '../sneaky'."""
    return "".join(c for c in str(s) if c.isalnum() or c in "-_")


def get_product_image_path(item_code: str = "", barcode: str = "") -> Path | None:
    """Return the path to this product's image if one exists, else None.

    Look-up order:
      1. <ItemCode>.<ext> for any supported extension
      2. bc_<Barcode>.<ext> as fallback
    First match wins."""
    if not PRODUCT_IMAGES_DIR.exists():
        return None
    ic = _safe_filename_part(item_code)
    bc = _safe_filename_part(barcode)
    for stem in [ic, f"bc_{bc}" if bc else ""]:
        if not stem:
            continue
        for ext in IMAGE_EXTS:
            candidate = PRODUCT_IMAGES_DIR / f"{stem}{ext}"
            if candidate.exists():
                return candidate
    return None


def save_product_image(item_code: str, barcode: str, name: str,
                       uploaded_bytes: bytes,
                       original_filename: str) -> Path:
    """Save raw image bytes under a stable filename. Returns the saved path.

    File naming priority: ItemCode → barcode → safe-name fallback.
    Extension comes from the original filename; falls back to .jpg."""
    PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    # Pick filename stem
    ic = _safe_filename_part(item_code)
    bc = _safe_filename_part(barcode)
    if ic:
        stem = ic
    elif bc:
        stem = f"bc_{bc}"
    else:
        # Last resort: derive from name (manual products, unlikely path)
        stem = "name_" + (_safe_filename_part(name) or "noname")[:40]

    # Determine extension from upload, default to .jpg
    ext = Path(original_filename).suffix.lower()
    if ext not in IMAGE_EXTS:
        ext = ".jpg"

    # Remove any older version of this product's image (any extension)
    for old_ext in IMAGE_EXTS:
        old = PRODUCT_IMAGES_DIR / f"{stem}{old_ext}"
        if old.exists():
            try:
                old.unlink()
            except OSError:
                pass

    dest = PRODUCT_IMAGES_DIR / f"{stem}{ext}"
    dest.write_bytes(uploaded_bytes)
    return dest


def delete_product_image(item_code: str = "", barcode: str = "") -> bool:
    """Delete this product's image if it exists. Returns True on deletion."""
    p = get_product_image_path(item_code, barcode)
    if p is None:
        return False
    try:
        p.unlink()
        return True
    except OSError:
        return False


# --- Auto-scan root directory for new images -----------------------------
# User-friendly drop zone: drop image files directly next to app.py and the
# system will move matching ones into product_images/ on startup or on
# demand. Files that don't match any product stay where they are so the
# user can see and fix them. We move (not copy) to avoid duplicates.
def scan_and_import_root_images(verbose: bool = False) -> dict:
    """Scan the working directory (where app.py lives) for image files
    whose names match an ItemCode or Barcode in the master, and move them
    into product_images/. Returns a summary dict with counts + lists.

    Files that don't match are left untouched, so the user can see exactly
    which filenames need fixing. Files inside product_images/ itself are
    skipped (already imported)."""
    summary = {"moved": [], "skipped_unmatched": [], "errors": []}

    # The "root" we scan is the current working directory — that's where
    # streamlit launches app.py from, so it's also where the user dragged
    # the images in.
    root = Path(".")
    if not root.exists():
        return summary

    products_df = load_products(_products_mtime(), _inventory_version(), _price_version())

    PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    # Iterate just the top level; don't recurse into subfolders. We don't
    # want to "discover" images that live in /backups/ etc.
    for entry in root.iterdir():
        if not entry.is_file():
            continue
        if entry.suffix.lower() not in IMAGE_EXTS:
            continue
        # Don't touch files inside product_images/ — already managed.
        # (We only iterate the top level, but be defensive anyway.)
        try:
            if entry.resolve().parent == PRODUCT_IMAGES_DIR.resolve():
                continue
        except OSError:
            continue

        m = match_filename_to_product(entry.name, products_df)
        if m is None:
            summary["skipped_unmatched"].append(entry.name)
            continue

        try:
            dest = save_product_image(
                item_code=m["item_code"],
                barcode=m["barcode"],
                name=m["name"],
                uploaded_bytes=entry.read_bytes(),
                original_filename=entry.name,
            )
            # Now that the image is safely in product_images/, remove the
            # source file. If unlink fails (perm denied, file in use), we
            # log it but the import already succeeded.
            try:
                entry.unlink()
            except OSError as e:
                summary["errors"].append(f"{entry.name}: copied but couldn't remove original — {e}")
            summary["moved"].append({
                "from": entry.name,
                "to":   dest.name,
                "matched_by": m["matched_by"],
                "name": m["name"],
            })
        except Exception as e:
            summary["errors"].append(f"{entry.name}: {e}")

    return summary


# --- Batch image upload --------------------------------------------------
# Pair uploaded files with products by filename. We accept either:
#   <ItemCode>.<ext>     (e.g. P001.jpg)
#   <Barcode>.<ext>      (e.g. 8801234567001.jpg)
# The stem (filename without extension) is matched against the products
# master. First exact match wins. Trailing/leading whitespace and case are
# ignored for ItemCode matching; Barcode matching is exact (digits only).
def match_filename_to_product(filename: str,
                              products_df: pd.DataFrame) -> dict | None:
    """Look up `filename` (e.g. 'P001.jpg') in the products master.
    Returns a dict with item_code, barcode, name, matched_by — or None."""
    if products_df is None or products_df.empty:
        return None
    stem = Path(filename).stem.strip()
    if not stem:
        return None

    # Try ItemCode first (case-insensitive)
    stem_lower = stem.lower()
    for _, row in products_df.iterrows():
        ic = str(row.get("ItemCode", "") or "").strip()
        if ic and ic.lower() == stem_lower:
            return {
                "item_code": ic,
                "barcode":   str(row.get("Barcode", "") or "").strip(),
                "name":      str(row.get("Name", "") or ""),
                "matched_by": "ItemCode",
            }

    # Then Barcode (exact)
    for _, row in products_df.iterrows():
        bc = str(row.get("Barcode", "") or "").strip()
        if bc and bc == stem:
            return {
                "item_code": str(row.get("ItemCode", "") or "").strip(),
                "barcode":   bc,
                "name":      str(row.get("Name", "") or ""),
                "matched_by": "Barcode",
            }

    return None


def plan_batch_image_upload(uploaded_files: list,
                            products_df: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    """Inspect a list of uploaded files and split them into
    (matched, unmatched). Each entry is a dict so the UI can render previews.

    `uploaded_files` items must have .name and .getvalue() (the standard
    Streamlit UploadedFile interface)."""
    matched: list[dict] = []
    unmatched: list[dict] = []
    for uf in uploaded_files:
        fname = getattr(uf, "name", "") or ""
        ext = Path(fname).suffix.lower()
        # Skip files with totally non-image extensions to avoid surprises.
        if ext and ext not in IMAGE_EXTS:
            unmatched.append({
                "filename": fname,
                "reason":   f"unsupported extension {ext}",
                "uf":       uf,
            })
            continue
        m = match_filename_to_product(fname, products_df)
        if m is None:
            unmatched.append({
                "filename": fname,
                "reason":   "no matching ItemCode or Barcode",
                "uf":       uf,
            })
        else:
            matched.append({
                "filename":   fname,
                "item_code":  m["item_code"],
                "barcode":    m["barcode"],
                "name":       m["name"],
                "matched_by": m["matched_by"],
                "uf":         uf,
            })
    return matched, unmatched


# --- Notification builders ------------------------------------------------
# Each helper returns (subject, body) ready to pass to notify(). Bodies are
# plain text — easy to read on phones, no HTML rendering issues.
def _format_lines_table(lines: list[dict]) -> str:
    """Render order lines as a fixed-width text table for the email body."""
    if not lines:
        return "  (no items)"
    # Compute column widths
    name_w = max(4, min(40, max(len(str(it.get("name", ""))) for it in lines)))
    out = []
    out.append(f"  {'Name':<{name_w}}  {'Cartons':>8}  {'Pcs':>6}  {'Unit':<8}")
    out.append(f"  {'-' * name_w}  {'-' * 8}  {'-' * 6}  {'-' * 8}")
    for it in lines:
        nm = str(it.get("name", ""))[:name_w]
        ct = it.get("qty_cartons", 0)
        pc = it.get("qty_pcs", 0)
        un = str(it.get("unit", ""))[:8]
        out.append(f"  {nm:<{name_w}}  {ct:>8}  {pc:>6}  {un:<8}")
    return "\n".join(out)


def build_new_order_email(order_id: str, branch: str,
                          lines: list[dict]) -> tuple[str, str]:
    subject = f"[新订单/New Order] {branch} · {order_id}"
    total_ct = sum(int(it.get("qty_cartons", 0)) for it in lines)
    total_pc = sum(int(it.get("qty_pcs", 0)) for it in lines)
    body = (
        f"📦 新订单 / New Order\n"
        f"=====================================\n"
        f"分店 / Branch:   {branch}\n"
        f"订单号 / Order:  {order_id}\n"
        f"时间 / Time:     {now_str()}\n"
        f"行数 / Lines:    {len(lines)}\n"
        f"合计 / Totals:   📦 {total_ct} cartons · 🔢 {total_pc} pcs\n"
        f"\n"
        f"商品明细 / Items:\n"
        f"{_format_lines_table(lines)}\n"
        f"\n"
        f"—— SUNSHINE 阳光集团 订货系统\n"
    )
    return subject, body


def build_new_order_excel(order_id: str, branch: str, lines: list[dict]) -> bytes:
    """Build an Excel attachment for new-order notification (per-order)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "NewOrder"
    ws.append([
        "Order ID",
        "Branch",
        "Name",
        "Unit",
        "Cartons",
        "Pcs",
    ])
    for it in lines:
        ws.append([
            order_id,
            branch,
            str(it.get("name", "") or ""),
            str(it.get("unit", "") or ""),
            int(it.get("qty_cartons", 0)),
            int(it.get("qty_pcs", 0)),
        ])
    _style_excel_header(ws, 6)
    _autosize(ws)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def build_supplier_order_excel(lines: list[dict], remarks: str, ts: str) -> bytes:
    """Excel attachment for warehouse → supplier order (product lines + remarks)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SupplierOrder"
    ws.append([
        "ItemCode",
        "Barcode",
        "Name",
        "Unit",
        "Cartons",
        "Pcs",
    ])
    for it in lines:
        ws.append([
            str(it.get("item_code", "") or ""),
            str(it.get("barcode", "") or ""),
            str(it.get("name", "") or ""),
            str(it.get("unit", "") or ""),
            int(it.get("qty_cartons", 0) or 0),
            int(it.get("qty_pcs", 0) or 0),
        ])
    if (remarks or "").strip():
        ws.append([])
        ws.append(["Remarks / 备注", (remarks or "").strip(), "", "", "", ""])
    ws.append([])
    ws.append(["Generated / 生成时间", ts, "", "", "", ""])
    _style_excel_header(ws, 6)
    _autosize(ws)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def _safe_excel_filename(title: str) -> str:
    s = (title or "").strip()
    for ch in '\\/:*?"<>|':
        s = s.replace(ch, "_")
    s = " ".join(s.split())
    if not s:
        s = "supplier_order"
    return s[:120]


def build_dispatched_email(order_id: str, branch: str,
                           lines: list[dict]) -> tuple[str, str]:
    """`lines` should each have qty_cartons / qty_pcs (=ordered) and
    dispatch_cartons / dispatch_pcs (=actually shipped)."""
    subject = f"[已发货/Dispatched] {branch} · {order_id}"
    name_w = max(4, min(40, max(len(str(it.get("name", ""))) for it in lines))) if lines else 20
    rows = [
        f"  {'Name':<{name_w}}  {'Ordered':>10}  {'Shipped':>10}  {'Note':<10}",
        f"  {'-' * name_w}  {'-' * 10}  {'-' * 10}  {'-' * 10}",
    ]
    has_short = False
    for it in lines:
        nm = str(it.get("name", ""))[:name_w]
        oct_ = int(it.get("qty_cartons", 0))
        opc  = int(it.get("qty_pcs", 0))
        dct  = int(it.get("dispatch_cartons", oct_) or 0)
        dpc  = int(it.get("dispatch_pcs", opc) or 0)
        flag = ""
        if dct < oct_ or dpc < opc:
            flag = "SHORT"
            has_short = True
        rows.append(
            f"  {nm:<{name_w}}  "
            f"{oct_:>4}c {opc:>3}p  "
            f"{dct:>4}c {dpc:>3}p  "
            f"{flag:<10}"
        )
    body = (
        f"🚚 已发货 / Dispatched\n"
        f"=====================================\n"
        f"分店 / Branch:   {branch}\n"
        f"订单号 / Order:  {order_id}\n"
        f"时间 / Time:     {now_str()}\n"
        f"\n"
        + ("⚠️  部分商品库存不足，发货数量少于订单数量 / "
           "Some items shipped short of ordered qty\n\n" if has_short else "")
        + "\n".join(rows)
        + f"\n\n—— SUNSHINE 阳光集团 订货系统\n"
    )
    return subject, body


def build_dispatched_excel(order_id: str, branch: str, lines: list[dict]) -> bytes:
    """Build an Excel attachment for dispatched notice (per-order)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispatched"
    ws.append([
        "Order ID",
        "Branch",
        "Name",
        "Unit",
        "Ordered Cartons",
        "Ordered Pcs",
        "Shipped Cartons",
        "Shipped Pcs",
        "Note",
    ])
    for it in lines:
        oct_ = int(it.get("qty_cartons", 0))
        opc = int(it.get("qty_pcs", 0))
        dct = int(it.get("dispatch_cartons", oct_) or 0)
        dpc = int(it.get("dispatch_pcs", opc) or 0)
        note = "SHORT" if (dct < oct_ or dpc < opc) else ""
        ws.append([
            order_id,
            branch,
            str(it.get("name", "") or ""),
            str(it.get("unit", "") or ""),
            oct_,
            opc,
            dct,
            dpc,
            note,
        ])
    _style_excel_header(ws, 9)
    _autosize(ws)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def build_shortage_email(order_id: str, branch: str,
                         shortages: list[dict]) -> tuple[str, str]:
    """`shortages` items should each have barcode, name, short_cartons, short_pcs."""
    subject = f"[缺货通知/Shortage] {branch} · {order_id}"
    rows = []
    for s in shortages:
        barcode = (s.get("barcode") or "").strip() or "-"
        rows.append(
            f"  • [{barcode}] {s.get('name', '')}: "
            f"📦 短 {s.get('short_cartons', 0)} cartons · "
            f"🔢 短 {s.get('short_pcs', 0)} pcs"
        )
    body = (
        f"🔔 缺货通知 / Shortage Reported\n"
        f"=====================================\n"
        f"分店 / Branch:   {branch}\n"
        f"订单号 / Order:  {order_id}\n"
        f"时间 / Time:     {now_str()}\n"
        f"\n"
        f"短少明细 / Items short:\n"
        + "\n".join(rows)
        + f"\n\n请尽快处理（补发或标记缺货）。\n"
        f"Please action — resend or mark out of stock.\n"
        f"\n—— SUNSHINE 阳光集团 订货系统\n"
    )
    return subject, body


# =========================================================================
# CSS
# =========================================================================
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap');

/* ----- App shell (full viewport) ----- */
.stApp {
    font-family: "Plus Jakarta Sans", ui-sans-serif, system-ui, sans-serif;
    background-color: #f1f5f9;
    background-image:
        radial-gradient(ellipse 120% 80% at 100% -10%, rgba(59, 130, 246, 0.09), transparent 50%),
        radial-gradient(ellipse 90% 60% at -15% 40%, rgba(14, 165, 233, 0.07), transparent 45%),
        radial-gradient(ellipse 70% 50% at 80% 100%, rgba(251, 191, 36, 0.06), transparent 40%);
    background-attachment: fixed;
}

section[data-testid="stMain"] > div {
    padding-left: 0.65rem;
    padding-right: 0.65rem;
}
@media (min-width: 768px) {
    section[data-testid="stMain"] > div {
        padding-left: 1.1rem;
        padding-right: 1.1rem;
    }
}

/* Main column: floating “page” card */
section[data-testid="stMain"] .block-container {
    padding-top: 2.5rem;
    padding-bottom: 2.75rem;
    max-width: 1180px;
    background: linear-gradient(
        165deg,
        rgba(255, 255, 255, 0.97) 0%,
        rgba(248, 250, 252, 0.96) 100%
    );
    backdrop-filter: saturate(180%) blur(12px);
    border-radius: 22px;
    border: 1px solid rgba(148, 163, 184, 0.35);
    box-shadow:
        0 0 0 1px rgba(255, 255, 255, 0.8) inset,
        0 1px 2px rgba(15, 23, 42, 0.04),
        0 24px 56px -16px rgba(30, 64, 175, 0.12),
        0 12px 32px -18px rgba(15, 23, 42, 0.08);
}
@media (max-width: 640px) {
    section[data-testid="stMain"] .block-container {
        padding-top: max(3.35rem, 2rem);
        padding-left: 0.75rem;
        padding-right: 0.75rem;
        padding-bottom: 2rem;
        border-radius: 14px;
        margin-left: 0.15rem;
        margin-right: 0.15rem;
    }
}

/* Markdown headings in main workspace */
section[data-testid="stMain"] h1,
section[data-testid="stMain"] h2,
section[data-testid="stMain"] h3 {
    color: #0f172a;
    font-weight: 700;
    letter-spacing: -0.02em;
}
section[data-testid="stMain"] h3 {
    font-size: 1.28rem;
    margin-top: 0.25rem;
    margin-bottom: 1rem;
    padding-bottom: 0.4rem;
    border-bottom: 2px solid rgba(25, 118, 210, 0.32);
}
section[data-testid="stMain"] h4 {
    font-size: 1.05rem;
    color: #334155;
    font-weight: 650;
    margin-top: 1.2rem;
    margin-bottom: 0.55rem;
    padding: 0.35rem 0 0.35rem 0.65rem;
    border-left: 3px solid #42a5f5;
    background: linear-gradient(90deg, rgba(66,165,245,0.08), transparent);
    border-radius: 0 8px 8px 0;
}

/* Page hero (HTML titles from render_page_heading) */
.page-hero {
    position: relative;
    overflow: hidden;
    background: linear-gradient(
        125deg,
        #0f172a 0%,
        #1e3a5f 42%,
        #1e40af 78%,
        #2563eb 100%
    );
    border: none;
    border-radius: 16px;
    padding: 22px 26px 22px 26px;
    margin: 0 0 1.5rem 0;
    box-shadow:
        0 4px 6px -1px rgba(15, 23, 42, 0.12),
        0 20px 40px -12px rgba(30, 58, 138, 0.45),
        0 0 0 1px rgba(255, 255, 255, 0.08) inset;
}
.page-hero::before {
    content: "";
    position: absolute;
    inset: 0;
    background: radial-gradient(
        ellipse 70% 120% at 100% 0%,
        rgba(251, 191, 36, 0.15),
        transparent 55%
    );
    pointer-events: none;
}
.page-hero::after {
    content: "";
    position: absolute;
    right: -20%;
    top: -40%;
    width: 55%;
    height: 160%;
    background: radial-gradient(circle, rgba(255,255,255,0.12) 0%, transparent 65%);
    pointer-events: none;
}
.page-hero-text {
    position: relative;
    z-index: 1;
}
.page-hero-title {
    margin: 0;
    font-size: 1.55rem;
    font-weight: 700;
    color: #f8fafc;
    letter-spacing: -0.03em;
    line-height: 1.2;
    text-shadow: 0 1px 2px rgba(0, 0, 0, 0.15);
}
.page-hero-desc {
    margin: 0.65rem 0 0 0;
    font-size: 0.92rem;
    color: rgba(226, 232, 240, 0.92);
    line-height: 1.6;
    max-width: 52rem;
}

/* In-page section strip (render_section_title) */
.section-title {
    margin: 1.55rem 0 0.85rem 0;
}
.section-title span {
    display: inline-flex;
    align-items: center;
    gap: 0.5rem;
    font-size: 1.06rem;
    font-weight: 700;
    color: #0f172a;
    letter-spacing: -0.02em;
    padding: 0.45rem 1rem 0.45rem 0.85rem;
    background: linear-gradient(90deg, rgba(241, 245, 249, 0.95), rgba(255, 255, 255, 0.4));
    border-radius: 10px;
    border: 1px solid rgba(148, 163, 184, 0.35);
    border-left: 4px solid #2563eb;
    box-shadow: 0 1px 3px rgba(15, 23, 42, 0.05);
}
.section-title span::before {
    content: "";
    width: 6px;
    height: 6px;
    border-radius: 50%;
    background: linear-gradient(135deg, #3b82f6, #60a5fa);
    flex-shrink: 0;
}

/* ---------- Blue gradient header ---------- */
.sunshine-header {
    position: relative;
    overflow: hidden;
    background: linear-gradient(
        115deg,
        #0c1929 0%,
        #132e52 35%,
        #1d4ed8 72%,
        #0ea5e9 100%
    );
    padding: 20px 24px;
    border-radius: 16px;
    color: #fff;
    margin-bottom: 18px;
    box-shadow:
        0 4px 6px -1px rgba(15, 23, 42, 0.15),
        0 16px 40px -8px rgba(29, 78, 216, 0.35);
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 10px;
    border: 1px solid rgba(255, 255, 255, 0.12);
}
.sunshine-header::after {
    content: "";
    position: absolute;
    inset: 0;
    background: radial-gradient(
        ellipse 80% 100% at 90% -20%,
        rgba(251, 191, 36, 0.18),
        transparent 50%
    );
    pointer-events: none;
}
.sunshine-header > div:first-child {
    position: relative;
    z-index: 1;
}
.sunshine-header .title-main {
    font-size: 23px;
    font-weight: 700;
    line-height: 1.2;
    letter-spacing: -0.02em;
    display: flex;
    align-items: center;
    gap: 10px;
}
.sunshine-header .title-sub {
    font-size: 13px;
    opacity: 0.88;
    margin-top: 4px;
    font-weight: 500;
}
.sunshine-header .title-context {
    position: relative;
    z-index: 1;
    font-size: 13px;
    opacity: 0.95;
    font-weight: 600;
    background: rgba(255, 255, 255, 0.14);
    padding: 6px 12px;
    border-radius: 999px;
    border: 1px solid rgba(255, 255, 255, 0.2);
    backdrop-filter: blur(6px);
}
.sunshine-header .title-meta {
    margin-top: 6px;
    max-width: 640px;
    color: rgba(255,255,255,.95);
}
.sunshine-header .ticker-wrap {
    overflow: hidden;
    white-space: nowrap;
    font-size: 12px;
    background: rgba(255,255,255,.16);
    border-radius: 8px;
    padding: 3px 8px;
}
.sunshine-header .ticker-text {
    display: inline-block;
    padding-left: 100%;
    animation: ticker-scroll 14s linear infinite;
}
@keyframes ticker-scroll {
    0% { transform: translateX(0); }
    100% { transform: translateX(-100%); }
}

/* ---------- Shortage badge ---------- */
.short-badge {
    background: #e53935; color: #fff;
    border-radius: 12px; padding: 2px 9px;
    font-size: 12px; font-weight: 700; margin-left: 6px;
    box-shadow: 0 0 0 2px rgba(255,255,255,.3);
    animation: pulse 1.6s infinite;
}
.short-badge-link {
    text-decoration: none;
    cursor: pointer;
}
.short-badge-link:hover .short-badge {
    filter: brightness(1.07);
}
@keyframes pulse {
    0%, 100% { transform: scale(1);   }
    50%      { transform: scale(1.08);}
}

/* ---------- Status pills ---------- */
.pill {
    display: inline-block; padding: 3px 10px;
    border-radius: 12px; font-size: 12px; font-weight: 600;
    color: #fff; line-height: 1.4;
}
.pill-Pending      { background: #fbc02d; color: #5d4500; }
.pill-Dispatched   { background: #1976d2; }
.pill-Received     { background: #388e3c; }
.pill-Open         { background: #fbc02d; color: #5d4500; }
.pill-Resending    { background: #1976d2; }
.pill-OutOfStock   { background: #e53935; }
.pill-Resolved     { background: #388e3c; }

/* ---------- Metric card ---------- */
.metric-card {
    background: linear-gradient(145deg, #ffffff 0%, #f8fafc 100%);
    border: 1px solid rgba(148, 163, 184, 0.35);
    border-left: 4px solid #2563eb;
    padding: 16px 18px;
    border-radius: 14px;
    box-shadow:
        0 1px 3px rgba(15, 23, 42, 0.04),
        0 8px 24px -8px rgba(30, 64, 175, 0.12);
    color: #0f172a;
}
.metric-card .label {
    font-size: 12px;
    font-weight: 600;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    color: #64748b;
}
.metric-card .value {
    font-size: 28px;
    font-weight: 700;
    color: #1d4ed8;
    margin-top: 6px;
    letter-spacing: -0.02em;
}
.metric-card.warn  { border-left-color: #eab308; }
.metric-card.warn  .value { color: #ca8a04; }
.metric-card.ok    { border-left-color: #16a34a; }
.metric-card.ok    .value { color: #15803d; }
.metric-card.alert { border-left-color: #dc2626; }
.metric-card.alert .value { color: #b91c1c; }

/* ---------- Sidebar ---------- */
section[data-testid="stSidebar"] {
    background: linear-gradient(195deg, #0f172a 0%, #1e293b 55%, #1e3a5f 100%);
    border-right: 1px solid rgba(148, 163, 184, 0.2);
    box-shadow: 4px 0 24px rgba(15, 23, 42, 0.12);
}
section[data-testid="stSidebar"] .block-container {
    padding-top: 1.6rem;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] span {
    color: rgba(241, 245, 249, 0.92) !important;
}
section[data-testid="stSidebar"] .stButton > button {
    width: 100%;
    text-align: left;
    justify-content: flex-start;
    border-radius: 12px;
    font-weight: 600;
    border: 1px solid rgba(148, 163, 184, 0.25) !important;
    background: rgba(255, 255, 255, 0.06) !important;
    color: #e2e8f0 !important;
    transition: transform 0.15s ease, box-shadow 0.15s ease, border-color 0.15s ease, background 0.15s ease;
}
section[data-testid="stSidebar"] .stButton > button:hover {
    transform: translateY(-1px);
    background: rgba(255, 255, 255, 0.12) !important;
    border-color: rgba(96, 165, 250, 0.45) !important;
    box-shadow: 0 4px 14px rgba(0, 0, 0, 0.2);
}
section[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%) !important;
    border: 1px solid rgba(147, 197, 253, 0.45) !important;
    color: #fff !important;
    box-shadow:
        0 2px 8px rgba(37, 99, 235, 0.35),
        0 0 0 1px rgba(255, 255, 255, 0.12) inset;
}
section[data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%) !important;
}
section[data-testid="stSidebar"] hr {
    border: none;
    border-top: 1px solid rgba(148, 163, 184, 0.25);
    margin: 1rem 0;
}

/* ---------- Sidebar branch identity card ---------- */
.branch-identity {
    margin: 6px 0 8px 0;
    padding: 10px 12px;
    border-radius: 10px;
    border: 1px solid #90caf9;
    background: linear-gradient(180deg, #e3f2fd 0%, #f5fbff 100%);
}
.branch-identity .k {
    font-size: 12px;
    color: #355070;
    font-weight: 600;
    margin-bottom: 4px;
}
.branch-identity .v {
    font-size: 20px;
    line-height: 1.25;
    font-weight: 800;
    color: #0d47a1;
    letter-spacing: 0.3px;
    word-break: break-word;
}
section[data-testid="stSidebar"] .branch-identity {
    border: 1px solid rgba(96, 165, 250, 0.35);
    background: linear-gradient(
        165deg,
        rgba(30, 58, 138, 0.5) 0%,
        rgba(15, 23, 42, 0.55) 100%
    );
    box-shadow: 0 4px 14px rgba(0, 0, 0, 0.15);
}
section[data-testid="stSidebar"] .branch-identity .k {
    color: #93c5fd;
}
section[data-testid="stSidebar"] .branch-identity .v {
    color: #f8fafc;
}
@media (max-width: 640px) {
    .branch-identity .v {
        font-size: 18px;
    }
}

/* ---------- Buttons (main) ---------- */
section[data-testid="stMain"] button[kind="primary"] {
    background: linear-gradient(90deg, #1565c0, #1e88e5) !important;
    border: none !important;
    border-radius: 10px !important;
    box-shadow: 0 2px 10px rgba(25, 118, 210, 0.3);
}
section[data-testid="stMain"] button[kind="primary"]:hover {
    background: linear-gradient(90deg, #0d47a1, #1565c0) !important;
}
section[data-testid="stMain"] button[kind="secondary"] {
    border-radius: 10px !important;
    border-color: #cbd5e1 !important;
}

/* ---------- Forms & data display ---------- */
section[data-testid="stMain"] .stTextInput input,
section[data-testid="stMain"] .stTextArea textarea {
    border-radius: 10px !important;
}
section[data-testid="stMain"] div[data-testid="stDataFrame"] {
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid rgba(148, 163, 184, 0.4);
    box-shadow: 0 2px 12px rgba(15, 23, 42, 0.05);
}

section[data-testid="stMain"] [data-baseweb="select"] > div {
    border-radius: 10px !important;
}
section[data-testid="stMain"] hr {
    margin: 1.5rem 0;
    border: none;
    height: 1px;
    background: linear-gradient(
        90deg,
        transparent,
        rgba(148, 163, 184, 0.55) 15%,
        rgba(59, 130, 246, 0.35) 50%,
        rgba(148, 163, 184, 0.55) 85%,
        transparent
    );
}

/* Alerts & notices */
section[data-testid="stMain"] div[data-testid="stAlert"] {
    border-radius: 12px !important;
    border: 1px solid rgba(148, 163, 184, 0.35) !important;
    box-shadow: 0 2px 12px rgba(15, 23, 42, 0.06);
}

/* Expanders */
section[data-testid="stMain"] details,
section[data-testid="stMain"] div[data-testid="stExpander"] {
    border-radius: 12px !important;
    border: 1px solid rgba(148, 163, 184, 0.35) !important;
    background: rgba(248, 250, 252, 0.72) !important;
    box-shadow: 0 1px 3px rgba(15, 23, 42, 0.04);
}
section[data-testid="stMain"] details summary {
    font-weight: 600 !important;
    color: #0f172a !important;
}

/* Tabs / pills / captions — calmer hierarchy */
section[data-testid="stMain"] [data-testid="stCaptionContainer"] {
    color: #64748b !important;
}

/* Radio groups (module switchers) */
section[data-testid="stMain"] [data-testid="stRadio"] label {
    font-weight: 500 !important;
}

/* Sidebar nav section — supply chain group */
.nav-section-head {
    font-size: 0.78rem;
    font-weight: 700;
    letter-spacing: 0.04em;
    text-transform: uppercase;
    color: #64748b;
    margin: 0.35rem 0 0.25rem 0;
    padding: 0.2rem 0.1rem;
}
.nav-section-gap {
    height: 0.35rem;
}

/* “Current page” strip under sidebar (see _render_active_page_hint) */
.active-page-hint {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 0.4rem 0.75rem;
    margin: 4px 0 14px 0;
    padding: 10px 14px;
    border-radius: 12px;
    background: linear-gradient(
        120deg,
        rgba(255, 255, 255, 0.92) 0%,
        rgba(241, 245, 249, 0.88) 100%
    );
    border: 1px solid rgba(148, 163, 184, 0.35);
    border-left: 4px solid var(--hint-accent, #2563eb);
    box-shadow: 0 2px 10px rgba(15, 23, 42, 0.05);
}
.active-page-hint__k {
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    color: #64748b;
}
.active-page-hint__v {
    font-size: 14px;
    font-weight: 700;
    color: #0f172a;
    letter-spacing: -0.01em;
}

/* ----- st.container(border=True): soft shadow + rounded corners ----- */
section[data-testid="stMain"] [data-testid="stVerticalBlockBorderWrapper"] {
    border-radius: 12px !important;
    overflow: hidden;
    box-shadow:
        0 1px 2px rgba(15, 23, 42, 0.04),
        0 4px 18px rgba(15, 23, 42, 0.06);
    transition: box-shadow 0.22s ease, transform 0.22s ease;
}
section[data-testid="stMain"] [data-testid="stVerticalBlockBorderWrapper"]:hover {
    box-shadow:
        0 2px 6px rgba(15, 23, 42, 0.08),
        0 12px 28px rgba(15, 23, 42, 0.12);
}

/* 分店下单商品卡片：默认 border 容器用了 overflow:hidden，会干扰横向 flex，改为可见 */
section[data-testid="stMain"] [data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) {
    overflow: visible !important;
}

/* ----- Branch order browse: product card（仅缩略图列居中，勿用 :first-child 以免误伤「箱数」列）
 * Streamlit ≥1.40 列节点为 data-testid="stColumn"；旧版为 "column"，下列双写兼容。
 * ----- */
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="column"]:has(.product-card-thumb-spacer-top),
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stColumn"]:has(.product-card-thumb-spacer-top) {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: flex-start;
    padding: 10px 6px 12px 8px;
}
.product-card-thumb-spacer-top { height: 10px; }
.product-card-thumb-spacer-bottom { height: 12px; }
.product-card-thumb-ph {
    width: 112px;
    height: 112px;
    border-radius: 12px;
    background: linear-gradient(145deg, #f1f5f9 0%, #e8eef5 100%);
    border: 1px dashed #cbd5e1;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 2rem;
    color: #94a3b8;
}
.product-card-title {
    font-size: 1.12rem;
    font-weight: 750;
    color: #0f172a;
    line-height: 1.38;
    margin: 0 0 12px 0;
    letter-spacing: -0.02em;
}
.product-card-meta {
    font-size: 0.84rem;
    color: #64748b;
    line-height: 1.55;
    margin: 0 0 14px 0;
}
.product-card-meta--empty {
    margin-bottom: 8px;
}
.product-card-stock {
    font-size: 0.88rem;
    color: #334155;
    background: linear-gradient(180deg, #f8fafc 0%, #f1f5f9 100%);
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 12px 14px;
    margin: 0 0 14px 0;
    line-height: 1.45;
}
.product-card-batch {
    font-size: 0.82rem;
    color: #047857;
    background: rgba(16, 185, 129, 0.09);
    border: 1px solid rgba(16, 185, 129, 0.25);
    border-radius: 8px;
    padding: 10px 12px;
    margin: 0 0 14px 0;
    line-height: 1.45;
}
.product-card-qty-section {
    margin-top: 2px;
    padding-top: 16px;
    border-top: 1px solid rgba(148, 163, 184, 0.4);
    margin-bottom: 8px;
}
.product-card-qty-section-title {
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 0.05em;
    text-transform: uppercase;
    color: #64748b;
}

/*
 * 分店下单商品卡片 · 手机端：禁止 Streamlit 把列叠成上下。
 * — 上图左、名称/编号/价格/库存永远在图右侧；
 * — 「订购数量」下箱数 / 个数同一行左右排列。
 */
section[data-testid="stMain"] [data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stHorizontalBlock"],
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stHorizontalBlock"] {
    display: flex !important;
    flex-wrap: nowrap !important;
    flex-direction: row !important;
    align-items: flex-start !important;
    gap: 0.5rem !important;
}
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stHorizontalBlock"] > [data-testid="column"],
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stHorizontalBlock"] > [data-testid="stColumn"] {
    min-width: 0 !important;
}
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="column"]:has(.product-card-thumb-spacer-top),
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stColumn"]:has(.product-card-thumb-spacer-top) {
    flex: 0 0 118px !important;
    max-width: 128px !important;
    width: 118px !important;
    box-sizing: border-box !important;
}
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="column"]:has(.product-card-title),
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stColumn"]:has(.product-card-title) {
    flex: 1 1 auto !important;
}
/* 订购数量行：两列等分（排除左侧缩略图列与右侧主信息列） */
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stHorizontalBlock"] > [data-testid="column"]:not(:has(.product-card-thumb-spacer-top)):not(:has(.product-card-title)),
[data-testid="stVerticalBlockBorderWrapper"]:has(.product-card-title) [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:not(:has(.product-card-thumb-spacer-top)):not(:has(.product-card-title)) {
    flex: 1 1 0 !important;
}

@media (max-width: 480px) {
    .product-card-title { font-size: 1.02rem; margin-bottom: 8px !important; }
    .product-card-meta { font-size: 0.8rem; margin-bottom: 8px !important; }
    .product-card-stock { padding: 8px 10px; margin-bottom: 10px !important; font-size: 0.84rem; }
    .product-card-qty-section { padding-top: 12px; }
}

/* Arrival notice item list — plain text only (no Markdown # → heading) */
.arrival-items-plain {
    margin: 0.4rem 0 0 0;
    padding-left: 1.2rem;
    font-size: 0.95rem;
    line-height: 1.55;
    font-weight: 500;
    color: #334155;
}
.arrival-items-plain li {
    margin-bottom: 0.35rem;
}

/* =======================================================================
   电商式购物车悬浮 / 吸底组件 (responsive cart dock)
   - 容器是 st.container(key="sunshine_cart_dock") → 渲染为 .st-key-sunshine_cart_dock
   - PC（≥992px）：右上角悬浮卡片，不与左侧蓝色侧边栏冲突
   - 手机（≤991px）：顶部吸顶横条（左摘要 + 右去结算）
   ======================================================================= */
/* 命中购物车 dock 的外层包裹元素：通过其内部"一定会渲染的可见内容"
   .sunshine-cart-summary 定位（与商品卡片用 .product-card-title 同理）。
   本页无其它含该内容的 border 包裹祖先，故唯一匹配。 */
[data-testid="stVerticalBlockBorderWrapper"]:has(.sunshine-cart-summary) {
    /* !important 必须：覆盖 Streamlit 默认 position，否则退化为随页面滚动。 */
    position: fixed !important;
    /* z-index 9999：高于 main 内容，但仍低于 Streamlit 移动端侧栏抽屉遮罩
       (≈999990)，因此打开蓝色侧边栏时不会被购物车横条盖住，避免层级冲突。 */
    z-index: 9999 !important;
    background: #ffffff !important;
    border: 1px solid #e2e8f0 !important;
    box-shadow: 0 12px 34px rgba(15, 23, 42, 0.20);
    /* 数量/金额刷新时的平滑过渡，减少 rerun 后的突兀感 */
    transition: box-shadow 0.2s ease;
}
[data-testid="stVerticalBlockBorderWrapper"]:has(.sunshine-cart-summary) .stButton > button {
    border-radius: 12px;
    font-weight: 700;
    white-space: nowrap;
}
.sunshine-cart-summary {
    display: flex;
    align-items: center;
    gap: 10px;
    flex-wrap: wrap;
    font-weight: 700;
    color: #0f172a;
    line-height: 1.15;
}
.sunshine-cart-summary .scd-ico { font-size: 1.4rem; }
.sunshine-cart-summary .scd-count {
    min-width: 26px;
    height: 26px;
    padding: 0 8px;
    border-radius: 13px;
    background: #2563eb;
    color: #fff;
    font-size: 0.92rem;
    display: inline-flex;
    align-items: center;
    justify-content: center;
    transition: background 0.2s ease;
}
.sunshine-cart-summary .scd-meta {
    font-size: 0.9rem;
    font-weight: 600;
    color: #475569;
}
.sunshine-cart-summary .scd-amount {
    font-size: 0.95rem;
    font-weight: 800;
    color: #0e7a4f;
}

/* PC（≥992px）：右上角悬浮窗，position:fixed 天然跟随视口滚动 */
@media (min-width: 992px) {
    [data-testid="stVerticalBlockBorderWrapper"]:has(.sunshine-cart-summary) {
        right: 28px !important;
        top: 16px !important;
        bottom: auto !important;
        left: auto !important;
        width: 320px !important;
        border-radius: 18px !important;
        padding: 14px 16px !important;
    }
    /* PC 端预留顶部空白：避免右上角悬浮窗压住顶部内容 */
    section[data-testid="stMain"] .block-container { padding-top: 90px !important; }
}

/* 手机（≤991px）：吸顶横条，始终锁在屏幕最顶部 */
@media (max-width: 991px) {
    [data-testid="stVerticalBlockBorderWrapper"]:has(.sunshine-cart-summary) {
        left: 0 !important;
        right: 0 !important;
        bottom: auto !important;
        top: 0 !important;
        width: 100% !important;
        border-radius: 0 0 16px 16px !important;
        padding: calc(10px + env(safe-area-inset-top, 0px)) 14px 10px !important;
    }
    [data-testid="stVerticalBlockBorderWrapper"]:has(.sunshine-cart-summary) .stButton > button { width: 100%; }
    /* 给主内容留出顶部空间，顶部内容不被吸顶横条遮挡 */
    section[data-testid="stMain"] .block-container { padding-top: 96px !important; }
}
</style>
"""


def inject_css() -> None:
    st.markdown(CSS, unsafe_allow_html=True)


def render_page_heading(title: str, subtitle: str | None = None) -> None:
    """Prominent title strip at the top of each feature page (escapes HTML)."""
    esc = html.escape(title)
    sub_html = ""
    if subtitle:
        sub_html = (
            f'<p class="page-hero-desc">{html.escape(subtitle)}</p>'
        )
    st.markdown(
        f'<div class="page-hero"><div class="page-hero-text">'
        f'<h2 class="page-hero-title">{esc}</h2>{sub_html}</div></div>',
        unsafe_allow_html=True,
    )


def render_section_title(title: str) -> None:
    """In-page section heading (export blocks, email settings, dashboard blocks)."""
    esc = html.escape(title)
    st.markdown(
        f'<div class="section-title"><span>{esc}</span></div>',
        unsafe_allow_html=True,
    )


# =========================================================================
# UI HELPERS
# =========================================================================
def render_header(
    context_label: str = "",
    *,
    show_brand_banner: bool = False,
) -> None:
    """Optional login banner only. Post-login chrome lives in the sidebar.

    Streamlit's top-right ⋮ menu cannot host custom actions (EN/中文); language
    is in ``render_sidebar_lang_switch`` instead.
    """
    if not show_brand_banner:
        return
    esc_ctx = html.escape(context_label) if context_label else ""
    ctx = f'<div class="title-context">{esc_ctx}</div>' if esc_ctx else ""
    st.markdown(
        f"""
        <div class="sunshine-header">
            <div>
                <div class="title-main">☀️ {html.escape(t('app_brand'))}</div>
                <div class="title-sub">{html.escape(t('app_subtitle'))}</div>
            </div>
            {ctx}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_lang_switch() -> None:
    """EN / 中文 in the sidebar + compact ticker (replaces the old top bar)."""
    c1, c2 = st.sidebar.columns(2)
    with c1:
        if st.button("EN", key="lang_en_sb", use_container_width=True):
            st.session_state.lang = "en"
            st.rerun()
    with c2:
        if st.button("中文", key="lang_zh_sb", use_container_width=True):
            st.session_state.lang = "zh"
            st.rerun()
    tick = header_ticker_text()
    if tick:
        st.sidebar.caption(f"📣 {tick}")


def render_lang_switch() -> None:
    _, en_col, zh_col = st.columns([8, 1, 1])
    with en_col:
        if st.button("EN", key="lang_en", use_container_width=True):
            st.session_state.lang = "en"
            st.rerun()
    with zh_col:
        if st.button("中文", key="lang_zh", use_container_width=True):
            st.session_state.lang = "zh"
            st.rerun()


def status_pill(status: str) -> str:
    css_status = status.replace(" ", "")
    label_key = STATUS_LABEL.get(status)
    label = t(label_key) if label_key else status
    return f'<span class="pill pill-{css_status}">{label}</span>'


def logout_button(key: str = "logout_btn") -> None:
    if st.button(f"🚪 {t('logout')}", key=key, use_container_width=True):
        try:
            audit_write(
                "logout",
                extra={"page": st.session_state.get("page")},
            )
        except Exception:
            pass
        try:
            import auth as sunshine_auth

            sunshine_auth.clear_login_cookie()
        except Exception:
            pass
        lang = st.session_state.get("lang", "en")
        last_role = st.session_state.get("role")
        logout_aid = st.session_state.get("account_id")
        logout_br = st.session_state.get("branch")
        if (
            last_role == Role.BRANCH
            and logout_aid is not None
            and logout_br
            and str(logout_br).strip()
        ):
            try:
                _delete_branch_cart_draft(int(logout_aid), str(logout_br).strip())
            except Exception:
                pass
        try:
            if URL_PAGE_QUERY_KEY in st.query_params:
                del st.query_params[URL_PAGE_QUERY_KEY]
        except Exception:
            pass
        st.session_state.clear()
        st.session_state.lang = lang
        # 主动退出：勿保留 last_role / URL ?p=，否则登录页像“退回上一屏”，
        # 且 Cookie+地址栏会恢复到上一业务页。
        st.session_state._explicit_logout = True
        st.session_state._sunshine_cookie_bootstrapped = True
        st.session_state._sunshine_cookie_restore_attempted = True
        st.rerun()


# =========================================================================
# SESSION
# =========================================================================
def init_session() -> None:
    defaults = {
        "lang": None,
        "role": None,
        "branch": None,
        "page": None,
        "cart": [],
        "last_role": None,
        "last_branch": None,
        "ai_chat": [],
        "ai_last_call_ts": 0.0,
        "account_id": None,
        "account_username": None,
        "account_perms": [],
        "branch_apply_submitted": False,
        "login_branch_context": None,
        "wh_supplier_sent": False,
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val
    # Cookie 自动登录在 route() 开头执行（需先挂载浏览器组件）


def _query_param_first(key: str) -> str | None:
    """Single query value for Streamlit query_params (string or one-element list)."""
    try:
        v = st.query_params.get(key)
    except Exception:
        return None
    if v is None:
        return None
    if isinstance(v, (list, tuple)):
        return str(v[0]).strip() if v else None
    s = str(v).strip()
    return s if s else None


def _sync_cookie_default_page(page_key: str) -> None:
    """当前页写入登录 Cookie，F5 时地址栏无 ?p= 也能回到上次页面。"""
    if not (page_key or "").strip():
        return
    if st.session_state.get("_cookie_sync_page") == page_key:
        return
    try:
        import auth as sunshine_auth

        sunshine_auth.touch_cookie_page(page_key)
        st.session_state._cookie_sync_page = page_key
    except Exception:
        pass


def _apply_url_page_to_session(role: str, allowed: frozenset[str]) -> None:
    """地址栏 ?p= 优先（含 F5 刷新、浏览器后退）。"""
    p = _query_param_first(URL_PAGE_QUERY_KEY)
    if not p or p not in allowed:
        return
    if role == Role.BRANCH:
        if p == "stock":  # 库存/临期页对所有分店账号开放，无需单独授权
            st.session_state.page = p
            return
        pmap = {
            "order": "order",
            "order_done": "order",
            "my_orders": "my_orders",
            "my_short": "my_short",
            "messages": "messages",
            "ai": "ai",
        }
        need = pmap.get(p, "order")
        if not has_branch_perm(need):
            return
    st.session_state.page = p


def _allowed_pages_for_role(role: str) -> frozenset[str]:
    if role == Role.WAREHOUSE:
        return WAREHOUSE_PAGE_KEYS
    if role == Role.ADMIN:
        return ADMIN_PAGE_KEYS
    if role == Role.BRANCH:
        keys: set[str] = set()
        for page_key, perm in (
            ("order", "order"),
            ("my_orders", "my_orders"),
            ("my_short", "my_short"),
            ("messages", "messages"),
            ("ai", "ai"),
        ):
            if has_branch_perm(perm):
                keys.add(page_key)
        keys.add("order_done")
        keys.add("stock")  # 库存/临期页对所有分店账号开放
        return frozenset(keys)
    return frozenset()


def _route_sync_page_from_url() -> None:
    """Align session page with ?p= when the user uses browser back/forward."""
    if not is_authenticated():
        return
    role = st.session_state.get("role")
    allowed = _allowed_pages_for_role(role)
    if not allowed:
        return
    prev = st.session_state.get("page")
    _apply_url_page_to_session(role, allowed)
    curr = st.session_state.get("page")
    if curr != prev and curr in allowed:
        st.session_state.pop("_nav_from_sidebar", None)
        st.rerun()


def _branch_nav_allowed_keys(nav_items: list[tuple[str, str]]) -> frozenset[str]:
    keys = {k for k, _ in nav_items}
    keys.add("order_done")
    return frozenset(keys)


def _set_session_page_for_app_nav(page_key: str) -> None:
    """Use when changing `page` from buttons (not sidebar) so URL/history stay aligned."""
    st.session_state.page = page_key
    st.session_state._nav_from_sidebar = True
    try:
        st.query_params[URL_PAGE_QUERY_KEY] = page_key
    except Exception:
        pass


WAREHOUSE_PAGE_KEYS = frozenset({
    "pending", "short_in", "history", "supplier", "inventory", "messages", "ai",
    "supplier_order", "order_success", "verify_inbound", "verify_success",
})
ADMIN_PAGE_KEYS = frozenset({
    "dashboard", "all_orders", "expiry_dash",
    "admin_suppliers", "supplier_order", "order_success",
    "verify_inbound", "verify_success",
    "images", "catalog", "product_master",
    "email", "export", "backup",
    # 仍可通过 URL 直达（未在折叠菜单展示）
    "shelf_mobile", "accounts", "audit", "inventory", "messages", "ai",
})


def is_authenticated() -> bool:
    role = st.session_state.get("role")
    if role == Role.BRANCH:
        if not st.session_state.get("branch"):
            return False
        if st.session_state.get("account_id") is None:
            return False
        return True
    return role in (Role.WAREHOUSE, Role.ADMIN)


def branch_effective_perms() -> frozenset[str]:
    """Permissions for the current branch session (account-backed only)."""
    if st.session_state.get("role") != Role.BRANCH:
        return frozenset(BRANCH_PERM_CODES)
    return frozenset(st.session_state.get("account_perms") or [])


def has_branch_perm(perm: str) -> bool:
    return perm in branch_effective_perms()


def first_allowed_branch_page() -> str:
    for p, _ in BRANCH_PERM_ORDER:
        if has_branch_perm(p):
            return p
    return BRANCH_PERM_ORDER[0][0]


def branch_perm_label(code: str) -> str:
    for p, key in BRANCH_PERM_ORDER:
        if p == code:
            return t(key)
    return code


def ensure_branch_page_allowed(page_key: str) -> None:
    if page_key == "stock":  # 库存/临期页对所有分店账号开放
        return
    pmap = {
        "order": "order",
        "order_done": "order",
        "my_orders": "my_orders",
        "my_short": "my_short",
        "messages": "messages",
        "ai": "ai",
    }
    need = pmap.get(page_key, "order")
    if has_branch_perm(need):
        return
    fallback = first_allowed_branch_page()
    if st.session_state.get("page") != fallback:
        _set_session_page_for_app_nav(fallback)
        st.warning(t("acct_no_access"))
        st.rerun()


# =========================================================================
# LOGIN PAGES
# =========================================================================
def page_pick_language() -> None:
    st.markdown(
        """
        <div class="sunshine-header" style="text-align:center; justify-content:center;">
            <div>
                <div class="title-main" style="justify-content:center;">
                    ☀️ SUNSHINE SHOPPING CENTER · 阳光集团
                </div>
                <div class="title-sub">Ordering System · 订货系统</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    render_section_title("Select Language / 选择语言")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🇬🇧 English", use_container_width=True, type="primary"):
            st.session_state.lang = "en"
            st.rerun()
    with c2:
        if st.button("🇨🇳 中文", use_container_width=True, type="primary"):
            st.session_state.lang = "zh"
            st.rerun()


def page_login() -> None:
    pending = st.session_state.get("pending_role")
    if pending is None and st.session_state.get("last_role") in (
        Role.BRANCH, Role.WAREHOUSE, Role.ADMIN
    ):
        pending = st.session_state.get("last_role")
        st.session_state.pending_role = pending

    if pending == Role.BRANCH and st.session_state.get("branch_apply_submitted"):
        render_header(show_brand_banner=True)
        render_lang_switch()
        render_page_heading(t("acct_apply_done_title"), None)
        bctx = st.session_state.get("login_branch_context")
        if bctx in BRANCHES:
            st.markdown(
                f"**{html.escape(t('branch'))}:** {html.escape(bctx)}"
            )
        st.success(t("acct_apply_ok"))
        if st.button(
            t("acct_back_branch_options"),
            type="primary",
            key="acct_apply_done_back",
            use_container_width=True,
        ):
            st.session_state.branch_apply_submitted = False
            st.rerun()
        return

    # 已选角色：进入该角色专属登录页（整页切换，不与角色选择堆叠）。
    if pending in (Role.BRANCH, Role.WAREHOUSE, Role.ADMIN):
        _render_role_login_page(pending)
        return

    # 未选角色：仅显示角色选择页。
    render_header(show_brand_banner=True)
    render_lang_switch()
    render_section_title(t("select_role"))
    if st.session_state.get("last_role"):
        last_role_label = {
            Role.BRANCH: t("role_branch"),
            Role.WAREHOUSE: t("role_warehouse"),
            Role.ADMIN: t("role_admin"),
        }.get(st.session_state.get("last_role"), st.session_state.get("last_role"))
        st.caption(f"{t('recent_login')}: {last_role_label}")
        if (
            st.session_state.get("last_role") == Role.BRANCH
            and st.session_state.get("last_branch")
        ):
            st.caption(f"{t('use_recent_branch')}: {st.session_state.get('last_branch')}")

    c1, c2, c3 = st.columns(3)
    if c1.button(t("role_branch"), use_container_width=True):
        st.session_state.pending_role = Role.BRANCH
        st.rerun()
    if c2.button(t("role_warehouse"), use_container_width=True):
        st.session_state.pending_role = Role.WAREHOUSE
        st.session_state.branch_apply_submitted = False
        st.session_state.login_branch_context = None
        st.rerun()
    if c3.button(t("role_admin"), use_container_width=True):
        st.session_state.pending_role = Role.ADMIN
        st.session_state.branch_apply_submitted = False
        st.session_state.login_branch_context = None
        st.rerun()


def _login_back_to_roles_button() -> None:
    """登录页顶部「返回选择角色」：清掉已选角色，回到角色选择页。"""
    if st.button(t("back_to_roles"), key="login_back_to_roles"):
        st.session_state.pop("pending_role", None)
        st.session_state.last_role = None
        st.session_state.branch_apply_submitted = False
        st.session_state.login_branch_context = None
        st.rerun()


def _render_role_login_page(pending: str) -> None:
    render_header(show_brand_banner=True)
    render_lang_switch()
    _login_back_to_roles_button()

    if pending == Role.WAREHOUSE:
        _password_login(Role.WAREHOUSE, WAREHOUSE_PASSWORD, default_page="pending")
        return
    if pending == Role.ADMIN:
        _password_login(
            Role.ADMIN,
            ADMIN_PASSWORD,
            default_page="dashboard",
            phone_allowlist=_admin_phone_allowlist(),
        )
        return

    if pending == Role.BRANCH:
        portal = st.session_state.get("login_branch_context")
        if portal not in BRANCHES:
            render_section_title(t("acct_enter_store_title"))
            st.caption(t("acct_enter_store_hint"))
            last_branch = st.session_state.get("last_branch")
            cols = st.columns(2)
            for i, branch in enumerate(BRANCHES):
                label = f"⭐ {branch}" if branch == last_branch else branch
                with cols[i % 2]:
                    if st.button(
                        label, key=f"portal_pick_{i}", use_container_width=True
                    ):
                        st.session_state.login_branch_context = branch
                        st.rerun()
            return

        st.markdown(
            f"<div style='margin-bottom:0.75rem;'>"
            f"<span style='font-size:1.05rem;font-weight:650;'>{html.escape(t('acct_current_store'))}: "
            f"</span><span style='font-size:1.05rem;font-weight:800;color:#0d47a1;'>"
            f"{html.escape(portal)}</span></div>",
            unsafe_allow_html=True,
        )
        if st.button(t("acct_change_store"), key="portal_change_store"):
            st.session_state.login_branch_context = None
            st.session_state.branch_apply_submitted = False
            st.rerun()

        tabs = st.tabs([t("acct_tab_login"), t("acct_tab_apply")])

        with tabs[0]:
            st.caption(t("acct_login_hint"))
            bu = st.text_input(t("acct_username"), key="branch_acct_user")
            bpw = st.text_input(t("password"), type="password", key="branch_acct_pw")
            if st.button(t("login"), type="primary", key="branch_acct_login"):
                code, row = account_try_login(bu, bpw)
                if code == "ok" and row is not None:
                    if row["branch"] != portal:
                        st.error(t("acct_wrong_branch"))
                    else:
                        st.session_state.role = Role.BRANCH
                        st.session_state.branch = row["branch"]
                        st.session_state.account_id = int(row["id"])
                        st.session_state.account_username = row["username"]
                        st.session_state.account_perms = _parse_permissions_json(
                            row["permissions"]
                        )
                        st.session_state.last_role = Role.BRANCH
                        st.session_state.last_branch = row["branch"]
                        first_pg = first_allowed_branch_page()
                        _set_session_page_for_app_nav(first_pg)
                        st.session_state.login_branch_context = None
                        st.session_state.pop("pending_role", None)
                        try:
                            import auth as sunshine_auth

                            sunshine_auth.persist_login_cookie(
                                role=Role.BRANCH,
                                branch=row["branch"],
                                account_id=int(row["id"]),
                                account_username=row["username"],
                                account_perms=st.session_state.account_perms,
                                default_page=first_pg,
                            )
                        except Exception:
                            pass
                        audit_write(
                            "login",
                            branch=row["branch"],
                            extra={"method": "branch_account", "account_id": int(row["id"])},
                        )
                        st.rerun()
                elif code == "pending":
                    st.warning(t("acct_pending_login"))
                elif code == "rejected":
                    st.error(t("acct_rejected_login"))
                elif code == "no_permissions":
                    st.error(t("acct_perm_need_admin"))
                else:
                    st.error(t("wrong_pw"))

        with tabs[1]:
            st.markdown(
                f"**{html.escape(t('branch'))}:** {html.escape(portal)}"
            )
            auser = st.text_input(t("acct_username"), key="acct_apply_user")
            dname = st.text_input(t("acct_display_name"), key="acct_apply_dname")
            phone = st.text_input(t("acct_phone"), key="acct_apply_phone")
            p1 = st.text_input(t("acct_password"), type="password", key="acct_apply_p1")
            p2 = st.text_input(t("acct_password2"), type="password", key="acct_apply_p2")
            if st.button(t("acct_apply_submit"), type="primary", key="acct_apply_submit"):
                if p1 != p2:
                    st.error(t("acct_pw_mismatch"))
                else:
                    ok, err_k = account_insert_application(
                        auser, p1, portal, dname, phone
                    )
                    if ok:
                        st.session_state.branch_apply_submitted = True
                        st.rerun()
                    else:
                        st.error(t(err_k))


def _password_login(
    role: str,
    expected: str,
    default_page: str,
    *,
    phone_allowlist: set[str] | None = None,
) -> None:
    st.divider()
    label = t("role_warehouse") if role == Role.WAREHOUSE else t("role_admin")
    render_section_title(label)
    need_phone = bool(phone_allowlist)
    if need_phone:
        st.caption(t("admin_login_phone_hint"))
        phone_in = st.text_input(
            t("admin_phone"),
            key=f"phone_{role}",
            placeholder="13800000000",
        )
    else:
        phone_in = ""
    pw = st.text_input(t("password"), type="password", key=f"pw_{role}")
    if st.button(t("login"), type="primary", key=f"login_{role}"):
        if need_phone:
            got = _normalize_phone_digits(phone_in)
            if not got or got not in (phone_allowlist or set()):
                st.error(t("wrong_phone"))
                return
        if pw == expected:
            st.session_state.role = role
            st.session_state.last_role = role
            _set_session_page_for_app_nav(default_page)
            st.session_state.pop("pending_role", None)
            try:
                import auth as sunshine_auth

                sunshine_auth.persist_login_cookie(
                    role=role,
                    default_page=default_page,
                )
            except Exception:
                pass
            extra: dict = {"method": "shared_password", "portal_role": role}
            if need_phone:
                extra["method"] = "admin_phone_password"
            audit_write("login", extra=extra)
            st.rerun()
        else:
            st.error(t("wrong_pw"))


def page_ai_assistant() -> None:
    render_page_heading(t("nav_ai"), t("ai_disclaimer"))
    if not GEMINI_API_KEY:
        st.warning(t("ai_missing_key"))
        return

    # Two-phase submit: click only queues; next run performs API inside spinner
    # so the user cannot double-click while the model is working.
    pending = st.session_state.get("_ai_pending")
    if isinstance(pending, dict):
        with st.spinner(t("ai_working")):
            q = (pending.get("q") or "").strip()
            with_ctx = bool(pending.get("with_ctx", True))
            st.session_state.pop("_ai_pending", None)
            if not q:
                st.warning(t("ai_empty_q"))
            else:
                now_ts = unix_ts()
                last_ts = float(st.session_state.get("ai_last_call_ts", 0.0))
                if now_ts - last_ts < 1.5:
                    st.warning(t("ai_rate_limit"))
                else:
                    st.session_state["ai_last_call_ts"] = now_ts
                    context_info = _build_ai_context() if with_ctx else ""
                    if with_ctx:
                        top_block = ""
                        try:
                            with db_conn() as conn:
                                rows = conn.execute(
                                    """
                                    SELECT name,
                                           SUM(COALESCE(dispatch_cartons, 0)) AS sum_ct,
                                           SUM(COALESCE(dispatch_pcs, 0)) AS sum_pc
                                    FROM orders
                                    WHERE status = 'Dispatched'
                                      AND dispatch_date IS NOT NULL
                                      AND datetime(dispatch_date)
                                          >= datetime('now', '-7 days')
                                    GROUP BY name
                                    ORDER BY sum_ct DESC, sum_pc DESC
                                    LIMIT 5
                                    """
                                ).fetchall()
                            if rows:
                                top_block = (
                                    "【近7日已发货出库量TOP5（按商品名汇总）】\n"
                                    + "\n".join(
                                        f"- {r['name']}: 共 {int(r['sum_ct'])} 箱 + "
                                        f"{int(r['sum_pc'])} 件"
                                        for r in rows
                                    )
                                )
                            else:
                                top_block = "【近7日已发货出库量TOP5】无记录"
                        except Exception as e:
                            log_exception("ai_ctx_weekly_dispatch_top5", e)
                            top_block = "【近7日已发货出库量TOP5】（读取失败，已跳过）"

                        short_block = ""
                        try:
                            with db_conn() as conn:
                                rows = conn.execute(
                                    """
                                    SELECT branch, name, short_cartons, short_pcs
                                    FROM shortages
                                    WHERE status = 'Open'
                                    ORDER BY branch, name
                                    """
                                ).fetchall()
                            if rows:
                                short_block = (
                                    "【当前待处理缺货 status=Open】\n"
                                    + "\n".join(
                                        f"- {r['branch']} | {r['name']} | "
                                        f"缺 {int(r['short_cartons'])} 箱 + "
                                        f"{int(r['short_pcs'])} 件"
                                        for r in rows
                                    )
                                )
                            else:
                                short_block = "【当前待处理缺货 status=Open】无记录"
                        except Exception as e:
                            log_exception("ai_ctx_open_shortages_list", e)
                            short_block = "【当前待处理缺货】（读取失败，已跳过）"

                        parts = [
                            p for p in (context_info.strip(), top_block, short_block) if p
                        ]
                        context_info = _safe_text("\n\n".join(parts), max_len=3200)

                    ok, ans = _call_gemini_api(q, context_info)
                    chat = st.session_state.setdefault("ai_chat", [])
                    chat.append({"q": q, "a": ans, "ok": ok, "ts": now_str()})
        st.rerun()

    # Context toggle must come before quick prompts so one-click shortcuts
    # read the same value as manual asks.
    with_ctx = st.checkbox(t("ai_with_ctx"), value=True, key="ai_with_ctx")

    q1, q2, q3 = st.columns([2, 2, 1])
    with q1:
        quick_shortage = st.button(
            t("ai_q_shortage"), use_container_width=True, key="ai_q_shortage_btn",
            disabled=bool(st.session_state.get("_ai_pending")),
        )
    with q2:
        quick_not_found = st.button(
            t("ai_q_not_found"), use_container_width=True, key="ai_q_nf_btn",
            disabled=bool(st.session_state.get("_ai_pending")),
        )
    with q3:
        if st.button(
            t("ai_clear_chat"), use_container_width=True, key="ai_clear_btn",
            disabled=bool(st.session_state.get("_ai_pending")),
        ):
            st.session_state["ai_chat"] = []
            st.rerun()

    quick_dispatch = st.button(
        t("ai_q_dispatch"), use_container_width=True, key="ai_q_dispatch_btn",
        disabled=bool(st.session_state.get("_ai_pending")),
    )

    chosen_quick = ""
    if quick_shortage:
        chosen_quick = t("ai_q_shortage")
    elif quick_not_found:
        chosen_quick = t("ai_q_not_found")
    elif quick_dispatch:
        chosen_quick = t("ai_q_dispatch")
    if chosen_quick:
        # Previous code only overwrote a Python variable; the text_area widget
        # state stayed empty, so "Ask AI" submitted blank. Queue same as manual.
        st.session_state["ai_input"] = chosen_quick
        st.session_state["_ai_pending"] = {
            "q": chosen_quick,
            "with_ctx": bool(with_ctx),
        }
        st.rerun()

    user_q = st.text_area(
        t("ai_question"), key="ai_input", height=110,
        disabled=bool(st.session_state.get("_ai_pending")),
    )

    ask_disabled = bool(st.session_state.get("_ai_pending"))
    if st.button(
        t("ai_ask_btn"),
        type="primary",
        use_container_width=True,
        key="ai_ask_submit",
        disabled=ask_disabled,
    ):
        q = (user_q or "").strip()
        if not q:
            st.warning(t("ai_empty_q"))
        else:
            st.session_state["_ai_pending"] = {
                "q": q,
                "with_ctx": bool(with_ctx),
            }
            st.rerun()

    chat = st.session_state.get("ai_chat", [])
    if not chat:
        return
    render_section_title(t("ai_answer"))
    for item in reversed(chat[-10:]):
        with st.container(border=True):
            st.caption(item.get("ts", ""))
            st.markdown(f"**Q:** {item.get('q', '')}")
            if item.get("ok"):
                st.markdown(item.get("a", ""))
            else:
                st.error(item.get("a", t("ai_error")))


def page_messages() -> None:
    unread = count_unread_notifications()
    subtitle = f"{unread} unread" if (st.session_state.get("lang") == "en") else f"未读 {unread} 条"
    render_page_heading(f"🔔 {t('msg_center_title')}", subtitle)

    c1, c2 = st.columns([1, 4])
    with c1:
        if st.button(t("msg_mark_all"), use_container_width=True, key="msg_mark_all"):
            mark_all_notifications_read()
            st.rerun()
    with c2:
        st.caption(f"{t('msg_center_title')} · {subtitle}")

    rows = load_notifications(limit=150)
    if not rows:
        st.info(t("msg_empty"))
        return

    for row in rows:
        is_read = bool(row["is_read"])
        badge = "✅" if is_read else "🟡"
        with st.container(border=True):
            st.markdown(f"**{badge} {row['title']}**")
            meta = f"{row['created_at'][:16]}"
            if row["order_id"]:
                meta += f" · {t('msg_related_order')}: {row['order_id']}"
            st.caption(meta)
            st.write(row["message"])
            if not is_read:
                if st.button(
                    t("msg_mark_read"),
                    key=f"msg_read_{row['id']}",
                    use_container_width=False,
                ):
                    mark_notification_read(int(row["id"]))
                    st.rerun()


def _render_arrival_items_list(items_text: str) -> None:
    """Render arrival SKU lines with uniform typography.

    Using ``st.markdown`` with ``- {line}`` breaks when a line starts with ``#``
    (Markdown headings) or other MD syntax — sizes become inconsistent.
    """
    lines = [ln.strip() for ln in (items_text or "").splitlines() if ln.strip()]
    if not lines:
        return
    st.markdown(
        "<ul class='arrival-items-plain'>"
        + "".join(f"<li>{html.escape(ln)}</li>" for ln in lines)
        + "</ul>",
        unsafe_allow_html=True,
    )


def _render_active_arrival_banner() -> None:
    row = get_active_stock_arrival()
    if row is None:
        return
    title = (row["title"] or "").strip() or t("nav_arrivals")
    notice = (row["notice"] or "").strip()
    items = (row["items_text"] or "").strip()
    created = str(row["created_at"] or "")[:16]
    with st.container(border=True):
        st.markdown(f"### 📦 {title}")
        st.caption(f"{t('arrival_priority_tip')} · {created}")
        if notice:
            st.info(notice)
        if items:
            _render_arrival_items_list(items)
        st.divider()


def _arrival_product_labels_from_df(df: pd.DataFrame) -> list[str]:
    """Build one label per row for arrival pickers (name + code + barcode)."""
    options: list[str] = []
    for _, r in df.iterrows():
        nm = str(r.get("Name", "") or "").strip()
        if not nm:
            continue
        code = str(r.get("ItemCode", "") or "").strip()
        bc = str(r.get("Barcode", "") or "").strip()
        parts = [nm]
        if code:
            parts.append(f"[{code}]")
        if bc:
            parts.append(f"({bc})")
        options.append(" ".join(parts))
    return options


@st.dialog("选择到货商品 · Pick products")
def _dialog_arrival_pick_products() -> None:
    """Modal: search / multi-select products and append lines to arrival_items_editor."""
    st.caption(t("arrival_pick_dialog_hint"))
    st.text_input(
        t("search_product"),
        key="arrival_dlg_search",
        placeholder=t("search_product"),
    )
    qstrip = (st.session_state.get("arrival_dlg_search") or "").strip()
    pick_df = (
        search_products(qstrip, limit=200)
        if qstrip
        else load_products(
            _products_mtime(), _inventory_version(), _price_version()
        ).head(200)
    )
    options = _arrival_product_labels_from_df(pick_df)
    if not options:
        st.info(t("no_results"))
    else:
        st.multiselect(
            t("arrival_pick_dialog_select"),
            options=options,
            key="arrival_dlg_multiselect",
        )
    c1, c2 = st.columns(2)
    with c1:
        do_add = st.button(
            t("arrival_pick_confirm"),
            type="primary",
            use_container_width=True,
            key="arrival_dlg_btn_ok",
        )
    with c2:
        do_close = st.button(
            t("arrival_pick_cancel"),
            use_container_width=True,
            key="arrival_dlg_btn_close",
        )
    if do_add:
        picked = list(st.session_state.get("arrival_dlg_multiselect") or [])
        if not picked:
            st.warning(t("arrival_pick_empty_sel"))
            return
        existing = str(st.session_state.get("arrival_items_editor", "") or "")
        existing_lines = [ln.strip() for ln in existing.splitlines() if ln.strip()]
        seen = set(existing_lines)
        appended = []
        for item in picked:
            if item not in seen:
                seen.add(item)
                appended.append(item)
        if appended:
            merged = existing_lines + appended
            st.session_state["arrival_items_editor"] = "\n".join(merged)
            st.session_state["_arrival_pick_flash"] = t("arrival_pick_appended").format(
                n=len(appended)
            )
        else:
            st.session_state["_arrival_pick_flash"] = t("arrival_pick_nothing_new")
        st.session_state["_arrival_dlg_open"] = False
        st.rerun()
    if do_close:
        st.session_state["_arrival_dlg_open"] = False
        st.rerun()


def page_admin_arrivals() -> None:
    render_page_heading(f"📦 {t('nav_arrivals')}")
    if st.session_state.get("_arrival_reset_editor"):
        st.session_state["arrival_items_editor"] = ""
        st.session_state.pop("arrival_pick_selected", None)
        st.session_state.pop("arrival_pick_kw", None)
        st.session_state.pop("arrival_dlg_multiselect", None)
        st.session_state.pop("arrival_dlg_search", None)
        st.session_state["_arrival_dlg_open"] = False
        st.session_state.pop("_arrival_reset_editor", None)
    current = get_active_stock_arrival()
    render_section_title(t("arrival_current"))
    if current is None:
        st.info(t("arrival_none"))
    else:
        st.markdown(f"**{(current['title'] or '').strip() or t('nav_arrivals')}**")
        st.caption(str(current["created_at"])[:16])
        if (current["notice"] or "").strip():
            st.write((current["notice"] or "").strip())
        if (current["items_text"] or "").strip():
            _render_arrival_items_list(str(current["items_text"]))

    st.divider()
    render_section_title(t("arrival_publish"))
    pick_flash = st.session_state.pop("_arrival_pick_flash", None)
    if pick_flash:
        st.success(pick_flash)
    st.caption(t("arrival_publish_hint_dialog"))
    if st.button(
        t("arrival_pick_open_btn"),
        key="arrival_open_pick_dlg",
        type="primary",
    ):
        st.session_state["_arrival_dlg_open"] = True
        st.session_state.pop("arrival_dlg_multiselect", None)
        st.session_state.pop("arrival_dlg_search", None)
    if st.session_state.get("_arrival_dlg_open"):
        _dialog_arrival_pick_products()

    if "arrival_items_editor" not in st.session_state:
        st.session_state["arrival_items_editor"] = ""

    with st.form("arrival_publish_form"):
        title = st.text_input(t("arrival_title"), value="")
        notice = st.text_area(t("arrival_notice"), value="", height=80)
        items_text = st.text_area(
            t("arrival_items"),
            key="arrival_items_editor",
            height=180,
        )
        ok = st.form_submit_button(t("arrival_publish"), type="primary")
    if ok:
        if not (items_text or "").strip():
            st.warning(t("arrival_items"))
            return
        publish_stock_arrival(title, notice, items_text)
        for b in BRANCHES:
            create_notification(
                "arrival_notice",
                f"新货到达 · {b}",
                f"{(title or '').strip() or '新货到达'}：请在下单前查看并优先下单。",
                target_role=Role.BRANCH,
                target_branch=b,
            )
        create_notification(
            "arrival_notice",
            "已发布到货通知",
            f"管理员发布了新的到货通知：{(title or '').strip() or '新货到达'}。",
            target_role=Role.WAREHOUSE,
        )
        st.session_state["_arrival_reset_editor"] = True
        st.success(t("arrival_published"))
        st.rerun()


# =========================================================================
# BRANCH PAGES
# =========================================================================
def page_branch_order() -> None:
    """Branch staff order placement.

    Two modes, switched by st.session_state.confirming:
      False (default) → browse / search / fill quantities / batch-add → cart
      True            → review cart, edit quantities, send to DB
    """
    if st.session_state.get("confirming"):
        _branch_order_confirm()
    else:
        _branch_order_browse()


# ----- Helpers ------------------------------------------------------------
# Why all this ceremony? Streamlit deletes a widget's session_state value
# the moment that widget is no longer rendered. So if a user fills qty for
# Product A on page 1 and clicks "Next ▶", the page-1 widgets disappear and
# Streamlit drops their values. The fix: every time a number_input changes,
# an on_change callback copies the new value into our own dict
# (_qty_snapshot) which we control. When the widget reappears later, we
# restore its initial value from the snapshot.
def _product_id(item_code: str, barcode: str, name: str) -> str:
    """Stable identity for a product across pages. Master rows can have
    blank ItemCode or Barcode, so we fall back to the name as last resort."""
    return f"{item_code}|{barcode}|{name}".lower()


def _qty_widget_keys(pid: str) -> tuple[str, str]:
    """Per-product widget keys used by Streamlit number_inputs."""
    return f"qty_w__{pid}__ct", f"qty_w__{pid}__pc"


def _on_qty_change(pid: str, which: str) -> None:
    """on_change callback: persist the widget's current value into our
    own snapshot before Streamlit can garbage-collect the widget state."""
    ct_key, pc_key = _qty_widget_keys(pid)
    snap = st.session_state.setdefault("_qty_snapshot", {})
    entry = snap.setdefault(pid, {"ct": 0, "pc": 0})
    if which == "ct":
        entry["ct"] = int(st.session_state.get(ct_key, 0) or 0)
    else:
        entry["pc"] = int(st.session_state.get(pc_key, 0) or 0)


def _register_product(pid: str, product: dict) -> None:
    """Remember product details (name / unit / price / etc.) keyed by pid,
    so that when batch-adding we know what each pid refers to."""
    reg = st.session_state.setdefault("_qty_registry", {})
    reg[pid] = product


def _get_qty(pid: str) -> tuple[int, int]:
    """Return (ct, pc) currently saved in the snapshot for this product."""
    snap = st.session_state.get("_qty_snapshot", {})
    entry = snap.get(pid, {"ct": 0, "pc": 0})
    return int(entry.get("ct", 0)), int(entry.get("pc", 0))


def _collect_qty_inputs() -> list[dict]:
    """Return all products with qty > 0, ready to be appended to the cart."""
    snap = st.session_state.get("_qty_snapshot", {})
    reg = st.session_state.get("_qty_registry", {})
    selected = []
    for pid, qty in snap.items():
        ct, pc = int(qty.get("ct", 0)), int(qty.get("pc", 0))
        if ct == 0 and pc == 0:
            continue
        product = reg.get(pid)
        if product is None:
            # Snapshot has qty but we never registered the product details.
            # Should not happen in practice, but skip gracefully if it does.
            continue
        selected.append({**product, "qty_cartons": ct, "qty_pcs": pc})
    return selected


def _clear_qty_inputs() -> None:
    """Reset both the snapshot and any live widget state."""
    st.session_state["_qty_snapshot"] = {}
    st.session_state["_qty_registry"] = {}
    # Also clear any number_input widgets that are still mounted.
    for k in [k for k in st.session_state if k.startswith("qty_w__")]:
        del st.session_state[k]


def _fmt_branch_catalog_price(v: object) -> str:
    """分店下单卡片上的价格展示：整数不显示小数尾随零。"""
    try:
        x = float(v or 0)
    except (TypeError, ValueError):
        return "0"
    if abs(x - round(x)) < 1e-9:
        return str(int(round(x)))
    return f"{x:.2f}"


def _render_branch_product_card(row: pd.Series) -> None:
    """Product row for branch order browse: thumbnail left, stacked info + qty right."""
    img_path = get_product_image_path(
        str(row.get("ItemCode", "")),
        str(row.get("Barcode", "")),
    )
    item_code = str(row.get("ItemCode", ""))
    barcode = str(row.get("Barcode", ""))
    name = str(row["Name"])
    pid = _product_id(item_code, barcode, name)
    ct_key, pc_key = _qty_widget_keys(pid)

    _register_product(pid, {
        "item_code": item_code,
        "barcode":   barcode,
        "name":      name,
        "unit":      str(row.get("Unit", "")),
        "price":     float(row.get("Price", 0) or 0),
        "is_manual": 0,
    })
    init_ct, init_pc = _get_qty(pid)
    if ct_key not in st.session_state:
        st.session_state[ct_key] = init_ct
    if pc_key not in st.session_state:
        st.session_state[pc_key] = init_pc

    # 使用横向 container 代替 st.columns，避免 Streamlit 1.56+ 默认 flex-wrap 在窄屏叠行；
    # key 仅用于稳定身份（CSS 仍用 :has(.product-card-title) 命中整张卡片）。
    _card_key = "bc_" + hashlib.sha256(pid.encode("utf-8")).hexdigest()[:28]
    with st.container(border=True, key=_card_key):
        with st.container(
            horizontal=True,
            gap="medium",
            vertical_alignment="top",
        ):
            with st.container(width=128):
                st.markdown(
                    '<div class="product-card-thumb-spacer-top"></div>',
                    unsafe_allow_html=True,
                )
                if img_path is not None:
                    st.image(str(img_path), width=112)
                else:
                    st.markdown(
                        '<div class="product-card-thumb-ph"><span>📦</span></div>',
                        unsafe_allow_html=True,
                    )
                st.markdown(
                    '<div class="product-card-thumb-spacer-bottom"></div>',
                    unsafe_allow_html=True,
                )

            with st.container():
                ic_raw = (item_code or "").strip()
                ic_core = ic_raw.lstrip("#").strip()
                if ic_core:
                    title_line = f"商品名：#{ic_core} {name}"
                else:
                    title_line = f"商品名：{name}"
                sku_line = (barcode or "").strip() or ic_raw
                st.markdown(
                    f'<div class="product-card-title">'
                    f"{html.escape(title_line)}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div class="product-card-meta">'
                    f"商品编号：{html.escape(sku_line) if sku_line else '—'}"
                    f"</div>",
                    unsafe_allow_html=True,
                )
                price_txt = _fmt_branch_catalog_price(row.get("Price", 0))
                st.markdown(
                    f'<div class="product-card-meta">'
                    f"商品价格：{html.escape(price_txt)}"
                    f"</div>",
                    unsafe_allow_html=True,
                )

                if bool(row.get("_has_stock_info", False)):
                    stock_ct = int(float(row.get("StockCartons", 0) or 0))
                    stock_pc = int(float(row.get("StockPcs", 0) or 0))
                    stock_total = int(float(row.get("StockTotal", 0) or 0))
                    if stock_ct > 0 or stock_pc > 0:
                        stock_line = (
                            f"🏬 库存：📦 {stock_ct} 箱 · 🔢 {stock_pc} 个"
                        )
                    else:
                        stock_line = f"🏬 库存：{stock_total}"
                    st.markdown(
                        f'<div class="product-card-stock">{html.escape(stock_line)}</div>',
                        unsafe_allow_html=True,
                    )

                prev_ct, prev_pc = _get_qty(pid)
                if prev_ct > 0 or prev_pc > 0:
                    batch_txt = (
                        f"✓ {t('saved_for_batch')}: 📦 {prev_ct} · 🔢 {prev_pc}"
                    )
                    st.markdown(
                        f'<div class="product-card-batch">{html.escape(batch_txt)}</div>',
                        unsafe_allow_html=True,
                    )

                st.markdown(
                    '<div class="product-card-qty-section">'
                    f'<span class="product-card-qty-section-title">'
                    f"{html.escape(t('order_qty_heading'))}</span></div>",
                    unsafe_allow_html=True,
                )
                with st.container(
                    horizontal=True,
                    gap="medium",
                    vertical_alignment="top",
                ):
                    with st.container():
                        st.number_input(
                            t("cartons"),
                            min_value=0,
                            step=1,
                            key=ct_key,
                            on_change=_on_qty_change,
                            args=(pid, "ct"),
                        )
                    with st.container():
                        st.number_input(
                            t("each_pcs"),
                            min_value=0,
                            step=1,
                            key=pc_key,
                            on_change=_on_qty_change,
                            args=(pid, "pc"),
                        )


def _enrich_cart_item(item: dict) -> dict:
    """补全购物车行的状态字段：分店ID + 生产日期（如有）。

    需求要求 cart 数据结构含：商品编号、商品名称、订购箱数、订购个数、
    分店ID、生产日期。前四者卡片已写入，这里统一补 branch / production_date。
    """
    item.setdefault("branch", st.session_state.get("branch", ""))
    item.setdefault("production_date", item.get("production_date", "") or "")
    return item


def _render_branch_cart_dock() -> None:
    """电商式购物车悬浮(PC)/吸底(手机)组件：实时摘要 + 去结算。

    用 st.container(key="sunshine_cart_dock") 渲染一个被 CSS 固定定位的容器，
    里面是真正的 Streamlit 按钮（点击触发 rerun 切换到结算页），
    因此既有原生交互、又有自定义悬浮布局。
    """
    cart = st.session_state.get("cart", [])
    # 暂存（商品卡片已填、尚未点"追加购物车"）的数量也并入悬浮车统计，
    # 这样用户边填边能在悬浮车看到实时变化；点"去结算"时再自动并入购物车。
    pending = _collect_qty_inputs()
    combined = list(cart) + list(pending)

    count = len(combined)
    total_ct = sum(int(it.get("qty_cartons", 0) or 0) for it in combined)
    total_pc = sum(int(it.get("qty_pcs", 0) or 0) for it in combined)
    total_units = total_ct + total_pc
    # 预估总金额：单价 × (箱数 + 个数)。手动商品价为 0 不计。
    total_amount = sum(
        float(it.get("price", 0) or 0)
        * (int(it.get("qty_cartons", 0) or 0) + int(it.get("qty_pcs", 0) or 0))
        for it in combined
    )
    amount_html = (
        f'<span class="scd-amount">💰 {total_amount:,.0f}</span>'
        if total_amount > 0
        else ""
    )

    # 用带边框容器（=stVerticalBlockBorderWrapper），CSS 通过
    # :has(.sunshine-cart-summary) 命中外层包裹元素并 position:fixed。
    # 命中点用"一定会渲染的可见内容"，比隐藏空标记 / .st-key- 类名都可靠。
    with st.container(key="sunshine_cart_dock", border=True):
        with st.container(
            horizontal=True,
            vertical_alignment="center",
            gap="small",
        ):
            st.markdown(
                '<div class="sunshine-cart-summary">'
                '<span class="scd-ico">🛒</span>'
                f'<span class="scd-count">{count}</span>'
                f'<span class="scd-meta">🧮 {total_units} · 📦 {total_ct} · 🔢 {total_pc}</span>'
                f"{amount_html}"
                "</div>",
                unsafe_allow_html=True,
            )
            if st.button(
                t("cart_checkout"),
                key="cart_dock_checkout",
                type="primary",
                disabled=(count == 0),
            ):
                # 先把暂存数量并入购物车，再进结算页，避免漏单。
                for item in pending:
                    st.session_state.cart.append(_enrich_cart_item(item))
                if pending:
                    _clear_qty_inputs()
                st.session_state["confirming"] = True
                st.rerun()

    # --- 终极固定方案：清除祖先的"包含块"属性 --------------------------
    # position:fixed 若失效（悬浮车随页面滚走而非钉在视口底部），几乎都是
    # 因为某个祖先元素带了 transform / filter / backdrop-filter / perspective /
    # will-change:transform / contain 等属性——这些会让 fixed 改为相对该祖先
    # 定位。CSS 难以预知是哪个祖先，故用 JS 从悬浮车向上遍历到 <body>，
    # 把这些属性逐个清除（不移动节点，原生按钮交互不受影响）。
    components.html(
        """
<script>
(function () {
  const doc = window.parent.document;
  const win = doc.defaultView;
  const PROPS = ['transform','filter','backdrop-filter','perspective','contain'];
  function neutralize() {
    const sum = doc.querySelector('.sunshine-cart-summary');
    if (!sum) return;
    const dock = sum.closest('[data-testid="stVerticalBlockBorderWrapper"]');
    if (!dock) return;
    let p = dock.parentElement;
    while (p && p !== doc.body && p.nodeType === 1) {
      const s = win.getComputedStyle(p);
      const wc = s.willChange || '';
      const needFix =
        s.transform !== 'none' ||
        s.filter !== 'none' ||
        (s.backdropFilter && s.backdropFilter !== 'none') ||
        s.perspective !== 'none' ||
        wc.indexOf('transform') !== -1 ||
        (s.contain && s.contain !== 'none' && s.contain !== 'normal');
      if (needFix) {
        p.style.setProperty('transform', 'none', 'important');
        p.style.setProperty('filter', 'none', 'important');
        p.style.setProperty('backdrop-filter', 'none', 'important');
        p.style.setProperty('perspective', 'none', 'important');
        p.style.setProperty('will-change', 'auto', 'important');
        p.style.setProperty('contain', 'none', 'important');
      }
      p = p.parentElement;
    }
  }
  neutralize();
  // Streamlit 会不断重渲染 DOM，用 MutationObserver 持续兜底。
  try {
    const mo = new MutationObserver(neutralize);
    mo.observe(doc.body, { childList: true, subtree: true });
  } catch (e) {}
})();
</script>
        """,
        height=0,
    )


# ----- Mode 1: browse / search / fill quantities --------------------------
def _branch_order_browse() -> None:
    render_page_heading(f"🛒 {t('nav_order')}")
    st.caption(t("branch_cart_sync_tip"))
    _render_active_arrival_banner()

    # --- Search ----------------------------------------------------------
    query_raw = st.text_input(
        t("search_product"),
        key="branch_search",
        placeholder=t("search_product"),
        help="Bluetooth scanner supported / 支持蓝牙扫码枪",
    )
    # Use live input directly. The previous debounce-based "stable query"
    # could get stuck on an older value because Streamlit has no timer-driven
    # rerun after user stops typing, which made searching appear broken.
    query = query_raw
    results = search_products(query, limit=500)

    # Reset pagination on query change
    last_query = st.session_state.get("_last_query")
    if last_query != query:
        st.session_state["_last_query"] = query
        st.session_state["search_page"] = 1
    if "search_page" not in st.session_state:
        st.session_state["search_page"] = 1

    PAGE_SIZE = 5

    if not results.empty:
        total = len(results)
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
        page = min(max(1, st.session_state["search_page"]), total_pages)
        st.session_state["search_page"] = page

        start_idx = (page - 1) * PAGE_SIZE
        end_idx = min(start_idx + PAGE_SIZE, total)
        page_results = results.iloc[start_idx:end_idx]

        st.caption(
            f"{t('showing')} {start_idx + 1}–{end_idx} {t('of')} {total} "
            f"{t('results_count')} · {t('page')} {page} {t('of')} {total_pages}"
        )

        for _, row in page_results.iterrows():
            _render_branch_product_card(row)

        # --- Pagination controls -----------------------------------------
        if total_pages > 1:
            st.markdown("")
            pc1, pc2, pc3 = st.columns([1, 2, 1])
            with pc1:
                if st.button(t("prev_page"), key="search_prev",
                             disabled=(page <= 1),
                             use_container_width=True):
                    st.session_state["search_page"] = page - 1
                    st.rerun()
            with pc2:
                st.markdown(
                    f"<div style='text-align:center;padding-top:6px;'>"
                    f"{t('page')} <b>{page}</b> {t('of')} <b>{total_pages}</b>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with pc3:
                if st.button(t("next_page"), key="search_next",
                             disabled=(page >= total_pages),
                             use_container_width=True):
                    st.session_state["search_page"] = page + 1
                    st.rerun()
    elif query:
        st.info(t("no_results"))

    # --- Manual add (still uses an instant form — single-item flow) ----
    with st.expander(t("manual_add")):
        with st.form("manual_form", clear_on_submit=True):
            mname = st.text_input(t("name"))
            c1, c2, c3 = st.columns(3)
            with c1: mct = st.number_input(t("cartons"), min_value=0, value=0, step=1)
            with c2: mpc = st.number_input(t("each_pcs"), min_value=0, value=0, step=1)
            with c3: mbar = st.text_input(t("barcode"))
            mprod = st.text_input(t("prod_date_opt"), placeholder="YYYY-MM-DD")
            if st.form_submit_button(t("add_to_cart"), type="primary"):
                if mname.strip() and (mct > 0 or mpc > 0):
                    st.session_state.cart.append(_enrich_cart_item({
                        "item_code": "",
                        "barcode":   mbar.strip(),
                        "name":      mname.strip(),
                        "unit":      "",
                        "price":     0.0,
                        "qty_cartons": int(mct),
                        "qty_pcs":     int(mpc),
                        "is_manual": 1,
                        "production_date": (mprod or "").strip(),
                    }))
                    st.success(f"✓ {mname}")
                    st.rerun()

    st.divider()

    # --- Batch action bar -----------------------------------------------
    selected = _collect_qty_inputs()
    cart = st.session_state.cart
    n_selected = len(selected)
    cart_count = len(cart)

    info_l, info_r = st.columns(2)
    with info_l:
        st.markdown(
            f"**🛒 {t('cart')}: {cart_count}** &nbsp;·&nbsp; "
            f"**📥 {t('items_count')}: {n_selected}**"
        )
    with info_r:
        if st.button(t("clear_qty"), key="clear_qty",
                     disabled=(n_selected == 0),
                     use_container_width=True):
            _clear_qty_inputs()
            st.success(t("qty_cleared"))
            st.rerun()

    # Big primary action: append all selected to cart
    if st.button(t("add_selected"), type="primary",
                 disabled=(n_selected == 0),
                 use_container_width=True):
        if n_selected == 0:
            st.warning(t("no_qty_selected"))
        else:
            for item in selected:
                st.session_state.cart.append(_enrich_cart_item(item))
            _clear_qty_inputs()
            st.success(t("added_n_items").format(n=n_selected))
            st.rerun()

    # 注：原先这里还有一个「去结算 / 购物车为空」按钮，与底部悬浮车的
    # 「去结算」功能重复，已移除，统一由悬浮车负责结算入口。

    # 电商式悬浮/吸底购物车（始终渲染，空车时按钮禁用）。
    _render_branch_cart_dock()


# ----- Mode 2: review cart, edit, then send -------------------------------
def _branch_order_confirm() -> None:
    render_page_heading(f"📝 {t('review_title')}")
    _render_active_arrival_banner()
    st.caption(t("review_subtitle"))

    cart = st.session_state.cart
    if not cart:
        # Edge case: cart got emptied while in confirm mode → bounce back
        st.session_state["confirming"] = False
        st.rerun()
        return

    # Render every cart line with editable qty + delete button
    for i, item in enumerate(cart):
        with st.container(border=True):
            top1, top2 = st.columns([4, 1])
            with top1:
                label = item["name"]
                if item.get("is_manual"):
                    label += f" 🖊️ {t('manual')}"
                st.markdown(f"**{label}**")
                sub = []
                if item["item_code"]: sub.append(f"📋 {item['item_code']}")
                if item["barcode"]:   sub.append(f"📊 {item['barcode']}")
                if item["unit"]:      sub.append(f"📦 {item['unit']}")
                if item.get("production_date"):
                    sub.append(f"🗓️ {item['production_date']}")
                if sub:
                    st.caption(" · ".join(sub))
            with top2:
                if st.button("🗑️", key=f"rm_confirm_{i}",
                             use_container_width=True):
                    st.session_state.cart.pop(i)
                    st.rerun()

            # Editable quantities — write back into cart on every render
            qc1, qc2 = st.columns(2)
            with qc1:
                new_ct = st.number_input(
                    t("cartons"), min_value=0,
                    value=int(item["qty_cartons"]), step=1,
                    key=f"confirm_ct_{i}",
                )
            with qc2:
                new_pc = st.number_input(
                    t("each_pcs"), min_value=0,
                    value=int(item["qty_pcs"]), step=1,
                    key=f"confirm_pc_{i}",
                )
            cart[i]["qty_cartons"] = int(new_ct)
            cart[i]["qty_pcs"]     = int(new_pc)

    # Drop any zero-quantity lines silently before showing summary
    cart[:] = [it for it in cart if it["qty_cartons"] > 0 or it["qty_pcs"] > 0]

    if not cart:
        st.warning(t("empty_cart"))
        if st.button(t("back_to_browse"), use_container_width=True):
            st.session_state["confirming"] = False
            st.rerun()
        return

    # Summary line
    st.divider()
    total_ct = sum(it["qty_cartons"] for it in cart)
    total_pc = sum(it["qty_pcs"] for it in cart)
    st.markdown(
        f"**{t('items_count')}: {len(cart)}** &nbsp;·&nbsp; "
        f"📦 **{total_ct}** {t('qty_cartons')} &nbsp;·&nbsp; "
        f"🔢 **{total_pc}** {t('qty_pcs')}"
    )

    # Action buttons
    bc1, bc2 = st.columns(2)
    with bc1:
        if st.button(t("back_to_browse"), use_container_width=True,
                     key="back_browse"):
            st.session_state["confirming"] = False
            st.rerun()
    with bc2:
        submit_busy = bool(st.session_state.get("_send_order_busy", False))
        if submit_busy:
            st.caption(f"⏳ {t('submit_busy')}")
        if st.button(
            t("send_order"),
            type="primary",
            use_container_width=True,
            key="send_order",
            disabled=submit_busy,
        ):
            if st.session_state.get("_send_order_busy", False):
                st.warning(t("submit_busy"))
                return
            st.session_state["_send_order_busy"] = True
            try:
                order_id = gen_order_id(st.session_state.branch)
                branch = st.session_state.branch
                # Snapshot the cart before clearing — we need it for the email
                cart_snapshot = [dict(it) for it in cart]

                # Defensive dedupe against accidental rapid double clicks.
                fp = _order_fingerprint(branch, cart_snapshot)
                last_fp = st.session_state.get("_last_order_submit_fp")
                last_ts = float(st.session_state.get("_last_order_submit_ts", 0.0))
                if last_fp == fp and (unix_ts() - last_ts) < 10:
                    st.warning(t("dup_submit_block"))
                    return

                with db_conn() as conn:
                    for item in cart_snapshot:
                        conn.execute("""
                            INSERT INTO orders
                            (order_id, branch, item_code, barcode, name, unit, price,
                             qty_cartons, qty_pcs, status, is_manual, order_date)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?, ?)
                        """, (
                            order_id, branch,
                            item["item_code"], item["barcode"], item["name"],
                            item["unit"], item["price"],
                            item["qty_cartons"], item["qty_pcs"],
                            item["is_manual"], now_str(),
                        ))

                audit_write(
                    "order_submit",
                    order_id=order_id,
                    branch=branch,
                    extra={"line_count": len(cart_snapshot)},
                )

                st.session_state["_last_order_submit_fp"] = fp
                st.session_state["_last_order_submit_ts"] = unix_ts()

                # Fire-and-forget email notification (warehouse + admin)
                try:
                    subject, body = build_new_order_email(
                        order_id, branch, cart_snapshot
                    )
                    new_order_xlsx = build_new_order_excel(
                        order_id, branch, cart_snapshot
                    )
                    notify(
                        "new_order",
                        subject,
                        body,
                        attachments=[(
                            f"{order_id}.xlsx",
                            new_order_xlsx,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )],
                    )
                    create_notification(
                        "new_order",
                        f"新订单待发货 · {branch}",
                        f"{branch} 提交了新订单 {order_id}，请在「待发货订单」处理。",
                        target_role=Role.WAREHOUSE,
                        order_id=order_id,
                    )
                    create_notification(
                        "new_order",
                        f"新订单提交 · {branch}",
                        f"{branch} 提交了新订单 {order_id}（{len(cart_snapshot)} 行）。",
                        target_role=Role.ADMIN,
                        order_id=order_id,
                    )
                except Exception as e:
                    log_exception("branch_submit_new_order_email", e)

                st.session_state.cart = []
                _persist_branch_cart()
                st.session_state["confirming"] = False
                st.session_state["last_submitted_order_id"] = order_id
                _set_session_page_for_app_nav("order_done")
                _clear_qty_inputs()
                st.rerun()
            finally:
                st.session_state["_send_order_busy"] = False


def page_branch_order_done() -> None:
    render_page_heading(f"✅ {t('order_sent_done')}")
    oid = st.session_state.get("last_submitted_order_id", "")
    if oid:
        st.success(f"{t('order_id')}: `{oid}`")
    st.info(t("order_sent_tip"))
    c1, c2 = st.columns(2)
    with c1:
        if st.button(t("create_new_order"), use_container_width=True, type="primary"):
            _set_session_page_for_app_nav("order")
            st.rerun()
    with c2:
        if st.button(t("view_my_orders_btn"), use_container_width=True):
            _set_session_page_for_app_nav("my_orders")
            st.rerun()


def page_branch_my_orders() -> None:
    render_page_heading(f"📋 {t('nav_my_orders')}")
    branch = st.session_state.branch
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE branch = ? ORDER BY order_date DESC",
            conn, params=(branch,),
        )

    if df.empty:
        st.info(t("no_data"))
        return

    for order_id, group in df.groupby("order_id", sort=False):
        status = group["status"].iloc[0]
        order_date = group["order_date"].iloc[0]

        with st.expander(
            f"📦 {order_id} · {order_date[:16]} · {len(group)} items",
            expanded=(status == OrderStatus.DISPATCHED),
        ):
            st.markdown(status_pill(status), unsafe_allow_html=True)
            st.markdown("")

            show = group[["name", "unit", "qty_cartons", "qty_pcs"]].rename(columns={
                "name": t("name"), "unit": t("unit"),
                "qty_cartons": t("qty_cartons"), "qty_pcs": t("qty_pcs"),
            })
            if status in (OrderStatus.DISPATCHED, OrderStatus.RECEIVED):
                # Show what the warehouse actually shipped vs what was ordered
                show[f"🚚 {t('qty_cartons')}"] = group["dispatch_cartons"].fillna("-").values
                show[f"🚚 {t('qty_pcs')}"] = group["dispatch_pcs"].fillna("-").values
                show[t("actual_cartons")] = group["actual_cartons"].fillna("-").values
                show[t("actual_pcs")] = group["actual_pcs"].fillna("-").values
            st.dataframe(show, use_container_width=True, hide_index=True)

            if status == OrderStatus.DISPATCHED:
                _render_receive_form(order_id, group)


def _render_receive_form(order_id: str, group: pd.DataFrame) -> None:
    render_section_title(f"✅ {t('receive')}")
    receive_q = st.text_input(
        f"🔍 {t('search')} ({t('barcode')}/{t('name')})",
        key=f"recv_q_{order_id}",
        help="Scanner supported / 支持扫码枪",
    )
    with st.form(f"recv_form_{order_id}"):
        actuals: dict[int, tuple[int, int]] = {}
        for _, row in group.iterrows():
            line_id = int(row["id"])
            if receive_q:
                q = receive_q.lower().strip()
                hay = f"{row['name']} {row['barcode']} {row['item_code']}".lower()
                if q not in hay:
                    continue

            # Use dispatch_cartons/pcs as the baseline if set; fall back to
            # qty_cartons/pcs for legacy orders dispatched before this feature.
            disp_ct = row["dispatch_cartons"]
            disp_pc = row["dispatch_pcs"]
            if pd.isna(disp_ct) or disp_ct is None:
                disp_ct = int(row["qty_cartons"])
            else:
                disp_ct = int(disp_ct)
            if pd.isna(disp_pc) or disp_pc is None:
                disp_pc = int(row["qty_pcs"])
            else:
                disp_pc = int(disp_pc)

            st.markdown(f"**{row['name']}**")
            # Show both: what was ordered AND what the warehouse dispatched.
            # If they differ, the difference is already a known short by the
            # warehouse — the branch only needs to compare against dispatch.
            if disp_ct != int(row["qty_cartons"]) or disp_pc != int(row["qty_pcs"]):
                st.caption(
                    f"📋 {t('ordered_qty')}: {row['qty_cartons']} {t('qty_cartons')} / "
                    f"{row['qty_pcs']} {t('qty_pcs')}  &nbsp;·&nbsp;  "
                    f"🚚 {t('dispatched')}: **{disp_ct}** {t('qty_cartons')} / "
                    f"**{disp_pc}** {t('qty_pcs')}"
                )
            else:
                st.caption(
                    f"🚚 {t('dispatched')}: {disp_ct} {t('qty_cartons')} / "
                    f"{disp_pc} {t('qty_pcs')}"
                )
            c1, c2 = st.columns(2)
            with c1:
                ac = st.number_input(
                    t("actual_cartons"), min_value=0,
                    value=disp_ct, step=1, key=f"ac_{line_id}",
                )
            with c2:
                ap = st.number_input(
                    t("actual_pcs"), min_value=0,
                    value=disp_pc, step=1, key=f"ap_{line_id}",
                )
            actuals[line_id] = (ac, ap, disp_ct, disp_pc)
        overall_remark = st.text_area(t("remarks"), key=f"rem_{order_id}")
        if st.form_submit_button(f"✅ {t('confirm_receive')}", type="primary"):
            any_short = False
            short_items: list[dict] = []  # collected for the email below
            branch_name = ""
            with db_conn() as conn:
                for _, row in group.iterrows():
                    line_id = int(row["id"])
                    if line_id in actuals:
                        ac, ap, disp_ct, disp_pc = actuals[line_id]
                    else:
                        # Filtered-out lines auto-confirm at the dispatched qty
                        disp_ct_raw = row["dispatch_cartons"]
                        disp_pc_raw = row["dispatch_pcs"]
                        disp_ct = (int(disp_ct_raw) if not (pd.isna(disp_ct_raw) or disp_ct_raw is None)
                                   else int(row["qty_cartons"]))
                        disp_pc = (int(disp_pc_raw) if not (pd.isna(disp_pc_raw) or disp_pc_raw is None)
                                   else int(row["qty_pcs"]))
                        ac, ap = disp_ct, disp_pc
                    conn.execute("""
                        UPDATE orders SET
                            status='Received',
                            actual_cartons=?, actual_pcs=?,
                            receive_remarks=?, receive_date=?
                        WHERE id=?
                    """, (ac, ap, overall_remark, now_str(), line_id))
                    # Shortage is computed against DISPATCHED quantity, not
                    # ordered quantity — the warehouse already accounts for
                    # any "couldn't ship" gap by entering a smaller dispatch
                    # number, so the shortage record is purely about
                    # transit/branch-level discrepancy.
                    short_ct = max(0, disp_ct - ac)
                    short_pc = max(0, disp_pc - ap)
                    if short_ct > 0 or short_pc > 0:
                        any_short = True
                        branch_name = row["branch"]
                        conn.execute("""
                            INSERT INTO shortages
                            (order_id, order_line_id, branch, item_code, barcode,
                             name, unit, short_cartons, short_pcs, status,
                             reported_date, branch_remarks)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Open', ?, ?)
                        """, (
                            order_id, line_id, row["branch"],
                            row["item_code"], row["barcode"], row["name"],
                            row["unit"], short_ct, short_pc,
                            now_str(), overall_remark,
                        ))
                        short_items.append({
                            "barcode": row["barcode"] or "",
                            "name": row["name"],
                            "short_cartons": short_ct,
                            "short_pcs": short_pc,
                        })
            audit_write(
                "receive_confirm",
                order_id=order_id,
                branch=branch_name or st.session_state.get("branch"),
                extra={
                    "line_count": int(len(group)),
                    "any_shortage": bool(any_short),
                },
            )
            # Send ONE shortage email summarizing all short lines for this
            # receipt — never one email per item.
            if any_short and short_items:
                try:
                    subject, body = build_shortage_email(
                        order_id, branch_name, short_items
                    )
                    notify("shortage", subject, body)
                    create_notification(
                        "shortage",
                        f"收货缺货上报 · {branch_name}",
                        f"订单 {order_id} 收货发现缺货 {len(short_items)} 项，请在「缺货通知」处理。",
                        target_role=Role.WAREHOUSE,
                        order_id=order_id,
                    )
                    create_notification(
                        "shortage",
                        f"缺货上报 · {branch_name}",
                        f"订单 {order_id} 收货发现缺货 {len(short_items)} 项。",
                        target_role=Role.ADMIN,
                        order_id=order_id,
                    )
                except Exception as e:
                    log_exception("branch_shortage_notify", e)
            create_notification(
                "received",
                f"订单已收货 · {branch_name or st.session_state.branch}",
                f"分店已确认订单 {order_id} 收货。请查看「出库历史」或订单明细。",
                target_role=Role.WAREHOUSE,
                order_id=order_id,
            )
            create_notification(
                "received",
                f"订单收货完成 · {branch_name or st.session_state.branch}",
                f"订单 {order_id} 已确认收货。",
                target_role=Role.ADMIN,
                order_id=order_id,
            )
            if any_short:
                st.warning(f"⚠️ {t('shortage_alert')}")
            st.success(f"✅ {t('receipt_done')}")
            st.rerun()


def page_branch_shortages() -> None:
    render_page_heading(f"🔔 {t('nav_my_short')}")
    branch = st.session_state.branch
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM shortages WHERE branch = ? ORDER BY reported_date DESC",
            conn, params=(branch,),
        )

    if df.empty:
        st.info(t("no_shortages"))
        return

    for _, row in df.iterrows():
        with st.container(border=True):
            c1, c2 = st.columns([4, 1])
            with c1:
                st.markdown(f"**{row['name']}** · {row['order_id']}")
                st.caption(
                    f"{t('shortage_qty')}: {row['short_cartons']} {t('qty_cartons')} / "
                    f"{row['short_pcs']} {t('qty_pcs')} · {row['reported_date'][:16]}"
                )
                if row["warehouse_reply"]:
                    st.info(f"💬 **{t('warehouse_reply')}:** {row['warehouse_reply']}")
            with c2:
                st.markdown(status_pill(row["status"]), unsafe_allow_html=True)
            if row["status"] == ShortageStatus.RESENDING:
                if st.button(t("confirm_resend"), key=f"cr_{row['id']}"):
                    with db_conn() as conn:
                        conn.execute(
                            "UPDATE shortages SET status='Resolved', resolved_date=? WHERE id=?",
                            (now_str(), row["id"]),
                        )
                    st.rerun()


# =========================================================================
# WAREHOUSE PAGES
# =========================================================================
def page_warehouse_pending() -> None:
    render_page_heading(f"📦 {t('nav_pending')}")
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE status='Pending' ORDER BY order_date ASC", conn,
        )

    if df.empty:
        st.info(t("no_pending"))
        return

    branch_options = sorted(df["branch"].dropna().unique().tolist())
    if "_pending_branches" not in st.session_state:
        st.session_state["_pending_branches"] = branch_options.copy()
    # Build helper date column once so filters and sort stay fast.
    date_series = pd.to_datetime(
        df["order_date"].fillna("").str.slice(0, 10),
        format="%Y-%m-%d",
        errors="coerce",
    )
    valid_dates = date_series.dropna()
    min_date = (
        valid_dates.min().date() if not valid_dates.empty else datetime.now().date()
    )
    max_date = (
        valid_dates.max().date() if not valid_dates.empty else datetime.now().date()
    )
    if "_pending_date_from" not in st.session_state:
        st.session_state["_pending_date_from"] = min_date
    if "_pending_date_to" not in st.session_state:
        st.session_state["_pending_date_to"] = max_date
    if "_pending_expand_mode" not in st.session_state:
        st.session_state["_pending_expand_mode"] = "smart"

    f1, f2, f3 = st.columns([2, 2, 2])
    with f1:
        pending_kw = st.text_input(
            t("filter_order_keyword"), key="pending_kw",
        ).strip().lower()
    with f2:
        selected_branches = st.multiselect(
            t("filter_branches"),
            options=branch_options,
            key="_pending_branches",
        )
    with f3:
        expand_match_only = st.checkbox(
            t("expand_match_only"), value=True, key="_pending_expand_only",
        )

    q1, q2, q3 = st.columns([1, 1, 1])
    with q1:
        if st.button(t("all_branches"), key="pending_all_branches", use_container_width=True):
            st.session_state["_pending_branches"] = branch_options.copy()
            st.rerun()
    with q2:
        if st.button(t("recent_7_days"), key="pending_recent_7", use_container_width=True):
            start_7 = max(min_date, max_date - timedelta(days=6))
            st.session_state["_pending_date_from"] = start_7
            st.session_state["_pending_date_to"] = max_date
            st.rerun()
    with q3:
        if st.button(t("all_dates"), key="pending_all_dates", use_container_width=True):
            st.session_state["_pending_date_from"] = min_date
            st.session_state["_pending_date_to"] = max_date
            st.rerun()

    d1, d2 = st.columns(2)
    with d1:
        date_from = st.date_input(
            t("filter_date_from"),
            value=st.session_state.get("_pending_date_from", min_date),
            min_value=min_date,
            max_value=max_date,
            key="_pending_date_from",
        )
    with d2:
        date_to = st.date_input(
            t("filter_date_to"),
            value=st.session_state.get("_pending_date_to", max_date),
            min_value=min_date,
            max_value=max_date,
            key="_pending_date_to",
        )
    if date_from > date_to:
        date_from, date_to = date_to, date_from

    e1, e2 = st.columns(2)
    with e1:
        if st.button(t("expand_all_orders"), key="pending_expand_all", use_container_width=True):
            st.session_state["_pending_expand_mode"] = "all"
            st.rerun()
    with e2:
        if st.button(t("collapse_all_orders"), key="pending_collapse_all", use_container_width=True):
            st.session_state["_pending_expand_mode"] = "none"
            st.rerun()

    filtered = df[df["branch"].isin(selected_branches)] if selected_branches else df.iloc[0:0]
    date_mask = (
        date_series.dt.date >= date_from
    ) & (
        date_series.dt.date <= date_to
    )
    filtered = filtered[date_mask.fillna(False)]
    if pending_kw:
        mask = (
            filtered["order_id"].fillna("").str.lower().str.contains(pending_kw, regex=False)
            | filtered["name"].fillna("").str.lower().str.contains(pending_kw, regex=False)
            | filtered["barcode"].fillna("").str.lower().str.contains(pending_kw, regex=False)
        )
        filtered = filtered[mask]

    selected_count = len(selected_branches)
    branch_part = (
        f"{selected_count}/{len(branch_options)} {t('branch')}"
        if branch_options else f"0 {t('branch')}"
    )
    date_part = f"{date_from} ~ {date_to}"
    keyword_part = pending_kw if pending_kw else "-"
    st.caption(
        f"🔎 {t('filter_summary')}: "
        f"🏪 {branch_part} · "
        f"📅 {date_part} · "
        f"🔤 {t('keyword_label')}: {keyword_part}"
    )

    if filtered.empty:
        st.info(t("no_data"))
        return

    # ----- Top-level "download all pending" --------------------------
    n_orders = filtered["order_id"].nunique()
    n_branches = filtered["branch"].nunique()
    cap_l, cap_r = st.columns([3, 2])
    with cap_l:
        st.caption(
            f"📦 {n_orders} {t('orders')} · 🏪 {n_branches} {t('branch')} · "
            f"{len(filtered)} {t('lines')}"
            if "orders" in T else f"📦 {n_orders} orders · {n_branches} branches · {len(filtered)} lines"
        )
    with cap_r:
        if st.button(t("dl_all_pending"), key="dl_all_pending",
                     use_container_width=True):
            bio = export_all_pending_picking()
            if bio is None:
                st.warning(t("no_data"))
            else:
                stamp = datetime.now().strftime("%Y%m%d_%H%M")
                st.session_state["_dl_all"] = (bio.getvalue(), f"picking_all_{stamp}.xlsx")
        # show download button if file ready
        if "_dl_all" in st.session_state:
            data, fname = st.session_state["_dl_all"]
            st.download_button(
                f"⬇️ {fname}", data=data, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_all_btn",
            )

    st.divider()

    # ----- Per-branch sections ---------------------------------------
    for branch in sorted(filtered["branch"].unique()):
        bdf = filtered[filtered["branch"] == branch]
        head_l, head_r = st.columns([3, 2])
        with head_l:
            render_section_title(
                f"🏪 {branch} ({bdf['order_id'].nunique()} orders · {len(bdf)} lines)"
            )
        with head_r:
            dl_key = f"dl_branch_{branch}"
            if st.button(t("dl_branch_all"), key=dl_key, use_container_width=True):
                bio = export_branch_pending_picking(branch)
                if bio is None:
                    st.warning(t("no_data"))
                else:
                    stamp = datetime.now().strftime("%Y%m%d_%H%M")
                    safe_branch = branch.replace(" ", "_").replace("/", "-")
                    st.session_state[f"_dl_b_{branch}"] = (
                        bio.getvalue(), f"picking_{safe_branch}_{stamp}.xlsx",
                    )
            if f"_dl_b_{branch}" in st.session_state:
                data, fname = st.session_state[f"_dl_b_{branch}"]
                st.download_button(
                    f"⬇️ {fname}", data=data, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key=f"dl_b_btn_{branch}",
                )

        # ----- Each individual order ---------------------------------
        for order_id, group in bdf.groupby("order_id", sort=False):
            expand_mode = st.session_state.get("_pending_expand_mode", "smart")
            if expand_mode == "all":
                expanded = True
            elif expand_mode == "none":
                expanded = False
            else:
                expanded = bool(expand_match_only and pending_kw)
            with st.expander(
                f"📦 {order_id} · {group['order_date'].iloc[0][:16]} · {len(group)} items",
                expanded=expanded,
            ):
                st.caption(f"💡 {t('dispatch_hint')}")
                quick1, quick2 = st.columns(2)
                with quick1:
                    if st.button(
                        t("dispatch_fill_ordered"),
                        key=f"fill_ordered_{order_id}",
                        use_container_width=True,
                    ):
                        for _, row in group.iterrows():
                            line_id = int(row["id"])
                            st.session_state[f"dc_{line_id}"] = int(row["qty_cartons"])
                            st.session_state[f"dp_qty_{line_id}"] = int(row["qty_pcs"])
                        st.rerun()
                with quick2:
                    if st.button(
                        t("dispatch_clear_all"),
                        key=f"clear_dispatch_{order_id}",
                        use_container_width=True,
                    ):
                        for _, row in group.iterrows():
                            line_id = int(row["id"])
                            st.session_state[f"dc_{line_id}"] = 0
                            st.session_state[f"dp_qty_{line_id}"] = 0
                        st.rerun()

                # Editable dispatch-quantity form for each line.
                # Default values = ordered quantities, so the common case
                # (full dispatch) is one click. Warehouse staff can lower
                # any number when stock is short.
                with st.form(f"dispatch_form_{order_id}"):
                    # Header row
                    h1, h2, h3, h4, h5 = st.columns([3, 2, 2, 2, 2])
                    with h1: st.markdown(f"**{t('name')}**")
                    with h2: st.markdown(f"**{t('barcode')}**")
                    with h3: st.markdown(f"**{t('ordered_qty')}**")
                    with h4: st.markdown(f"**{t('dispatch_cartons')}**")
                    with h5: st.markdown(f"**{t('dispatch_pcs')}**")
                    st.divider()

                    dispatch_inputs: dict[int, tuple[int, int]] = {}
                    for _, row in group.iterrows():
                        line_id = int(row["id"])
                        ordered_ct = int(row["qty_cartons"])
                        ordered_pc = int(row["qty_pcs"])

                        c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 2])
                        with c1:
                            label = row["name"]
                            if row["item_code"]:
                                st.markdown(f"**{label}**")
                                st.caption(f"📋 {row['item_code']}")
                            else:
                                st.markdown(f"**{label}**")
                        with c2:
                            st.markdown(f"`{row['barcode'] or '-'}`")
                            if row["unit"]:
                                st.caption(f"📦 {row['unit']}")
                        with c3:
                            st.markdown(
                                f"📦 {ordered_ct} {t('qty_cartons')}<br>"
                                f"🔢 {ordered_pc} {t('qty_pcs')}",
                                unsafe_allow_html=True,
                            )
                        with c4:
                            dc = st.number_input(
                                t("dispatch_cartons"),
                                min_value=0, value=ordered_ct, step=1,
                                key=f"dc_{line_id}",
                                label_visibility="collapsed",
                            )
                        with c5:
                            dp = st.number_input(
                                t("dispatch_pcs"),
                                min_value=0, value=ordered_pc, step=1,
                                key=f"dp_qty_{line_id}",
                                label_visibility="collapsed",
                            )
                        dispatch_inputs[line_id] = (int(dc), int(dp))
                        changed = (int(dc) != ordered_ct) or (int(dp) != ordered_pc)
                        if changed:
                            st.caption(f"🟡 {t('dispatch_changed_row')}")

                        # Inline warning when the staff is shipping less
                        # than ordered — helpful but non-blocking.
                        if int(dc) < ordered_ct or int(dp) < ordered_pc:
                            st.caption(
                                f"⚠️ {t('less_than_ordered')}: "
                                f"📦 {ordered_ct - int(dc)} · "
                                f"🔢 {ordered_pc - int(dp)}"
                            )

                    st.divider()

                    # Action row inside the form
                    ac1, ac2 = st.columns(2)
                    with ac1:
                        # Picking-slip download stays as a regular button OUTSIDE
                        # the form — but we can't put it inside this form easily.
                        # Show a placeholder instructional caption.
                        st.caption(
                            f"💡 {t('dl_this_order')} → "
                            f"{t('back_to_browse').replace('◀', '').strip()}"
                        )
                    with ac2:
                        if st.form_submit_button(
                            f"🚚 {t('mark_dispatched')}",
                            type="primary", use_container_width=True,
                        ):
                            busy_orders = st.session_state.setdefault(
                                "_dispatch_busy_orders", {}
                            )
                            if busy_orders.get(order_id):
                                st.warning(t("dispatch_busy"))
                            else:
                                busy_orders[order_id] = True
                                try:
                                    stamp = now_str()
                                    with db_conn() as conn:
                                        insufficient: list[str] = []
                                        # 防跨行超卖：同一商品若出现在多行，按"累计需求"对
                                        # 可用库存做递减校验；无库存记录的商品按 0 计，
                                        # 任何正向发货都会被拦截（避免发出不存在的库存）。
                                        avail: dict[str, list[int]] = {}
                                        for _, row in group.iterrows():
                                            line_id = int(row["id"])
                                            dc, dp = dispatch_inputs.get(line_id, (0, 0))
                                            if int(dc) == 0 and int(dp) == 0:
                                                continue
                                            ic = str(row.get("item_code", "") or "")
                                            bc = str(row.get("barcode", "") or "")
                                            nm = str(row.get("name", "") or "")
                                            key = _inventory_item_key(ic, bc, nm)
                                            if key not in avail:
                                                inv_row = _get_inventory_row(conn, ic, bc, nm)
                                                avail[key] = [
                                                    int(inv_row["stock_cartons"] or 0) if inv_row else 0,
                                                    int(inv_row["stock_pcs"] or 0) if inv_row else 0,
                                                ]
                                            rem_ct, rem_pc = avail[key]
                                            if int(dc) > rem_ct or int(dp) > rem_pc:
                                                insufficient.append(
                                                    f"{nm} (剩余 {rem_ct}/{rem_pc}, 发货 {int(dc)}/{int(dp)})"
                                                )
                                            else:
                                                avail[key][0] = rem_ct - int(dc)
                                                avail[key][1] = rem_pc - int(dp)
                                        if insufficient:
                                            rows = []
                                        else:
                                            # Write dispatch quantities for each line, and
                                            # flip status + dispatch_date in one transaction.
                                            for line_id, (dc, dp) in dispatch_inputs.items():
                                                conn.execute("""
                                                    UPDATE orders SET
                                                        status='Dispatched',
                                                        dispatch_date=?,
                                                        dispatch_cartons=?,
                                                        dispatch_pcs=?
                                                    WHERE id=?
                                                """, (stamp, dc, dp, line_id))
                                            # Deduct inventory by dispatched quantity.
                                            for _, row in group.iterrows():
                                                line_id = int(row["id"])
                                                dc, dp = dispatch_inputs.get(line_id, (0, 0))
                                                if int(dc) == 0 and int(dp) == 0:
                                                    continue
                                                _apply_inventory_change(
                                                    conn,
                                                    txn_type="OUT",
                                                    item_code=str(row.get("item_code", "") or ""),
                                                    barcode=str(row.get("barcode", "") or ""),
                                                    name=str(row.get("name", "") or ""),
                                                    unit=str(row.get("unit", "") or ""),
                                                    change_ct=-int(dc),
                                                    change_pc=-int(dp),
                                                    order_id=order_id,
                                                    operator="warehouse_dispatch",
                                                )
                                            conn.execute(
                                                "INSERT INTO shipments "
                                                "(order_id, branch, dispatch_date, dispatched_by) "
                                                "VALUES (?, ?, ?, 'warehouse')",
                                                (order_id, branch, stamp),
                                            )
                                            # Re-read the rows we just updated so the email
                                            # body reflects the saved dispatch quantities.
                                            rows = conn.execute(
                                                "SELECT name, unit, qty_cartons, qty_pcs, "
                                                "       dispatch_cartons, dispatch_pcs "
                                                "FROM orders WHERE order_id = ?",
                                                (order_id,),
                                            ).fetchall()
                                    if insufficient:
                                        st.error(
                                            f"{t('inv_low_stock')}: " + "；".join(insufficient[:5])
                                        )
                                        st.stop()
                                    try:
                                        lines_for_email = [dict(r) for r in rows]
                                        subject, body = build_dispatched_email(
                                            order_id, branch, lines_for_email
                                        )
                                        dispatch_xlsx = build_dispatched_excel(
                                            order_id, branch, lines_for_email
                                        )
                                        # Optional: CC the branch's specific email
                                        cfg = load_email_config()
                                        branch_email = (cfg.get("branch_emails") or {}
                                                        ).get(branch, "").strip()
                                        extra = [branch_email] if branch_email else []
                                        notify(
                                            "dispatched",
                                            subject,
                                            body,
                                            extra_to=extra,
                                            attachments=[(
                                                f"{order_id}.xlsx",
                                                dispatch_xlsx,
                                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            )],
                                        )
                                        create_notification(
                                            "dispatched",
                                            f"订单已发货 · {branch}",
                                            f"订单 {order_id} 已由仓库发货，请在「我的订单」确认收货。",
                                            target_role=Role.BRANCH,
                                            target_branch=branch,
                                            order_id=order_id,
                                        )
                                        create_notification(
                                            "dispatched",
                                            f"仓库已发货 · {branch}",
                                            f"订单 {order_id} 已标记发货。",
                                            target_role=Role.ADMIN,
                                            order_id=order_id,
                                        )
                                    except Exception as e:
                                        log_exception("warehouse_dispatch_notify", e)
                                    st.session_state.pop(f"_dl_o_{order_id}", None)
                                    st.success(f"✅ {t('shipment_marked')}")
                                    st.rerun()
                                finally:
                                    busy_orders[order_id] = False

                # ----- Picking-slip download (separate, outside form) -----
                dl_o_key = f"dl_o_{order_id}"
                if st.button(t("dl_this_order"), key=dl_o_key,
                             use_container_width=True):
                    bio = export_single_order_picking(order_id)
                    if bio is None:
                        st.warning(t("no_data"))
                    else:
                        st.session_state[f"_dl_o_{order_id}"] = (
                            bio.getvalue(),
                            f"picking_{order_id}.xlsx",
                        )
                if f"_dl_o_{order_id}" in st.session_state:
                    data, fname = st.session_state[f"_dl_o_{order_id}"]
                    st.download_button(
                        f"⬇️ {fname}", data=data, file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key=f"dl_o_btn_{order_id}",
                    )


def page_warehouse_shortages() -> None:
    render_page_heading(f"🔔 {t('nav_short_in')}")
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM shortages WHERE status='Open' ORDER BY reported_date DESC", conn,
        )

    if df.empty:
        st.info(t("no_shortages"))
        return

    for _, row in df.iterrows():
        with st.container(border=True):
            st.markdown(f"**{row['name']}** · 🏪 {row['branch']}")
            st.caption(
                f"{t('order_id')}: {row['order_id']} · "
                f"{t('shortage_qty')}: {row['short_cartons']} {t('qty_cartons')} / "
                f"{row['short_pcs']} {t('qty_pcs')} · {row['reported_date'][:16]}"
            )
            if row["branch_remarks"]:
                st.caption(f"📝 {row['branch_remarks']}")
            reply = st.text_input(t("warehouse_reply"), key=f"wr_{row['id']}")
            c1, c2 = st.columns(2)
            with c1:
                if st.button(f"♻️ {t('resend')}", key=f"rs_{row['id']}",
                             use_container_width=True):
                    with db_conn() as conn:
                        conn.execute(
                            "UPDATE shortages SET status='Resending', warehouse_reply=? WHERE id=?",
                            (reply or "Will resend", row["id"]),
                        )
                    create_notification(
                        "shortage_reply",
                        f"缺货处理更新 · {row['branch']}",
                        f"订单 {row['order_id']} 的缺货项「{row['name']}」已标记为补发中。",
                        target_role=Role.BRANCH,
                        target_branch=row["branch"],
                        order_id=row["order_id"],
                    )
                    st.rerun()
            with c2:
                if st.button(f"❌ {t('mark_oos')}", key=f"oos_{row['id']}",
                             use_container_width=True):
                    with db_conn() as conn:
                        conn.execute(
                            "UPDATE shortages SET status='Out of Stock', warehouse_reply=?, "
                            "resolved_date=? WHERE id=?",
                            (reply or "Out of stock", now_str(), row["id"]),
                        )
                    create_notification(
                        "shortage_reply",
                        f"缺货处理更新 · {row['branch']}",
                        f"订单 {row['order_id']} 的缺货项「{row['name']}」已标记为缺货。",
                        target_role=Role.BRANCH,
                        target_branch=row["branch"],
                        order_id=row["order_id"],
                    )
                    st.rerun()


def page_warehouse_dispatch_history() -> None:
    """Browse dispatched orders by date. Joins shipments with orders to show
    what was sent on which day and which branch received it."""
    render_page_heading(f"📜 {t('dispatch_history')}")

    # ----- Date filter -------------------------------------------------
    with db_conn() as conn:
        # Pull list of distinct dispatch dates so we know what's available
        date_rows = conn.execute(
            "SELECT DISTINCT date(dispatch_date) AS d "
            "FROM shipments WHERE dispatch_date IS NOT NULL "
            "ORDER BY d DESC"
        ).fetchall()
    available_dates = [r["d"] for r in date_rows]

    if not available_dates:
        st.info(t("no_dispatch_history"))
        return

    # Default to the most recent date
    default_date = st.session_state.get("_dh_date") or available_dates[0]

    fc1, fc2, fc3 = st.columns([2, 1, 2])
    with fc1:
        # Use date_input with min/max so only valid dates can be picked
        try:
            picked = st.date_input(
                t("select_date"),
                value=datetime.strptime(default_date, "%Y-%m-%d").date(),
                key="_dh_date_input",
            )
            picked_date = picked.strftime("%Y-%m-%d")
        except Exception:
            picked_date = default_date
    with fc2:
        if st.button(t("today_btn"), key="_dh_today",
                     use_container_width=True):
            st.session_state["_dh_date"] = today_str()
            st.rerun()
    with fc3:
        show_all = st.checkbox(t("all_dates"), key="_dh_all")

    # ----- Query dispatched orders for the chosen date(s) --------------
    with db_conn() as conn:
        if show_all:
            ship_df = pd.read_sql_query(
                "SELECT * FROM shipments ORDER BY dispatch_date DESC", conn,
            )
        else:
            ship_df = pd.read_sql_query(
                "SELECT * FROM shipments WHERE date(dispatch_date) = ? "
                "ORDER BY dispatch_date DESC",
                conn, params=(picked_date,),
            )

    if ship_df.empty:
        st.info(t("no_dispatch_history"))
        return

    # ----- Summary metrics ---------------------------------------------
    n_orders = ship_df["order_id"].nunique()
    n_branches = ship_df["branch"].nunique()
    label_when = t("all_dates") if show_all else picked_date
    st.caption(
        f"📅 {label_when} · 🚚 {n_orders} {t('orders')} · "
        f"🏪 {n_branches} {t('branch')}"
    )

    # ----- Group by branch, then list each dispatched order -----------
    for branch in sorted(ship_df["branch"].unique()):
        bdf = ship_df[ship_df["branch"] == branch]
        render_section_title(f"🏪 {branch} ({len(bdf)} {t('orders')})")

        for _, ship in bdf.iterrows():
            order_id = ship["order_id"]
            dispatch_date = ship["dispatch_date"]

            # Pull the line items + downstream receipt info for this order
            with db_conn() as conn:
                lines = pd.read_sql_query(
                    "SELECT name, item_code, barcode, unit, "
                    "       qty_cartons, qty_pcs, "
                    "       dispatch_cartons, dispatch_pcs, "
                    "       actual_cartons, actual_pcs, "
                    "       status, receive_date "
                    "FROM orders WHERE order_id = ? ORDER BY name",
                    conn, params=(order_id,),
                )

            # Helper: effective dispatched qty for a row (fallback to ordered
            # qty for legacy rows that have no dispatch_* set).
            def _disp(r, col_dispatch, col_qty):
                val = r[col_dispatch]
                if pd.isna(val) or val is None:
                    return int(r[col_qty])
                return int(val)

            # Determine receipt status across all lines
            statuses = set(lines["status"].tolist())
            if statuses == {OrderStatus.RECEIVED}:
                # All received — check whether any short vs DISPATCHED qty
                any_short = any(
                    (pd.notna(r["actual_cartons"]) and
                     r["actual_cartons"] < _disp(r, "dispatch_cartons", "qty_cartons")) or
                    (pd.notna(r["actual_pcs"]) and
                     r["actual_pcs"] < _disp(r, "dispatch_pcs", "qty_pcs"))
                    for _, r in lines.iterrows()
                )
                receipt_label = (
                    t("partial_received") if any_short else t("fully_received")
                )
                receipt_color = "#fbc02d" if any_short else "#388e3c"
            else:
                receipt_label = t("awaiting_receipt")
                receipt_color = "#1976d2"

            with st.expander(
                f"📦 {order_id} · 🚚 {dispatch_date[:16]} · "
                f"{len(lines)} items · {receipt_label}"
            ):
                # Receipt-status pill
                st.markdown(
                    f"<span class='pill' style='background:{receipt_color};'>"
                    f"{receipt_label}</span>",
                    unsafe_allow_html=True,
                )
                st.markdown("")

                show = lines[[
                    "name", "item_code", "barcode", "unit",
                    "qty_cartons", "qty_pcs",
                    "dispatch_cartons", "dispatch_pcs",
                    "actual_cartons", "actual_pcs",
                ]].fillna("-").rename(columns={
                    "name":             t("name"),
                    "item_code":        t("item_code"),
                    "barcode":          t("barcode"),
                    "unit":             t("unit"),
                    "qty_cartons":      f"📋 {t('qty_cartons')}",
                    "qty_pcs":          f"📋 {t('qty_pcs')}",
                    "dispatch_cartons": f"🚚 {t('qty_cartons')}",
                    "dispatch_pcs":     f"🚚 {t('qty_pcs')}",
                    "actual_cartons":   f"✅ {t('qty_cartons')}",
                    "actual_pcs":       f"✅ {t('qty_pcs')}",
                })
                st.dataframe(show, use_container_width=True, hide_index=True)


def _warehouse_supplier_email_body(
    title_line: str,
    lines: list[dict],
    remarks: str,
    ts: str,
) -> str:
    preview = "\n".join(
        f"  · {str(x.get('name', '') or '')[:55]}  "
        f"📦{int(x.get('qty_cartons', 0) or 0)}  "
        f"🔢{int(x.get('qty_pcs', 0) or 0)}"
        for x in lines[:30]
    )
    if len(lines) > 30:
        preview += f"\n  … (+{len(lines) - 30} more / 更多行)"
    return (
        "SUNSHINE · Supplier purchase request / 供货商订货\n"
        + ("=" * 52)
        + "\n"
        f"Time / 时间: {ts}\n"
        f"Title / 标题: {title_line}\n"
        f"Lines / 行数: {len(lines)}\n\n"
        "Items / 商品（摘要）:\n"
        + (preview or "  (none)")
        + "\n\n"
        "Full line list is in the attached Excel file / 完整明细见附件 Excel。\n\n"
        "Remarks / 备注:\n"
        + ((remarks or "").strip() or "-")
        + "\n\n—— Warehouse staff / 仓库员工\n"
    )


def _wh_supplier_merge_line(cart: list[dict], line: dict) -> None:
    key = (
        str(line.get("item_code", "") or ""),
        str(line.get("barcode", "") or ""),
        str(line.get("name", "") or ""),
    )
    for x in cart:
        if (
            str(x.get("item_code", "") or ""),
            str(x.get("barcode", "") or ""),
            str(x.get("name", "") or ""),
        ) == key:
            x["qty_cartons"] = int(x.get("qty_cartons", 0) or 0) + int(
                line.get("qty_cartons", 0) or 0
            )
            x["qty_pcs"] = int(x.get("qty_pcs", 0) or 0) + int(
                line.get("qty_pcs", 0) or 0
            )
            return
    cart.append(line)


def page_warehouse_supplier_order() -> None:
    """Warehouse: submit a restock / supplier order; notify admin + email."""
    if "wh_supplier_cart" not in st.session_state:
        st.session_state.wh_supplier_cart = []
    if "wh_supplier_search_q" not in st.session_state:
        st.session_state.wh_supplier_search_q = ""

    if st.session_state.get("wh_supplier_sent"):
        render_page_heading(t("wh_supplier_done_title"), None)
        st.success(t("wh_supplier_done_msg"))
        if st.button(
            t("wh_supplier_back"),
            type="primary",
            use_container_width=True,
            key="wh_supplier_back",
        ):
            st.session_state.wh_supplier_sent = False
            st.rerun()
        return

    render_page_heading(
        f"📝 {t('nav_supplier_order')}", t("wh_supplier_subtitle")
    )
    subj = st.text_input(
        t("wh_supplier_subject"),
        key="wh_s_subj",
        placeholder=t("wh_supplier_default_subj"),
    )
    remarks = st.text_area(t("remarks"), key="wh_s_rem", height=90)

    st.divider()
    st.markdown(f"**{t('wh_supplier_details')}**")
    st.caption(t("wh_supplier_search_commit_hint"))
    with st.form("wh_supplier_search_form", clear_on_submit=False):
        sq1, sq2 = st.columns([5, 1])
        with sq1:
            q_raw = st.text_input(
                t("search_product"),
                key="wh_sup_q",
                placeholder=t("search_product"),
            )
        with sq2:
            st.markdown(
                "<div style='padding-top:28px;'></div>",
                unsafe_allow_html=True,
            )
            do_search = st.form_submit_button(
                t("search"), type="primary", use_container_width=True
            )
    if do_search:
        st.session_state.wh_supplier_search_q = (q_raw or "").strip()
    q = str(st.session_state.get("wh_supplier_search_q") or "").strip()
    if len(q) < 1:
        st.caption(t("wh_supplier_search_hint"))
        results = pd.DataFrame()
    else:
        results = search_products(q, limit=200)

    if not results.empty:
        n = len(results)
        labels: list[str] = []
        for _, row in results.iterrows():
            nm = str(row.get("Name", "") or "")[:45]
            ic = str(row.get("ItemCode", "") or "")
            bc = str(row.get("Barcode", "") or "")
            labels.append(f"{nm}  |  {ic}  |  {bc}")

        pick_idx = st.selectbox(
            t("wh_supplier_pick"),
            options=list(range(n)),
            format_func=lambda i: labels[i],
            key="wh_sup_pick_idx",
        )
        ac1, ac2, ac3 = st.columns([1, 1, 2])
        with ac1:
            add_ct = st.number_input(
                t("cartons"), min_value=0, value=0, step=1, key="wh_sup_add_ct"
            )
        with ac2:
            add_pc = st.number_input(
                t("each_pcs"), min_value=0, value=0, step=1, key="wh_sup_add_pc"
            )
        with ac3:
            st.markdown("")  # vertical spacer
            if st.button(t("wh_supplier_add_line"), type="secondary"):
                if add_ct <= 0 and add_pc <= 0:
                    st.warning(t("no_qty_selected"))
                else:
                    row = results.iloc[int(pick_idx)]
                    line = {
                        "item_code": str(row.get("ItemCode", "") or ""),
                        "barcode": str(row.get("Barcode", "") or ""),
                        "name": str(row.get("Name", "") or ""),
                        "unit": str(row.get("Unit", "") or ""),
                        "qty_cartons": int(add_ct),
                        "qty_pcs": int(add_pc),
                    }
                    _wh_supplier_merge_line(st.session_state.wh_supplier_cart, line)
                    st.success(t("added_n_items").format(n=1))
                    st.rerun()
    elif q:
        st.info(t("no_results"))

    cart: list[dict] = st.session_state.wh_supplier_cart
    st.divider()
    st.markdown(
        f"**{t('wh_supplier_cart')}** &nbsp;·&nbsp; **{len(cart)}**"
    )
    if cart:
        for i, ln in enumerate(list(cart)):
            nm = str(ln.get("name", "") or "")
            ct_ = int(ln.get("qty_cartons", 0) or 0)
            pc_ = int(ln.get("qty_pcs", 0) or 0)
            bit = str(ln.get("item_code", "") or "") or str(
                ln.get("barcode", "") or ""
            )
            row_cols = st.columns([5, 1])
            with row_cols[0]:
                st.markdown(
                    f"{html.escape(nm)} &nbsp;·&nbsp; 📦 **{ct_}** &nbsp;·&nbsp; "
                    f"🔢 **{pc_}**"
                    + (f" &nbsp;`{html.escape(bit)}`" if bit else "")
                )
            with row_cols[1]:
                if st.button(
                    t("wh_supplier_remove"),
                    key=f"wh_sup_rm_{i}",
                    use_container_width=True,
                ):
                    cart.pop(i)
                    st.rerun()
    else:
        st.caption("—")

    st.divider()
    if st.button(t("wh_supplier_send"), type="primary", key="wh_sup_send"):
        if not cart:
            st.error(t("wh_supplier_need_lines"))
            return
        title_line = (subj or "").strip() or t("wh_supplier_default_subj")
        ts = now_str()
        body = _warehouse_supplier_email_body(
            title_line, list(cart), remarks or "", ts
        )
        subj_out = f"[{t('app_brand')}] {t('nav_supplier_order')}: {title_line[:100]}"
        fname = f"{_safe_excel_filename(title_line)}.xlsx"
        try:
            xlsx_bytes = build_supplier_order_excel(list(cart), remarks or "", ts)
        except Exception as e:
            log_exception("build_supplier_order_excel", e)
            st.error(str(e))
            return
        try:
            notify(
                "supplier_order",
                subj_out,
                body,
                attachments=[
                    (
                        fname,
                        xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet",
                    )
                ],
            )
        except Exception as e:
            log_exception("notify_supplier_order", e)
        try:
            summ = f"{len(cart)} lines · {title_line[:120]}"
            create_notification(
                "supplier_order",
                f"{t('wh_notif_inbox_title')}: {title_line[:160]}",
                summ[:900] + ("…" if len(summ) > 900 else ""),
                target_role=Role.ADMIN,
            )
        except Exception as e:
            log_exception("create_notification_supplier_order", e)
        try:
            audit_write(
                "supplier_order",
                extra={
                    "title": title_line,
                    "lines": len(cart),
                },
            )
        except Exception:
            pass
        st.session_state.wh_supplier_cart = []
        st.session_state.wh_supplier_search_q = ""
        st.session_state.pop("wh_sup_q", None)
        st.session_state.wh_supplier_sent = True
        st.rerun()


# =========================================================================
# ADMIN PAGES
# =========================================================================
def page_admin_accounts() -> None:
    render_page_heading(t("nav_accounts"), "")

    with st.expander("➕ " + t("acct_create_user"), expanded=False):
        c1, c2 = st.columns(2)
        with c1:
            new_u = st.text_input(t("acct_username"), key="adm_new_u")
            new_b = st.selectbox(
                t("select_branch"), list(BRANCHES), key="adm_new_b"
            )
        with c2:
            new_p1 = st.text_input(
                t("acct_password"), type="password", key="adm_new_p1"
            )
            new_p2 = st.text_input(
                t("acct_password2"), type="password", key="adm_new_p2"
            )
        new_perms = st.multiselect(
            t("acct_perms_pick"),
            [p for p, _ in BRANCH_PERM_ORDER],
            default=[p for p, _ in BRANCH_PERM_ORDER],
            format_func=branch_perm_label,
            key="adm_new_perms",
        )
        if st.button(t("acct_create_user"), key="adm_create_submit"):
            if new_p1 != new_p2:
                st.error(t("acct_pw_mismatch"))
            else:
                ok, ek = account_create_direct(new_u, new_p1, new_b, new_perms)
                if ok:
                    st.success(t("acct_updated"))
                    st.rerun()
                else:
                    st.error(t(ek))

    st.divider()
    st.subheader(t("acct_pending_hdr"))
    pending_rows = account_list_by_status(ACCOUNT_STATUS_PENDING)
    if not pending_rows:
        st.caption(t("no_data"))
    for row in pending_rows:
        rid = int(row["id"])
        st.markdown(
            f"**{html.escape(row['username'])}** · "
            f"{html.escape(row['branch'])} · "
            f"{html.escape((row['display_name'] or '').strip() or '—')}"
        )
        st.caption(
            f"{t('acct_created')}: {row['created_at']}"
        )
        ap_sel = st.multiselect(
            t("acct_perms_pick"),
            [p for p, _ in BRANCH_PERM_ORDER],
            default=[p for p, _ in BRANCH_PERM_ORDER],
            format_func=branch_perm_label,
            key=f"pend_perms_{rid}",
        )
        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button(t("acct_approve"), key=f"appr_{rid}"):
                account_approve(rid, ap_sel)
                st.success(t("acct_updated"))
                st.rerun()
        with c2:
            rj_note = st.text_input(
                t("acct_note_reject"), key=f"rj_{rid}"
            )
        with c3:
            if st.button(t("acct_reject"), key=f"rjbtn_{rid}"):
                account_set_status(rid, ACCOUNT_STATUS_REJECTED, rj_note)
                st.success(t("acct_updated"))
                st.rerun()
        st.divider()

    st.subheader(t("acct_approved_hdr"))
    appr = account_list_by_status(ACCOUNT_STATUS_APPROVED)
    if not appr:
        st.caption(t("no_data"))
    for row in appr:
        rid = int(row["id"])
        with st.expander(
            f"{html.escape(row['username'])} — {html.escape(row['branch'])}"
        ):
            cur = _parse_permissions_json(row["permissions"])
            ed = st.multiselect(
                t("acct_perms_pick"),
                [p for p, _ in BRANCH_PERM_ORDER],
                default=cur or [p for p, _ in BRANCH_PERM_ORDER],
                format_func=branch_perm_label,
                key=f"ed_perms_{rid}",
            )
            if st.button(t("acct_save_perms"), key=f"savep_{rid}"):
                if not ed:
                    st.error(t("acct_perm_need_admin"))
                else:
                    account_update_permissions(rid, ed)
                    st.success(t("acct_updated"))
                    st.rerun()
            st.divider()
            npw1 = st.text_input(
                t("acct_reset_pw"), type="password", key=f"npw1_{rid}"
            )
            npw2 = st.text_input(
                t("acct_password2"), type="password", key=f"npw2_{rid}"
            )
            if st.button(t("acct_pw_apply"), key=f"setpw_{rid}"):
                if len(npw1) < 6:
                    st.error(t("acct_pw_short"))
                elif npw1 != npw2:
                    st.error(t("acct_pw_mismatch"))
                else:
                    account_set_password(rid, npw1)
                    st.success(t("acct_updated"))
                    st.rerun()


def page_admin_audit_log() -> None:
    render_page_heading(t("nav_audit_log"), t("audit_log_subtitle"))
    c1, c2, c3 = st.columns(3)
    with c1:
        d0 = st.date_input(
            t("audit_filter_from"),
            value=datetime.now().date() - timedelta(days=7),
            key="adm_audit_d0",
        )
    with c2:
        d1 = st.date_input(
            t("audit_filter_to"),
            value=datetime.now().date(),
            key="adm_audit_d1",
        )
    with c3:
        br_sel = st.selectbox(
            t("audit_filter_branch"),
            [t("audit_all_branches")] + list(BRANCHES),
            key="adm_audit_br",
        )
    c4, c5, c6 = st.columns(3)
    with c4:
        ev_pick = st.multiselect(
            t("audit_filter_events"),
            list(AUDIT_EVENT_TYPES),
            default=list(AUDIT_EVENT_TYPES),
            format_func=_audit_event_label,
            key="adm_audit_ev",
        )
    with c5:
        ord_q = (st.text_input(t("audit_filter_order"), key="adm_audit_oid") or "").strip()
    with c6:
        usr_q = (st.text_input(t("audit_filter_user"), key="adm_audit_usr") or "").strip()

    s0, s1 = d0.isoformat(), d1.isoformat()
    clauses: list[str] = [
        "date(created_at) >= date(?)",
        "date(created_at) <= date(?)",
    ]
    par: list[object] = [s0, s1]
    if ev_pick:
        clauses.append(
            "event_type IN (%s)" % ",".join("?" * len(ev_pick))
        )
        par.extend(ev_pick)
    if br_sel and br_sel != t("audit_all_branches"):
        clauses.append("branch = ?")
        par.append(br_sel)
    if ord_q:
        clauses.append("order_id LIKE ?")
        par.append(f"%{ord_q}%")
    if usr_q:
        clauses.append("COALESCE(username,'') LIKE ?")
        par.append(f"%{usr_q}%")
    sql = (
        "SELECT * FROM audit_log WHERE "
        + " AND ".join(clauses)
        + " ORDER BY created_at DESC LIMIT 2000"
    )
    with db_conn() as conn:
        rows = conn.execute(sql, par).fetchall()
    if not rows:
        st.info(t("audit_no_rows"))
        return
    out = []
    for r in rows:
        ev = r["event_type"]
        out.append(
            {
                t("audit_col_time"): r["created_at"],
                t("audit_col_event"): _audit_event_label(ev),
                t("audit_col_role"): r["role"] or "",
                t("audit_col_account"): (
                    str(int(r["account_id"]))
                    if r["account_id"] is not None
                    else ""
                ),
                t("audit_col_user"): r["username"] or "",
                t("audit_col_branch"): r["branch"] or "",
                t("audit_col_order"): r["order_id"] or "",
                t("audit_col_detail"): _format_audit_detail_display(
                    r["detail"], str(ev)
                ),
            }
        )
    display_df = pd.DataFrame(out)
    st.dataframe(display_df, use_container_width=True, hide_index=True)

    csv_bytes = display_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label=t("audit_export_csv"),
        data=csv_bytes,
        file_name=f"audit_log_{s0}_{s1}.csv",
        mime="text/csv",
        key="adm_audit_dl",
    )


def page_admin_shelf_mobile() -> None:
    """Mobile-friendly catalog builder: wipe all, then enter shelf lines."""
    render_page_heading(t("nav_shelf_mobile"), t("shelf_mobile_subtitle"))

    rev = int(st.session_state.get("shelf_rev", 0))
    boot_key = f"_shelf_boot_df_{rev}"
    if boot_key not in st.session_state:
        st.session_state[boot_key] = _shelf_bootstrap_dataframe()

    with st.expander(t("shelf_wipe_title"), expanded=False):
        st.markdown(t("shelf_wipe_blurb"))
        wipe_ack = (st.text_input(
            t("shelf_wipe_confirm"),
            key=f"shelf_wipe_ack_{rev}",
            placeholder="CLEAR / 清空",
        ) or "").strip()
        wipe_ok = wipe_ack.upper() == "CLEAR" or wipe_ack == "清空"
        if st.button(
            "⚠️ " + t("shelf_wipe_title"),
            type="primary",
            disabled=not wipe_ok,
            key=f"shelf_wipe_go_{rev}",
            use_container_width=True,
        ):
            admin_wipe_all_products_and_stock()
            st.session_state.shelf_rev = rev + 1
            for k in list(st.session_state.keys()):
                if str(k).startswith("_shelf_boot_df_"):
                    del st.session_state[k]
            audit_write(
                "catalog_reset",
                detail="admin wiped products.xlsx + inventory + prices",
            )
            st.success(t("shelf_wipe_ok"))
            st.rerun()

    st.caption(t("shelf_editor_hint"))
    init_df = st.session_state[boot_key]
    show_cols = [c for c in _product_master_excel_columns() if c != "StockTotal"]
    for c in show_cols:
        if c not in init_df.columns:
            init_df[c] = 0.0 if c in (
                "Price", "PcsPerCarton", "StockCartons", "StockPcs", "StockTotal",
            ) else ""
    init_df = init_df[show_cols]
    edited = st.data_editor(
        init_df,
        key=f"shelf_grid_{rev}",
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "ItemCode": st.column_config.TextColumn(t("item_code"), width="small"),
            "Barcode": st.column_config.TextColumn(t("barcode"), width="medium"),
            "Name": st.column_config.TextColumn(t("name"), width="large"),
            "Unit": st.column_config.TextColumn(t("unit"), width="small"),
            "Category": st.column_config.TextColumn("Category", width="small"),
            "Price": st.column_config.NumberColumn(
                t("price"),
                format="%.2f",
                min_value=0.0,
                step=0.01,
            ),
            "PcsPerCarton": st.column_config.NumberColumn(
                t("shelf_pcs_per_carton"),
                min_value=0,
                step=1,
                format="%d",
            ),
            "StockCartons": st.column_config.NumberColumn(
                t("grid_inv_stock_ct"),
                min_value=0,
                step=1,
                format="%d",
            ),
            "StockPcs": st.column_config.NumberColumn(
                t("grid_inv_stock_pc"),
                min_value=0,
                step=1,
                format="%d",
            ),
        },
        hide_index=True,
    )

    if st.button(
        t("shelf_save_btn"),
        type="primary",
        key=f"shelf_save_{rev}",
        use_container_width=True,
    ):
        n_prod, inv_lines = admin_save_shelf_catalog(edited)
        if n_prod <= 0:
            st.warning(t("shelf_need_name"))
        else:
            st.session_state.shelf_rev = rev + 1
            for k in list(st.session_state.keys()):
                if str(k).startswith("_shelf_boot_df_"):
                    del st.session_state[k]
            audit_write(
                "catalog_save",
                detail=json.dumps(
                    {"products_saved": n_prod, "inventory_writes": inv_lines},
                    ensure_ascii=False,
                ),
            )
            st.success(t("shelf_save_ok"))
            st.rerun()


def page_admin_dashboard() -> None:
    render_page_heading(f"📊 {t('nav_dashboard')}")
    today = today_str()

    with db_conn() as conn:
        n_today = conn.execute(
            "SELECT COUNT(DISTINCT order_id) FROM orders WHERE date(order_date) = ?",
            (today,),
        ).fetchone()[0]
        n_pending = conn.execute(
            "SELECT COUNT(DISTINCT order_id) FROM orders WHERE status='Pending'",
        ).fetchone()[0]
        n_dispatched = conn.execute(
            "SELECT COUNT(DISTINCT order_id) FROM orders WHERE status='Dispatched'",
        ).fetchone()[0]
        n_short = conn.execute(
            "SELECT COUNT(*) FROM shortages WHERE status='Open'",
        ).fetchone()[0]

    cards = [
        ("📊", t("card_today"),      n_today,     ""),
        ("⏳", t("card_pending"),    n_pending,   "warn"),
        ("🚚", t("card_dispatched"), n_dispatched,"ok"),
        ("🔔", t("card_short"),      n_short,     "alert"),
    ]
    cols = st.columns(4)
    for col, (icon, label, val, mod) in zip(cols, cards):
        with col:
            st.markdown(f"""
            <div class="metric-card {mod}">
                <div class="label">{icon} {label}</div>
                <div class="value">{val}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("")
    render_section_title(f"🏪 {t('branch_status')}")
    with db_conn() as conn:
        bdf = pd.read_sql_query("""
            SELECT branch,
                   COUNT(DISTINCT CASE WHEN status='Pending'    THEN order_id END) AS pending,
                   COUNT(DISTINCT CASE WHEN status='Dispatched' THEN order_id END) AS dispatched,
                   COUNT(DISTINCT CASE WHEN status='Received'   THEN order_id END) AS received
            FROM orders GROUP BY branch
        """, conn)
        sdf = pd.read_sql_query(
            "SELECT branch, COUNT(*) AS open_short FROM shortages "
            "WHERE status='Open' GROUP BY branch", conn,
        )
    merged = (
        pd.DataFrame({"branch": BRANCHES})
        .merge(bdf, on="branch", how="left")
        .merge(sdf, on="branch", how="left")
        .fillna(0)
    )
    for col in ["pending", "dispatched", "received", "open_short"]:
        merged[col] = merged[col].astype(int)
    merged.columns = [t("branch"), t("st_pending"), t("st_dispatched"),
                      t("st_received"), t("card_short")]
    st.dataframe(merged, use_container_width=True, hide_index=True)

    render_section_title(f"📋 {t('latest_orders')}")
    with db_conn() as conn:
        latest = pd.read_sql_query("""
            SELECT order_id, branch, order_date, status, COUNT(*) AS lines
            FROM orders
            GROUP BY order_id, branch, order_date, status
            ORDER BY order_date DESC LIMIT 10
        """, conn)
    if latest.empty:
        st.caption(t("no_data"))
    else:
        latest.columns = [t("order_id"), t("branch"), t("order_date"),
                          t("status"), "Lines"]
        st.dataframe(latest, use_container_width=True, hide_index=True)

    render_section_title(f"🔔 {t('latest_short')}")
    with db_conn() as conn:
        ls = pd.read_sql_query("""
            SELECT branch, name, short_cartons, short_pcs, status, reported_date
            FROM shortages ORDER BY reported_date DESC LIMIT 10
        """, conn)
    if ls.empty:
        st.caption(t("no_data"))
    else:
        ls.columns = [t("branch"), t("name"), t("qty_cartons"),
                      t("qty_pcs"), t("status"), t("order_date")]
        st.dataframe(ls, use_container_width=True, hide_index=True)


def page_admin_all_orders() -> None:
    render_page_heading(f"📋 {t('nav_all_orders')}")
    bs = st.multiselect(t("branch"), BRANCHES, default=BRANCHES, key="adm_b")
    ss = st.multiselect(t("status"), list(OrderStatus.ALL),
                        default=list(OrderStatus.ALL), key="adm_s")
    if not bs or not ss:
        st.info(t("no_data"))
        return
    qb = ",".join("?" * len(bs))
    qs = ",".join("?" * len(ss))
    with db_conn() as conn:
        df = pd.read_sql_query(
            f"SELECT * FROM orders WHERE branch IN ({qb}) AND status IN ({qs}) "
            "ORDER BY order_date DESC", conn, params=bs + ss,
        )
    if df.empty:
        st.info(t("no_data"))
        return
    for order_id, group in df.groupby("order_id", sort=False):
        status = group["status"].iloc[0]
        with st.expander(
            f"📦 {order_id} · 🏪 {group['branch'].iloc[0]} · "
            f"{group['order_date'].iloc[0][:16]} · {len(group)} items"
        ):
            st.markdown(status_pill(status), unsafe_allow_html=True)
            st.markdown("")
            show = group[["name", "unit",
                          "qty_cartons", "qty_pcs",
                          "dispatch_cartons", "dispatch_pcs",
                          "actual_cartons", "actual_pcs"]].fillna("-")
            show.columns = [
                t("name"), t("unit"),
                f"📋 {t('qty_cartons')}", f"📋 {t('qty_pcs')}",
                f"🚚 {t('qty_cartons')}", f"🚚 {t('qty_pcs')}",
                f"✅ {t('qty_cartons')}", f"✅ {t('qty_pcs')}",
            ]
            st.dataframe(show, use_container_width=True, hide_index=True)


# =========================================================================
# EXCEL EXPORTS
# =========================================================================
def _style_excel_header(ws, last_col: int) -> None:
    fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    border = Border(*(Side(style="thin", color="999999"),) * 4)
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 8
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


# --- Picking-slip exports for warehouse staff -----------------------------
def _attach_categories(df: pd.DataFrame) -> pd.DataFrame:
    """Add a 'Category' column to an order-lines dataframe by looking each
    line up in the products master. Match on ItemCode first (exact), then
    Barcode (exact). Unmatched (e.g. manually added items) → 'General'."""
    if df.empty:
        return df.assign(Category="General")
    products = load_products(_products_mtime(), _inventory_version(), _price_version())
    if products.empty:
        return df.assign(Category="General")

    # Build two lookups: ItemCode → Category, Barcode → Category
    by_code = {}
    by_barcode = {}
    for _, p in products.iterrows():
        cat = str(p.get("Category", "")).strip() or "General"
        ic = str(p.get("ItemCode", "")).strip()
        bc = str(p.get("Barcode", "")).strip()
        if ic:
            by_code[ic] = cat
        if bc:
            by_barcode[bc] = cat

    def _lookup(row) -> str:
        ic = str(row.get("item_code", "") or "").strip()
        bc = str(row.get("barcode", "") or "").strip()
        if ic and ic in by_code:
            return by_code[ic]
        if bc and bc in by_barcode:
            return by_barcode[bc]
        return "General"

    out = df.copy()
    out["Category"] = out.apply(_lookup, axis=1)
    return out


def _build_picking_slip_sheet(ws, df: pd.DataFrame, title_block: list[tuple[str, str]]) -> None:
    """Render a printable picking slip into ws, grouped by Category.

    Columns kept (5):  Barcode | Name | Cartons | Pcs | ✓
    Columns dropped:   Order ID, Order Date, Item Code, Unit

    Each category gets a banner row (light-blue, full-width, bold) listing
    the count + total cartons/pcs for that category. Categories appear in
    a fixed warehouse-zone order (Frozen first → Chilled → ... → General),
    with any unrecognized categories appended alphabetically at the end.
    Items without a Category in products.xlsx fall back to 'General'.
    """
    # ----- Title rows --------------------------------------------------
    for i, (label, value) in enumerate(title_block, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=i, column=2, value=value).font = Font(size=11)

    cur_row = len(title_block) + 2  # one blank row gap below title block

    # ----- Style helpers ----------------------------------------------
    blue_fill   = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    banner_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    white_font  = Font(bold=True, color="FFFFFF", size=11)
    cat_font    = Font(bold=True, color="0D47A1", size=12)
    align_c     = Alignment(horizontal="center", vertical="center")
    align_l     = Alignment(horizontal="left",   vertical="center")
    border      = Border(*(Side(style="thin", color="999999"),) * 4)

    headers = [t("barcode"), t("name"), t("qty_cartons"), t("qty_pcs"), "✓"]
    n_cols = len(headers)

    # ----- Group by category ------------------------------------------
    df_cat = _attach_categories(df)

    # Preferred display order: arranged so warehouse staff can walk a sensible
    # picking route (cold first, then ambient by zone). Anything not listed
    # here gets appended alphabetically at the end.
    PREFERRED = ["Frozen", "Chilled", "Dairy", "Fruit", "Vegetable",
                 "Rice", "Beverage", "General"]
    present = list(df_cat["Category"].unique())
    ordered_cats = [c for c in PREFERRED if c in present] + sorted(
        c for c in present if c not in PREFERRED
    )

    for cat in ordered_cats:
        sub = df_cat[df_cat["Category"] == cat]
        if sub.empty:
            continue

        # ---- Banner row: full-width category header + totals ---------
        n_lines  = len(sub)
        total_ct = int(sub["qty_cartons"].sum())
        total_pc = int(sub["qty_pcs"].sum())
        banner = ws.cell(
            row=cur_row, column=1,
            value=f"  📦  {cat}   ·   {n_lines} items   ·   "
                  f"{total_ct} cartons   ·   {total_pc} pcs",
        )
        banner.font = cat_font
        banner.fill = banner_fill
        banner.alignment = align_l
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=n_cols,
        )
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # ---- Column-header row (per group, so each group is self-
        # contained when the warehouse staff splits the printout) -----
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=cur_row, column=col_idx, value=h)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = align_c
            cell.border = border
        cur_row += 1

        # ---- Data rows ---------------------------------------------
        for r in sub.itertuples(index=False):
            ws.cell(row=cur_row, column=1,
                    value=str(getattr(r, "barcode", "") or ""))
            ws.cell(row=cur_row, column=2,
                    value=str(getattr(r, "name", "") or ""))
            ws.cell(row=cur_row, column=3,
                    value=int(getattr(r, "qty_cartons", 0) or 0))
            ws.cell(row=cur_row, column=4,
                    value=int(getattr(r, "qty_pcs", 0) or 0))
            ws.cell(row=cur_row, column=5, value="")  # tick box
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=cur_row, column=col_idx).border = border
            cur_row += 1

        # ---- Spacer between groups -------------------------------
        cur_row += 1

    _autosize(ws)


def export_single_order_picking(order_id: str):
    """Single order picking slip — one Excel for printing."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE order_id = ? ORDER BY name",
            conn, params=(order_id,),
        )
    if df.empty:
        return None
    branch = df["branch"].iloc[0]
    order_date = df["order_date"].iloc[0]
    wb = Workbook()
    ws = wb.active
    ws.title = "Picking Slip"
    title_block = [
        (f"📋 {t('picking_slip')}", "SUNSHINE / 阳光集团"),
        (t("branch"),     branch),
        (t("order_id"),   order_id),
        (t("order_date"), order_date),
        (t("total_items"), str(len(df))),
    ]
    _build_picking_slip_sheet(ws, df, title_block)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_branch_pending_picking(branch: str):
    """All pending orders for one branch, one sheet per order_id."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE branch = ? AND status = 'Pending' "
            "ORDER BY order_date ASC, name",
            conn, params=(branch,),
        )
    if df.empty:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    for order_id, group in df.groupby("order_id", sort=False):
        # Excel sheet names <= 31 chars, no special chars
        safe = order_id[-25:].replace(":", "-").replace("/", "-")
        ws = wb.create_sheet(safe)
        title_block = [
            (f"📋 {t('picking_slip')}", "SUNSHINE / 阳光集团"),
            (t("branch"),     branch),
            (t("order_id"),   order_id),
            (t("order_date"), group["order_date"].iloc[0]),
            (t("total_items"), str(len(group))),
        ]
        _build_picking_slip_sheet(ws, group, title_block)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_all_pending_picking():
    """All branches, all pending orders. One sheet per branch (consolidated)."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE status = 'Pending' "
            "ORDER BY branch, order_date ASC, name",
            conn,
        )
    if df.empty:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    for branch in BRANCHES:
        sub = df[df["branch"] == branch]
        if sub.empty:
            continue
        ws = wb.create_sheet(branch[:30].replace("/", "-"))
        title_block = [
            (f"📋 {t('picking_slip')}", "SUNSHINE / 阳光集团"),
            (t("branch"), branch),
            (t("order_date"), datetime.now().strftime("%Y-%m-%d %H:%M")),
            (t("total_items"), str(len(sub))),
        ]
        _build_picking_slip_sheet(ws, sub, title_block)
    if not wb.sheetnames:
        return None
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_picking_list(date_from: str, date_to: str):
    with db_conn() as conn:
        df = pd.read_sql_query("""
            SELECT * FROM orders
            WHERE date(order_date) BETWEEN ? AND ?
            ORDER BY branch, name
        """, conn, params=(date_from, date_to))
    if df.empty:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    headers = [t("order_id"), t("order_date"), t("status"), t("item_code"),
               t("barcode"), t("name"), t("unit"), t("qty_cartons"), t("qty_pcs")]
    for branch in BRANCHES:
        sub = df[df["branch"] == branch]
        if sub.empty:
            continue
        ws = wb.create_sheet(branch[:30].replace("/", "-"))
        ws.append(headers)
        for _, r in sub.iterrows():
            ws.append([r["order_id"], r["order_date"], r["status"],
                       r["item_code"], r["barcode"], r["name"], r["unit"],
                       r["qty_cartons"], r["qty_pcs"]])
        _style_excel_header(ws, len(headers))
        _autosize(ws)
    if not wb.sheetnames:
        return None
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_reconciliation(date_from: str, date_to: str):
    with db_conn() as conn:
        df = pd.read_sql_query("""
            SELECT branch, order_id, order_date, item_code, barcode, name, unit,
                   qty_cartons, qty_pcs,
                   dispatch_cartons, dispatch_pcs,
                   actual_cartons, actual_pcs, status
            FROM orders
            WHERE date(order_date) BETWEEN ? AND ?
              AND status IN ('Dispatched', 'Received')
            ORDER BY branch, order_id, name
        """, conn, params=(date_from, date_to))
    if df.empty:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Reconciliation"
    headers = [
        t("branch"), t("order_id"), t("order_date"), t("item_code"), t("barcode"),
        t("name"), t("unit"),
        f"📋 {t('qty_cartons')}", f"📋 {t('qty_pcs')}",
        f"🚚 {t('qty_cartons')}", f"🚚 {t('qty_pcs')}",
        f"✅ {t('qty_cartons')}", f"✅ {t('qty_pcs')}",
        f"Δ {t('qty_cartons')}", f"Δ {t('qty_pcs')}",
        t("status"),
    ]
    ws.append(headers)
    red    = PatternFill("solid", start_color="FFCDD2", end_color="FFCDD2")
    yellow = PatternFill("solid", start_color="FFF59D", end_color="FFF59D")

    def _eff_dispatch(row):
        """Effective dispatched qty: prefer dispatch_*, fallback to qty_*."""
        d_ct = row["dispatch_cartons"]
        d_pc = row["dispatch_pcs"]
        d_ct = int(row["qty_cartons"]) if pd.isna(d_ct) or d_ct is None else int(d_ct)
        d_pc = int(row["qty_pcs"])     if pd.isna(d_pc) or d_pc is None else int(d_pc)
        return d_ct, d_pc

    for _, r in df.iterrows():
        d_ct, d_pc = _eff_dispatch(r)
        ac = r["actual_cartons"]; ap = r["actual_pcs"]
        if pd.isna(ac) or pd.isna(ap):
            diff_ct = ""; diff_pc = ""; tag = "Pending"
            ac_v = "-" if pd.isna(ac) else int(ac)
            ap_v = "-" if pd.isna(ap) else int(ap)
        else:
            ac = int(ac); ap = int(ap)
            # Δ = received vs DISPATCHED (transit/branch shortage), not vs ordered
            diff_ct = ac - d_ct
            diff_pc = ap - d_pc
            tag = "SHORT" if (diff_ct < 0 or diff_pc < 0) else (
                  "OVER"  if (diff_ct > 0 or diff_pc > 0) else "OK")
            ac_v, ap_v = ac, ap
        ws.append([
            r["branch"], r["order_id"], r["order_date"],
            r["item_code"], r["barcode"], r["name"], r["unit"],
            r["qty_cartons"], r["qty_pcs"],   # 📋 ordered
            d_ct, d_pc,                        # 🚚 dispatched (effective)
            ac_v, ap_v,                        # ✅ actually received
            diff_ct, diff_pc,                  # Δ
            tag,
        ])
        row_idx = ws.max_row
        if tag == "SHORT":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red
        elif tag == "OVER":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow
    _style_excel_header(ws, len(headers))
    _autosize(ws)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_shortage_report(date_from: str, date_to: str):
    with db_conn() as conn:
        df = pd.read_sql_query("""
            SELECT * FROM shortages
            WHERE date(reported_date) BETWEEN ? AND ?
            ORDER BY reported_date DESC
        """, conn, params=(date_from, date_to))
    if df.empty:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Shortages"
    headers = [t("branch"), t("order_id"), t("name"), t("item_code"), t("barcode"),
               t("unit"),
               f"{t('shortage_qty')} {t('qty_cartons')}",
               f"{t('shortage_qty')} {t('qty_pcs')}",
               t("status"), t("warehouse_reply"),
               t("order_date"), t("remarks")]
    ws.append(headers)
    for _, r in df.iterrows():
        ws.append([r["branch"], r["order_id"], r["name"], r["item_code"], r["barcode"],
                   r["unit"], r["short_cartons"], r["short_pcs"],
                   r["status"], r["warehouse_reply"] or "",
                   r["reported_date"], r["branch_remarks"] or ""])
    _style_excel_header(ws, len(headers))
    _autosize(ws)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_inventory_report(date_from: str = "", date_to: str = ""):
    """Export current inventory snapshot (all SKUs)."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            """
            SELECT item_code, barcode, name, unit, stock_cartons, stock_pcs, updated_at
            FROM inventory
            ORDER BY name ASC
            """,
            conn,
        )
    if df.empty:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Inventory"
    headers = [
        t("item_code"), t("barcode"), t("name"), t("unit"),
        t("qty_cartons"), t("qty_pcs"), t("order_date"),
    ]
    ws.append(headers)
    for _, r in df.iterrows():
        ws.append([
            r["item_code"], r["barcode"], r["name"], r["unit"],
            int(r["stock_cartons"] or 0), int(r["stock_pcs"] or 0), r["updated_at"],
        ])
    _style_excel_header(ws, len(headers))
    _autosize(ws)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def page_admin_images() -> None:
    """Upload / view / remove product images. Manager-only."""
    render_page_heading(f"🖼️ {t('img_title')}")
    st.caption(t("img_subtitle"))

    # =====================================================================
    # Auto-scan root folder. The system already imported on startup; this
    # section shows that summary and lets the user re-scan without
    # restarting (useful when they drop new files mid-session).
    # =====================================================================
    with st.expander(t("img_scan_title"), expanded=False):
        st.caption(t("img_scan_hint"))

        # Show the startup-scan summary if we still have it cached. This is
        # what was moved automatically when the user (re)started the app.
        startup_summary = st.session_state.get("_startup_scan_summary")

        # On-demand re-scan button
        if st.button(t("img_scan_now"), key="img_scan_now_btn",
                     type="primary", use_container_width=True):
            new_summary = scan_and_import_root_images()
            st.session_state["_startup_scan_summary"] = new_summary
            st.rerun()

        # Render the summary (whether from startup or from "Scan now")
        if startup_summary:
            moved = startup_summary.get("moved", [])
            unmatched = startup_summary.get("skipped_unmatched", [])
            errors = startup_summary.get("errors", [])

            if not moved and not unmatched and not errors:
                st.info(t("img_scan_none"))
            else:
                mc1, mc2, mc3 = st.columns(3)
                with mc1:
                    st.metric(f"✅ {t('img_scan_moved')}", len(moved))
                with mc2:
                    st.metric(f"⚠️ {t('img_scan_left')}", len(unmatched))
                with mc3:
                    st.metric(f"❌ {t('img_scan_errors')}", len(errors))

                if moved:
                    rows = [{
                        t("img_filename"):  m["from"],
                        t("img_matched_to"): f"{m['name']}  ({m['matched_by']})",
                    } for m in moved]
                    st.dataframe(pd.DataFrame(rows),
                                 use_container_width=True, hide_index=True)

                if unmatched:
                    st.warning(
                        f"⚠️ {len(unmatched)} file(s) in root folder do not "
                        f"match any product. Rename them or add the products "
                        f"to products.xlsx, then click '{t('img_scan_now')}'."
                    )
                    st.dataframe(
                        pd.DataFrame({t("img_filename"): unmatched}),
                        use_container_width=True, hide_index=True,
                    )

                if errors:
                    for e in errors:
                        st.error(f"❌ {e}")

    # =====================================================================
    # Batch upload by filename — naming convention: <ItemCode>.jpg or
    # <Barcode>.jpg. Files that don't match anything in products.xlsx are
    # surfaced separately so the user can fix names and re-upload.
    # =====================================================================
    with st.expander(t("img_batch_title"), expanded=False):
        st.caption(t("img_batch_hint"))

        batch_uploads = st.file_uploader(
            t("img_batch_choose"),
            type=[e.lstrip(".") for e in IMAGE_EXTS],
            accept_multiple_files=True,
            key="img_batch_files",
        )

        if batch_uploads:
            # Build a pairing preview from the current products master.
            products_df = load_products(_products_mtime(), _inventory_version(), _price_version())
            matched, unmatched = plan_batch_image_upload(
                batch_uploads, products_df,
            )

            render_section_title(t("img_batch_preview"))
            mc1, mc2 = st.columns(2)
            with mc1:
                st.metric(f"✅ {t('img_matched')}", len(matched))
            with mc2:
                st.metric(f"⚠️ {t('img_unmatched')}", len(unmatched))

            # ---- Matched files preview ----
            if matched:
                preview_rows = [{
                    t("img_filename"):  m["filename"],
                    t("img_matched_to"): f"{m['name']}  ({m['matched_by']}: "
                                         f"{m['item_code'] or m['barcode']})",
                } for m in matched]
                st.dataframe(
                    pd.DataFrame(preview_rows),
                    use_container_width=True, hide_index=True,
                )

            # ---- Unmatched files preview ----
            if unmatched:
                st.warning(
                    f"⚠️ {len(unmatched)} file(s) could not be matched. "
                    f"Rename and re-upload to fix."
                )
                un_rows = [{
                    t("img_filename"):    u["filename"],
                    t("img_match_reason"): u["reason"],
                } for u in unmatched]
                st.dataframe(
                    pd.DataFrame(un_rows),
                    use_container_width=True, hide_index=True,
                )

            # ---- Confirm & save (only for matched ones) ----
            if matched:
                cc1, cc2 = st.columns([3, 1])
                with cc1:
                    if st.button(
                        t("img_batch_confirm").format(n=len(matched)),
                        type="primary", use_container_width=True,
                        key="img_batch_go",
                    ):
                        # Process one at a time so a single bad file doesn't
                        # block the rest. We collect results to show a final
                        # summary.
                        prog = st.progress(0.0, text="...")
                        ok_count = 0
                        fail_list: list[tuple[str, str]] = []
                        total = len(matched)
                        for i, m in enumerate(matched):
                            try:
                                save_product_image(
                                    item_code=m["item_code"],
                                    barcode=m["barcode"],
                                    name=m["name"],
                                    uploaded_bytes=m["uf"].getvalue(),
                                    original_filename=m["filename"],
                                )
                                ok_count += 1
                            except Exception as e:
                                fail_list.append((m["filename"], str(e)))
                            prog.progress(
                                (i + 1) / total,
                                text=f"{i + 1}/{total}  ·  {m['filename']}",
                            )
                        prog.empty()

                        if ok_count:
                            st.success(
                                t("img_batch_done").format(n=ok_count)
                            )
                        if fail_list:
                            st.error(t("img_batch_failed").format(n=len(fail_list)))
                            for fn, err in fail_list:
                                st.caption(f"❌ {fn}: {err}")
                        # Reset the uploader so the file list clears
                        st.session_state.pop("img_batch_files", None)
                        st.rerun()
                with cc2:
                    if st.button(t("img_batch_cancel"),
                                 use_container_width=True,
                                 key="img_batch_cancel"):
                        st.session_state.pop("img_batch_files", None)
                        st.rerun()

    st.divider()

    # ----- Search a product to upload for ------------------------------
    query = st.text_input(
        t("img_search"), key="img_search_q",
        placeholder=t("search_product"),
        help="Bluetooth scanner supported / 支持扫码枪",
    )
    results = search_products(query, limit=50)

    if results.empty and query:
        st.info(t("no_results"))
        return
    if results.empty:
        return

    PAGE_SIZE = 10
    total = len(results)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = st.session_state.get("_img_page", 1)
    last_q = st.session_state.get("_img_last_q")
    if last_q != query:
        page = 1
        st.session_state["_img_last_q"] = query
    page = min(max(1, page), total_pages)
    st.session_state["_img_page"] = page

    start = (page - 1) * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)
    page_rows = results.iloc[start:end]

    st.caption(
        f"{t('showing')} {start + 1}–{end} {t('of')} {total} "
        f"{t('results_count')} · {t('page')} {page} {t('of')} {total_pages}"
    )

    # ----- Render each product with image preview + uploader ----------
    for idx, row in page_rows.iterrows():
        item_code = str(row.get("ItemCode", "") or "")
        barcode   = str(row.get("Barcode", "") or "")
        name      = str(row.get("Name", "") or "")
        with st.container(border=True):
            cl, cm, cr = st.columns([1, 3, 2])

            # ---- Current image preview (or placeholder) ----
            img_path = get_product_image_path(item_code, barcode)
            with cl:
                if img_path is not None:
                    st.image(str(img_path), width=120)
                else:
                    st.markdown(
                        "<div style='width:120px;height:120px;"
                        "background:#f0f0f0;border:1px dashed #bbb;"
                        "border-radius:6px;display:flex;align-items:center;"
                        "justify-content:center;color:#999;font-size:28px;'>"
                        "📷</div>",
                        unsafe_allow_html=True,
                    )

            # ---- Product info ----
            with cm:
                st.markdown(f"**{name}**")
                bits = []
                if item_code: bits.append(f"📋 {item_code}")
                if barcode:   bits.append(f"📊 {barcode}")
                if bits:
                    st.caption(" · ".join(bits))
                if img_path is not None:
                    st.caption(f"✅ {t('img_has_image')}: `{img_path.name}`")
                else:
                    st.caption(f"⚪ {t('img_no_image')}")

            # ---- Upload + delete ----
            with cr:
                # Stable widget keys per product so the input survives
                # rerenders without colliding with other rows.
                upload_key = f"img_up_{item_code}_{barcode}_{idx}"
                up = st.file_uploader(
                    t("img_upload"),
                    type=[e.lstrip(".") for e in IMAGE_EXTS],
                    key=upload_key,
                    label_visibility="collapsed",
                )
                if up is not None:
                    try:
                        save_product_image(
                            item_code, barcode, name,
                            up.getvalue(), up.name,
                        )
                        st.success(f"✅ {t('img_uploaded')}")
                        # Reset the uploader so the success message doesn't
                        # re-fire on the next rerun
                        st.session_state.pop(upload_key, None)
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

                if img_path is not None:
                    if st.button(t("img_delete"),
                                 key=f"img_del_{item_code}_{barcode}_{idx}",
                                 use_container_width=True):
                        if delete_product_image(item_code, barcode):
                            st.success(t("img_deleted"))
                            st.rerun()

    # ----- Pagination -------------------------------------------------
    if total_pages > 1:
        pc1, pc2, pc3 = st.columns([1, 2, 1])
        with pc1:
            if st.button(t("prev_page"), key="img_prev",
                         disabled=(page <= 1),
                         use_container_width=True):
                st.session_state["_img_page"] = page - 1
                st.rerun()
        with pc2:
            st.markdown(
                f"<div style='text-align:center;padding-top:6px;'>"
                f"{t('page')} <b>{page}</b> {t('of')} <b>{total_pages}</b>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with pc3:
            if st.button(t("next_page"), key="img_next",
                         disabled=(page >= total_pages),
                         use_container_width=True):
                st.session_state["_img_page"] = page + 1
                st.rerun()


def page_admin_backup() -> None:
    """Local DB backup & restore management for admins."""
    render_page_heading(f"💾 {t('backup_title')}")
    st.caption(t("backup_subtitle"))

    # ----- Current DB info ---------------------------------------------
    if DB_PATH.exists():
        size_kb = round(DB_PATH.stat().st_size / 1024, 1)
        with db_conn() as conn:
            n_orders = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
            n_ships  = conn.execute("SELECT COUNT(*) FROM shipments").fetchone()[0]
            n_short  = conn.execute("SELECT COUNT(*) FROM shortages").fetchone()[0]
        st.markdown(
            f"**📂 {t('current_db_info')}:** `{DB_PATH}` &nbsp;·&nbsp; "
            f"**{size_kb} KB** &nbsp;·&nbsp; "
            f"**{t('backup_counts')}:** {n_orders} / {n_ships} / {n_short}"
        )

    # ----- Action buttons ----------------------------------------------
    a1, a2, a3 = st.columns(3)
    with a1:
        if st.button(t("backup_now"), type="primary",
                     use_container_width=True, key="bk_now"):
            try:
                p = backup_database("manual")
                if p is None:
                    st.warning(t("no_data"))
                else:
                    st.success(f"{t('backup_done')}: `{p.name}`")
                    st.rerun()
            except Exception as e:
                st.error(f"{t('backup_failed')}: {e}")

    with a2:
        # Inline export-SQL — runs on click then surfaces a download_button
        if st.button(t("export_sql"), use_container_width=True, key="exp_sql"):
            data = export_db_dump_sql()
            if not data:
                st.warning(t("no_data"))
            else:
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.session_state["_sql_dump"] = (data, f"orders_dump_{stamp}.sql")
        if "_sql_dump" in st.session_state:
            data, fname = st.session_state["_sql_dump"]
            st.download_button(
                f"⬇️ {fname}",
                data=data, file_name=fname, mime="application/sql",
                use_container_width=True, key="dl_sql_dump",
            )

    with a3:
        # Import from SQL dump (text). File upload reuses st.file_uploader.
        with st.expander(t("import_sql")):
            up = st.file_uploader(t("upload_sql"), type=["sql", "txt"],
                                  key="imp_sql_file")
            if up is not None:
                if st.button(f"♻️ {t('import_sql')}", key="do_imp_sql",
                             type="primary"):
                    try:
                        safety = restore_from_sql_dump(up.getvalue())
                        st.success(
                            f"{t('import_done')}`{safety.name if safety else 'n/a'}`"
                        )
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

    st.divider()

    # ----- Backup list -------------------------------------------------
    render_section_title(t("available_backups"))
    backups = list_backups()
    if not backups:
        st.info(t("no_backups"))
        return

    # If a restore is awaiting confirmation, show the dialog inline
    pending = st.session_state.get("_pending_restore")
    if pending:
        st.warning(f"{t('confirm_restore_q')}\n\n📁 `{pending}`")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button(t("yes_restore"), type="primary",
                         use_container_width=True, key="confirm_restore"):
                try:
                    safety = restore_from_backup(Path(pending))
                    st.success(
                        f"{t('restore_done')}`{safety.name if safety else 'n/a'}`"
                    )
                    st.session_state.pop("_pending_restore", None)
                    st.rerun()
                except Exception as e:
                    st.error(str(e))
                    st.session_state.pop("_pending_restore", None)
        with cc2:
            if st.button(t("cancel"), use_container_width=True, key="cancel_restore"):
                st.session_state.pop("_pending_restore", None)
                st.rerun()
        st.divider()

    # Render each backup as a row with download / restore / delete buttons
    for b in backups:
        with st.container(border=True):
            r1, r2 = st.columns([3, 2])
            with r1:
                # Color the row name based on label so manual / pre_restore
                # snapshots stand out from the auto-snapshots
                icon = "💾"
                if "_manual_" in b["name"]:
                    icon = "📌"
                elif "_pre_restore_" in b["name"]:
                    icon = "🛟"
                st.markdown(f"{icon} **`{b['name']}`**")
                st.caption(
                    f"📅 {b['created']} &nbsp;·&nbsp; "
                    f"💽 {b['size_kb']} KB &nbsp;·&nbsp; "
                    f"📊 {b['counts']['orders']} / "
                    f"{b['counts']['shipments']} / "
                    f"{b['counts']['shortages']}"
                )
            with r2:
                bc1, bc2, bc3 = st.columns(3)
                # Download — read bytes then surface download_button
                with bc1:
                    dl_key = f"dl_bk_{b['name']}"
                    if st.button(t("download_backup"), key=dl_key,
                                 use_container_width=True):
                        data = b["path"].read_bytes()
                        st.session_state[f"_dl_bk_{b['name']}"] = data
                    if f"_dl_bk_{b['name']}" in st.session_state:
                        st.download_button(
                            "⬇️", data=st.session_state[f"_dl_bk_{b['name']}"],
                            file_name=b["name"],
                            mime="application/x-sqlite3",
                            use_container_width=True,
                            key=f"dl2_bk_{b['name']}",
                        )
                with bc2:
                    if st.button(t("restore_backup"),
                                 key=f"rst_bk_{b['name']}",
                                 use_container_width=True):
                        st.session_state["_pending_restore"] = str(b["path"])
                        st.rerun()
                with bc3:
                    if st.button(t("delete_backup"),
                                 key=f"del_bk_{b['name']}",
                                 use_container_width=True):
                        try:
                            b["path"].unlink()
                            # Drop any cached download bytes for this file
                            st.session_state.pop(f"_dl_bk_{b['name']}", None)
                            st.rerun()
                        except OSError as e:
                            st.error(str(e))


def page_admin_email() -> None:
    """Configure SMTP, per-event recipients, and test sending."""
    render_page_heading(f"📧 {t('email_title')}")
    st.caption(t("email_subtitle"))
    if RESEND_API_KEY:
        st.success(t("email_resend_active"))
    elif os.getenv("RAILWAY_ENVIRONMENT_NAME", "").strip():
        st.warning(t("email_railway_smtp_hint"))

    cfg = load_email_config()

    # ----- Master toggle ------------------------------------------------
    enabled = st.checkbox(
        t("email_enabled"), value=bool(cfg.get("enabled", False)),
        key="em_enabled",
    )

    # ----- SMTP settings ------------------------------------------------
    render_section_title(f"🔧 {t('smtp_settings')}")
    s1, s2 = st.columns(2)
    with s1:
        smtp_host = st.text_input(
            t("smtp_host"), value=cfg.get("smtp_host", ""),
            placeholder="smtp.gmail.com", key="em_host",
        )
        smtp_user = st.text_input(
            t("smtp_user"), value=cfg.get("smtp_user", ""),
            key="em_user",
        )
        from_addr = st.text_input(
            t("smtp_from"), value=cfg.get("from_addr", ""),
            placeholder="orders@example.com", key="em_from",
        )
    with s2:
        smtp_port = st.number_input(
            t("smtp_port"), min_value=1, max_value=65535,
            value=int(cfg.get("smtp_port", 587)), step=1, key="em_port",
        )
        smtp_password = st.text_input(
            t("smtp_password"), value=cfg.get("smtp_password", ""),
            type="password", key="em_pw",
        )
        use_tls = st.checkbox(
            t("smtp_use_tls"), value=bool(cfg.get("use_tls", True)),
            key="em_tls",
            help="Port 465 → implicit SSL (this option ignored). "
                 "Port 587 → STARTTLS (this option recommended).",
        )

    # ----- Per-event recipients -----------------------------------------
    render_section_title(f"🔔 {t('event_settings')}")
    ev_cfg = cfg.get("events", {})
    new_events: dict[str, dict] = {}
    for key, label_key in [
        ("new_order",  "ev_new_order"),
        ("dispatched", "ev_dispatched"),
        ("shortage",   "ev_shortage"),
        ("supplier_order", "ev_supplier_order"),
    ]:
        cur = ev_cfg.get(key, {"enabled": True, "to": []})
        with st.container(border=True):
            ec1, ec2 = st.columns([1, 3])
            with ec1:
                ev_on = st.checkbox(
                    t(label_key),
                    value=bool(cur.get("enabled", True)),
                    key=f"em_ev_{key}",
                )
            with ec2:
                to_text = st.text_area(
                    t("recipients"),
                    value="\n".join(cur.get("to", [])),
                    key=f"em_to_{key}",
                    height=80,
                )
            new_events[key] = {
                "enabled": bool(ev_on),
                "to": [
                    line.strip() for line in to_text.splitlines()
                    if line.strip()
                ],
            }

    # ----- Per-branch recipient (optional) -----------------------------
    render_section_title(f"🏪 {t('branch_recipients')}")
    st.caption(t("branch_email_hint"))
    branch_emails = cfg.get("branch_emails", {}) or {}
    new_branch_emails: dict[str, str] = {}
    for branch in BRANCHES:
        new_branch_emails[branch] = st.text_input(
            branch, value=branch_emails.get(branch, ""),
            key=f"em_br_{branch}", placeholder="(optional)",
        )

    # ----- Save ---------------------------------------------------------
    save_col, _ = st.columns([1, 3])
    with save_col:
        if st.button(t("save_email_cfg"), type="primary",
                     use_container_width=True, key="em_save"):
            new_cfg = {
                "enabled":       bool(enabled),
                "smtp_host":     smtp_host.strip(),
                "smtp_port":     int(smtp_port),
                "smtp_user":     smtp_user.strip(),
                "smtp_password": smtp_password,
                "from_addr":     from_addr.strip(),
                "use_tls":       bool(use_tls),
                "events":        new_events,
                "branch_emails": {k: v.strip() for k, v in
                                  new_branch_emails.items() if v.strip()},
            }
            save_email_config(new_cfg)
            st.success(t("email_cfg_saved"))
            st.rerun()

    st.divider()

    # ----- Test send ---------------------------------------------------
    render_section_title(f"🧪 {t('test_email')}")
    tc1, tc2 = st.columns([3, 1])
    with tc1:
        test_to = st.text_input(
            t("test_to"), value="", key="em_test_to",
            placeholder=cfg.get("from_addr", ""),
        )
    with tc2:
        st.markdown("&nbsp;", unsafe_allow_html=True)  # vertical spacer
        if st.button(t("test_email"), key="em_test_btn",
                     use_container_width=True):
            if not test_to.strip():
                st.warning("⚠️ recipient empty")
            else:
                # Use the *current form values* (not the saved ones), so
                # admin can test before saving.
                test_cfg = {
                    "smtp_host":     smtp_host.strip(),
                    "smtp_port":     int(smtp_port),
                    "smtp_user":     smtp_user.strip(),
                    "smtp_password": smtp_password,
                    "from_addr":     from_addr.strip(),
                    "use_tls":       bool(use_tls),
                }
                ok, err = send_test_email(test_cfg, test_to.strip())
                if ok:
                    st.success(t("test_sent_ok"))
                else:
                    st.error(f"{t('test_sent_fail')}: {err}")

    st.divider()

    # ----- Activity log ------------------------------------------------
    render_section_title(f"📜 {t('email_log_title')}")
    log = list(reversed(load_email_log()))  # newest first
    if not log:
        st.info(t("email_log_empty"))
    else:
        # Show last 50 entries; deeper history kept on disk
        rows = []
        for entry in log[:50]:
            rows.append({
                t("log_time"):   entry.get("ts", ""),
                t("log_event"):  entry.get("event", ""),
                t("log_to"):     ", ".join(entry.get("to", [])),
                "Subject":       entry.get("subject", ""),
                t("log_status"): "✅ " + t("log_ok") if entry.get("ok")
                                 else "❌ " + t("log_fail"),
                "Error":         entry.get("error", ""),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     hide_index=True)
        if st.button(t("log_clear"), key="em_log_clear"):
            try:
                if EMAIL_LOG_PATH.exists():
                    EMAIL_LOG_PATH.unlink()
                st.rerun()
            except Exception as e:
                st.error(str(e))


def _persist_excel_import_feedback(
    storage_key: str, outcome: ExcelSheetImportOutcome
) -> None:
    st.session_state[storage_key] = {
        "written": outcome.n_written,
        "skipped": outcome.n_skipped_benign,
        "failed": outcome.n_failed,
        "errors": list(outcome.failure_messages[:120]),
    }


def _render_excel_import_feedback_panel(storage_key: str) -> None:
    blob = st.session_state.get(storage_key)
    if not blob:
        return
    n_written = int(blob.get("written") or 0)
    n_fail = int(blob.get("failed") or 0)
    skipped = int(blob.get("skipped") or 0)
    if n_fail > 0:
        st.warning(t("import_finished_with_errors"))
    elif n_written > 0:
        st.success(t("import_finished_toast").format(n=n_written))
    elif skipped > 0:
        st.info(t("import_finished_nothing_changed"))
    else:
        st.info(t("import_finished_no_effect_rows"))
    mc1, mc2, mc3 = st.columns(3)
    with mc1:
        st.metric(t("import_metric_written"), n_written)
    with mc2:
        st.metric(t("import_metric_skipped"), skipped)
    with mc3:
        st.metric(t("import_metric_failed"), n_fail)
    if skipped > 0 and n_fail == 0:
        st.caption(t("import_summary_skipped_hint"))
    errs = list(blob.get("errors") or [])
    if errs:
        with st.expander(
            t("import_error_detail"),
            expanded=(len(errs) <= 12),
        ):
            for line in errs:
                st.markdown(
                    html.escape(str(line)),
                    unsafe_allow_html=True,
                )


def page_admin_product_catalog() -> None:
    """管理员商品主档管理页面 - 从 Excel 导入商品到数据库"""
    zh = st.session_state.get("lang") == "zh"
    render_page_heading("📋 商品主档管理" if zh else "📋 Product Catalog")

    st.info("📌 " + ("在此导入商品主档 Excel，导入后所有分店下单页面将显示这些商品。" if zh
            else "Import product catalog Excel here. After import, all branch order pages will show these products."))

    # --- 当前状态 ---
    try:
        with db_conn() as conn:
            count = conn.execute("SELECT COUNT(*) FROM product_catalog").fetchone()[0]
        if count > 0:
            st.success(f"✅ {'当前数据库中有' if zh else 'Database has'} **{count}** {'条商品记录' if zh else 'product records'}")
        else:
            st.warning("⚠️ " + ("数据库中没有商品记录，请导入商品主档。" if zh else "No products in database. Please import a catalog."))
    except Exception:
        st.warning("⚠️ " + ("商品表尚未创建，导入后将自动创建。" if zh else "Product table not yet created. Will be created on import."))
        count = 0

    # --- 导入 Excel ---
    st.subheader("📥 " + ("导入商品 Excel" if zh else "Import Product Excel"))
    st.caption("支持列名：商品编号/ItemCode, 条码/Barcode, 名称/Name, 零售价/价格/Price, 单位/Unit, 类别/Category, 每箱个数/PcsPerCarton"
               if zh else "Supported columns: ItemCode, Barcode, Name, Price, Unit, Category, PcsPerCarton")

    uploaded = st.file_uploader(
        "选择 Excel 文件" if zh else "Choose Excel file",
        type=["xlsx", "xls"],
        key="catalog_uploader",
    )

    replace_mode = st.checkbox(
        "🔄 " + ("清空旧数据后导入（替换模式）" if zh else "Clear old data before import (replace mode)"),
        value=True,
    )

    if uploaded and st.button("🚀 " + ("开始导入" if zh else "Start Import"), type="primary"):
        try:
            raw_df = pd.read_excel(uploaded, dtype=str)
            raw_df = raw_df.rename(columns=lambda c: str(c).replace("\ufeff", "").strip())

            # Column mapping
            col_lower = {str(c).strip().lower(): c for c in raw_df.columns}
            renames = {}

            ic_aliases = ["商品编号", "货号", "编号", "内部编号", "sku", "itemcode", "item_code"]
            for a in ic_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "ItemCode":
                    renames[hit] = "ItemCode"
                    break

            bc_aliases = ["条码", "商品条码", "条形码", "barcode", "ean", "upc"]
            for a in bc_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "Barcode":
                    renames[hit] = "Barcode"
                    break

            nm_aliases = ["商品名称", "名称", "品名", "name", "description"]
            for a in nm_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "Name":
                    renames[hit] = "Name"
                    break

            price_aliases = ["零售价", "价格", "单价", "售价", "price", "unitprice", "unit_price",
                             "listprice", "批发价", "进货价", "单价(元)", "价格(元)", "售价(元)"]
            for a in price_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "Price":
                    renames[hit] = "Price"
                    break

            unit_aliases = ["单位", "unit", "计量单位"]
            for a in unit_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "Unit":
                    renames[hit] = "Unit"
                    break

            cat_aliases = ["类别", "分类", "category", "商品分类"]
            for a in cat_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "Category":
                    renames[hit] = "Category"
                    break

            ppc_aliases = ["每箱个数", "箱规", "箱装数", "pcspercarton", "packsize", "qtypercarton"]
            for a in ppc_aliases:
                hit = col_lower.get(a.lower())
                if hit and hit != "PcsPerCarton":
                    renames[hit] = "PcsPerCarton"
                    break

            if renames:
                raw_df = raw_df.rename(columns=renames)

            # Ensure required columns
            if "Name" not in raw_df.columns:
                st.error("❌ " + ("找不到商品名称列! 请确保 Excel 中有 名称 或 Name 列。" if zh
                         else "Cannot find Name column! Ensure Excel has a 'Name' column."))
                return

            for col in ["ItemCode", "Barcode", "Unit", "Category"]:
                if col not in raw_df.columns:
                    raw_df[col] = ""
            if "Price" not in raw_df.columns:
                raw_df["Price"] = 0.0
            if "PcsPerCarton" not in raw_df.columns:
                raw_df["PcsPerCarton"] = 0.0

            # Clean data
            raw_df["Name"] = raw_df["Name"].fillna("").astype(str).str.strip()
            raw_df = raw_df[raw_df["Name"] != ""]  # Remove rows without name

            if raw_df.empty:
                st.error("❌ " + ("Excel 中没有有效的商品数据（所有行的名称为空）。" if zh
                         else "No valid product data in Excel (all Name fields are empty)."))
                return

            raw_df["Price"] = pd.to_numeric(raw_df["Price"], errors="coerce").fillna(0.0)
            raw_df["PcsPerCarton"] = pd.to_numeric(raw_df["PcsPerCarton"], errors="coerce").fillna(0.0)

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            with db_conn() as conn:
                if replace_mode:
                    conn.execute("DELETE FROM product_catalog")

                inserted = 0
                for _, row in raw_df.iterrows():
                    conn.execute(
                        "INSERT INTO product_catalog (item_code, barcode, name, unit, price, category, pcs_per_carton, created_at, updated_at) "
                        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                        (
                            str(row.get("ItemCode", "") or "").strip(),
                            str(row.get("Barcode", "") or "").strip(),
                            str(row["Name"]).strip(),
                            str(row.get("Unit", "") or "").strip(),
                            float(row.get("Price", 0) or 0),
                            str(row.get("Category", "General") or "General").strip(),
                            float(row.get("PcsPerCarton", 0) or 0),
                            now, now,
                        )
                    )
                    inserted += 1

            # Clear product cache
            load_products.clear()

            st.success(f"🎉 {'导入成功！共导入' if zh else 'Import successful! Imported'} **{inserted}** {'条商品记录。' if zh else ' product records.'}")
            st.balloons()

        except Exception as e:
            st.error(f"❌ {'导入失败' if zh else 'Import failed'}: {e}")

    # --- 查看当前商品 ---
    if count > 0:
        st.subheader("📊 " + ("当前商品预览（前 20 条）" if zh else "Current Products Preview (first 20)"))
        try:
            with db_conn() as conn:
                preview = conn.execute(
                    "SELECT item_code, barcode, name, price, category FROM product_catalog LIMIT 20"
                ).fetchall()
            preview_df = pd.DataFrame([dict(r) for r in preview])
            preview_df.columns = ["商品编号" if zh else "ItemCode",
                                  "条码" if zh else "Barcode",
                                  "名称" if zh else "Name",
                                  "价格" if zh else "Price",
                                  "类别" if zh else "Category"]
            st.dataframe(preview_df, use_container_width=True)
        except Exception as e:
            st.error(str(e))

        # --- 清空按钮 ---
        st.subheader("⚠️ " + ("危险操作" if zh else "Danger Zone"))
        confirm = st.text_input("输入 CLEAR 确认清空" if zh else "Type CLEAR to confirm", key="catalog_clear_confirm")
        if st.button("🗑️ " + ("清空全部商品" if zh else "Clear All Products"), type="secondary"):
            if confirm.strip() == "CLEAR":
                with db_conn() as conn:
                    conn.execute("DELETE FROM product_catalog")
                load_products.clear()
                st.success("已清空" if zh else "Cleared")
                st.rerun()
            else:
                st.warning("请输入 CLEAR 确认" if zh else "Please type CLEAR to confirm")


def page_admin_price_management() -> None:
    render_section_title(f"💰 {t('nav_price')}")
    price_import_feedback_key = "_price_import_feedback"
    with st.expander(t("excel_batch_import"), expanded=False):
        c1, c2 = st.columns([2, 4])
        with c1:
            st.download_button(
                label=t("download_price_template"),
                data=_build_price_import_template_bytes(),
                file_name="价格导入模板.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="price_download_template",
            )
            price_file = st.file_uploader(
                t("select_import_file"),
                type=["xlsx", "xls"],
                key="price_import_file",
                help="Excel: ItemCode / Barcode / Name / Price",
            )
            if st.button(
                t("price_import_from_products"),
                type="primary",
                use_container_width=True,
                key="price_import_btn",
            ):
                if price_file is None:
                    st.warning(t("import_file_required"))
                else:
                    up_df = _load_products_sheet_from_upload(price_file)
                    if up_df.empty:
                        st.warning(t("import_file_invalid"))
                    else:
                        nm = up_df["Name"].astype(str).str.strip()
                        with_name = up_df[nm != ""]
                        if not with_name.empty:
                            pv = pd.to_numeric(
                                with_name["Price"], errors="coerce"
                            ).fillna(0)
                            if (pv == 0).all():
                                st.warning(t("price_import_no_numeric_prices"))
                        outcome = import_prices_from_products_detailed(df=up_df)
                        _persist_excel_import_feedback(
                            price_import_feedback_key,
                            outcome,
                        )
                        st.session_state.pop("price_import_file", None)
                        st.rerun()
        with c2:
            _render_excel_import_feedback_panel(price_import_feedback_key)
            st.caption(t("price_hint"))

    if "price_applied_q" not in st.session_state:
        st.session_state["price_applied_q"] = ""

    with st.form("price_search_form"):
        st.text_input(
            t("grid_keyword_search"),
            key="price_kw_form_field",
            placeholder="2151",
        )
        submitted = st.form_submit_button(
            t("grid_search_apply"),
            use_container_width=True,
        )
    if submitted:
        st.session_state["price_applied_q"] = (
            st.session_state.get("price_kw_form_field", "") or ""
        ).strip()

    active_q = st.session_state["price_applied_q"].strip()
    if not active_q:
        st.info(t("grid_price_enter_keyword"))
        st.caption(t("grid_price_click_edit_hint"))
        return

    products = search_products(active_q, limit=500)
    if products.empty:
        st.warning(t("no_data"))
        st.caption(t("grid_price_click_edit_hint"))
        return

    sheet_map = _sheet_price_lookup()
    rows: list[dict] = []
    base_prices: dict[str, float] = {}
    for _, row in products.iterrows():
        nm = str(row.get("Name", "") or "").strip()
        if not nm:
            continue
        ic = str(row.get("ItemCode", "") or "").strip()
        bc = str(row.get("Barcode", "") or "").strip()
        k = _inventory_item_key(ic, bc, nm)
        eff = float(row.get("Price", 0) or 0)
        lst = float(sheet_map.get(k, eff))
        base_prices[k] = eff
        rows.append({
            "ItemCode": ic,
            "Barcode": bc,
            "Name": nm,
            "Unit": str(row.get("Unit", "") or "").strip(),
            "ListPrice": round(lst, 2),
            "Price": round(eff, 2),
        })
    if not rows:
        st.info(t("no_data"))
        return

    edit_df = pd.DataFrame(rows)
    st.caption(t("grid_price_click_edit_hint"))
    edited = st.data_editor(
        edit_df,
        key="price_management_grid",
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        disabled=["ItemCode", "Barcode", "Name", "Unit", "ListPrice"],
        column_config={
            "ItemCode": st.column_config.TextColumn(t("item_code"), width="small"),
            "Barcode": st.column_config.TextColumn(t("barcode"), width="medium"),
            "Name": st.column_config.TextColumn(t("name"), width="large"),
            "Unit": st.column_config.TextColumn(t("unit"), width="small"),
            "ListPrice": st.column_config.NumberColumn(
                t("price_sheet_col"),
                format="%.2f",
            ),
            "Price": st.column_config.NumberColumn(
                t("price_current_col"),
                format="%.2f",
                min_value=0.0,
                step=0.01,
            ),
        },
    )
    if st.button(t("grid_price_save"), type="primary", key="price_save_grid"):
        n = 0
        with db_conn() as conn:
            for _, row in edited.iterrows():
                ic = str(row.get("ItemCode", "") or "").strip()
                bc = str(row.get("Barcode", "") or "").strip()
                nm = str(row.get("Name", "") or "").strip()
                if not nm:
                    continue
                k = _inventory_item_key(ic, bc, nm)
                new_p = float(row.get("Price", 0) or 0)
                old_p = float(base_prices.get(k, new_p))
                if abs(new_p - old_p) < 1e-9:
                    continue
                _upsert_product_price(
                    conn,
                    item_code=ic,
                    barcode=bc,
                    name=nm,
                    price=new_p,
                    operator="admin_price_edit",
                )
                n += 1
        if n == 0:
            st.warning(t("grid_price_no_changes"))
        else:
            load_products.clear()
            st.success(t("price_import_done").format(n=n))
            st.rerun()

    with st.expander(t("price_full_list_expander"), expanded=False):
        full = load_products(_products_mtime(), _inventory_version(), _price_version())
        if full.empty:
            st.info(t("no_data"))
        else:
            with db_conn() as conn:
                price_rows = conn.execute(
                    "SELECT item_key, updated_at FROM product_prices"
                ).fetchall()
            updated_by_key = {
                str(r["item_key"]): str(r["updated_at"] or "")
                for r in price_rows
            }
            summary_rows = []
            for _, row in full.iterrows():
                name = str(row.get("Name", "") or "").strip()
                if not name:
                    continue
                ic = str(row.get("ItemCode", "") or "").strip()
                bc = str(row.get("Barcode", "") or "").strip()
                key = _inventory_item_key(ic, bc, name)
                summary_rows.append({
                    t("item_code"): ic,
                    t("barcode"): bc,
                    t("name"): name,
                    t("price"): float(row.get("Price", 0) or 0),
                    t("updated_at"): updated_by_key.get(key, ""),
                })
            st.dataframe(
                pd.DataFrame(summary_rows),
                use_container_width=True,
                hide_index=True,
            )


def _render_inventory_workspace(
    *,
    operator_key: str,
    key_prefix: str = "",
) -> None:
    """Import / adjust / current snapshot — shared by admin hub and warehouse."""
    pf = key_prefix
    inv_feedback_key = f"{pf}_inv_excel_import_feedback"
    c1, c2 = st.columns([2, 4])
    with c1:
        st.download_button(
            label=t("download_inv_template"),
            data=_build_inventory_import_template_bytes(),
            file_name="库存导入模板.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{pf}inv_download_template",
        )
        inv_file = st.file_uploader(
            t("select_import_file"),
            type=["xlsx", "xls"],
            key=f"{pf}inv_import_file",
            help="Excel: ItemCode / Barcode / Name / StockCartons / StockPcs / StockTotal",
        )
        if st.button(
            t("inv_import_append"),
            type="primary",
            use_container_width=True,
            key=f"{pf}inv_import_append_btn",
        ):
            if inv_file is None:
                st.warning(t("import_file_required"))
            else:
                up_df = _load_products_sheet_from_upload(inv_file)
                if up_df.empty:
                    st.warning(t("import_file_invalid"))
                else:
                    outcome = import_inventory_from_products_detailed(
                        overwrite=False, df=up_df
                    )
                    _persist_excel_import_feedback(inv_feedback_key, outcome)
                    st.session_state.pop(f"{pf}inv_import_file", None)
                    st.rerun()
        with st.expander(t("inv_import_overwrite_section")):
            st.caption(t("inv_import_overwrite_hint"))
            if st.button(
                t("inv_import_overwrite"),
                type="secondary",
                use_container_width=True,
                key=f"{pf}inv_import_ow_btn",
            ):
                if inv_file is None:
                    st.warning(t("import_file_required"))
                else:
                    up_df = _load_products_sheet_from_upload(inv_file)
                    if up_df.empty:
                        st.warning(t("import_file_invalid"))
                    else:
                        outcome = import_inventory_from_products_detailed(
                            overwrite=True, df=up_df
                        )
                        _persist_excel_import_feedback(inv_feedback_key, outcome)
                        st.session_state.pop(f"{pf}inv_import_file", None)
                        st.rerun()
    with c2:
        _render_excel_import_feedback_panel(inv_feedback_key)
        st.caption(t("inv_import_append_hint"))
        st.caption("库存以数据库为准；发货时会自动扣减。")

    st.divider()
    render_section_title(t("inv_adjust"))
    sq = f"{pf}inv_applied_q"
    fk = f"{pf}inv_kw_form_field"
    if sq not in st.session_state:
        st.session_state[sq] = ""

    with st.form(f"{pf}inv_search_form"):
        st.text_input(
            t("grid_keyword_search"),
            key=fk,
            placeholder="2151",
        )
        inv_sub = st.form_submit_button(
            t("grid_search_apply"),
            use_container_width=True,
        )
    if inv_sub:
        st.session_state[sq] = (st.session_state.get(fk, "") or "").strip()

    active_inv_q = st.session_state[sq].strip()
    if not active_inv_q:
        st.info(t("grid_price_enter_keyword"))
    else:
        inv_products = search_products(active_inv_q, limit=500)
        if inv_products.empty:
            st.warning(t("no_data"))
        else:
            st.caption(t("grid_inv_delta_hint"))
            inv_rows: list[dict] = []
            for _, r in inv_products.iterrows():
                nm = str(r.get("Name", "") or "").strip()
                if not nm:
                    continue
                ic = str(r.get("ItemCode", "") or "").strip()
                bc = str(r.get("Barcode", "") or "").strip()
                inv_rows.append({
                    "ItemCode": ic,
                    "Barcode": bc,
                    "Name": nm,
                    "Unit": str(r.get("Unit", "") or "").strip(),
                    "StockCartons": int(
                        float(r.get("StockCartons", 0) or 0)
                    ),
                    "StockPcs": int(float(r.get("StockPcs", 0) or 0)),
                    "DeltaCt": 0,
                    "DeltaPc": 0,
                })
            if not inv_rows:
                st.info(t("no_data"))
            else:
                inv_edit = pd.DataFrame(inv_rows)
                inv_edited = st.data_editor(
                    inv_edit,
                    key=f"{pf}inv_grid",
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    disabled=[
                        "ItemCode",
                        "Barcode",
                        "Name",
                        "Unit",
                        "StockCartons",
                        "StockPcs",
                    ],
                    column_config={
                        "ItemCode": st.column_config.TextColumn(
                            t("item_code"), width="small"
                        ),
                        "Barcode": st.column_config.TextColumn(
                            t("barcode"), width="medium"
                        ),
                        "Name": st.column_config.TextColumn(
                            t("name"), width="large"
                        ),
                        "Unit": st.column_config.TextColumn(
                            t("unit"), width="small"
                        ),
                        "StockCartons": st.column_config.NumberColumn(
                            t("grid_inv_stock_ct"),
                            format="%d",
                        ),
                        "StockPcs": st.column_config.NumberColumn(
                            t("grid_inv_stock_pc"),
                            format="%d",
                        ),
                        "DeltaCt": st.column_config.NumberColumn(
                            t("grid_inv_delta_ct"),
                            format="%d",
                            step=1,
                        ),
                        "DeltaPc": st.column_config.NumberColumn(
                            t("grid_inv_delta_pc"),
                            format="%d",
                            step=1,
                        ),
                    },
                )
                if st.button(
                    t("grid_inv_apply_rows"),
                    type="primary",
                    key=f"{pf}inv_apply_grid_btn",
                ):
                    n = 0
                    with db_conn() as conn:
                        for _, row in inv_edited.iterrows():
                            dct = int(float(row.get("DeltaCt", 0) or 0))
                            dpc = int(float(row.get("DeltaPc", 0) or 0))
                            if dct == 0 and dpc == 0:
                                continue
                            nm = str(row.get("Name", "") or "").strip()
                            if not nm:
                                continue
                            ic = str(row.get("ItemCode", "") or "").strip()
                            bc = str(row.get("Barcode", "") or "").strip()
                            ut = str(row.get("Unit", "") or "").strip()
                            _apply_inventory_change(
                                conn,
                                txn_type="ADJUST",
                                item_code=ic if ic != "-" else "",
                                barcode=bc if bc != "-" else "",
                                name=nm,
                                unit=ut,
                                change_ct=dct,
                                change_pc=dpc,
                                operator=operator_key,
                            )
                            n += 1
                    if n == 0:
                        st.warning(t("grid_inv_no_delta"))
                    else:
                        st.success(t("inv_adjust_ok"))
                        st.rerun()

    st.divider()
    render_section_title(t("inv_current"))
    with db_conn() as conn:
        inv_df = pd.read_sql_query(
            "SELECT item_code, barcode, name, stock_cartons, stock_pcs, updated_at "
            "FROM inventory ORDER BY name ASC LIMIT 1000",
            conn,
        )
    if inv_df.empty:
        st.info(t("no_data"))
    else:
        inv_df.columns = [
            t("item_code"),
            t("barcode"),
            t("name"),
            t("qty_cartons"),
            t("qty_pcs"),
            t("order_date"),
        ]
        st.dataframe(inv_df, use_container_width=True, hide_index=True)


def page_warehouse_inventory() -> None:
    """Warehouse: same inventory tools as admin (import / adjust / view current)."""
    render_page_heading(f"📦 {t('nav_inventory')}")
    _render_inventory_workspace(
        operator_key="warehouse_adjust",
        key_prefix="wh_inv_",
    )


def page_admin_inventory() -> None:
    render_page_heading(f"📦 {t('inv_title')}")
    module_options = [
        t("inv_title"),
        t("nav_price"),
        t("nav_dispatch"),
        t("nav_all_orders"),
        t("nav_images"),
    ]
    module = st.radio(
        t("choose_module"),
        options=module_options,
        horizontal=True,
        key="admin_inventory_module",
    )
    if module == t("nav_price"):
        page_admin_price_management()
        return
    if module == t("nav_dispatch"):
        page_warehouse_pending()
        return
    if module == t("nav_all_orders"):
        page_admin_all_orders()
        return
    if module == t("nav_images"):
        page_admin_images()
        return

    _render_inventory_workspace(
        operator_key="admin_adjust",
        key_prefix="",
    )


def page_admin_export() -> None:
    render_page_heading(f"📥 {t('nav_export')}")
    today = datetime.now().date()
    c1, c2 = st.columns(2)
    with c1: d1 = st.date_input(t("from_date"), value=today, key="exp_d1")
    with c2: d2 = st.date_input(t("to_date"),   value=today, key="exp_d2")
    d1s = d1.strftime("%Y-%m-%d"); d2s = d2.strftime("%Y-%m-%d")

    for label_key, fn, fname in [
        ("exp_picking", export_picking_list,    "picking_list"),
        ("exp_recon",   export_reconciliation,  "reconciliation"),
        ("exp_short",   export_shortage_report, "shortage_report"),
        ("exp_inventory", export_inventory_report, "inventory_report"),
    ]:
        st.divider()
        render_section_title(t(label_key))
        if st.button(f"{t('generate')} — {t(label_key)}", key=f"gen_{fname}"):
            bio = fn(d1s, d2s)
            if bio is None:
                st.warning(t("no_data"))
            else:
                st.download_button(
                    f"⬇️ {t('download')} {fname}_{d1s}_{d2s}.xlsx",
                    data=bio,
                    file_name=f"{fname}_{d1s}_{d2s}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


def page_admin_messages_center() -> None:
    render_page_heading(f"📨 {t('nav_messages')} · {t('admin_messages_hub')}")
    module_options = [
        t("nav_messages"),
        t("nav_short_mgmt"),
        t("nav_arrivals"),
        t("nav_email"),
    ]
    module = st.radio(
        t("choose_module"),
        options=module_options,
        horizontal=True,
        key="admin_messages_module",
    )
    if module == t("nav_short_mgmt"):
        page_warehouse_shortages()
        return
    if module == t("nav_arrivals"):
        page_admin_arrivals()
        return
    if module == t("nav_email"):
        page_admin_email()
        return
    page_messages()


# =========================================================================
# WORKSPACE ROUTERS
# =========================================================================
def _render_sidebar_nav_sections(
    sections: list[tuple[str | None, list[tuple[str, str]]]],
) -> list[tuple[str, str]]:
    """分组侧栏导航；返回扁平列表供当前页提示条使用。"""
    flat: list[tuple[str, str]] = []
    for title, items in sections:
        if title:
            st.sidebar.markdown(
                f'<div class="nav-section-head">{html.escape(title)}</div>',
                unsafe_allow_html=True,
            )
        _render_sidebar_nav(items)
        flat.extend(items)
        if title:
            st.sidebar.markdown('<div class="nav-section-gap"></div>', unsafe_allow_html=True)
    return flat


def _render_sidebar_nav(items: list[tuple[str, str]]) -> None:
    for page_key, label in items:
        is_active = st.session_state.page == page_key
        show_label = f"👉 {label}" if is_active else label
        if st.sidebar.button(
            show_label, key=f"nav_{page_key}",
            use_container_width=True,
            type="primary" if is_active else "secondary",
        ):
            st.session_state.page = page_key
            st.session_state._nav_from_sidebar = True
            try:
                st.query_params[URL_PAGE_QUERY_KEY] = page_key
            except Exception:
                pass
            try:
                import auth as sunshine_auth

                sunshine_auth.touch_cookie_page(page_key)
            except Exception:
                pass
            # Leaving any page should drop the order-confirm modal state.
            st.session_state["confirming"] = False
            st.rerun()


def _nav_messages_label() -> str:
    unread = count_unread_notifications()
    base = t("nav_messages")
    return f"{base} ({unread})" if unread > 0 else base


def _page_theme_color(page_key: str) -> str:
    color_map = {
        "order": "#1565c0",
        "order_done": "#2e7d32",
        "my_orders": "#6a1b9a",
        "my_short": "#ef6c00",
        "pending": "#00838f",
        "short_in": "#c62828",
        "history": "#455a64",
        "supplier": "#00695c",
        "supplier_order": "#00695c",
        "verify_inbound": "#00838f",
        "admin_suppliers": "#004d40",
        "order_success": "#2e7d32",
        "verify_success": "#2e7d32",
        "dashboard": "#3949ab",
        "shelf_mobile": "#00695c",
        "accounts": "#5e35b1",
        "audit": "#37474f",
        "all_orders": "#5d4037",
        "dispatch": "#0277bd",
        "short_mgmt": "#d84315",
        "arrivals": "#00897b",
        "inventory": "#2e7d32",
        "export": "#546e7a",
        "images": "#8e24aa",
        "backup": "#6d4c41",
        "email": "#7b1fa2",
        "product_master": "#1565c0",
        "messages": "#1976d2",
        "ai": "#283593",
    }
    return color_map.get(page_key, "#1976d2")


def _render_active_page_hint(items: list[tuple[str, str]]) -> None:
    page_key = st.session_state.get("page") or ""
    label_map = {k: v for k, v in items}
    label = label_map.get(page_key, page_key)
    if not label:
        return
    color = _page_theme_color(page_key)
    st.markdown(
        f'<div class="active-page-hint" style="--hint-accent:{color};">'
        f'<span class="active-page-hint__k">当前页面</span>'
        f'<span class="active-page-hint__v">{html.escape(label)}</span></div>',
        unsafe_allow_html=True,
    )


def _stock_fill_from_product_row(mrow) -> None:
    """把商品库一行写入入库表单的 session 字段，并标记为"已锁定"。"""
    st.session_state["stk_name"] = str(mrow.get("Name", "") or "")
    st.session_state["stk_code"] = str(
        mrow.get("ItemCode") or mrow.get("Barcode") or ""
    ).strip()
    st.session_state["stk_unit"] = str(mrow.get("Unit", "") or "")
    st.session_state["_stk_locked"] = True


def _stock_series_col(df: pd.DataFrame, col: str) -> pd.Series:
    """安全取列并规整为去空白字符串 Series（列缺失时返回空串）。"""
    if col in df.columns:
        return df[col].fillna("").astype(str).str.strip()
    return pd.Series([""] * len(df), index=df.index)


def _on_stock_search_enter() -> None:
    """搜索框回车 / 扫码枪录入：若能唯一精确匹配编号或条码，直接填入并清空搜索框，
    实现"扫一个填一个"的流式录入。回调在 rerun 前执行，故可安全改写控件 session。"""
    q = (st.session_state.get("stock_pick_q") or "").strip()
    if not q:
        return
    df = search_products(q, limit=50)
    if df is None or df.empty:
        return
    bc = _stock_series_col(df, "Barcode")
    ic = _stock_series_col(df, "ItemCode")
    exact = df[(bc == q) | (ic == q)]
    row = None
    if len(exact) == 1:
        row = exact.iloc[0]
    elif len(df) == 1:
        row = df.iloc[0]
    if row is not None:
        _stock_fill_from_product_row(row)
        # 清空搜索/下拉，准备下一次扫码
        st.session_state["stock_pick_q"] = ""
        st.session_state["stock_pick_sel"] = "—"


def _reset_stock_pick() -> None:
    """复位入库表单：清空锁定与所有相关字段（用 on_click 回调，在控件重建前执行，
    因此可安全删除 widget 的 session key，避免"实例化后修改"报错）。"""
    for k in ("stk_name", "stk_code", "stk_unit",
              "stock_pick_q", "stock_pick_sel"):
        st.session_state.pop(k, None)
    st.session_state["_stk_locked"] = False


def _on_stock_select() -> None:
    """下拉选择即填入（取消"填入表单"中转按钮）。"""
    label = st.session_state.get("stock_pick_sel", "")
    mp = st.session_state.get("_stk_opt_map", {})
    if label and label != "—" and label in mp:
        d = mp[label]
        st.session_state["stk_name"] = d["name"]
        st.session_state["stk_code"] = d["code"]
        st.session_state["stk_unit"] = d["unit"]
        st.session_state["_stk_locked"] = True


def page_branch_stock() -> None:
    """分店库存/临期页：录入批次（入库）+ 查看本店当前批次与临期/过期提醒。"""
    branch = st.session_state.get("branch") or ""
    render_page_heading(t("nav_stock"), t("stock_subtitle"))
    warn_days = expiry_warn_days()

    # ---- 商品来源：即搜即填 / 扫码即填（搜索器在 st.form 之外）----
    # 表单内控件提交前不会 rerun，无法联动；故搜索 + 选择放在表单外，
    # 通过 on_change 回调把选中商品写入表单字段的 session key。
    st.markdown(f"#### {t('stock_pick_title')}")
    st.text_input(
        t("stock_pick_ph"), key="stock_pick_q",
        placeholder=t("stock_pick_ph"),
        help="支持扫码枪：扫码后回车自动匹配并填入 / Scanner: scan + Enter to auto-fill",
        on_change=_on_stock_search_enter,
    )
    pick_q = (st.session_state.get("stock_pick_q") or "").strip()
    if pick_q:
        matches = search_products(pick_q, limit=50)
        if matches is not None and not matches.empty:
            options = ["—"]
            opt_map: dict[str, dict] = {}
            for _, mrow in matches.iterrows():
                code_disp = str(mrow.get("ItemCode") or mrow.get("Barcode") or "").strip()
                label = f"{mrow.get('Name', '')}" + (f"｜{code_disp}" if code_disp else "")
                if label in opt_map:                      # 避免重复标签互相覆盖
                    label = f"{label} #{len(opt_map)}"
                options.append(label)
                opt_map[label] = {
                    "name": str(mrow.get("Name", "") or ""),
                    "code": code_disp,
                    "unit": str(mrow.get("Unit", "") or ""),
                }
            st.session_state["_stk_opt_map"] = opt_map   # 供 on_change 回调取用
            # 选择即填入：on_change 回调直接写 session，无需中转按钮
            st.selectbox(
                t("stock_pick_select"), options,
                key="stock_pick_sel",
                on_change=_on_stock_select,
            )
        else:
            st.caption(t("no_results"))

    locked = bool(st.session_state.get("_stk_locked", False))
    if locked:
        lc1, lc2 = st.columns([3, 1])
        lc1.success(f"✓ {st.session_state.get('stk_name', '')}")
        # 用 on_click 回调复位，避免在主流程里改写已实例化的 widget key
        lc2.button(t("stock_pick_clear"), key="stk_clear",
                   on_click=_reset_stock_pick, use_container_width=True)

    # ---- 入库表单（核心：必须填过期日期）----
    with st.form("branch_stock_add", clear_on_submit=True):
        st.markdown(f"#### {t('stock_add_title')}")
        if locked:
            st.caption(f"✅ {t('stock_locked_hint')}")
        # 注意：表单内 disabled 控件提交时不回传值（同 HTML 原生行为），
        # 会导致校验"商品名称为空"而无法保存。故这里保持可编辑（自动带入，
        # 允许微调），既消除该 Bug，又满足"自动填充 + 可微调"。
        name = st.text_input(t("stock_name"), key="stk_name")
        c1, c2 = st.columns(2)
        code = c1.text_input(t("stock_code"), key="stk_code")
        unit = c2.text_input(t("stock_unit"), key="stk_unit")
        c3, c4, c5 = st.columns(3)
        qty_ct = c3.number_input(t("stock_qty_ct"), min_value=0, step=1, value=0)
        qty_pc = c4.number_input(t("stock_qty_pc"), min_value=0, step=1, value=0)
        # 过期日期默认留空，等待店员录入（避免误用默认值）。
        expire = c5.date_input(t("stock_expire"), value=None, format="YYYY-MM-DD")
        batch_no = st.text_input(t("stock_batch_no"))
        submitted = st.form_submit_button(
            t("stock_add_btn"), type="primary", use_container_width=True
        )
    if submitted:
        if expire is None:
            st.error(t("stock_need_expire"))
        elif int(qty_ct) <= 0 and int(qty_pc) <= 0:
            st.error(t("stock_need_qty"))
        elif not (name or "").strip():
            st.error(t("stock_name"))
        else:
            # 编号若为 8 位以上纯数字按条码处理，否则当作商品编号。
            code_s = (code or "").strip()
            is_barcode = code_s.isdigit() and len(code_s) >= 8
            try:
                add_branch_batch(
                    branch=branch,
                    name=name,
                    item_code="" if is_barcode else code_s,
                    barcode=code_s if is_barcode else "",
                    unit=unit,
                    qty_cartons=int(qty_ct),
                    qty_pcs=int(qty_pc),
                    expire_date=expire.strftime("%Y-%m-%d"),
                    batch_no=batch_no,
                    received_by=st.session_state.get("account_username") or "",
                )
                # 复位锁定，准备下一条录入（_stk_locked 非 widget key，安全；
                # 表单字段由 clear_on_submit 自动清空）。
                st.session_state["_stk_locked"] = False
                st.success(t("stock_add_ok"))
                st.rerun()
            except ValueError as e:
                st.error(str(e))

    st.divider()

    # ---- 当前批次 + 临期提醒 ----
    with db_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM branch_inventory_batches
            WHERE branch = ? AND status = 'active'
              AND (qty_cartons > 0 OR qty_pcs > 0)
            ORDER BY expire_date ASC
            """,
            (branch,),
        ).fetchall()

    if not rows:
        st.info(t("stock_none"))
        return

    today = datetime.now().date()

    def _days_left(r) -> int | None:
        exp = _parse_ymd(r["expire_date"])
        return (exp.date() - today).days if exp else None

    # 临期/过期高亮（<= 预警天数 或 已过期）
    warn_rows = [r for r in rows if (_days_left(r) is not None and _days_left(r) <= warn_days)]
    if warn_rows:
        lines = []
        for r in warn_rows:
            dl = _days_left(r)
            tag = t("stock_expired") if dl < 0 else f"{dl} {t('stock_days_left')}"
            lines.append(f"- **{r['name']}** · {r['expire_date']} · {tag}")
        st.warning(f"#### ⚠️ {t('stock_warn_title')}\n" + "\n".join(lines))

    st.markdown(f"#### {t('stock_list_title')}")
    table = []
    for r in rows:
        dl = _days_left(r)
        if dl is None:
            status_txt = ""
        elif dl < 0:
            status_txt = t("stock_expired")
        else:
            status_txt = f"{dl} {t('stock_days_left')}"
        qty = []
        if int(r["qty_cartons"] or 0) > 0:
            qty.append(f"{int(r['qty_cartons'])} {t('stock_qty_ct')}")
        if int(r["qty_pcs"] or 0) > 0:
            qty.append(f"{int(r['qty_pcs'])} {t('stock_qty_pc')}")
        table.append({
            t("stock_name"): r["name"],
            t("stock_code"): (r["item_code"] or r["barcode"] or ""),
            t("stock_qty_ct"): " / ".join(qty),
            t("stock_expire"): r["expire_date"],
            t("stock_warn_title"): status_txt,
        })
    st.dataframe(pd.DataFrame(table), use_container_width=True, hide_index=True)


def render_branch() -> None:
    render_sidebar_lang_switch()

    st.sidebar.markdown(
        (
            "<div class='branch-identity'>"
            f"<div class='k'>{html.escape(t('current_branch'))}</div>"
            f"<div class='v'>{html.escape(st.session_state.branch)}</div>"
            "</div>"
        ),
        unsafe_allow_html=True,
    )
    st.sidebar.divider()

    full_nav = [
        ("order", t("nav_order"), "order"),
        ("my_orders", t("nav_my_orders"), "my_orders"),
        ("my_short", t("nav_my_short"), "my_short"),
        ("messages", _nav_messages_label(), "messages"),
        ("ai", t("nav_ai"), "ai"),
    ]
    nav_items = [(a, b) for a, b, p in full_nav if has_branch_perm(p)]
    if not nav_items:
        st.error(t("acct_perm_need_admin"))
        logout_button("logout_branch_noperm")
        return

    # 库存/临期页对所有分店账号开放（不依赖逐个授权）。
    nav_items.append(("stock", t("nav_stock")))

    allowed = _branch_nav_allowed_keys(nav_items)
    _apply_url_page_to_session(Role.BRANCH, allowed)

    ensure_branch_page_allowed(st.session_state.get("page") or nav_items[0][0])

    _hydrate_branch_cart_if_needed()

    _render_sidebar_nav(nav_items)
    st.sidebar.divider()
    with st.sidebar:
        logout_button("logout_branch")
    _render_active_page_hint(nav_items)

    pages = {
        "order":     page_branch_order,
        "order_done": page_branch_order_done,
        "my_orders": page_branch_my_orders,
        "my_short":  page_branch_shortages,
        "messages":  page_messages,
        "ai":        page_ai_assistant,
        "stock":     page_branch_stock,
    }
    pages.get(st.session_state.page, page_branch_order)()
    _sync_cookie_default_page(st.session_state.page)
    _persist_branch_cart()


def _sc_apply_inventory_receive(
    *,
    item_code: str,
    barcode: str,
    name: str,
    unit: str,
    change_ct: int,
    change_pc: int,
    order_id: str,
    operator: str,
) -> tuple[int, int]:
    with db_conn() as conn:
        return _apply_inventory_change(
            conn,
            "PURCHASE_RECEIVE",
            item_code,
            barcode,
            name,
            unit,
            change_ct,
            change_pc,
            order_id=order_id,
            operator=operator,
        )


def render_warehouse() -> None:
    import main as supply_main

    render_sidebar_lang_switch()

    st.sidebar.markdown(f"**{t('role_warehouse')}**")
    st.sidebar.divider()
    _apply_url_page_to_session(Role.WAREHOUSE, WAREHOUSE_PAGE_KEYS)
    if (st.session_state.get("page") or "") not in WAREHOUSE_PAGE_KEYS:
        st.session_state.page = "pending"
    nav_items = supply_main.render_role_sidebar(
        Role.WAREHOUSE, messages_label=_nav_messages_label(),
    )
    st.sidebar.divider()
    with st.sidebar:
        logout_button("logout_warehouse")
    _render_active_page_hint(nav_items)

    page_key = st.session_state.page or "pending"
    if supply_main.dispatch_supply_chain_page(
        page_key,
        render_page_heading=render_page_heading,
        search_products=search_products,
        apply_inventory_receive=_sc_apply_inventory_receive,
        audit_write=audit_write,
    ):
        return
    pages = {
        "pending":  page_warehouse_pending,
        "short_in": page_warehouse_shortages,
        "history":  page_warehouse_dispatch_history,
        "supplier": page_warehouse_supplier_order,
        "inventory": page_warehouse_inventory,
        "messages": page_messages,
        "ai":       page_ai_assistant,
    }
    pages.get(page_key, page_warehouse_pending)()
    _sync_cookie_default_page(page_key)


def render_admin() -> None:
    import main as supply_main

    render_sidebar_lang_switch()

    st.sidebar.markdown(f"**{t('role_admin')}**")
    st.sidebar.divider()
    _apply_url_page_to_session(Role.ADMIN, ADMIN_PAGE_KEYS)
    if (st.session_state.get("page") or "") not in ADMIN_PAGE_KEYS:
        st.session_state.page = "dashboard"
    nav_items = supply_main.render_role_sidebar(
        Role.ADMIN, messages_label=_nav_messages_label(),
    )
    st.sidebar.divider()
    with st.sidebar:
        logout_button("logout_admin")
    _render_active_page_hint(nav_items)

    page_key = st.session_state.page or "dashboard"
    if supply_main.dispatch_supply_chain_page(
        page_key,
        render_page_heading=render_page_heading,
        search_products=search_products,
        apply_inventory_receive=_sc_apply_inventory_receive,
        audit_write=audit_write,
    ):
        return
    pages = {
        "dashboard":  page_admin_dashboard,
        "all_orders": page_admin_all_orders,
        "expiry_dash": page_admin_expiry_dashboard,
        "catalog":    page_admin_product_catalog,
        "product_master": page_admin_product_catalog,
        "shelf_mobile": page_admin_shelf_mobile,
        "accounts":   page_admin_accounts,
        "audit":      page_admin_audit_log,
        "inventory":  page_admin_inventory,
        "images":     page_admin_images,
        "email":      page_admin_email,
        "export":     page_admin_export,
        "backup":     page_admin_backup,
        "messages":   page_admin_messages_center,
        "ai":         page_ai_assistant,
    }
    pages.get(page_key, page_admin_dashboard)()
    _sync_cookie_default_page(page_key)


# =========================================================================
# ROUTER
# =========================================================================
def route() -> None:
    try:
        import auth as sunshine_auth

        sunshine_auth.prepare_auth_from_cookie()
    except Exception:
        pass

    nav = None
    try:
        nav = st.query_params.get("nav")
    except Exception:
        try:
            qp = st.experimental_get_query_params()
            vals = qp.get("nav") or []
            nav = vals[0] if vals else None
        except Exception:
            nav = None
    if nav == "messages" and is_authenticated():
        try:
            st.query_params.clear()
        except Exception:
            try:
                st.experimental_set_query_params()
            except Exception:
                pass
        _set_session_page_for_app_nav("messages")
    if st.session_state.lang is None:
        page_pick_language()
        return
    if not is_authenticated():
        page_login()
        return
    _route_sync_page_from_url()
    role = st.session_state.role
    if role == Role.BRANCH:
        render_branch()
    elif role == Role.WAREHOUSE:
        render_warehouse()
    elif role == Role.ADMIN:
        render_admin()
    else:
        st.error("Unknown role; please log out.")
        logout_button("logout_unknown")


# =========================================================================
# FIRST-LOAD SPLASH · 品牌化动态加载页
# =========================================================================
# 全屏覆盖层：科技蓝微渐变 + 毛玻璃 + 双环旋转 + 品牌文字 + 进度条流光。
# 用 position:fixed 覆盖整个视口，盖住 Streamlit 默认的灰色骨架屏。
# 仅在每个会话「首次进入 / 刷新页面」时出现一次（由 _app_warmed 守卫），
# 页面内切菜单、加减购物车等 rerun 不会再触发，保证操作流畅不被打断。
LOADING_SPLASH_HTML = """
<style>
@keyframes sunsp-spin   { to { transform: rotate(360deg); } }
@keyframes sunsp-pulse  { 0%,100% { transform: scale(.92); opacity:.6; }
                          50%      { transform: scale(1.08); opacity:1; } }
@keyframes sunsp-fade   { from { opacity: 0; } to { opacity: 1; } }
@keyframes sunsp-bar    { 0% { left: -45%; } 100% { left: 100%; } }
@keyframes sunsp-float  { 0%,100% { transform: translateY(0); }
                          50%      { transform: translateY(-6px); } }

#sunshine-splash {
    position: fixed;
    inset: 0;
    z-index: 2147483646;            /* 盖住一切（含侧栏/骨架屏） */
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    gap: 22px;
    text-align: center;
    color: #eaf2ff;
    font-family: "Segoe UI", system-ui, -apple-system, "PingFang SC",
                 "Microsoft YaHei", sans-serif;
    background:
        radial-gradient(1100px 560px at 50% -12%, rgba(64,140,255,.30), transparent 60%),
        radial-gradient(800px 460px at 88% 110%, rgba(255,196,84,.16), transparent 55%),
        linear-gradient(135deg, #081a3c 0%, #0c2a63 46%, #11409a 100%);
    -webkit-backdrop-filter: blur(16px) saturate(165%);
    backdrop-filter: blur(16px) saturate(165%);
    animation: sunsp-fade .35s ease both;
}
#sunshine-splash .sunsp-ring-wrap {
    position: relative;
    width: 124px;
    height: 124px;
    animation: sunsp-float 3s ease-in-out infinite;
}
#sunshine-splash .sunsp-ring {
    position: absolute;
    inset: 0;
    border-radius: 50%;
    border: 4px solid rgba(255,255,255,.10);
    border-top-color: #5aa6ff;
    border-right-color: #9cc8ff;
    box-shadow: 0 0 34px rgba(90,166,255,.45);
    animation: sunsp-spin 1.05s linear infinite;
}
#sunshine-splash .sunsp-ring.inner {
    inset: 17px;
    border-width: 3px;
    border-top-color: #ffd36b;
    border-right-color: transparent;
    border-bottom-color: rgba(255,211,107,.35);
    box-shadow: none;
    animation: sunsp-spin 1.6s linear infinite reverse;
}
#sunshine-splash .sunsp-emoji {
    position: absolute;
    inset: 0;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 2.5rem;
    animation: sunsp-pulse 1.7s ease-in-out infinite;
}
#sunshine-splash .sunsp-title {
    font-size: 1.4rem;
    font-weight: 800;
    letter-spacing: .4px;
    background: linear-gradient(90deg, #ffffff 0%, #bcd8ff 100%);
    -webkit-background-clip: text;
    background-clip: text;
    color: transparent;
}
#sunshine-splash .sunsp-sub {
    font-size: .94rem;
    font-weight: 500;
    color: #aecbf3;
    max-width: 80vw;
}
#sunshine-splash .sunsp-sub-en {
    font-size: .8rem;
    color: #7e9fd0;
    margin-top: -8px;
}
#sunshine-splash .sunsp-progress {
    position: relative;
    width: 260px;
    max-width: 70vw;
    height: 4px;
    border-radius: 99px;
    background: rgba(255,255,255,.12);
    overflow: hidden;
}
#sunshine-splash .sunsp-progress::after {
    content: "";
    position: absolute;
    top: 0;
    left: -45%;
    width: 45%;
    height: 100%;
    border-radius: 99px;
    background: linear-gradient(90deg, transparent, #5aa6ff, #9cc8ff, transparent);
    animation: sunsp-bar 1.15s ease-in-out infinite;
}
</style>
<div id="sunshine-splash">
    <div class="sunsp-ring-wrap">
        <div class="sunsp-ring"></div>
        <div class="sunsp-ring inner"></div>
        <div class="sunsp-emoji">🛒</div>
    </div>
    <div class="sunsp-title">☀️ SUNSHINE 订货系统正在初始化…</div>
    <div class="sunsp-sub">正在为您同步各分店实时库存，请稍候…</div>
    <div class="sunsp-sub-en">Syncing live inventory across stores · please wait</div>
    <div class="sunsp-progress"></div>
</div>
"""


def _warm_app_caches() -> None:
    """首屏一次性预热重型缓存（商品主档）。

    load_products 带 @st.cache_data，首次调用真正读库/读表（较慢），
    之后命中缓存几乎瞬时——这样炫酷加载页只会在首次进入时出现。"""
    try:
        load_products(_products_mtime(), _inventory_version(), _price_version())
    except Exception as e:
        log_exception("warm_app_caches", e)


def maybe_run_first_load_splash() -> None:
    """仅在每个会话首次进入 / 刷新页面时，显示一次品牌加载页并预热缓存。

    用 st.empty() 占位渲染全屏覆盖层，预热完成后调用 .empty() 移除，
    随后正常渲染页面内容，形成「加载中 → 加载完成」的平滑过渡。
    """
    if st.session_state.get("_app_warmed"):
        return
    splash = st.empty()
    splash.markdown(LOADING_SPLASH_HTML, unsafe_allow_html=True)
    started = time.monotonic()
    _warm_app_caches()
    # 保证动画至少完整呈现 ~0.6s，避免缓存极快命中时「一闪而过」。
    remaining = 0.6 - (time.monotonic() - started)
    if remaining > 0:
        time.sleep(remaining)
    st.session_state["_app_warmed"] = True
    splash.empty()


# =========================================================================
# ENTRY
# =========================================================================
def main() -> None:
    st.set_page_config(
        page_title="SUNSHINE · 阳光集团 订货系统",
        page_icon="☀️",
        layout="wide",
        initial_sidebar_state="auto",
    )
    inject_css()
    init_db()
    init_session()

    # 首屏品牌化加载页：仅本会话首次进入/刷新时出现一次，并预热商品主档缓存。
    # 之后的菜单切换、加减购物车等 rerun 都会命中缓存、跳过此动画，操作流畅。
    maybe_run_first_load_splash()

    # Take a one-time backup snapshot per Streamlit process so even if a
    # code change later corrupts the DB, we have a "fresh process start"
    # restore point. The throttled auto-snapshot logic kicks in beyond that.
    if not st.session_state.get("_startup_backup_done"):
        try:
            auto_snapshot_if_due()
        except Exception as e:
            log_exception("startup_auto_snapshot", e)
        st.session_state["_startup_backup_done"] = True

    # Auto-import any image files dropped into the root folder. Stash the
    # summary in session_state so the admin Images page can show what got
    # moved at startup. Done once per process (same pattern as backup).
    if not st.session_state.get("_startup_scan_done"):
        try:
            st.session_state["_startup_scan_summary"] = \
                scan_and_import_root_images()
        except Exception as e:
            st.session_state["_startup_scan_summary"] = {
                "moved": [], "skipped_unmatched": [],
                "errors": [f"scan failed: {e}"],
            }
        st.session_state["_startup_scan_done"] = True

    route()


if __name__ == "__main__":
    main()


