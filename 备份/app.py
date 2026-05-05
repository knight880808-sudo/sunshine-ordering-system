"""
SUNSHINE SHOPPING CENTER 订货系统 / Ordering System
====================================================
Single-file Streamlit app.

Run:
    streamlit run app.py --server.port 8502 --server.address 0.0.0.0
"""
from __future__ import annotations

import io
import hashlib
import json
import os
import smtplib
import sqlite3
import ssl
import threading
import traceback
from contextlib import contextmanager
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =========================================================================
# CONFIG / CONSTANTS
# =========================================================================
DB_PATH = Path("orders.db")
PRODUCTS_PATH = Path("products.xlsx")

# Product image folder. File naming convention (in priority order):
#   images/<ItemCode>.<ext>   e.g. images/P001.jpg
#   images/<Barcode>.<ext>    e.g. images/8992730950194.png
# Supported extensions: jpg, jpeg, png, webp.
# Missing images are silently ignored — the UI shows a placeholder.
# Recommended: keep each image under 200KB and ≤ 800px on the longest edge
# so a 10-item search page stays snappy on slow phone connections.
IMAGES_DIR = Path("images")

# Local DB backup directory. Used for both auto-snapshots (taken on startup
# and throttled by BACKUP_MIN_INTERVAL_MINUTES) and manual admin backups.
BACKUP_DIR = Path("backups")
BACKUP_RETAIN = 30                    # keep at most this many auto-snapshots
BACKUP_MIN_INTERVAL_MINUTES = 60      # auto-snapshot at most once per hour

# Email notification config & log live in their own JSON files (NOT in the
# SQLite DB) so backups of orders.db don't accidentally include SMTP creds.
EMAIL_CONFIG_PATH = Path("email_config.json")
EMAIL_LOG_PATH = Path("email_log.json")
EMAIL_LOG_KEEP = 200                  # keep last N log entries
APP_LOG_PATH = Path("app_runtime.log")

# Product images live as plain files in this folder. Filename convention:
#   <ItemCode>.<ext>          e.g. P001.jpg
#   bc_<Barcode>.<ext>        e.g. bc_8801234567001.jpg  (when no ItemCode)
# This keeps images out of the SQLite DB (so backups stay small) and out of
# any added schema (so we don't need to migrate). Supported extensions: jpg
# jpeg png gif webp — whatever the phone camera or browser produces.
PRODUCT_IMAGES_DIR = Path("product_images")
IMAGE_EXTS = (".jpg", ".jpeg", ".png", ".gif", ".webp")

BRANCHES: list[str] = [
    "NAMBER ONE STORE",
    "SUNSHINE MARKET",
    "SUNSHINE PS",
    "SUNSHINE FU-SANMA",
    "SUNSHINE FU-NEMO",
    "CHEN STORE-SARAKATA",
    "CHEN STORE-CHAPI",
]

WAREHOUSE_PASSWORD = os.getenv("SUNSHINE_WAREHOUSE_PASSWORD", "sunshine888")
ADMIN_PASSWORD = os.getenv("SUNSHINE_ADMIN_PASSWORD", "sunshine")


class Role:
    BRANCH = "branch"
    WAREHOUSE = "warehouse"
    ADMIN = "admin"


class OrderStatus:
    PENDING = "Pending"
    DISPATCHED = "Dispatched"
    RECEIVED = "Received"
    ALL = (PENDING, DISPATCHED, RECEIVED)


class ShortageStatus:
    OPEN = "Open"
    RESENDING = "Resending"
    OUT_OF_STOCK = "Out of Stock"
    RESOLVED = "Resolved"
    ALL = (OPEN, RESENDING, OUT_OF_STOCK, RESOLVED)


# =========================================================================
# I18N
# =========================================================================
T: dict[str, dict[str, str]] = {
    # Branding
    "app_brand":      {"en": "SUNSHINE SHOPPING CENTER", "zh": "阳光集团"},
    "app_subtitle":   {"en": "Ordering System",          "zh": "订货系统"},
    # Auth
    "select_role":    {"en": "Select your role",         "zh": "请选择您的角色"},
    "role_branch":    {"en": "🛒 Branch Staff",           "zh": "🛒 分店员工"},
    "role_warehouse": {"en": "📦 Warehouse Staff",        "zh": "📦 仓库员工"},
    "role_admin":     {"en": "🏭 Administrator",          "zh": "🏭 管理员"},
    "select_branch":  {"en": "Select Branch",            "zh": "选择分店"},
    "password":       {"en": "Password",                 "zh": "密码"},
    "login":          {"en": "Login",                    "zh": "登录"},
    "logout":         {"en": "Logout",                   "zh": "退出登录"},
    "wrong_pw":       {"en": "Wrong password",           "zh": "密码错误"},
    # Nav
    "nav_order":      {"en": "🛒 Place Order",           "zh": "🛒 下单"},
    "nav_my_orders":  {"en": "📋 My Orders",             "zh": "📋 我的订单"},
    "nav_my_short":   {"en": "🔔 My Shortages",          "zh": "🔔 我的缺货"},
    "nav_pending":    {"en": "📦 Pending Dispatch",      "zh": "📦 待发货订单"},
    "nav_short_in":   {"en": "🔔 Shortage Notifications","zh": "🔔 缺货通知"},
    "nav_dashboard":  {"en": "📊 Dashboard",              "zh": "📊 管理概览"},
    "nav_all_orders": {"en": "📋 All Orders",             "zh": "📋 所有订单"},
    "nav_dispatch":   {"en": "📦 Dispatch",               "zh": "📦 出库发货"},
    "nav_short_mgmt": {"en": "🔔 Shortages",              "zh": "🔔 缺货通知"},
    "nav_export":     {"en": "📥 Export Reports",         "zh": "📥 导出报表"},
    # Statuses
    "st_pending":     {"en": "Pending",                  "zh": "待发货"},
    "st_dispatched":  {"en": "Dispatched",               "zh": "已发货"},
    "st_received":    {"en": "Received",                 "zh": "已收货"},
    "st_open":        {"en": "Open",                     "zh": "待处理"},
    "st_resending":   {"en": "Resending",                "zh": "补发中"},
    "st_oos":         {"en": "Out of Stock",             "zh": "缺货"},
    "st_resolved":    {"en": "Resolved",                 "zh": "已解决"},
    # Misc
    "current_branch": {"en": "Current branch",           "zh": "当前分店"},
    "search_product": {"en": "Search by name / barcode / code (scanner OK)",
                       "zh": "按名称/条码/编号搜索（支持扫码枪）"},
    "no_results":     {"en": "No products found",        "zh": "未找到商品"},
    "manual_add":     {"en": "+ Add product not in list","zh": "+ 添加列表外商品"},
    "manual":         {"en": "Manual",                   "zh": "手动"},
    "name":           {"en": "Name",                     "zh": "名称"},
    "unit":           {"en": "Unit",                     "zh": "单位"},
    "item_code":      {"en": "Item Code",                "zh": "商品编号"},
    "barcode":        {"en": "Barcode",                  "zh": "条码"},
    "cartons":        {"en": "Cartons",                  "zh": "箱数"},
    "each_pcs":       {"en": "Each pcs",                 "zh": "个数"},
    "qty_cartons":    {"en": "Cartons",                  "zh": "箱"},
    "qty_pcs":        {"en": "Pcs",                      "zh": "个"},
    "add_to_cart":    {"en": "Add to Cart",              "zh": "加入购物车"},
    "cart":           {"en": "Cart",                     "zh": "购物车"},
    "empty_cart":     {"en": "Cart is empty",            "zh": "购物车为空"},
    "submit_order":   {"en": "Submit Order",             "zh": "提交订单"},
    "order_submitted":{"en": "Order submitted",          "zh": "订单已提交"},
    "submit_busy":    {"en": "Submission in progress, please wait",
                       "zh": "正在提交，请稍候"},
    "dup_submit_block":{"en": "Duplicate submission blocked. Please refresh and check.",
                        "zh": "已拦截重复提交，请刷新后确认结果"},
    "order_id":       {"en": "Order ID",                 "zh": "订单号"},
    "order_date":     {"en": "Order Date",               "zh": "下单日期"},
    "remarks":        {"en": "Remarks",                  "zh": "备注"},
    "search":         {"en": "Search",                   "zh": "搜索"},
    "status":         {"en": "Status",                   "zh": "状态"},
    "branch":         {"en": "Branch",                   "zh": "分店"},
    "no_data":        {"en": "No data",                  "zh": "暂无数据"},
    "no_pending":     {"en": "No pending orders",        "zh": "暂无待发货订单"},
    "no_shortages":   {"en": "No shortages",             "zh": "暂无缺货"},
    "receive":        {"en": "Receive Goods",            "zh": "收货确认"},
    "actual_cartons": {"en": "Actual Cartons",           "zh": "实收箱数"},
    "actual_pcs":     {"en": "Actual Pcs",               "zh": "实收个数"},
    "ordered":        {"en": "Ordered",                  "zh": "下单"},
    "dispatched":     {"en": "Dispatched",               "zh": "发货"},
    "confirm_receive":{"en": "Confirm Receipt",          "zh": "确认收货"},
    "receipt_done":   {"en": "Receipt confirmed",        "zh": "收货已确认"},
    "shortage_alert": {"en": "Shortage detected, warehouse notified",
                       "zh": "检测到缺货，已通知仓库"},
    "mark_dispatched":{"en": "Mark as Dispatched",       "zh": "标记已发货"},
    "shipment_marked":{"en": "Marked as dispatched",     "zh": "已标记发货"},
    "dispatch_busy":  {"en": "Dispatch is processing, please wait",
                       "zh": "正在处理发货，请稍候"},
    "shortage_qty":   {"en": "Shortage Qty",             "zh": "缺货数量"},
    "warehouse_reply":{"en": "Warehouse Reply",          "zh": "仓库回复"},
    "resend":         {"en": "Resend",                   "zh": "补发"},
    "mark_oos":       {"en": "Mark Out of Stock",        "zh": "标记缺货"},
    "confirm_resend": {"en": "Confirm Resend Received",  "zh": "确认补发已收到"},
    "card_today":     {"en": "Orders Today",             "zh": "今日订单"},
    "card_pending":   {"en": "Pending Dispatch",         "zh": "待发货"},
    "card_dispatched":{"en": "Dispatched",               "zh": "已发货"},
    "card_short":     {"en": "Open Shortages",           "zh": "待处理缺货"},
    "branch_status":  {"en": "Branch Status",            "zh": "各分店状态"},
    "latest_orders":  {"en": "Latest Orders",            "zh": "最新订单"},
    "latest_short":   {"en": "Latest Shortages",         "zh": "最新缺货"},
    "from_date":      {"en": "From",                     "zh": "起始日期"},
    "to_date":        {"en": "To",                       "zh": "结束日期"},
    "exp_picking":    {"en": "Picking List (per branch)","zh": "捡货单（各店分列）"},
    "exp_recon":      {"en": "Reconciliation (Dispatched vs Received)",
                       "zh": "对账单（发货 vs 实收）"},
    "exp_short":      {"en": "Shortage Report",          "zh": "缺货报告"},
    "generate":       {"en": "Generate",                 "zh": "生成"},
    "download":       {"en": "Download",                 "zh": "下载"},
    "tbd":            {"en": "(to be implemented)",      "zh": "（待实现）"},
    # Pagination
    "page":           {"en": "Page",                     "zh": "页"},
    "of":             {"en": "of",                       "zh": "/"},
    "prev_page":      {"en": "◀ Previous",               "zh": "◀ 上一页"},
    "next_page":      {"en": "Next ▶",                   "zh": "下一页 ▶"},
    "showing":        {"en": "Showing",                  "zh": "显示"},
    "items":          {"en": "items",                    "zh": "条"},
    "results_count":  {"en": "results",                  "zh": "结果"},
    "per_page":       {"en": "per page",                 "zh": "每页"},
    # Dispatch history
    "nav_dispatch_history":  {"en": "📜 Dispatch History",
                              "zh": "📜 出库历史"},
    "dispatch_history":      {"en": "Dispatch History",  "zh": "出库历史"},
    "select_date":           {"en": "Select Date",       "zh": "选择日期"},
    "today_btn":             {"en": "Today",             "zh": "今天"},
    "all_dates":             {"en": "All Dates",         "zh": "所有日期"},
    "dispatch_date":         {"en": "Dispatch Date",     "zh": "出库日期"},
    "dispatch_time":         {"en": "Dispatch Time",     "zh": "出库时间"},
    "no_dispatch_history":   {"en": "No dispatch history",
                              "zh": "暂无出库记录"},
    "dispatched_orders":     {"en": "Dispatched Orders", "zh": "已发货订单"},
    "received_status":       {"en": "Receipt status",    "zh": "收货状态"},
    "fully_received":        {"en": "Fully received",    "zh": "已全部收货"},
    "partial_received":      {"en": "Partially received","zh": "部分收货"},
    "awaiting_receipt":      {"en": "Awaiting receipt",  "zh": "待收货"},
    # Picking-list downloads in warehouse view
    "dl_picking":     {"en": "🖨️ Picking List",          "zh": "🖨️ 捡货单"},
    "dl_this_order":  {"en": "🖨️ Download Picking Slip", "zh": "🖨️ 下载捡货单"},
    "dl_branch_all":  {"en": "🖨️ Download All for Branch", "zh": "🖨️ 下载本店全部"},
    "dl_all_pending": {"en": "🖨️ Download All Pending",  "zh": "🖨️ 下载全部待发货"},
    "picking_slip":   {"en": "Picking Slip",             "zh": "捡货单"},
    "lines":          {"en": "Lines",                    "zh": "行数"},
    "total_items":    {"en": "Total Lines",              "zh": "合计行数"},
    "orders":         {"en": "orders",                   "zh": "订单"},
    # Batch add / confirmation flow
    "add_selected":      {"en": "📥 Add Selected to Cart",
                          "zh": "📥 追加购物车（已选商品）"},
    "added_n_items":     {"en": "Added {n} items to cart",
                          "zh": "已追加 {n} 项到购物车"},
    "no_qty_selected":   {"en": "No items have a quantity. Enter cartons or pcs first.",
                          "zh": "没有商品填写数量。请先输入箱数或个数。"},
    "review_cart":       {"en": "📝 Review & Send Order",
                          "zh": "📝 确认并发送订单"},
    "review_title":      {"en": "Review Your Order",
                          "zh": "确认订单"},
    "review_subtitle":   {"en": "Check each line, edit quantities or remove items, then send.",
                          "zh": "请核对每行数量，可修改或删除，然后发送。"},
    "send_order":        {"en": "📤 Send Order",
                          "zh": "📤 发送订单"},
    "back_to_browse":    {"en": "◀ Back to browsing",
                          "zh": "◀ 返回继续选购"},
    "items_count":       {"en": "Items in this order",
                          "zh": "本次订单商品数"},
    "clear_qty":         {"en": "Clear all entered quantities",
                          "zh": "清空已输入数量"},
    "qty_cleared":       {"en": "Quantities cleared",
                          "zh": "数量已清空"},
    "saved_for_batch":   {"en": "Saved",
                          "zh": "已保存"},
    # Warehouse dispatch quantity input
    "dispatch_qty":      {"en": "Dispatch Qty",
                          "zh": "发货数量"},
    "dispatch_cartons":  {"en": "Dispatch Cartons",
                          "zh": "发货箱数"},
    "dispatch_pcs":      {"en": "Dispatch Pcs",
                          "zh": "发货个数"},
    "dispatch_hint":     {"en": "Enter actual quantities being shipped. "
                                "Defaults to ordered amount.",
                          "zh": "请输入实际发货数量，默认等于订单数量。"},
    "less_than_ordered": {"en": "Short of order",
                          "zh": "少于订单"},
    "ordered_qty":       {"en": "Ordered",
                          "zh": "订货"},
    # Database backup / restore
    "nav_backup":        {"en": "💾 Database Backup",
                          "zh": "💾 数据库备份"},
    "backup_title":      {"en": "Database Backup & Restore",
                          "zh": "数据库备份与恢复"},
    "backup_subtitle":   {"en": "Snapshots are stored locally in the "
                                "backups/ folder. Auto-snapshot runs on "
                                "startup and at most once per hour.",
                          "zh": "备份文件保存在 backups/ 目录。"
                                "启动时和每小时最多一次会自动备份。"},
    "backup_now":        {"en": "💾 Backup Now",
                          "zh": "💾 立即备份"},
    "backup_done":       {"en": "Backup created",
                          "zh": "备份已创建"},
    "backup_failed":     {"en": "Backup failed",
                          "zh": "备份失败"},
    "available_backups": {"en": "Available Backups",
                          "zh": "可用备份"},
    "no_backups":        {"en": "No backups yet",
                          "zh": "尚无备份"},
    "backup_file":       {"en": "File",
                          "zh": "文件"},
    "backup_size":       {"en": "Size (KB)",
                          "zh": "大小 (KB)"},
    "backup_created":    {"en": "Created",
                          "zh": "创建时间"},
    "backup_counts":     {"en": "Records (orders / shipments / shortages)",
                          "zh": "记录数 (订单 / 发货 / 缺货)"},
    "backup_actions":    {"en": "Actions",
                          "zh": "操作"},
    "download_backup":   {"en": "⬇️ Download",
                          "zh": "⬇️ 下载"},
    "restore_backup":    {"en": "♻️ Restore",
                          "zh": "♻️ 恢复"},
    "delete_backup":     {"en": "🗑️ Delete",
                          "zh": "🗑️ 删除"},
    "confirm_restore_q": {"en": "⚠️ Restore from this backup? "
                                "Current data will be replaced. "
                                "A safety snapshot of current data will "
                                "be created first.",
                          "zh": "⚠️ 确认从此备份恢复？"
                                "当前数据将被覆盖。"
                                "恢复前会自动备份当前数据库。"},
    "yes_restore":       {"en": "✅ Yes, restore now",
                          "zh": "✅ 是，立即恢复"},
    "cancel":            {"en": "Cancel",
                          "zh": "取消"},
    "restore_done":      {"en": "Restore complete. Safety snapshot saved as: ",
                          "zh": "恢复完成。安全备份保存为: "},
    "export_sql":        {"en": "📤 Export as SQL Dump",
                          "zh": "📤 导出 SQL 文件"},
    "import_sql":        {"en": "📥 Restore from SQL Dump",
                          "zh": "📥 从 SQL 文件恢复"},
    "upload_sql":        {"en": "Upload .sql file",
                          "zh": "上传 .sql 文件"},
    "import_done":       {"en": "Import complete. Safety snapshot: ",
                          "zh": "导入完成。安全备份: "},
    "current_db_info":   {"en": "Current Database",
                          "zh": "当前数据库"},
    # Product images
    "nav_images":        {"en": "🖼️ Product Images",
                          "zh": "🖼️ 商品图片"},
    "img_title":         {"en": "Product Image Management",
                          "zh": "商品图片管理"},
    "img_subtitle":      {"en": "Upload a photo for each product. "
                                "On phones, the upload button opens the camera.",
                          "zh": "为每个商品上传图片。"
                                "在手机上点击上传按钮会自动调用相机。"},
    "img_search":        {"en": "Find product to upload image",
                          "zh": "搜索要上传图片的商品"},
    "img_upload":        {"en": "📷 Take photo / Choose image",
                          "zh": "📷 拍照 / 选择图片"},
    "img_uploaded":      {"en": "Image saved",
                          "zh": "图片已保存"},
    "img_delete":        {"en": "🗑️ Remove image",
                          "zh": "🗑️ 删除图片"},
    "img_deleted":       {"en": "Image removed",
                          "zh": "图片已删除"},
    "img_has_image":     {"en": "Has image",
                          "zh": "已有图片"},
    "img_no_image":      {"en": "No image yet",
                          "zh": "暂无图片"},
    "img_show_in_search":{"en": "Show product images in search results",
                          "zh": "在搜索结果中显示商品图片"},
    # Batch image upload
    "img_batch_title":   {"en": "🚀 Batch upload by filename",
                          "zh": "🚀 批量上传（按文件名配对）"},
    "img_batch_hint":    {"en": "Name files <ItemCode>.jpg or <Barcode>.jpg "
                                "to auto-match. Multiple files allowed.",
                          "zh": "文件命名为 <商品编号>.jpg 或 <条码>.jpg 即可自动配对。"
                                "支持多文件同时上传。"},
    "img_batch_choose":  {"en": "Choose multiple image files",
                          "zh": "选择多个图片文件"},
    "img_batch_preview": {"en": "Pairing preview",
                          "zh": "配对预览"},
    "img_matched":       {"en": "Matched",
                          "zh": "已匹配"},
    "img_unmatched":     {"en": "Unmatched",
                          "zh": "未匹配"},
    "img_filename":      {"en": "File name",
                          "zh": "文件名"},
    "img_matched_to":    {"en": "Matched to",
                          "zh": "匹配到"},
    "img_match_reason":  {"en": "Reason",
                          "zh": "原因"},
    "img_batch_confirm": {"en": "✅ Confirm and upload {n} images",
                          "zh": "✅ 确认上传 {n} 张图片"},
    "img_batch_done":    {"en": "{n} images uploaded successfully",
                          "zh": "成功上传 {n} 张图片"},
    "img_batch_failed":  {"en": "{n} files failed to save",
                          "zh": "{n} 个文件保存失败"},
    "img_batch_cancel":  {"en": "Cancel",
                          "zh": "取消"},
    # Auto-scan root directory for images
    "img_scan_title":    {"en": "📂 Auto-import from app folder",
                          "zh": "📂 自动导入应用根目录"},
    "img_scan_hint":     {"en": "Drop image files (named <ItemCode>.jpg or "
                                "<Barcode>.jpg) into the same folder as "
                                "app.py. They will be moved into "
                                "product_images/ on next start, or click "
                                "below to do it now.",
                          "zh": "把图片文件（命名为 <商品编号>.jpg 或 "
                                "<条码>.jpg）拖到 app.py 所在的根目录。"
                                "下次启动时会自动归档到 product_images/，"
                                "或点击下方按钮立即归档。"},
    "img_scan_now":      {"en": "🔍 Scan & import now",
                          "zh": "🔍 立即扫描并归档"},
    "img_scan_moved":    {"en": "Files imported on startup",
                          "zh": "启动时归档"},
    "img_scan_left":     {"en": "Files left in root (no match)",
                          "zh": "根目录留下（无法匹配）"},
    "img_scan_none":     {"en": "Nothing to import — root folder is clean.",
                          "zh": "无需导入——根目录干净。"},
    "img_scan_errors":   {"en": "Errors",
                          "zh": "错误"},
    # Email notifications
    "nav_email":         {"en": "📧 Email Notifications",
                          "zh": "📧 邮件通知"},
    "email_title":       {"en": "Email Notification Settings",
                          "zh": "邮件通知设置"},
    "email_subtitle":    {"en": "Send automatic email alerts for new orders, "
                                "dispatches, and shortages. Sending is "
                                "asynchronous — UI is never blocked.",
                          "zh": "为新订单、发货、缺货事件自动发送邮件通知。"
                                "邮件发送是异步的，不会阻塞页面操作。"},
    "email_enabled":     {"en": "Enable email notifications",
                          "zh": "启用邮件通知"},
    "smtp_settings":     {"en": "SMTP Settings",
                          "zh": "SMTP 设置"},
    "smtp_host":         {"en": "SMTP Host",
                          "zh": "SMTP 服务器"},
    "smtp_port":         {"en": "SMTP Port",
                          "zh": "SMTP 端口"},
    "smtp_user":         {"en": "SMTP Username",
                          "zh": "SMTP 用户名"},
    "smtp_password":     {"en": "SMTP Password",
                          "zh": "SMTP 密码"},
    "smtp_from":         {"en": "From Address",
                          "zh": "发件人地址"},
    "smtp_use_tls":      {"en": "Use STARTTLS",
                          "zh": "使用 STARTTLS"},
    "event_settings":    {"en": "Per-event Settings",
                          "zh": "事件设置"},
    "ev_new_order":      {"en": "📦 New Order",
                          "zh": "📦 新订单"},
    "ev_dispatched":     {"en": "🚚 Dispatched",
                          "zh": "🚚 已发货"},
    "ev_shortage":       {"en": "🔔 Shortage",
                          "zh": "🔔 缺货"},
    "recipients":        {"en": "Recipients (one per line)",
                          "zh": "收件人（每行一个邮箱）"},
    "branch_recipients": {"en": "Per-branch Email (optional)",
                          "zh": "分店专属邮箱（可选）"},
    "branch_email_hint": {"en": "If set, dispatched-event emails for that "
                                "branch will additionally CC this address.",
                          "zh": "若填写，发货邮件会额外抄送给该分店的此邮箱。"},
    "save_email_cfg":    {"en": "💾 Save Settings",
                          "zh": "💾 保存设置"},
    "email_cfg_saved":   {"en": "Settings saved",
                          "zh": "设置已保存"},
    "test_email":        {"en": "📧 Send Test Email",
                          "zh": "📧 发送测试邮件"},
    "test_to":           {"en": "Send test to",
                          "zh": "测试邮件发送至"},
    "test_sent_ok":      {"en": "✅ Test email sent successfully",
                          "zh": "✅ 测试邮件发送成功"},
    "test_sent_fail":    {"en": "❌ Test email failed",
                          "zh": "❌ 测试邮件发送失败"},
    "email_log_title":   {"en": "Recent Email Activity",
                          "zh": "近期邮件记录"},
    "email_log_empty":   {"en": "No email activity yet",
                          "zh": "暂无邮件记录"},
    "log_time":          {"en": "Time",
                          "zh": "时间"},
    "log_event":         {"en": "Event",
                          "zh": "事件"},
    "log_to":            {"en": "Recipients",
                          "zh": "收件人"},
    "log_status":        {"en": "Status",
                          "zh": "状态"},
    "log_ok":            {"en": "OK",
                          "zh": "成功"},
    "log_fail":          {"en": "Failed",
                          "zh": "失败"},
    "log_clear":         {"en": "🗑️ Clear log",
                          "zh": "🗑️ 清空记录"},
}

STATUS_LABEL: dict[str, str] = {
    OrderStatus.PENDING:    "st_pending",
    OrderStatus.DISPATCHED: "st_dispatched",
    OrderStatus.RECEIVED:   "st_received",
    ShortageStatus.OPEN:         "st_open",
    ShortageStatus.RESENDING:    "st_resending",
    ShortageStatus.OUT_OF_STOCK: "st_oos",
    ShortageStatus.RESOLVED:     "st_resolved",
}


def t(key: str) -> str:
    lang = st.session_state.get("lang", "en")
    entry = T.get(key)
    if not entry:
        return key
    return entry.get(lang) or entry.get("en") or key


# =========================================================================
# DATABASE
# =========================================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS orders (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        TEXT    NOT NULL,
    branch          TEXT    NOT NULL,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    price           REAL    DEFAULT 0,
    qty_cartons     INTEGER NOT NULL DEFAULT 0,
    qty_pcs         INTEGER NOT NULL DEFAULT 0,
    is_manual       INTEGER NOT NULL DEFAULT 0,
    status          TEXT    NOT NULL DEFAULT 'Pending',
    order_date      TEXT    NOT NULL,
    dispatch_date   TEXT,
    dispatch_cartons INTEGER,                           -- what warehouse actually shipped
    dispatch_pcs    INTEGER,                            -- (may differ from qty_* if low stock)
    receive_date    TEXT,
    actual_cartons  INTEGER,
    actual_pcs      INTEGER,
    receive_remarks TEXT
);
CREATE INDEX IF NOT EXISTS idx_orders_order_id ON orders(order_id);
CREATE INDEX IF NOT EXISTS idx_orders_branch   ON orders(branch);
CREATE INDEX IF NOT EXISTS idx_orders_status   ON orders(status);
CREATE INDEX IF NOT EXISTS idx_orders_date     ON orders(order_date);

CREATE TABLE IF NOT EXISTS shipments (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        TEXT    NOT NULL,
    branch          TEXT    NOT NULL,
    dispatch_date   TEXT    NOT NULL,
    dispatched_by   TEXT
);
CREATE INDEX IF NOT EXISTS idx_shipments_order_id ON shipments(order_id);

CREATE TABLE IF NOT EXISTS shortages (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id        TEXT    NOT NULL,
    order_line_id   INTEGER NOT NULL,
    branch          TEXT    NOT NULL,
    item_code       TEXT,
    barcode         TEXT,
    name            TEXT    NOT NULL,
    unit            TEXT,
    short_cartons   INTEGER NOT NULL DEFAULT 0,
    short_pcs       INTEGER NOT NULL DEFAULT 0,
    status          TEXT    NOT NULL DEFAULT 'Open',
    warehouse_reply TEXT,
    branch_remarks  TEXT,
    reported_date   TEXT    NOT NULL,
    resolved_date   TEXT
);
CREATE INDEX IF NOT EXISTS idx_short_status ON shortages(status);
CREATE INDEX IF NOT EXISTS idx_short_branch ON shortages(branch);
CREATE INDEX IF NOT EXISTS idx_short_order  ON shortages(order_id);
"""


@contextmanager
def db_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_db() -> None:
    with db_conn() as conn:
        conn.executescript(SCHEMA_SQL)
        # Migration: older orders.db files won't have dispatch_cartons/pcs.
        # SQLite's CREATE TABLE IF NOT EXISTS leaves an existing table alone,
        # so add the columns explicitly when missing. This is idempotent.
        existing_cols = {
            r[1] for r in conn.execute("PRAGMA table_info(orders)").fetchall()
        }
        if "dispatch_cartons" not in existing_cols:
            conn.execute("ALTER TABLE orders ADD COLUMN dispatch_cartons INTEGER")
        if "dispatch_pcs" not in existing_cols:
            conn.execute("ALTER TABLE orders ADD COLUMN dispatch_pcs INTEGER")


def count_open_shortages() -> int:
    with db_conn() as conn:
        return conn.execute(
            "SELECT COUNT(*) FROM shortages WHERE status = ?",
            (ShortageStatus.OPEN,),
        ).fetchone()[0]


# =========================================================================
# BACKUP / RESTORE
# =========================================================================
# Strategy:
#   - On startup: take one snapshot (so even a "code change broke the app"
#     scenario leaves us with a fresh restore point).
#   - On meaningful writes: auto-snapshot, but throttled so we don't make
#     hundreds of files during a busy ordering session.
#   - Manual backup button in the admin panel for explicit checkpoints.
#   - Restore is admin-only and requires explicit confirmation. Before
#     overwriting the live DB, the *current* DB is itself backed up under
#     a "pre_restore" name so the restore is reversible.
#
# We use SQLite's native backup() API rather than file copy, so we get a
# transactionally consistent snapshot even if writes are happening
# concurrently. backup() pages the source into the destination atomically.
def _ensure_backup_dir() -> None:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def backup_database(label: str = "auto") -> Path | None:
    """Snapshot the live DB to backups/orders_<label>_<timestamp>.db.

    Uses sqlite3's native backup API for atomic consistency. Returns the
    backup file path, or None if the source DB doesn't exist yet."""
    if not DB_PATH.exists():
        return None
    _ensure_backup_dir()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = "".join(c for c in label if c.isalnum() or c in "-_") or "auto"
    dest = BACKUP_DIR / f"orders_{safe_label}_{ts}.db"
    # If a backup with the exact same timestamp already exists (two backups
    # in the same second), append a counter so we never silently overwrite.
    if dest.exists():
        n = 1
        while True:
            cand = BACKUP_DIR / f"orders_{safe_label}_{ts}_{n}.db"
            if not cand.exists():
                dest = cand
                break
            n += 1

    src = sqlite3.connect(DB_PATH)
    try:
        dst = sqlite3.connect(dest)
        try:
            src.backup(dst)  # atomic page-level copy with consistent view
        finally:
            dst.close()
    finally:
        src.close()

    return dest


def _prune_old_auto_backups() -> int:
    """Keep only the most recent BACKUP_RETAIN auto-* snapshots.
    Manual / pre_restore backups are never pruned. Returns count deleted."""
    if not BACKUP_DIR.exists():
        return 0
    auto = sorted(
        BACKUP_DIR.glob("orders_auto_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    to_delete = auto[BACKUP_RETAIN:]
    for p in to_delete:
        try:
            p.unlink()
        except OSError:
            pass
    return len(to_delete)


def auto_snapshot_if_due() -> Path | None:
    """Throttled auto-snapshot. Skips if the most recent auto-snapshot was
    less than BACKUP_MIN_INTERVAL_MINUTES ago."""
    if not DB_PATH.exists():
        return None
    _ensure_backup_dir()
    auto = sorted(
        BACKUP_DIR.glob("orders_auto_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if auto:
        age_min = (datetime.now().timestamp() - auto[0].stat().st_mtime) / 60.0
        if age_min < BACKUP_MIN_INTERVAL_MINUTES:
            return None
    path = backup_database("auto")
    _prune_old_auto_backups()
    return path


def list_backups() -> list[dict]:
    """Return all backups with a quick row-count peek per table.
    Sorted newest first. Each entry: {path, name, size_kb, created, counts}."""
    if not BACKUP_DIR.exists():
        return []
    out = []
    files = sorted(
        BACKUP_DIR.glob("orders_*.db"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    for p in files:
        stat = p.stat()
        # Peek table counts. Use a try/except so a corrupt backup just
        # shows as "?" rather than crashing the listing.
        counts = {"orders": "?", "shipments": "?", "shortages": "?"}
        try:
            c = sqlite3.connect(p)
            try:
                for table in counts:
                    try:
                        n = c.execute(
                            f"SELECT COUNT(*) FROM {table}"
                        ).fetchone()[0]
                        counts[table] = int(n)
                    except sqlite3.DatabaseError:
                        counts[table] = "?"
            finally:
                c.close()
        except sqlite3.DatabaseError:
            pass
        out.append({
            "path": p,
            "name": p.name,
            "size_kb": round(stat.st_size / 1024, 1),
            "created": datetime.fromtimestamp(stat.st_mtime).strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            "counts": counts,
        })
    return out


def restore_from_backup(backup_path: Path) -> Path:
    """Restore the live DB from a backup file. Before overwriting, the
    current DB is itself snapshotted as `pre_restore_*` so the operation
    is reversible. Returns path to the safety snapshot."""
    if not backup_path.exists():
        raise FileNotFoundError(f"Backup not found: {backup_path}")

    # Sanity check: backup must look like one of our DB files
    try:
        c = sqlite3.connect(backup_path)
        try:
            tables = {
                r[0] for r in c.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
            }
        finally:
            c.close()
    except sqlite3.DatabaseError as e:
        raise ValueError(f"Backup file is not a valid SQLite DB: {e}")

    if not {"orders", "shipments", "shortages"}.issubset(tables):
        raise ValueError(
            f"Backup is missing required tables. Found: {tables}"
        )

    # Safety snapshot of whatever is currently live
    safety = backup_database("pre_restore")

    # Overwrite the live DB with the backup. Using sqlite's backup API into
    # the live file gives us an atomic, transactionally clean restore even
    # if the live DB is in use by another connection.
    src = sqlite3.connect(backup_path)
    try:
        dst = sqlite3.connect(DB_PATH)
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()

    return safety


def export_db_dump_sql() -> bytes:
    """Export the live DB as a portable SQL text dump. Useful for migrating
    across SQLite versions or inspecting the data in a text editor."""
    if not DB_PATH.exists():
        return b""
    conn = sqlite3.connect(DB_PATH)
    try:
        lines = list(conn.iterdump())
    finally:
        conn.close()
    return "\n".join(lines).encode("utf-8")


def restore_from_sql_dump(sql_bytes: bytes) -> Path:
    """Replace the live DB with one built from a SQL dump.
    Pre-restore safety backup is taken first."""
    text = sql_bytes.decode("utf-8") if isinstance(sql_bytes, (bytes, bytearray)) else sql_bytes
    safety = backup_database("pre_restore")
    # Replace the live DB. Build into a temp file then move into place so
    # we never leave a half-written orders.db.
    tmp_path = DB_PATH.with_suffix(".db.restoring")
    if tmp_path.exists():
        tmp_path.unlink()
    conn = sqlite3.connect(tmp_path)
    try:
        conn.executescript(text)
        conn.commit()
    finally:
        conn.close()
    # Atomic replace
    if DB_PATH.exists():
        DB_PATH.unlink()
    tmp_path.rename(DB_PATH)
    return safety


# =========================================================================
# EMAIL NOTIFICATIONS
# =========================================================================
# Three event types:
#   - new_order   (branch submitted an order  → notify warehouse + admin)
#   - dispatched  (warehouse marked dispatch  → notify branch + admin)
#   - shortage    (branch reported short qty  → notify warehouse + admin)
#
# Design constraints:
#   1. Sending must be ASYNCHRONOUS — never block a UI button click on SMTP.
#      Failures degrade silently (logged, never raised to the user).
#   2. Config is in a JSON file, NOT the SQLite DB, so backups of orders.db
#      don't carry SMTP creds.
#   3. Per-event "enabled" toggles + per-event recipient lists.
#   4. Every send attempt — success or failure — gets a row in email_log.json.
#      Admin page reads & displays this so the operator can see what's flowing.
#   5. Test-send button so the operator can verify SMTP without triggering a
#      real order.

EVENT_KEYS = ("new_order", "dispatched", "shortage")


def _default_email_config() -> dict:
    return {
        "enabled": False,
        "smtp_host": "",
        "smtp_port": 587,
        "smtp_user": "",
        "smtp_password": "",
        "from_addr": "",
        "use_tls": True,
        "events": {
            "new_order":  {"enabled": True, "to": []},
            "dispatched": {"enabled": True, "to": []},
            "shortage":   {"enabled": True, "to": []},
        },
        # Optional per-branch recipient — a "dispatched" event for a given
        # branch will additionally CC the address listed here.
        "branch_emails": {},
    }


def load_email_config() -> dict:
    """Load config from disk; merge with defaults so a missing field never
    crashes the page."""
    base = _default_email_config()
    if not EMAIL_CONFIG_PATH.exists():
        return base
    try:
        with EMAIL_CONFIG_PATH.open("r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return base
    # Shallow-merge top-level, deep-merge events
    merged = {**base, **data}
    merged_events = base["events"].copy()
    for k, v in (data.get("events") or {}).items():
        if k in merged_events:
            merged_events[k] = {**merged_events[k], **v}
    merged["events"] = merged_events
    merged["branch_emails"] = data.get("branch_emails", {}) or {}
    return merged


def save_email_config(cfg: dict) -> None:
    with EMAIL_CONFIG_PATH.open("w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)


# --- Logging --------------------------------------------------------------
_log_lock = threading.Lock()


def _append_email_log(entry: dict) -> None:
    """Append one log entry, keeping only the most recent EMAIL_LOG_KEEP."""
    with _log_lock:
        log = []
        if EMAIL_LOG_PATH.exists():
            try:
                with EMAIL_LOG_PATH.open("r", encoding="utf-8") as f:
                    log = json.load(f)
                if not isinstance(log, list):
                    log = []
            except Exception:
                log = []
        log.append(entry)
        log = log[-EMAIL_LOG_KEEP:]
        try:
            with EMAIL_LOG_PATH.open("w", encoding="utf-8") as f:
                json.dump(log, f, indent=2, ensure_ascii=False)
        except Exception:
            pass


def load_email_log() -> list[dict]:
    if not EMAIL_LOG_PATH.exists():
        return []
    try:
        with EMAIL_LOG_PATH.open("r", encoding="utf-8") as f:
            log = json.load(f)
        if not isinstance(log, list):
            return []
        return log
    except Exception:
        return []


# --- Sending --------------------------------------------------------------
def _send_smtp(cfg: dict, to: list[str], subject: str, body: str) -> tuple[bool, str]:
    """Synchronous send — DO NOT call from UI code. Returns (ok, error_msg)."""
    if not cfg.get("smtp_host") or not cfg.get("from_addr"):
        return False, "SMTP not fully configured"
    if not to:
        return False, "No recipients"

    msg = EmailMessage()
    msg["From"] = cfg["from_addr"]
    msg["To"] = ", ".join(to)
    msg["Subject"] = subject
    msg.set_content(body)

    host = cfg["smtp_host"]
    port = int(cfg.get("smtp_port") or 587)
    user = cfg.get("smtp_user") or ""
    pw = cfg.get("smtp_password") or ""
    use_tls = bool(cfg.get("use_tls", True))

    try:
        # Port 465 = implicit SSL; everything else uses STARTTLS when enabled
        if port == 465:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(host, port, context=ctx, timeout=15) as s:
                if user:
                    s.login(user, pw)
                s.send_message(msg)
        else:
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.ehlo()
                if use_tls:
                    s.starttls(context=ssl.create_default_context())
                    s.ehlo()
                if user:
                    s.login(user, pw)
                s.send_message(msg)
        return True, ""
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def _send_async(cfg: dict, to: list[str], subject: str, body: str,
                event: str) -> None:
    """Background-thread wrapper around _send_smtp. Logs every attempt."""
    def _worker():
        ok, err = _send_smtp(cfg, to, subject, body)
        _append_email_log({
            "ts":      now_str(),
            "event":   event,
            "to":      to,
            "subject": subject,
            "ok":      ok,
            "error":   err if not ok else "",
        })
    t = threading.Thread(target=_worker, daemon=True)
    t.start()


def notify(event: str, subject: str, body: str,
           extra_to: list[str] | None = None) -> None:
    """Public entry point used by business code. Non-blocking, never raises.

    `event` must be one of EVENT_KEYS. Recipients = config.events[event].to
    plus any `extra_to` (e.g. a specific branch's address)."""
    if event not in EVENT_KEYS:
        return
    try:
        cfg = load_email_config()
    except Exception as e:
        log_exception("notify_load_email_config", e)
        return
    if not cfg.get("enabled"):
        return
    ev = cfg.get("events", {}).get(event, {})
    if not ev.get("enabled", True):
        return
    to: list[str] = []
    for addr in (ev.get("to") or []) + (extra_to or []):
        addr = (addr or "").strip()
        if addr and addr not in to:
            to.append(addr)
    if not to:
        return
    _send_async(cfg, to, subject, body, event)


def send_test_email(cfg: dict, to_addr: str) -> tuple[bool, str]:
    """Synchronous test send used by the admin UI. Logs the attempt and
    returns (ok, error_msg) so the page can show the result immediately."""
    subject = "[SUNSHINE 阳光集团] Test email / 测试邮件"
    body = (
        "This is a test email from the SUNSHINE ordering system.\n\n"
        "If you can read this, your SMTP configuration is working.\n\n"
        "—— SUNSHINE 阳光集团 订货系统"
    )
    ok, err = _send_smtp(cfg, [to_addr], subject, body)
    _append_email_log({
        "ts":      now_str(),
        "event":   "test",
        "to":      [to_addr],
        "subject": subject,
        "ok":      ok,
        "error":   err if not ok else "",
    })
    return ok, err


# =========================================================================
# PRODUCTS
# =========================================================================
@st.cache_data(ttl=300)
def load_products() -> pd.DataFrame:
    """Load product master from products.xlsx; auto-create demo if missing.

    Recognized columns: ItemCode, Barcode, Name, Unit, Price, Category.
    Category is optional. When present, it's used to group items in the
    warehouse picking-slip exports (Frozen / Rice / Beverage / etc.).
    Items without a category fall back to 'General'."""
    if not PRODUCTS_PATH.exists():
        demo = pd.DataFrame({
            "ItemCode": ["P001", "P002", "P003", "P004", "P005"],
            "Barcode":  ["8801234567001", "8801234567002", "8801234567003",
                         "8801234567004", "8801234567005"],
            "Name":     ["Demo Rice 5kg", "Demo Sugar 1kg", "Demo Cooking Oil 1L",
                         "Demo Soap Bar", "Demo Soft Drink"],
            "Unit":     ["bag", "bag", "bottle", "pc", "can"],
            "Price":    [12.5, 3.2, 4.8, 1.5, 2.0],
            "Category": ["Rice", "General", "General", "General", "Beverage"],
        })
        try:
            demo.to_excel(PRODUCTS_PATH, index=False)
        except Exception as e:
            log_exception("products_autocreate_demo", e)
        return demo
    try:
        # Read ItemCode and Barcode as strings from the get-go. Otherwise
        # pandas infers them as floats (especially large numeric barcodes
        # like 8801234567001), and "8801234567001.0" never matches anything.
        df = pd.read_excel(
            PRODUCTS_PATH,
            dtype={"ItemCode": str, "Barcode": str, "Category": str, "Unit": str},
        )
        for col in ["ItemCode", "Barcode", "Name", "Unit", "Price", "Category"]:
            if col not in df.columns:
                df[col] = ""
        for c in ["ItemCode", "Barcode", "Name", "Unit", "Category"]:
            df[c] = df[c].astype(str).fillna("").replace({"nan": "", "None": ""})
        # Belt-and-braces: if any cell still ended up like '8801234567001.0',
        # strip a trailing ".0" so matching works.
        for c in ["ItemCode", "Barcode"]:
            df[c] = df[c].str.replace(r"\.0$", "", regex=True)
        df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0.0)
        # Items with no category fall back to "General" so grouping always works
        df.loc[df["Category"].str.strip() == "", "Category"] = "General"
        return df
    except Exception as e:
        st.error(f"Error loading products.xlsx: {e}")
        return pd.DataFrame(columns=["ItemCode", "Barcode", "Name", "Unit",
                                     "Price", "Category"])


def search_products(query: str, limit: int = 80) -> pd.DataFrame:
    df = load_products()
    if df.empty:
        return df
    q = (query or "").strip().lower()
    if not q:
        return df.head(limit)
    mask = (
        df["Name"].str.lower().str.contains(q, na=False, regex=False)
        | df["Barcode"].str.lower().str.contains(q, na=False, regex=False)
        | df["ItemCode"].str.lower().str.contains(q, na=False, regex=False)
    )
    return df[mask].head(limit)


# =========================================================================
# UTILITIES
# =========================================================================
def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def unix_ts() -> float:
    return datetime.now().timestamp()


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def _append_app_log(level: str, where: str, detail: str) -> None:
    """Best-effort local runtime log. Never raises."""
    line = f"{now_str()} [{level}] [{where}] {detail}\n"
    try:
        with APP_LOG_PATH.open("a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass


def log_exception(where: str, exc: Exception) -> None:
    _append_app_log("ERROR", where, f"{type(exc).__name__}: {exc}")


def _order_fingerprint(branch: str, cart_snapshot: list[dict]) -> str:
    """Stable hash to prevent accidental double-submission by rapid clicks."""
    payload = json.dumps(
        {"branch": branch, "cart": cart_snapshot},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def gen_order_id(branch: str) -> str:
    safe = "".join(c for c in branch if c.isalnum()).upper()[:10]
    return f"{safe}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"


# --- Product images -------------------------------------------------------
# Simple file-on-disk approach. No DB schema, no compression, no resizing —
# just save the raw upload under a predictable filename. The browser scales
# the image at render time, so this is "fast enough" for a few hundred
# products. If we ever need to optimize, we can add Pillow compression
# without changing the public helpers below.
def _safe_filename_part(s: str) -> str:
    """Strip path separators / weird characters from a string before
    using it as a filename. Defends against weird ItemCode values like
    'A/B' or '../sneaky'."""
    return "".join(c for c in str(s) if c.isalnum() or c in "-_")


def get_product_image_path(item_code: str = "", barcode: str = "") -> Path | None:
    """Return the path to this product's image if one exists, else None.

    Look-up order:
      1. <ItemCode>.<ext> for any supported extension
      2. bc_<Barcode>.<ext> as fallback
    First match wins."""
    if not PRODUCT_IMAGES_DIR.exists():
        return None
    ic = _safe_filename_part(item_code)
    bc = _safe_filename_part(barcode)
    for stem in [ic, f"bc_{bc}" if bc else ""]:
        if not stem:
            continue
        for ext in IMAGE_EXTS:
            candidate = PRODUCT_IMAGES_DIR / f"{stem}{ext}"
            if candidate.exists():
                return candidate
    return None


def save_product_image(item_code: str, barcode: str, name: str,
                       uploaded_bytes: bytes,
                       original_filename: str) -> Path:
    """Save raw image bytes under a stable filename. Returns the saved path.

    File naming priority: ItemCode → barcode → safe-name fallback.
    Extension comes from the original filename; falls back to .jpg."""
    PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    # Pick filename stem
    ic = _safe_filename_part(item_code)
    bc = _safe_filename_part(barcode)
    if ic:
        stem = ic
    elif bc:
        stem = f"bc_{bc}"
    else:
        # Last resort: derive from name (manual products, unlikely path)
        stem = "name_" + (_safe_filename_part(name) or "noname")[:40]

    # Determine extension from upload, default to .jpg
    ext = Path(original_filename).suffix.lower()
    if ext not in IMAGE_EXTS:
        ext = ".jpg"

    # Remove any older version of this product's image (any extension)
    for old_ext in IMAGE_EXTS:
        old = PRODUCT_IMAGES_DIR / f"{stem}{old_ext}"
        if old.exists():
            try:
                old.unlink()
            except OSError:
                pass

    dest = PRODUCT_IMAGES_DIR / f"{stem}{ext}"
    dest.write_bytes(uploaded_bytes)
    return dest


def delete_product_image(item_code: str = "", barcode: str = "") -> bool:
    """Delete this product's image if it exists. Returns True on deletion."""
    p = get_product_image_path(item_code, barcode)
    if p is None:
        return False
    try:
        p.unlink()
        return True
    except OSError:
        return False


# --- Auto-scan root directory for new images -----------------------------
# User-friendly drop zone: drop image files directly next to app.py and the
# system will move matching ones into product_images/ on startup or on
# demand. Files that don't match any product stay where they are so the
# user can see and fix them. We move (not copy) to avoid duplicates.
def scan_and_import_root_images(verbose: bool = False) -> dict:
    """Scan the working directory (where app.py lives) for image files
    whose names match an ItemCode or Barcode in the master, and move them
    into product_images/. Returns a summary dict with counts + lists.

    Files that don't match are left untouched, so the user can see exactly
    which filenames need fixing. Files inside product_images/ itself are
    skipped (already imported)."""
    summary = {"moved": [], "skipped_unmatched": [], "errors": []}

    # The "root" we scan is the current working directory — that's where
    # streamlit launches app.py from, so it's also where the user dragged
    # the images in.
    root = Path(".")
    if not root.exists():
        return summary

    products_df = load_products()

    PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    # Iterate just the top level; don't recurse into subfolders. We don't
    # want to "discover" images that live in /backups/ etc.
    for entry in root.iterdir():
        if not entry.is_file():
            continue
        if entry.suffix.lower() not in IMAGE_EXTS:
            continue
        # Don't touch files inside product_images/ — already managed.
        # (We only iterate the top level, but be defensive anyway.)
        try:
            if entry.resolve().parent == PRODUCT_IMAGES_DIR.resolve():
                continue
        except OSError:
            continue

        m = match_filename_to_product(entry.name, products_df)
        if m is None:
            summary["skipped_unmatched"].append(entry.name)
            continue

        try:
            dest = save_product_image(
                item_code=m["item_code"],
                barcode=m["barcode"],
                name=m["name"],
                uploaded_bytes=entry.read_bytes(),
                original_filename=entry.name,
            )
            # Now that the image is safely in product_images/, remove the
            # source file. If unlink fails (perm denied, file in use), we
            # log it but the import already succeeded.
            try:
                entry.unlink()
            except OSError as e:
                summary["errors"].append(f"{entry.name}: copied but couldn't remove original — {e}")
            summary["moved"].append({
                "from": entry.name,
                "to":   dest.name,
                "matched_by": m["matched_by"],
                "name": m["name"],
            })
        except Exception as e:
            summary["errors"].append(f"{entry.name}: {e}")

    return summary


# --- Batch image upload --------------------------------------------------
# Pair uploaded files with products by filename. We accept either:
#   <ItemCode>.<ext>     (e.g. P001.jpg)
#   <Barcode>.<ext>      (e.g. 8801234567001.jpg)
# The stem (filename without extension) is matched against the products
# master. First exact match wins. Trailing/leading whitespace and case are
# ignored for ItemCode matching; Barcode matching is exact (digits only).
def match_filename_to_product(filename: str,
                              products_df: pd.DataFrame) -> dict | None:
    """Look up `filename` (e.g. 'P001.jpg') in the products master.
    Returns a dict with item_code, barcode, name, matched_by — or None."""
    if products_df is None or products_df.empty:
        return None
    stem = Path(filename).stem.strip()
    if not stem:
        return None

    # Try ItemCode first (case-insensitive)
    stem_lower = stem.lower()
    for _, row in products_df.iterrows():
        ic = str(row.get("ItemCode", "") or "").strip()
        if ic and ic.lower() == stem_lower:
            return {
                "item_code": ic,
                "barcode":   str(row.get("Barcode", "") or "").strip(),
                "name":      str(row.get("Name", "") or ""),
                "matched_by": "ItemCode",
            }

    # Then Barcode (exact)
    for _, row in products_df.iterrows():
        bc = str(row.get("Barcode", "") or "").strip()
        if bc and bc == stem:
            return {
                "item_code": str(row.get("ItemCode", "") or "").strip(),
                "barcode":   bc,
                "name":      str(row.get("Name", "") or ""),
                "matched_by": "Barcode",
            }

    return None


def plan_batch_image_upload(uploaded_files: list,
                            products_df: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    """Inspect a list of uploaded files and split them into
    (matched, unmatched). Each entry is a dict so the UI can render previews.

    `uploaded_files` items must have .name and .getvalue() (the standard
    Streamlit UploadedFile interface)."""
    matched: list[dict] = []
    unmatched: list[dict] = []
    for uf in uploaded_files:
        fname = getattr(uf, "name", "") or ""
        ext = Path(fname).suffix.lower()
        # Skip files with totally non-image extensions to avoid surprises.
        if ext and ext not in IMAGE_EXTS:
            unmatched.append({
                "filename": fname,
                "reason":   f"unsupported extension {ext}",
                "uf":       uf,
            })
            continue
        m = match_filename_to_product(fname, products_df)
        if m is None:
            unmatched.append({
                "filename": fname,
                "reason":   "no matching ItemCode or Barcode",
                "uf":       uf,
            })
        else:
            matched.append({
                "filename":   fname,
                "item_code":  m["item_code"],
                "barcode":    m["barcode"],
                "name":       m["name"],
                "matched_by": m["matched_by"],
                "uf":         uf,
            })
    return matched, unmatched


# --- Notification builders ------------------------------------------------
# Each helper returns (subject, body) ready to pass to notify(). Bodies are
# plain text — easy to read on phones, no HTML rendering issues.
def _format_lines_table(lines: list[dict]) -> str:
    """Render order lines as a fixed-width text table for the email body."""
    if not lines:
        return "  (no items)"
    # Compute column widths
    name_w = max(4, min(40, max(len(str(it.get("name", ""))) for it in lines)))
    out = []
    out.append(f"  {'Name':<{name_w}}  {'Cartons':>8}  {'Pcs':>6}  {'Unit':<8}")
    out.append(f"  {'-' * name_w}  {'-' * 8}  {'-' * 6}  {'-' * 8}")
    for it in lines:
        nm = str(it.get("name", ""))[:name_w]
        ct = it.get("qty_cartons", 0)
        pc = it.get("qty_pcs", 0)
        un = str(it.get("unit", ""))[:8]
        out.append(f"  {nm:<{name_w}}  {ct:>8}  {pc:>6}  {un:<8}")
    return "\n".join(out)


def build_new_order_email(order_id: str, branch: str,
                          lines: list[dict]) -> tuple[str, str]:
    subject = f"[新订单/New Order] {branch} · {order_id}"
    total_ct = sum(int(it.get("qty_cartons", 0)) for it in lines)
    total_pc = sum(int(it.get("qty_pcs", 0)) for it in lines)
    body = (
        f"📦 新订单 / New Order\n"
        f"=====================================\n"
        f"分店 / Branch:   {branch}\n"
        f"订单号 / Order:  {order_id}\n"
        f"时间 / Time:     {now_str()}\n"
        f"行数 / Lines:    {len(lines)}\n"
        f"合计 / Totals:   📦 {total_ct} cartons · 🔢 {total_pc} pcs\n"
        f"\n"
        f"商品明细 / Items:\n"
        f"{_format_lines_table(lines)}\n"
        f"\n"
        f"—— SUNSHINE 阳光集团 订货系统\n"
    )
    return subject, body


def build_dispatched_email(order_id: str, branch: str,
                           lines: list[dict]) -> tuple[str, str]:
    """`lines` should each have qty_cartons / qty_pcs (=ordered) and
    dispatch_cartons / dispatch_pcs (=actually shipped)."""
    subject = f"[已发货/Dispatched] {branch} · {order_id}"
    name_w = max(4, min(40, max(len(str(it.get("name", ""))) for it in lines))) if lines else 20
    rows = [
        f"  {'Name':<{name_w}}  {'Ordered':>10}  {'Shipped':>10}  {'Note':<10}",
        f"  {'-' * name_w}  {'-' * 10}  {'-' * 10}  {'-' * 10}",
    ]
    has_short = False
    for it in lines:
        nm = str(it.get("name", ""))[:name_w]
        oct_ = int(it.get("qty_cartons", 0))
        opc  = int(it.get("qty_pcs", 0))
        dct  = int(it.get("dispatch_cartons", oct_) or 0)
        dpc  = int(it.get("dispatch_pcs", opc) or 0)
        flag = ""
        if dct < oct_ or dpc < opc:
            flag = "SHORT"
            has_short = True
        rows.append(
            f"  {nm:<{name_w}}  "
            f"{oct_:>4}c {opc:>3}p  "
            f"{dct:>4}c {dpc:>3}p  "
            f"{flag:<10}"
        )
    body = (
        f"🚚 已发货 / Dispatched\n"
        f"=====================================\n"
        f"分店 / Branch:   {branch}\n"
        f"订单号 / Order:  {order_id}\n"
        f"时间 / Time:     {now_str()}\n"
        f"\n"
        + ("⚠️  部分商品库存不足，发货数量少于订单数量 / "
           "Some items shipped short of ordered qty\n\n" if has_short else "")
        + "\n".join(rows)
        + f"\n\n—— SUNSHINE 阳光集团 订货系统\n"
    )
    return subject, body


def build_shortage_email(order_id: str, branch: str,
                         shortages: list[dict]) -> tuple[str, str]:
    """`shortages` items should each have name, short_cartons, short_pcs."""
    subject = f"[缺货通知/Shortage] {branch} · {order_id}"
    rows = []
    for s in shortages:
        rows.append(
            f"  • {s.get('name', '')}: "
            f"📦 短 {s.get('short_cartons', 0)} cartons · "
            f"🔢 短 {s.get('short_pcs', 0)} pcs"
        )
    body = (
        f"🔔 缺货通知 / Shortage Reported\n"
        f"=====================================\n"
        f"分店 / Branch:   {branch}\n"
        f"订单号 / Order:  {order_id}\n"
        f"时间 / Time:     {now_str()}\n"
        f"\n"
        f"短少明细 / Items short:\n"
        + "\n".join(rows)
        + f"\n\n请尽快处理（补发或标记缺货）。\n"
        f"Please action — resend or mark out of stock.\n"
        f"\n—— SUNSHINE 阳光集团 订货系统\n"
    )
    return subject, body


# =========================================================================
# CSS
# =========================================================================
CSS = """
<style>
/* Push content down so the Streamlit toolbar doesn't overlap our header. */
.block-container { padding-top: 3.5rem; padding-bottom: 2rem; max-width: 1200px; }
@media (max-width: 640px) {
    .block-container { padding-left: .6rem; padding-right: .6rem; padding-top: 3rem; }
}

/* ---------- Blue gradient header ---------- */
.sunshine-header {
    background: linear-gradient(90deg, #1565c0 0%, #1976d2 50%, #42a5f5 100%);
    padding: 18px 22px;
    border-radius: 10px;
    color: #fff;
    margin-bottom: 18px;
    box-shadow: 0 2px 8px rgba(25, 118, 210, .25);
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 8px;
}
.sunshine-header .title-main {
    font-size: 22px; font-weight: 700; line-height: 1.2;
    display: flex; align-items: center; gap: 10px;
}
.sunshine-header .title-sub {
    font-size: 13px; opacity: .92; margin-top: 2px;
}
.sunshine-header .title-context {
    font-size: 13px; opacity: .9;
    background: rgba(255,255,255,.15);
    padding: 4px 10px; border-radius: 6px;
}

/* ---------- Shortage badge ---------- */
.short-badge {
    background: #e53935; color: #fff;
    border-radius: 12px; padding: 2px 9px;
    font-size: 12px; font-weight: 700; margin-left: 6px;
    box-shadow: 0 0 0 2px rgba(255,255,255,.3);
    animation: pulse 1.6s infinite;
}
@keyframes pulse {
    0%, 100% { transform: scale(1);   }
    50%      { transform: scale(1.08);}
}

/* ---------- Status pills ---------- */
.pill {
    display: inline-block; padding: 3px 10px;
    border-radius: 12px; font-size: 12px; font-weight: 600;
    color: #fff; line-height: 1.4;
}
.pill-Pending      { background: #fbc02d; color: #5d4500; }
.pill-Dispatched   { background: #1976d2; }
.pill-Received     { background: #388e3c; }
.pill-Open         { background: #fbc02d; color: #5d4500; }
.pill-Resending    { background: #1976d2; }
.pill-OutOfStock   { background: #e53935; }
.pill-Resolved     { background: #388e3c; }

/* ---------- Metric card ---------- */
.metric-card {
    background: #fff;
    border: 1px solid #e0e0e0;
    border-left: 4px solid #1976d2;
    padding: 14px 16px; border-radius: 6px;
    box-shadow: 0 1px 3px rgba(0,0,0,.05);
    color: #222;
}
.metric-card .label { font-size: 13px; color: #666; }
.metric-card .value { font-size: 28px; font-weight: 700; color: #1976d2; margin-top: 4px; }
.metric-card.warn  { border-left-color: #fbc02d; } .metric-card.warn  .value { color: #f57c00; }
.metric-card.ok    { border-left-color: #388e3c; } .metric-card.ok    .value { color: #2e7d32; }
.metric-card.alert { border-left-color: #e53935; } .metric-card.alert .value { color: #c62828; }

/* ---------- Sidebar buttons left-aligned ---------- */
section[data-testid="stSidebar"] .stButton > button {
    width: 100%; text-align: left; justify-content: flex-start;
}
</style>
"""


def inject_css() -> None:
    st.markdown(CSS, unsafe_allow_html=True)


# =========================================================================
# UI HELPERS
# =========================================================================
def render_header(context_label: str = "") -> None:
    open_short = count_open_shortages()
    badge = f'<span class="short-badge">🔔 {open_short}</span>' if open_short else ""
    ctx = f'<div class="title-context">{context_label}</div>' if context_label else ""
    st.markdown(
        f"""
        <div class="sunshine-header">
            <div>
                <div class="title-main">☀️ {t('app_brand')}{badge}</div>
                <div class="title-sub">{t('app_subtitle')}</div>
            </div>
            {ctx}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_lang_switch() -> None:
    _, en_col, zh_col = st.columns([8, 1, 1])
    with en_col:
        if st.button("EN", key="lang_en", use_container_width=True):
            st.session_state.lang = "en"
            st.rerun()
    with zh_col:
        if st.button("中文", key="lang_zh", use_container_width=True):
            st.session_state.lang = "zh"
            st.rerun()


def status_pill(status: str) -> str:
    css_status = status.replace(" ", "")
    label_key = STATUS_LABEL.get(status)
    label = t(label_key) if label_key else status
    return f'<span class="pill pill-{css_status}">{label}</span>'


def logout_button(key: str = "logout_btn") -> None:
    if st.button(f"🚪 {t('logout')}", key=key, use_container_width=True):
        lang = st.session_state.get("lang", "en")
        st.session_state.clear()
        st.session_state.lang = lang
        st.rerun()


# =========================================================================
# SESSION
# =========================================================================
def init_session() -> None:
    defaults = {
        "lang": None,
        "role": None,
        "branch": None,
        "page": None,
        "cart": [],
    }
    for key, val in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = val


def is_authenticated() -> bool:
    role = st.session_state.get("role")
    if role == Role.BRANCH:
        return bool(st.session_state.get("branch"))
    return role in (Role.WAREHOUSE, Role.ADMIN)


# =========================================================================
# LOGIN PAGES
# =========================================================================
def page_pick_language() -> None:
    st.markdown(
        """
        <div class="sunshine-header" style="text-align:center; justify-content:center;">
            <div>
                <div class="title-main" style="justify-content:center;">
                    ☀️ SUNSHINE SHOPPING CENTER · 阳光集团
                </div>
                <div class="title-sub">Ordering System · 订货系统</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.subheader("Select Language / 选择语言")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("🇬🇧 English", use_container_width=True, type="primary"):
            st.session_state.lang = "en"
            st.rerun()
    with c2:
        if st.button("🇨🇳 中文", use_container_width=True, type="primary"):
            st.session_state.lang = "zh"
            st.rerun()


def page_login() -> None:
    render_header()
    render_lang_switch()
    st.subheader(t("select_role"))

    pending = st.session_state.get("pending_role")

    c1, c2, c3 = st.columns(3)
    if c1.button(t("role_branch"), use_container_width=True):
        st.session_state.pending_role = Role.BRANCH
        st.rerun()
    if c2.button(t("role_warehouse"), use_container_width=True):
        st.session_state.pending_role = Role.WAREHOUSE
        st.rerun()
    if c3.button(t("role_admin"), use_container_width=True):
        st.session_state.pending_role = Role.ADMIN
        st.rerun()

    if pending == Role.BRANCH:
        st.divider()
        st.markdown(f"### {t('select_branch')}")
        branch = st.selectbox(t("select_branch"), [""] + BRANCHES, key="branch_pick")
        if branch and st.button(t("login"), type="primary", key="login_branch"):
            st.session_state.role = Role.BRANCH
            st.session_state.branch = branch
            st.session_state.page = "order"
            st.session_state.pop("pending_role", None)
            st.rerun()

    elif pending == Role.WAREHOUSE:
        _password_login(Role.WAREHOUSE, WAREHOUSE_PASSWORD, default_page="pending")

    elif pending == Role.ADMIN:
        _password_login(Role.ADMIN, ADMIN_PASSWORD, default_page="dashboard")


def _password_login(role: str, expected: str, default_page: str) -> None:
    st.divider()
    label = t("role_warehouse") if role == Role.WAREHOUSE else t("role_admin")
    st.markdown(f"### {label}")
    pw = st.text_input(t("password"), type="password", key=f"pw_{role}")
    if st.button(t("login"), type="primary", key=f"login_{role}"):
        if pw == expected:
            st.session_state.role = role
            st.session_state.page = default_page
            st.session_state.pop("pending_role", None)
            st.rerun()
        else:
            st.error(t("wrong_pw"))


# =========================================================================
# BRANCH PAGES
# =========================================================================
def page_branch_order() -> None:
    """Branch staff order placement.

    Two modes, switched by st.session_state.confirming:
      False (default) → browse / search / fill quantities / batch-add → cart
      True            → review cart, edit quantities, send to DB
    """
    if st.session_state.get("confirming"):
        _branch_order_confirm()
    else:
        _branch_order_browse()


# ----- Helpers ------------------------------------------------------------
# Why all this ceremony? Streamlit deletes a widget's session_state value
# the moment that widget is no longer rendered. So if a user fills qty for
# Product A on page 1 and clicks "Next ▶", the page-1 widgets disappear and
# Streamlit drops their values. The fix: every time a number_input changes,
# an on_change callback copies the new value into our own dict
# (_qty_snapshot) which we control. When the widget reappears later, we
# restore its initial value from the snapshot.
def _product_id(item_code: str, barcode: str, name: str) -> str:
    """Stable identity for a product across pages. Master rows can have
    blank ItemCode or Barcode, so we fall back to the name as last resort."""
    return f"{item_code}|{barcode}|{name}".lower()


def _qty_widget_keys(pid: str) -> tuple[str, str]:
    """Per-product widget keys used by Streamlit number_inputs."""
    return f"qty_w__{pid}__ct", f"qty_w__{pid}__pc"


def _on_qty_change(pid: str, which: str) -> None:
    """on_change callback: persist the widget's current value into our
    own snapshot before Streamlit can garbage-collect the widget state."""
    ct_key, pc_key = _qty_widget_keys(pid)
    snap = st.session_state.setdefault("_qty_snapshot", {})
    entry = snap.setdefault(pid, {"ct": 0, "pc": 0})
    if which == "ct":
        entry["ct"] = int(st.session_state.get(ct_key, 0) or 0)
    else:
        entry["pc"] = int(st.session_state.get(pc_key, 0) or 0)


def _register_product(pid: str, product: dict) -> None:
    """Remember product details (name / unit / price / etc.) keyed by pid,
    so that when batch-adding we know what each pid refers to."""
    reg = st.session_state.setdefault("_qty_registry", {})
    reg[pid] = product


def _get_qty(pid: str) -> tuple[int, int]:
    """Return (ct, pc) currently saved in the snapshot for this product."""
    snap = st.session_state.get("_qty_snapshot", {})
    entry = snap.get(pid, {"ct": 0, "pc": 0})
    return int(entry.get("ct", 0)), int(entry.get("pc", 0))


def _collect_qty_inputs() -> list[dict]:
    """Return all products with qty > 0, ready to be appended to the cart."""
    snap = st.session_state.get("_qty_snapshot", {})
    reg = st.session_state.get("_qty_registry", {})
    selected = []
    for pid, qty in snap.items():
        ct, pc = int(qty.get("ct", 0)), int(qty.get("pc", 0))
        if ct == 0 and pc == 0:
            continue
        product = reg.get(pid)
        if product is None:
            # Snapshot has qty but we never registered the product details.
            # Should not happen in practice, but skip gracefully if it does.
            continue
        selected.append({**product, "qty_cartons": ct, "qty_pcs": pc})
    return selected


def _clear_qty_inputs() -> None:
    """Reset both the snapshot and any live widget state."""
    st.session_state["_qty_snapshot"] = {}
    st.session_state["_qty_registry"] = {}
    # Also clear any number_input widgets that are still mounted.
    for k in [k for k in st.session_state if k.startswith("qty_w__")]:
        del st.session_state[k]


# ----- Mode 1: browse / search / fill quantities --------------------------
def _branch_order_browse() -> None:
    st.markdown(f"### 🛒 {t('nav_order')}")

    # --- Search ----------------------------------------------------------
    query = st.text_input(
        t("search_product"),
        key="branch_search",
        placeholder=t("search_product"),
        help="Bluetooth scanner supported / 支持蓝牙扫码枪",
    )
    results = search_products(query, limit=500)

    # Reset pagination on query change
    last_query = st.session_state.get("_last_query")
    if last_query != query:
        st.session_state["_last_query"] = query
        st.session_state["search_page"] = 1
    if "search_page" not in st.session_state:
        st.session_state["search_page"] = 1

    PAGE_SIZE = 10

    if not results.empty:
        total = len(results)
        total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
        page = min(max(1, st.session_state["search_page"]), total_pages)
        st.session_state["search_page"] = page

        start_idx = (page - 1) * PAGE_SIZE
        end_idx = min(start_idx + PAGE_SIZE, total)
        page_results = results.iloc[start_idx:end_idx]

        st.caption(
            f"{t('showing')} {start_idx + 1}–{end_idx} {t('of')} {total} "
            f"{t('results_count')} · {t('page')} {page} {t('of')} {total_pages}"
        )

        for idx, row in page_results.iterrows():
            with st.container(border=True):
                # Three columns: thumbnail | info | qty inputs
                c0, c1, c2 = st.columns([1, 3, 2])
                # ---- Thumbnail (left) ----
                with c0:
                    img_path = get_product_image_path(
                        str(row.get("ItemCode", "")),
                        str(row.get("Barcode", "")),
                    )
                    if img_path is not None:
                        st.image(str(img_path), width=90)
                    else:
                        # Lightweight placeholder so layout is stable
                        st.markdown(
                            "<div style='width:90px;height:90px;"
                            "background:#f0f0f0;border:1px dashed #bbb;"
                            "border-radius:4px;display:flex;"
                            "align-items:center;justify-content:center;"
                            "color:#bbb;font-size:24px;'>📦</div>",
                            unsafe_allow_html=True,
                        )
                with c1:
                    st.markdown(f"**{row['Name']}**")
                    parts = []
                    if row.get("ItemCode"): parts.append(f"📋 {row['ItemCode']}")
                    if row.get("Barcode"):  parts.append(f"📊 {row['Barcode']}")
                    if row.get("Unit"):     parts.append(f"📦 {row['Unit']}")
                    try:
                        if float(row.get("Price", 0)) > 0:
                            parts.append(f"💰 {float(row['Price']):.2f}")
                    except Exception:
                        pass
                    if parts:
                        st.caption(" · ".join(parts))
                    # Show "already entered" badge if user filled this on
                    # a previous page so they can see prior input survived
                    prev_ct, prev_pc = _get_qty(
                        _product_id(
                            str(row.get("ItemCode", "")),
                            str(row.get("Barcode", "")),
                            str(row["Name"]),
                        )
                    )
                    if prev_ct > 0 or prev_pc > 0:
                        st.caption(
                            f"✓ {t('saved_for_batch')}: "
                            f"📦 {prev_ct} · 🔢 {prev_pc}"
                        )
                with c2:
                    item_code = str(row.get("ItemCode", ""))
                    barcode   = str(row.get("Barcode", ""))
                    name      = str(row["Name"])
                    pid = _product_id(item_code, barcode, name)
                    ct_key, pc_key = _qty_widget_keys(pid)

                    # Register product details so we can rebuild cart later
                    _register_product(pid, {
                        "item_code": item_code,
                        "barcode":   barcode,
                        "name":      name,
                        "unit":      str(row.get("Unit", "")),
                        "price":     float(row.get("Price", 0) or 0),
                        "is_manual": 0,
                    })

                    # Restore from snapshot — if the widget is being mounted
                    # for the first time this rerun, seed it from our dict.
                    init_ct, init_pc = _get_qty(pid)
                    if ct_key not in st.session_state:
                        st.session_state[ct_key] = init_ct
                    if pc_key not in st.session_state:
                        st.session_state[pc_key] = init_pc

                    cc1, cc2 = st.columns(2)
                    with cc1:
                        # NOTE: we do NOT pass `value=` because that would
                        # fight Streamlit's session_state-driven default.
                        st.number_input(
                            t("cartons"), min_value=0, step=1, key=ct_key,
                            on_change=_on_qty_change, args=(pid, "ct"),
                        )
                    with cc2:
                        st.number_input(
                            t("each_pcs"), min_value=0, step=1, key=pc_key,
                            on_change=_on_qty_change, args=(pid, "pc"),
                        )

        # --- Pagination controls -----------------------------------------
        if total_pages > 1:
            st.markdown("")
            pc1, pc2, pc3 = st.columns([1, 2, 1])
            with pc1:
                if st.button(t("prev_page"), key="search_prev",
                             disabled=(page <= 1),
                             use_container_width=True):
                    st.session_state["search_page"] = page - 1
                    st.rerun()
            with pc2:
                st.markdown(
                    f"<div style='text-align:center;padding-top:6px;'>"
                    f"{t('page')} <b>{page}</b> {t('of')} <b>{total_pages}</b>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with pc3:
                if st.button(t("next_page"), key="search_next",
                             disabled=(page >= total_pages),
                             use_container_width=True):
                    st.session_state["search_page"] = page + 1
                    st.rerun()
    elif query:
        st.info(t("no_results"))

    # --- Manual add (still uses an instant form — single-item flow) ----
    with st.expander(t("manual_add")):
        with st.form("manual_form", clear_on_submit=True):
            mname = st.text_input(t("name"))
            c1, c2, c3, c4 = st.columns(4)
            with c1: munit = st.text_input(t("unit"), value="pc")
            with c2: mct = st.number_input(t("cartons"), min_value=0, value=0, step=1)
            with c3: mpc = st.number_input(t("each_pcs"), min_value=0, value=0, step=1)
            with c4: mbar = st.text_input(t("barcode"))
            if st.form_submit_button(t("add_to_cart"), type="primary"):
                if mname.strip() and (mct > 0 or mpc > 0):
                    st.session_state.cart.append({
                        "item_code": "",
                        "barcode":   mbar.strip(),
                        "name":      mname.strip(),
                        "unit":      munit.strip(),
                        "price":     0.0,
                        "qty_cartons": int(mct),
                        "qty_pcs":     int(mpc),
                        "is_manual": 1,
                    })
                    st.success(f"✓ {mname}")
                    st.rerun()

    st.divider()

    # --- Batch action bar -----------------------------------------------
    selected = _collect_qty_inputs()
    cart = st.session_state.cart
    n_selected = len(selected)
    cart_count = len(cart)

    info_l, info_r = st.columns(2)
    with info_l:
        st.markdown(
            f"**🛒 {t('cart')}: {cart_count}** &nbsp;·&nbsp; "
            f"**📥 {t('items_count')}: {n_selected}**"
        )
    with info_r:
        if st.button(t("clear_qty"), key="clear_qty",
                     disabled=(n_selected == 0),
                     use_container_width=True):
            _clear_qty_inputs()
            st.success(t("qty_cleared"))
            st.rerun()

    # Big primary action: append all selected to cart
    if st.button(t("add_selected"), type="primary",
                 disabled=(n_selected == 0),
                 use_container_width=True):
        if n_selected == 0:
            st.warning(t("no_qty_selected"))
        else:
            for item in selected:
                st.session_state.cart.append(item)
            _clear_qty_inputs()
            st.success(t("added_n_items").format(n=n_selected))
            st.rerun()

    # Secondary action: go to confirm page (only if cart non-empty)
    if cart_count > 0:
        if st.button(t("review_cart"), use_container_width=True,
                     key="go_review"):
            st.session_state["confirming"] = True
            st.rerun()
    else:
        st.info(t("empty_cart"))


# ----- Mode 2: review cart, edit, then send -------------------------------
def _branch_order_confirm() -> None:
    st.markdown(f"### 📝 {t('review_title')}")
    st.caption(t("review_subtitle"))

    cart = st.session_state.cart
    if not cart:
        # Edge case: cart got emptied while in confirm mode → bounce back
        st.session_state["confirming"] = False
        st.rerun()
        return

    # Render every cart line with editable qty + delete button
    for i, item in enumerate(cart):
        with st.container(border=True):
            top1, top2 = st.columns([4, 1])
            with top1:
                label = item["name"]
                if item.get("is_manual"):
                    label += f" 🖊️ {t('manual')}"
                st.markdown(f"**{label}**")
                sub = []
                if item["item_code"]: sub.append(f"📋 {item['item_code']}")
                if item["barcode"]:   sub.append(f"📊 {item['barcode']}")
                if item["unit"]:      sub.append(f"📦 {item['unit']}")
                if sub:
                    st.caption(" · ".join(sub))
            with top2:
                if st.button("🗑️", key=f"rm_confirm_{i}",
                             use_container_width=True):
                    st.session_state.cart.pop(i)
                    st.rerun()

            # Editable quantities — write back into cart on every render
            qc1, qc2 = st.columns(2)
            with qc1:
                new_ct = st.number_input(
                    t("cartons"), min_value=0,
                    value=int(item["qty_cartons"]), step=1,
                    key=f"confirm_ct_{i}",
                )
            with qc2:
                new_pc = st.number_input(
                    t("each_pcs"), min_value=0,
                    value=int(item["qty_pcs"]), step=1,
                    key=f"confirm_pc_{i}",
                )
            cart[i]["qty_cartons"] = int(new_ct)
            cart[i]["qty_pcs"]     = int(new_pc)

    # Drop any zero-quantity lines silently before showing summary
    cart[:] = [it for it in cart if it["qty_cartons"] > 0 or it["qty_pcs"] > 0]

    if not cart:
        st.warning(t("empty_cart"))
        if st.button(t("back_to_browse"), use_container_width=True):
            st.session_state["confirming"] = False
            st.rerun()
        return

    # Summary line
    st.divider()
    total_ct = sum(it["qty_cartons"] for it in cart)
    total_pc = sum(it["qty_pcs"] for it in cart)
    st.markdown(
        f"**{t('items_count')}: {len(cart)}** &nbsp;·&nbsp; "
        f"📦 **{total_ct}** {t('qty_cartons')} &nbsp;·&nbsp; "
        f"🔢 **{total_pc}** {t('qty_pcs')}"
    )

    # Action buttons
    bc1, bc2 = st.columns(2)
    with bc1:
        if st.button(t("back_to_browse"), use_container_width=True,
                     key="back_browse"):
            st.session_state["confirming"] = False
            st.rerun()
    with bc2:
        submit_busy = bool(st.session_state.get("_send_order_busy", False))
        if submit_busy:
            st.caption(f"⏳ {t('submit_busy')}")
        if st.button(
            t("send_order"),
            type="primary",
            use_container_width=True,
            key="send_order",
            disabled=submit_busy,
        ):
            if st.session_state.get("_send_order_busy", False):
                st.warning(t("submit_busy"))
                return
            st.session_state["_send_order_busy"] = True
            try:
                order_id = gen_order_id(st.session_state.branch)
                branch = st.session_state.branch
                # Snapshot the cart before clearing — we need it for the email
                cart_snapshot = [dict(it) for it in cart]

                # Defensive dedupe against accidental rapid double clicks.
                fp = _order_fingerprint(branch, cart_snapshot)
                last_fp = st.session_state.get("_last_order_submit_fp")
                last_ts = float(st.session_state.get("_last_order_submit_ts", 0.0))
                if last_fp == fp and (unix_ts() - last_ts) < 10:
                    st.warning(t("dup_submit_block"))
                    return

                with db_conn() as conn:
                    for item in cart_snapshot:
                        conn.execute("""
                            INSERT INTO orders
                            (order_id, branch, item_code, barcode, name, unit, price,
                             qty_cartons, qty_pcs, status, is_manual, order_date)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pending', ?, ?)
                        """, (
                            order_id, branch,
                            item["item_code"], item["barcode"], item["name"],
                            item["unit"], item["price"],
                            item["qty_cartons"], item["qty_pcs"],
                            item["is_manual"], now_str(),
                        ))

                st.session_state["_last_order_submit_fp"] = fp
                st.session_state["_last_order_submit_ts"] = unix_ts()

                # Fire-and-forget email notification (warehouse + admin)
                try:
                    subject, body = build_new_order_email(
                        order_id, branch, cart_snapshot
                    )
                    notify("new_order", subject, body)
                except Exception as e:
                    log_exception("branch_submit_new_order_email", e)

                st.session_state.cart = []
                st.session_state["confirming"] = False
                _clear_qty_inputs()
                st.success(f"✅ {t('order_submitted')} — {t('order_id')}: `{order_id}`")
                st.balloons()
            finally:
                st.session_state["_send_order_busy"] = False


def page_branch_my_orders() -> None:
    st.markdown(f"### 📋 {t('nav_my_orders')}")
    branch = st.session_state.branch
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE branch = ? ORDER BY order_date DESC",
            conn, params=(branch,),
        )

    if df.empty:
        st.info(t("no_data"))
        return

    for order_id, group in df.groupby("order_id", sort=False):
        status = group["status"].iloc[0]
        order_date = group["order_date"].iloc[0]

        with st.expander(
            f"📦 {order_id} · {order_date[:16]} · {len(group)} items",
            expanded=(status == OrderStatus.DISPATCHED),
        ):
            st.markdown(status_pill(status), unsafe_allow_html=True)
            st.markdown("")

            show = group[["name", "unit", "qty_cartons", "qty_pcs"]].rename(columns={
                "name": t("name"), "unit": t("unit"),
                "qty_cartons": t("qty_cartons"), "qty_pcs": t("qty_pcs"),
            })
            if status in (OrderStatus.DISPATCHED, OrderStatus.RECEIVED):
                # Show what the warehouse actually shipped vs what was ordered
                show[f"🚚 {t('qty_cartons')}"] = group["dispatch_cartons"].fillna("-").values
                show[f"🚚 {t('qty_pcs')}"] = group["dispatch_pcs"].fillna("-").values
                show[t("actual_cartons")] = group["actual_cartons"].fillna("-").values
                show[t("actual_pcs")] = group["actual_pcs"].fillna("-").values
            st.dataframe(show, use_container_width=True, hide_index=True)

            if status == OrderStatus.DISPATCHED:
                _render_receive_form(order_id, group)


def _render_receive_form(order_id: str, group: pd.DataFrame) -> None:
    st.markdown(f"#### ✅ {t('receive')}")
    receive_q = st.text_input(
        f"🔍 {t('search')} ({t('barcode')}/{t('name')})",
        key=f"recv_q_{order_id}",
        help="Scanner supported / 支持扫码枪",
    )
    with st.form(f"recv_form_{order_id}"):
        actuals: dict[int, tuple[int, int]] = {}
        for _, row in group.iterrows():
            line_id = int(row["id"])
            if receive_q:
                q = receive_q.lower().strip()
                hay = f"{row['name']} {row['barcode']} {row['item_code']}".lower()
                if q not in hay:
                    continue

            # Use dispatch_cartons/pcs as the baseline if set; fall back to
            # qty_cartons/pcs for legacy orders dispatched before this feature.
            disp_ct = row["dispatch_cartons"]
            disp_pc = row["dispatch_pcs"]
            if pd.isna(disp_ct) or disp_ct is None:
                disp_ct = int(row["qty_cartons"])
            else:
                disp_ct = int(disp_ct)
            if pd.isna(disp_pc) or disp_pc is None:
                disp_pc = int(row["qty_pcs"])
            else:
                disp_pc = int(disp_pc)

            st.markdown(f"**{row['name']}**")
            # Show both: what was ordered AND what the warehouse dispatched.
            # If they differ, the difference is already a known short by the
            # warehouse — the branch only needs to compare against dispatch.
            if disp_ct != int(row["qty_cartons"]) or disp_pc != int(row["qty_pcs"]):
                st.caption(
                    f"📋 {t('ordered_qty')}: {row['qty_cartons']} {t('qty_cartons')} / "
                    f"{row['qty_pcs']} {t('qty_pcs')}  &nbsp;·&nbsp;  "
                    f"🚚 {t('dispatched')}: **{disp_ct}** {t('qty_cartons')} / "
                    f"**{disp_pc}** {t('qty_pcs')}"
                )
            else:
                st.caption(
                    f"🚚 {t('dispatched')}: {disp_ct} {t('qty_cartons')} / "
                    f"{disp_pc} {t('qty_pcs')}"
                )
            c1, c2 = st.columns(2)
            with c1:
                ac = st.number_input(
                    t("actual_cartons"), min_value=0,
                    value=disp_ct, step=1, key=f"ac_{line_id}",
                )
            with c2:
                ap = st.number_input(
                    t("actual_pcs"), min_value=0,
                    value=disp_pc, step=1, key=f"ap_{line_id}",
                )
            actuals[line_id] = (ac, ap, disp_ct, disp_pc)
        overall_remark = st.text_area(t("remarks"), key=f"rem_{order_id}")
        if st.form_submit_button(f"✅ {t('confirm_receive')}", type="primary"):
            any_short = False
            short_items: list[dict] = []  # collected for the email below
            branch_name = ""
            with db_conn() as conn:
                for _, row in group.iterrows():
                    line_id = int(row["id"])
                    if line_id in actuals:
                        ac, ap, disp_ct, disp_pc = actuals[line_id]
                    else:
                        # Filtered-out lines auto-confirm at the dispatched qty
                        disp_ct_raw = row["dispatch_cartons"]
                        disp_pc_raw = row["dispatch_pcs"]
                        disp_ct = (int(disp_ct_raw) if not (pd.isna(disp_ct_raw) or disp_ct_raw is None)
                                   else int(row["qty_cartons"]))
                        disp_pc = (int(disp_pc_raw) if not (pd.isna(disp_pc_raw) or disp_pc_raw is None)
                                   else int(row["qty_pcs"]))
                        ac, ap = disp_ct, disp_pc
                    conn.execute("""
                        UPDATE orders SET
                            status='Received',
                            actual_cartons=?, actual_pcs=?,
                            receive_remarks=?, receive_date=?
                        WHERE id=?
                    """, (ac, ap, overall_remark, now_str(), line_id))
                    # Shortage is computed against DISPATCHED quantity, not
                    # ordered quantity — the warehouse already accounts for
                    # any "couldn't ship" gap by entering a smaller dispatch
                    # number, so the shortage record is purely about
                    # transit/branch-level discrepancy.
                    short_ct = max(0, disp_ct - ac)
                    short_pc = max(0, disp_pc - ap)
                    if short_ct > 0 or short_pc > 0:
                        any_short = True
                        branch_name = row["branch"]
                        conn.execute("""
                            INSERT INTO shortages
                            (order_id, order_line_id, branch, item_code, barcode,
                             name, unit, short_cartons, short_pcs, status,
                             reported_date, branch_remarks)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'Open', ?, ?)
                        """, (
                            order_id, line_id, row["branch"],
                            row["item_code"], row["barcode"], row["name"],
                            row["unit"], short_ct, short_pc,
                            now_str(), overall_remark,
                        ))
                        short_items.append({
                            "name": row["name"],
                            "short_cartons": short_ct,
                            "short_pcs": short_pc,
                        })
            # Send ONE shortage email summarizing all short lines for this
            # receipt — never one email per item.
            if any_short and short_items:
                try:
                    subject, body = build_shortage_email(
                        order_id, branch_name, short_items
                    )
                    notify("shortage", subject, body)
                except Exception as e:
                    log_exception("branch_shortage_notify", e)
            if any_short:
                st.warning(f"⚠️ {t('shortage_alert')}")
            st.success(f"✅ {t('receipt_done')}")
            st.rerun()


def page_branch_shortages() -> None:
    st.markdown(f"### 🔔 {t('nav_my_short')}")
    branch = st.session_state.branch
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM shortages WHERE branch = ? ORDER BY reported_date DESC",
            conn, params=(branch,),
        )

    if df.empty:
        st.info(t("no_shortages"))
        return

    for _, row in df.iterrows():
        with st.container(border=True):
            c1, c2 = st.columns([4, 1])
            with c1:
                st.markdown(f"**{row['name']}** · {row['order_id']}")
                st.caption(
                    f"{t('shortage_qty')}: {row['short_cartons']} {t('qty_cartons')} / "
                    f"{row['short_pcs']} {t('qty_pcs')} · {row['reported_date'][:16]}"
                )
                if row["warehouse_reply"]:
                    st.info(f"💬 **{t('warehouse_reply')}:** {row['warehouse_reply']}")
            with c2:
                st.markdown(status_pill(row["status"]), unsafe_allow_html=True)
            if row["status"] == ShortageStatus.RESENDING:
                if st.button(t("confirm_resend"), key=f"cr_{row['id']}"):
                    with db_conn() as conn:
                        conn.execute(
                            "UPDATE shortages SET status='Resolved', resolved_date=? WHERE id=?",
                            (now_str(), row["id"]),
                        )
                    st.rerun()


# =========================================================================
# WAREHOUSE PAGES
# =========================================================================
def page_warehouse_pending() -> None:
    st.markdown(f"### 📦 {t('nav_pending')}")
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE status='Pending' ORDER BY order_date ASC", conn,
        )

    if df.empty:
        st.info(t("no_pending"))
        return

    # ----- Top-level "download all pending" --------------------------
    n_orders = df["order_id"].nunique()
    n_branches = df["branch"].nunique()
    cap_l, cap_r = st.columns([3, 2])
    with cap_l:
        st.caption(
            f"📦 {n_orders} {t('orders')} · 🏪 {n_branches} {t('branch')} · "
            f"{len(df)} {t('lines')}"
            if "orders" in T else f"📦 {n_orders} orders · {n_branches} branches · {len(df)} lines"
        )
    with cap_r:
        if st.button(t("dl_all_pending"), key="dl_all_pending",
                     use_container_width=True):
            bio = export_all_pending_picking()
            if bio is None:
                st.warning(t("no_data"))
            else:
                stamp = datetime.now().strftime("%Y%m%d_%H%M")
                st.session_state["_dl_all"] = (bio.getvalue(), f"picking_all_{stamp}.xlsx")
        # show download button if file ready
        if "_dl_all" in st.session_state:
            data, fname = st.session_state["_dl_all"]
            st.download_button(
                f"⬇️ {fname}", data=data, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_all_btn",
            )

    st.divider()

    # ----- Per-branch sections ---------------------------------------
    for branch in sorted(df["branch"].unique()):
        bdf = df[df["branch"] == branch]
        head_l, head_r = st.columns([3, 2])
        with head_l:
            st.markdown(f"#### 🏪 {branch} ({bdf['order_id'].nunique()} orders · {len(bdf)} lines)")
        with head_r:
            dl_key = f"dl_branch_{branch}"
            if st.button(t("dl_branch_all"), key=dl_key, use_container_width=True):
                bio = export_branch_pending_picking(branch)
                if bio is None:
                    st.warning(t("no_data"))
                else:
                    stamp = datetime.now().strftime("%Y%m%d_%H%M")
                    safe_branch = branch.replace(" ", "_").replace("/", "-")
                    st.session_state[f"_dl_b_{branch}"] = (
                        bio.getvalue(), f"picking_{safe_branch}_{stamp}.xlsx",
                    )
            if f"_dl_b_{branch}" in st.session_state:
                data, fname = st.session_state[f"_dl_b_{branch}"]
                st.download_button(
                    f"⬇️ {fname}", data=data, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key=f"dl_b_btn_{branch}",
                )

        # ----- Each individual order ---------------------------------
        for order_id, group in bdf.groupby("order_id", sort=False):
            with st.expander(
                f"📦 {order_id} · {group['order_date'].iloc[0][:16]} · {len(group)} items"
            ):
                st.caption(f"💡 {t('dispatch_hint')}")

                # Editable dispatch-quantity form for each line.
                # Default values = ordered quantities, so the common case
                # (full dispatch) is one click. Warehouse staff can lower
                # any number when stock is short.
                with st.form(f"dispatch_form_{order_id}"):
                    # Header row
                    h1, h2, h3, h4, h5 = st.columns([3, 2, 2, 2, 2])
                    with h1: st.markdown(f"**{t('name')}**")
                    with h2: st.markdown(f"**{t('barcode')}**")
                    with h3: st.markdown(f"**{t('ordered_qty')}**")
                    with h4: st.markdown(f"**{t('dispatch_cartons')}**")
                    with h5: st.markdown(f"**{t('dispatch_pcs')}**")
                    st.divider()

                    dispatch_inputs: dict[int, tuple[int, int]] = {}
                    for _, row in group.iterrows():
                        line_id = int(row["id"])
                        ordered_ct = int(row["qty_cartons"])
                        ordered_pc = int(row["qty_pcs"])

                        c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 2])
                        with c1:
                            label = row["name"]
                            if row["item_code"]:
                                st.markdown(f"**{label}**")
                                st.caption(f"📋 {row['item_code']}")
                            else:
                                st.markdown(f"**{label}**")
                        with c2:
                            st.markdown(f"`{row['barcode'] or '-'}`")
                            if row["unit"]:
                                st.caption(f"📦 {row['unit']}")
                        with c3:
                            st.markdown(
                                f"📦 {ordered_ct} {t('qty_cartons')}<br>"
                                f"🔢 {ordered_pc} {t('qty_pcs')}",
                                unsafe_allow_html=True,
                            )
                        with c4:
                            dc = st.number_input(
                                t("dispatch_cartons"),
                                min_value=0, value=ordered_ct, step=1,
                                key=f"dc_{line_id}",
                                label_visibility="collapsed",
                            )
                        with c5:
                            dp = st.number_input(
                                t("dispatch_pcs"),
                                min_value=0, value=ordered_pc, step=1,
                                key=f"dp_qty_{line_id}",
                                label_visibility="collapsed",
                            )
                        dispatch_inputs[line_id] = (int(dc), int(dp))

                        # Inline warning when the staff is shipping less
                        # than ordered — helpful but non-blocking.
                        if int(dc) < ordered_ct or int(dp) < ordered_pc:
                            st.caption(
                                f"⚠️ {t('less_than_ordered')}: "
                                f"📦 {ordered_ct - int(dc)} · "
                                f"🔢 {ordered_pc - int(dp)}"
                            )

                    st.divider()

                    # Action row inside the form
                    ac1, ac2 = st.columns(2)
                    with ac1:
                        # Picking-slip download stays as a regular button OUTSIDE
                        # the form — but we can't put it inside this form easily.
                        # Show a placeholder instructional caption.
                        st.caption(
                            f"💡 {t('dl_this_order')} → "
                            f"{t('back_to_browse').replace('◀', '').strip()}"
                        )
                    with ac2:
                        if st.form_submit_button(
                            f"🚚 {t('mark_dispatched')}",
                            type="primary", use_container_width=True,
                        ):
                            busy_orders = st.session_state.setdefault(
                                "_dispatch_busy_orders", {}
                            )
                            if busy_orders.get(order_id):
                                st.warning(t("dispatch_busy"))
                            else:
                                busy_orders[order_id] = True
                                try:
                                    stamp = now_str()
                                    with db_conn() as conn:
                                        # Write dispatch quantities for each line, and
                                        # flip status + dispatch_date in one transaction.
                                        for line_id, (dc, dp) in dispatch_inputs.items():
                                            conn.execute("""
                                                UPDATE orders SET
                                                    status='Dispatched',
                                                    dispatch_date=?,
                                                    dispatch_cartons=?,
                                                    dispatch_pcs=?
                                                WHERE id=?
                                            """, (stamp, dc, dp, line_id))
                                        conn.execute(
                                            "INSERT INTO shipments "
                                            "(order_id, branch, dispatch_date, dispatched_by) "
                                            "VALUES (?, ?, ?, 'warehouse')",
                                            (order_id, branch, stamp),
                                        )
                                        # Re-read the rows we just updated so the email
                                        # body reflects the saved dispatch quantities.
                                        rows = conn.execute(
                                            "SELECT name, unit, qty_cartons, qty_pcs, "
                                            "       dispatch_cartons, dispatch_pcs "
                                            "FROM orders WHERE order_id = ?",
                                            (order_id,),
                                        ).fetchall()
                                    try:
                                        lines_for_email = [dict(r) for r in rows]
                                        subject, body = build_dispatched_email(
                                            order_id, branch, lines_for_email
                                        )
                                        # Optional: CC the branch's specific email
                                        cfg = load_email_config()
                                        branch_email = (cfg.get("branch_emails") or {}
                                                        ).get(branch, "").strip()
                                        extra = [branch_email] if branch_email else []
                                        notify("dispatched", subject, body, extra_to=extra)
                                    except Exception as e:
                                        log_exception("warehouse_dispatch_notify", e)
                                    st.session_state.pop(f"_dl_o_{order_id}", None)
                                    st.success(f"✅ {t('shipment_marked')}")
                                    st.rerun()
                                finally:
                                    busy_orders[order_id] = False

                # ----- Picking-slip download (separate, outside form) -----
                dl_o_key = f"dl_o_{order_id}"
                if st.button(t("dl_this_order"), key=dl_o_key,
                             use_container_width=True):
                    bio = export_single_order_picking(order_id)
                    if bio is None:
                        st.warning(t("no_data"))
                    else:
                        st.session_state[f"_dl_o_{order_id}"] = (
                            bio.getvalue(),
                            f"picking_{order_id}.xlsx",
                        )
                if f"_dl_o_{order_id}" in st.session_state:
                    data, fname = st.session_state[f"_dl_o_{order_id}"]
                    st.download_button(
                        f"⬇️ {fname}", data=data, file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, key=f"dl_o_btn_{order_id}",
                    )


def page_warehouse_shortages() -> None:
    st.markdown(f"### 🔔 {t('nav_short_in')}")
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM shortages WHERE status='Open' ORDER BY reported_date DESC", conn,
        )

    if df.empty:
        st.info(t("no_shortages"))
        return

    for _, row in df.iterrows():
        with st.container(border=True):
            st.markdown(f"**{row['name']}** · 🏪 {row['branch']}")
            st.caption(
                f"{t('order_id')}: {row['order_id']} · "
                f"{t('shortage_qty')}: {row['short_cartons']} {t('qty_cartons')} / "
                f"{row['short_pcs']} {t('qty_pcs')} · {row['reported_date'][:16]}"
            )
            if row["branch_remarks"]:
                st.caption(f"📝 {row['branch_remarks']}")
            reply = st.text_input(t("warehouse_reply"), key=f"wr_{row['id']}")
            c1, c2 = st.columns(2)
            with c1:
                if st.button(f"♻️ {t('resend')}", key=f"rs_{row['id']}",
                             use_container_width=True):
                    with db_conn() as conn:
                        conn.execute(
                            "UPDATE shortages SET status='Resending', warehouse_reply=? WHERE id=?",
                            (reply or "Will resend", row["id"]),
                        )
                    st.rerun()
            with c2:
                if st.button(f"❌ {t('mark_oos')}", key=f"oos_{row['id']}",
                             use_container_width=True):
                    with db_conn() as conn:
                        conn.execute(
                            "UPDATE shortages SET status='Out of Stock', warehouse_reply=?, "
                            "resolved_date=? WHERE id=?",
                            (reply or "Out of stock", now_str(), row["id"]),
                        )
                    st.rerun()


def page_warehouse_dispatch_history() -> None:
    """Browse dispatched orders by date. Joins shipments with orders to show
    what was sent on which day and which branch received it."""
    st.markdown(f"### 📜 {t('dispatch_history')}")

    # ----- Date filter -------------------------------------------------
    with db_conn() as conn:
        # Pull list of distinct dispatch dates so we know what's available
        date_rows = conn.execute(
            "SELECT DISTINCT date(dispatch_date) AS d "
            "FROM shipments WHERE dispatch_date IS NOT NULL "
            "ORDER BY d DESC"
        ).fetchall()
    available_dates = [r["d"] for r in date_rows]

    if not available_dates:
        st.info(t("no_dispatch_history"))
        return

    # Default to the most recent date
    default_date = st.session_state.get("_dh_date") or available_dates[0]

    fc1, fc2, fc3 = st.columns([2, 1, 2])
    with fc1:
        # Use date_input with min/max so only valid dates can be picked
        try:
            picked = st.date_input(
                t("select_date"),
                value=datetime.strptime(default_date, "%Y-%m-%d").date(),
                key="_dh_date_input",
            )
            picked_date = picked.strftime("%Y-%m-%d")
        except Exception:
            picked_date = default_date
    with fc2:
        if st.button(t("today_btn"), key="_dh_today",
                     use_container_width=True):
            st.session_state["_dh_date"] = today_str()
            st.rerun()
    with fc3:
        show_all = st.checkbox(t("all_dates"), key="_dh_all")

    # ----- Query dispatched orders for the chosen date(s) --------------
    with db_conn() as conn:
        if show_all:
            ship_df = pd.read_sql_query(
                "SELECT * FROM shipments ORDER BY dispatch_date DESC", conn,
            )
        else:
            ship_df = pd.read_sql_query(
                "SELECT * FROM shipments WHERE date(dispatch_date) = ? "
                "ORDER BY dispatch_date DESC",
                conn, params=(picked_date,),
            )

    if ship_df.empty:
        st.info(t("no_dispatch_history"))
        return

    # ----- Summary metrics ---------------------------------------------
    n_orders = ship_df["order_id"].nunique()
    n_branches = ship_df["branch"].nunique()
    label_when = t("all_dates") if show_all else picked_date
    st.caption(
        f"📅 {label_when} · 🚚 {n_orders} {t('orders')} · "
        f"🏪 {n_branches} {t('branch')}"
    )

    # ----- Group by branch, then list each dispatched order -----------
    for branch in sorted(ship_df["branch"].unique()):
        bdf = ship_df[ship_df["branch"] == branch]
        st.markdown(f"#### 🏪 {branch} ({len(bdf)} {t('orders')})")

        for _, ship in bdf.iterrows():
            order_id = ship["order_id"]
            dispatch_date = ship["dispatch_date"]

            # Pull the line items + downstream receipt info for this order
            with db_conn() as conn:
                lines = pd.read_sql_query(
                    "SELECT name, item_code, barcode, unit, "
                    "       qty_cartons, qty_pcs, "
                    "       dispatch_cartons, dispatch_pcs, "
                    "       actual_cartons, actual_pcs, "
                    "       status, receive_date "
                    "FROM orders WHERE order_id = ? ORDER BY name",
                    conn, params=(order_id,),
                )

            # Helper: effective dispatched qty for a row (fallback to ordered
            # qty for legacy rows that have no dispatch_* set).
            def _disp(r, col_dispatch, col_qty):
                val = r[col_dispatch]
                if pd.isna(val) or val is None:
                    return int(r[col_qty])
                return int(val)

            # Determine receipt status across all lines
            statuses = set(lines["status"].tolist())
            if statuses == {OrderStatus.RECEIVED}:
                # All received — check whether any short vs DISPATCHED qty
                any_short = any(
                    (pd.notna(r["actual_cartons"]) and
                     r["actual_cartons"] < _disp(r, "dispatch_cartons", "qty_cartons")) or
                    (pd.notna(r["actual_pcs"]) and
                     r["actual_pcs"] < _disp(r, "dispatch_pcs", "qty_pcs"))
                    for _, r in lines.iterrows()
                )
                receipt_label = (
                    t("partial_received") if any_short else t("fully_received")
                )
                receipt_color = "#fbc02d" if any_short else "#388e3c"
            else:
                receipt_label = t("awaiting_receipt")
                receipt_color = "#1976d2"

            with st.expander(
                f"📦 {order_id} · 🚚 {dispatch_date[:16]} · "
                f"{len(lines)} items · {receipt_label}"
            ):
                # Receipt-status pill
                st.markdown(
                    f"<span class='pill' style='background:{receipt_color};'>"
                    f"{receipt_label}</span>",
                    unsafe_allow_html=True,
                )
                st.markdown("")

                show = lines[[
                    "name", "item_code", "barcode", "unit",
                    "qty_cartons", "qty_pcs",
                    "dispatch_cartons", "dispatch_pcs",
                    "actual_cartons", "actual_pcs",
                ]].fillna("-").rename(columns={
                    "name":             t("name"),
                    "item_code":        t("item_code"),
                    "barcode":          t("barcode"),
                    "unit":             t("unit"),
                    "qty_cartons":      f"📋 {t('qty_cartons')}",
                    "qty_pcs":          f"📋 {t('qty_pcs')}",
                    "dispatch_cartons": f"🚚 {t('qty_cartons')}",
                    "dispatch_pcs":     f"🚚 {t('qty_pcs')}",
                    "actual_cartons":   f"✅ {t('qty_cartons')}",
                    "actual_pcs":       f"✅ {t('qty_pcs')}",
                })
                st.dataframe(show, use_container_width=True, hide_index=True)


# =========================================================================
# ADMIN PAGES
# =========================================================================
def page_admin_dashboard() -> None:
    st.markdown(f"### 📊 {t('nav_dashboard')}")
    today = today_str()

    with db_conn() as conn:
        n_today = conn.execute(
            "SELECT COUNT(DISTINCT order_id) FROM orders WHERE date(order_date) = ?",
            (today,),
        ).fetchone()[0]
        n_pending = conn.execute(
            "SELECT COUNT(DISTINCT order_id) FROM orders WHERE status='Pending'",
        ).fetchone()[0]
        n_dispatched = conn.execute(
            "SELECT COUNT(DISTINCT order_id) FROM orders WHERE status='Dispatched'",
        ).fetchone()[0]
        n_short = conn.execute(
            "SELECT COUNT(*) FROM shortages WHERE status='Open'",
        ).fetchone()[0]

    cards = [
        ("📊", t("card_today"),      n_today,     ""),
        ("⏳", t("card_pending"),    n_pending,   "warn"),
        ("🚚", t("card_dispatched"), n_dispatched,"ok"),
        ("🔔", t("card_short"),      n_short,     "alert"),
    ]
    cols = st.columns(4)
    for col, (icon, label, val, mod) in zip(cols, cards):
        with col:
            st.markdown(f"""
            <div class="metric-card {mod}">
                <div class="label">{icon} {label}</div>
                <div class="value">{val}</div>
            </div>
            """, unsafe_allow_html=True)

    st.markdown("")
    st.markdown(f"#### 🏪 {t('branch_status')}")
    with db_conn() as conn:
        bdf = pd.read_sql_query("""
            SELECT branch,
                   COUNT(DISTINCT CASE WHEN status='Pending'    THEN order_id END) AS pending,
                   COUNT(DISTINCT CASE WHEN status='Dispatched' THEN order_id END) AS dispatched,
                   COUNT(DISTINCT CASE WHEN status='Received'   THEN order_id END) AS received
            FROM orders GROUP BY branch
        """, conn)
        sdf = pd.read_sql_query(
            "SELECT branch, COUNT(*) AS open_short FROM shortages "
            "WHERE status='Open' GROUP BY branch", conn,
        )
    merged = (
        pd.DataFrame({"branch": BRANCHES})
        .merge(bdf, on="branch", how="left")
        .merge(sdf, on="branch", how="left")
        .fillna(0)
    )
    for col in ["pending", "dispatched", "received", "open_short"]:
        merged[col] = merged[col].astype(int)
    merged.columns = [t("branch"), t("st_pending"), t("st_dispatched"),
                      t("st_received"), t("card_short")]
    st.dataframe(merged, use_container_width=True, hide_index=True)

    st.markdown(f"#### 📋 {t('latest_orders')}")
    with db_conn() as conn:
        latest = pd.read_sql_query("""
            SELECT order_id, branch, order_date, status, COUNT(*) AS lines
            FROM orders
            GROUP BY order_id, branch, order_date, status
            ORDER BY order_date DESC LIMIT 10
        """, conn)
    if latest.empty:
        st.caption(t("no_data"))
    else:
        latest.columns = [t("order_id"), t("branch"), t("order_date"),
                          t("status"), "Lines"]
        st.dataframe(latest, use_container_width=True, hide_index=True)

    st.markdown(f"#### 🔔 {t('latest_short')}")
    with db_conn() as conn:
        ls = pd.read_sql_query("""
            SELECT branch, name, short_cartons, short_pcs, status, reported_date
            FROM shortages ORDER BY reported_date DESC LIMIT 10
        """, conn)
    if ls.empty:
        st.caption(t("no_data"))
    else:
        ls.columns = [t("branch"), t("name"), t("qty_cartons"),
                      t("qty_pcs"), t("status"), t("order_date")]
        st.dataframe(ls, use_container_width=True, hide_index=True)


def page_admin_all_orders() -> None:
    st.markdown(f"### 📋 {t('nav_all_orders')}")
    bs = st.multiselect(t("branch"), BRANCHES, default=BRANCHES, key="adm_b")
    ss = st.multiselect(t("status"), list(OrderStatus.ALL),
                        default=list(OrderStatus.ALL), key="adm_s")
    if not bs or not ss:
        st.info(t("no_data"))
        return
    qb = ",".join("?" * len(bs))
    qs = ",".join("?" * len(ss))
    with db_conn() as conn:
        df = pd.read_sql_query(
            f"SELECT * FROM orders WHERE branch IN ({qb}) AND status IN ({qs}) "
            "ORDER BY order_date DESC", conn, params=bs + ss,
        )
    if df.empty:
        st.info(t("no_data"))
        return
    for order_id, group in df.groupby("order_id", sort=False):
        status = group["status"].iloc[0]
        with st.expander(
            f"📦 {order_id} · 🏪 {group['branch'].iloc[0]} · "
            f"{group['order_date'].iloc[0][:16]} · {len(group)} items"
        ):
            st.markdown(status_pill(status), unsafe_allow_html=True)
            st.markdown("")
            show = group[["name", "unit",
                          "qty_cartons", "qty_pcs",
                          "dispatch_cartons", "dispatch_pcs",
                          "actual_cartons", "actual_pcs"]].fillna("-")
            show.columns = [
                t("name"), t("unit"),
                f"📋 {t('qty_cartons')}", f"📋 {t('qty_pcs')}",
                f"🚚 {t('qty_cartons')}", f"🚚 {t('qty_pcs')}",
                f"✅ {t('qty_cartons')}", f"✅ {t('qty_pcs')}",
            ]
            st.dataframe(show, use_container_width=True, hide_index=True)


# =========================================================================
# EXCEL EXPORTS
# =========================================================================
def _style_excel_header(ws, last_col: int) -> None:
    fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    align = Alignment(horizontal="center", vertical="center")
    border = Border(*(Side(style="thin", color="999999"),) * 4)
    for col_idx in range(1, last_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 8
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


# --- Picking-slip exports for warehouse staff -----------------------------
def _attach_categories(df: pd.DataFrame) -> pd.DataFrame:
    """Add a 'Category' column to an order-lines dataframe by looking each
    line up in the products master. Match on ItemCode first (exact), then
    Barcode (exact). Unmatched (e.g. manually added items) → 'General'."""
    if df.empty:
        return df.assign(Category="General")
    products = load_products()
    if products.empty:
        return df.assign(Category="General")

    # Build two lookups: ItemCode → Category, Barcode → Category
    by_code = {}
    by_barcode = {}
    for _, p in products.iterrows():
        cat = str(p.get("Category", "")).strip() or "General"
        ic = str(p.get("ItemCode", "")).strip()
        bc = str(p.get("Barcode", "")).strip()
        if ic:
            by_code[ic] = cat
        if bc:
            by_barcode[bc] = cat

    def _lookup(row) -> str:
        ic = str(row.get("item_code", "") or "").strip()
        bc = str(row.get("barcode", "") or "").strip()
        if ic and ic in by_code:
            return by_code[ic]
        if bc and bc in by_barcode:
            return by_barcode[bc]
        return "General"

    out = df.copy()
    out["Category"] = out.apply(_lookup, axis=1)
    return out


def _build_picking_slip_sheet(ws, df: pd.DataFrame, title_block: list[tuple[str, str]]) -> None:
    """Render a printable picking slip into ws, grouped by Category.

    Columns kept (5):  Barcode | Name | Cartons | Pcs | ✓
    Columns dropped:   Order ID, Order Date, Item Code, Unit

    Each category gets a banner row (light-blue, full-width, bold) listing
    the count + total cartons/pcs for that category. Categories appear in
    a fixed warehouse-zone order (Frozen first → Chilled → ... → General),
    with any unrecognized categories appended alphabetically at the end.
    Items without a Category in products.xlsx fall back to 'General'.
    """
    # ----- Title rows --------------------------------------------------
    for i, (label, value) in enumerate(title_block, start=1):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=11)
        ws.cell(row=i, column=2, value=value).font = Font(size=11)

    cur_row = len(title_block) + 2  # one blank row gap below title block

    # ----- Style helpers ----------------------------------------------
    blue_fill   = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    banner_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    white_font  = Font(bold=True, color="FFFFFF", size=11)
    cat_font    = Font(bold=True, color="0D47A1", size=12)
    align_c     = Alignment(horizontal="center", vertical="center")
    align_l     = Alignment(horizontal="left",   vertical="center")
    border      = Border(*(Side(style="thin", color="999999"),) * 4)

    headers = [t("barcode"), t("name"), t("qty_cartons"), t("qty_pcs"), "✓"]
    n_cols = len(headers)

    # ----- Group by category ------------------------------------------
    df_cat = _attach_categories(df)

    # Preferred display order: arranged so warehouse staff can walk a sensible
    # picking route (cold first, then ambient by zone). Anything not listed
    # here gets appended alphabetically at the end.
    PREFERRED = ["Frozen", "Chilled", "Dairy", "Fruit", "Vegetable",
                 "Rice", "Beverage", "General"]
    present = list(df_cat["Category"].unique())
    ordered_cats = [c for c in PREFERRED if c in present] + sorted(
        c for c in present if c not in PREFERRED
    )

    for cat in ordered_cats:
        sub = df_cat[df_cat["Category"] == cat]
        if sub.empty:
            continue

        # ---- Banner row: full-width category header + totals ---------
        n_lines  = len(sub)
        total_ct = int(sub["qty_cartons"].sum())
        total_pc = int(sub["qty_pcs"].sum())
        banner = ws.cell(
            row=cur_row, column=1,
            value=f"  📦  {cat}   ·   {n_lines} items   ·   "
                  f"{total_ct} cartons   ·   {total_pc} pcs",
        )
        banner.font = cat_font
        banner.fill = banner_fill
        banner.alignment = align_l
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=n_cols,
        )
        ws.row_dimensions[cur_row].height = 22
        cur_row += 1

        # ---- Column-header row (per group, so each group is self-
        # contained when the warehouse staff splits the printout) -----
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=cur_row, column=col_idx, value=h)
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = align_c
            cell.border = border
        cur_row += 1

        # ---- Data rows ---------------------------------------------
        for r in sub.itertuples(index=False):
            ws.cell(row=cur_row, column=1,
                    value=str(getattr(r, "barcode", "") or ""))
            ws.cell(row=cur_row, column=2,
                    value=str(getattr(r, "name", "") or ""))
            ws.cell(row=cur_row, column=3,
                    value=int(getattr(r, "qty_cartons", 0) or 0))
            ws.cell(row=cur_row, column=4,
                    value=int(getattr(r, "qty_pcs", 0) or 0))
            ws.cell(row=cur_row, column=5, value="")  # tick box
            for col_idx in range(1, n_cols + 1):
                ws.cell(row=cur_row, column=col_idx).border = border
            cur_row += 1

        # ---- Spacer between groups -------------------------------
        cur_row += 1

    _autosize(ws)


def export_single_order_picking(order_id: str):
    """Single order picking slip — one Excel for printing."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE order_id = ? ORDER BY name",
            conn, params=(order_id,),
        )
    if df.empty:
        return None
    branch = df["branch"].iloc[0]
    order_date = df["order_date"].iloc[0]
    wb = Workbook()
    ws = wb.active
    ws.title = "Picking Slip"
    title_block = [
        (f"📋 {t('picking_slip')}", "SUNSHINE / 阳光集团"),
        (t("branch"),     branch),
        (t("order_id"),   order_id),
        (t("order_date"), order_date),
        (t("total_items"), str(len(df))),
    ]
    _build_picking_slip_sheet(ws, df, title_block)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_branch_pending_picking(branch: str):
    """All pending orders for one branch, one sheet per order_id."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE branch = ? AND status = 'Pending' "
            "ORDER BY order_date ASC, name",
            conn, params=(branch,),
        )
    if df.empty:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    for order_id, group in df.groupby("order_id", sort=False):
        # Excel sheet names <= 31 chars, no special chars
        safe = order_id[-25:].replace(":", "-").replace("/", "-")
        ws = wb.create_sheet(safe)
        title_block = [
            (f"📋 {t('picking_slip')}", "SUNSHINE / 阳光集团"),
            (t("branch"),     branch),
            (t("order_id"),   order_id),
            (t("order_date"), group["order_date"].iloc[0]),
            (t("total_items"), str(len(group))),
        ]
        _build_picking_slip_sheet(ws, group, title_block)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_all_pending_picking():
    """All branches, all pending orders. One sheet per branch (consolidated)."""
    with db_conn() as conn:
        df = pd.read_sql_query(
            "SELECT * FROM orders WHERE status = 'Pending' "
            "ORDER BY branch, order_date ASC, name",
            conn,
        )
    if df.empty:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    for branch in BRANCHES:
        sub = df[df["branch"] == branch]
        if sub.empty:
            continue
        ws = wb.create_sheet(branch[:30].replace("/", "-"))
        title_block = [
            (f"📋 {t('picking_slip')}", "SUNSHINE / 阳光集团"),
            (t("branch"), branch),
            (t("order_date"), datetime.now().strftime("%Y-%m-%d %H:%M")),
            (t("total_items"), str(len(sub))),
        ]
        _build_picking_slip_sheet(ws, sub, title_block)
    if not wb.sheetnames:
        return None
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_picking_list(date_from: str, date_to: str):
    with db_conn() as conn:
        df = pd.read_sql_query("""
            SELECT * FROM orders
            WHERE date(order_date) BETWEEN ? AND ?
            ORDER BY branch, name
        """, conn, params=(date_from, date_to))
    if df.empty:
        return None
    wb = Workbook()
    wb.remove(wb.active)
    headers = [t("order_id"), t("order_date"), t("status"), t("item_code"),
               t("barcode"), t("name"), t("unit"), t("qty_cartons"), t("qty_pcs")]
    for branch in BRANCHES:
        sub = df[df["branch"] == branch]
        if sub.empty:
            continue
        ws = wb.create_sheet(branch[:30].replace("/", "-"))
        ws.append(headers)
        for _, r in sub.iterrows():
            ws.append([r["order_id"], r["order_date"], r["status"],
                       r["item_code"], r["barcode"], r["name"], r["unit"],
                       r["qty_cartons"], r["qty_pcs"]])
        _style_excel_header(ws, len(headers))
        _autosize(ws)
    if not wb.sheetnames:
        return None
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_reconciliation(date_from: str, date_to: str):
    with db_conn() as conn:
        df = pd.read_sql_query("""
            SELECT branch, order_id, order_date, item_code, barcode, name, unit,
                   qty_cartons, qty_pcs,
                   dispatch_cartons, dispatch_pcs,
                   actual_cartons, actual_pcs, status
            FROM orders
            WHERE date(order_date) BETWEEN ? AND ?
              AND status IN ('Dispatched', 'Received')
            ORDER BY branch, order_id, name
        """, conn, params=(date_from, date_to))
    if df.empty:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Reconciliation"
    headers = [
        t("branch"), t("order_id"), t("order_date"), t("item_code"), t("barcode"),
        t("name"), t("unit"),
        f"📋 {t('qty_cartons')}", f"📋 {t('qty_pcs')}",
        f"🚚 {t('qty_cartons')}", f"🚚 {t('qty_pcs')}",
        f"✅ {t('qty_cartons')}", f"✅ {t('qty_pcs')}",
        f"Δ {t('qty_cartons')}", f"Δ {t('qty_pcs')}",
        t("status"),
    ]
    ws.append(headers)
    red    = PatternFill("solid", start_color="FFCDD2", end_color="FFCDD2")
    yellow = PatternFill("solid", start_color="FFF59D", end_color="FFF59D")

    def _eff_dispatch(row):
        """Effective dispatched qty: prefer dispatch_*, fallback to qty_*."""
        d_ct = row["dispatch_cartons"]
        d_pc = row["dispatch_pcs"]
        d_ct = int(row["qty_cartons"]) if pd.isna(d_ct) or d_ct is None else int(d_ct)
        d_pc = int(row["qty_pcs"])     if pd.isna(d_pc) or d_pc is None else int(d_pc)
        return d_ct, d_pc

    for _, r in df.iterrows():
        d_ct, d_pc = _eff_dispatch(r)
        ac = r["actual_cartons"]; ap = r["actual_pcs"]
        if pd.isna(ac) or pd.isna(ap):
            diff_ct = ""; diff_pc = ""; tag = "Pending"
            ac_v = "-" if pd.isna(ac) else int(ac)
            ap_v = "-" if pd.isna(ap) else int(ap)
        else:
            ac = int(ac); ap = int(ap)
            # Δ = received vs DISPATCHED (transit/branch shortage), not vs ordered
            diff_ct = ac - d_ct
            diff_pc = ap - d_pc
            tag = "SHORT" if (diff_ct < 0 or diff_pc < 0) else (
                  "OVER"  if (diff_ct > 0 or diff_pc > 0) else "OK")
            ac_v, ap_v = ac, ap
        ws.append([
            r["branch"], r["order_id"], r["order_date"],
            r["item_code"], r["barcode"], r["name"], r["unit"],
            r["qty_cartons"], r["qty_pcs"],   # 📋 ordered
            d_ct, d_pc,                        # 🚚 dispatched (effective)
            ac_v, ap_v,                        # ✅ actually received
            diff_ct, diff_pc,                  # Δ
            tag,
        ])
        row_idx = ws.max_row
        if tag == "SHORT":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = red
        elif tag == "OVER":
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow
    _style_excel_header(ws, len(headers))
    _autosize(ws)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def export_shortage_report(date_from: str, date_to: str):
    with db_conn() as conn:
        df = pd.read_sql_query("""
            SELECT * FROM shortages
            WHERE date(reported_date) BETWEEN ? AND ?
            ORDER BY reported_date DESC
        """, conn, params=(date_from, date_to))
    if df.empty:
        return None
    wb = Workbook(); ws = wb.active; ws.title = "Shortages"
    headers = [t("branch"), t("order_id"), t("name"), t("item_code"), t("barcode"),
               t("unit"),
               f"{t('shortage_qty')} {t('qty_cartons')}",
               f"{t('shortage_qty')} {t('qty_pcs')}",
               t("status"), t("warehouse_reply"),
               t("order_date"), t("remarks")]
    ws.append(headers)
    for _, r in df.iterrows():
        ws.append([r["branch"], r["order_id"], r["name"], r["item_code"], r["barcode"],
                   r["unit"], r["short_cartons"], r["short_pcs"],
                   r["status"], r["warehouse_reply"] or "",
                   r["reported_date"], r["branch_remarks"] or ""])
    _style_excel_header(ws, len(headers))
    _autosize(ws)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio


def page_admin_images() -> None:
    """Upload / view / remove product images. Manager-only."""
    st.markdown(f"### 🖼️ {t('img_title')}")
    st.caption(t("img_subtitle"))

    # =====================================================================
    # Auto-scan root folder. The system already imported on startup; this
    # section shows that summary and lets the user re-scan without
    # restarting (useful when they drop new files mid-session).
    # =====================================================================
    with st.expander(t("img_scan_title"), expanded=False):
        st.caption(t("img_scan_hint"))

        # Show the startup-scan summary if we still have it cached. This is
        # what was moved automatically when the user (re)started the app.
        startup_summary = st.session_state.get("_startup_scan_summary")

        # On-demand re-scan button
        if st.button(t("img_scan_now"), key="img_scan_now_btn",
                     type="primary", use_container_width=True):
            new_summary = scan_and_import_root_images()
            st.session_state["_startup_scan_summary"] = new_summary
            st.rerun()

        # Render the summary (whether from startup or from "Scan now")
        if startup_summary:
            moved = startup_summary.get("moved", [])
            unmatched = startup_summary.get("skipped_unmatched", [])
            errors = startup_summary.get("errors", [])

            if not moved and not unmatched and not errors:
                st.info(t("img_scan_none"))
            else:
                mc1, mc2, mc3 = st.columns(3)
                with mc1:
                    st.metric(f"✅ {t('img_scan_moved')}", len(moved))
                with mc2:
                    st.metric(f"⚠️ {t('img_scan_left')}", len(unmatched))
                with mc3:
                    st.metric(f"❌ {t('img_scan_errors')}", len(errors))

                if moved:
                    rows = [{
                        t("img_filename"):  m["from"],
                        t("img_matched_to"): f"{m['name']}  ({m['matched_by']})",
                    } for m in moved]
                    st.dataframe(pd.DataFrame(rows),
                                 use_container_width=True, hide_index=True)

                if unmatched:
                    st.warning(
                        f"⚠️ {len(unmatched)} file(s) in root folder do not "
                        f"match any product. Rename them or add the products "
                        f"to products.xlsx, then click '{t('img_scan_now')}'."
                    )
                    st.dataframe(
                        pd.DataFrame({t("img_filename"): unmatched}),
                        use_container_width=True, hide_index=True,
                    )

                if errors:
                    for e in errors:
                        st.error(f"❌ {e}")

    # =====================================================================
    # Batch upload by filename — naming convention: <ItemCode>.jpg or
    # <Barcode>.jpg. Files that don't match anything in products.xlsx are
    # surfaced separately so the user can fix names and re-upload.
    # =====================================================================
    with st.expander(t("img_batch_title"), expanded=False):
        st.caption(t("img_batch_hint"))

        batch_uploads = st.file_uploader(
            t("img_batch_choose"),
            type=[e.lstrip(".") for e in IMAGE_EXTS],
            accept_multiple_files=True,
            key="img_batch_files",
        )

        if batch_uploads:
            # Build a pairing preview from the current products master.
            products_df = load_products()
            matched, unmatched = plan_batch_image_upload(
                batch_uploads, products_df,
            )

            st.markdown(f"#### {t('img_batch_preview')}")
            mc1, mc2 = st.columns(2)
            with mc1:
                st.metric(f"✅ {t('img_matched')}", len(matched))
            with mc2:
                st.metric(f"⚠️ {t('img_unmatched')}", len(unmatched))

            # ---- Matched files preview ----
            if matched:
                preview_rows = [{
                    t("img_filename"):  m["filename"],
                    t("img_matched_to"): f"{m['name']}  ({m['matched_by']}: "
                                         f"{m['item_code'] or m['barcode']})",
                } for m in matched]
                st.dataframe(
                    pd.DataFrame(preview_rows),
                    use_container_width=True, hide_index=True,
                )

            # ---- Unmatched files preview ----
            if unmatched:
                st.warning(
                    f"⚠️ {len(unmatched)} file(s) could not be matched. "
                    f"Rename and re-upload to fix."
                )
                un_rows = [{
                    t("img_filename"):    u["filename"],
                    t("img_match_reason"): u["reason"],
                } for u in unmatched]
                st.dataframe(
                    pd.DataFrame(un_rows),
                    use_container_width=True, hide_index=True,
                )

            # ---- Confirm & save (only for matched ones) ----
            if matched:
                cc1, cc2 = st.columns([3, 1])
                with cc1:
                    if st.button(
                        t("img_batch_confirm").format(n=len(matched)),
                        type="primary", use_container_width=True,
                        key="img_batch_go",
                    ):
                        # Process one at a time so a single bad file doesn't
                        # block the rest. We collect results to show a final
                        # summary.
                        prog = st.progress(0.0, text="...")
                        ok_count = 0
                        fail_list: list[tuple[str, str]] = []
                        total = len(matched)
                        for i, m in enumerate(matched):
                            try:
                                save_product_image(
                                    item_code=m["item_code"],
                                    barcode=m["barcode"],
                                    name=m["name"],
                                    uploaded_bytes=m["uf"].getvalue(),
                                    original_filename=m["filename"],
                                )
                                ok_count += 1
                            except Exception as e:
                                fail_list.append((m["filename"], str(e)))
                            prog.progress(
                                (i + 1) / total,
                                text=f"{i + 1}/{total}  ·  {m['filename']}",
                            )
                        prog.empty()

                        if ok_count:
                            st.success(
                                t("img_batch_done").format(n=ok_count)
                            )
                        if fail_list:
                            st.error(t("img_batch_failed").format(n=len(fail_list)))
                            for fn, err in fail_list:
                                st.caption(f"❌ {fn}: {err}")
                        # Reset the uploader so the file list clears
                        st.session_state.pop("img_batch_files", None)
                        st.rerun()
                with cc2:
                    if st.button(t("img_batch_cancel"),
                                 use_container_width=True,
                                 key="img_batch_cancel"):
                        st.session_state.pop("img_batch_files", None)
                        st.rerun()

    st.divider()

    # ----- Search a product to upload for ------------------------------
    query = st.text_input(
        t("img_search"), key="img_search_q",
        placeholder=t("search_product"),
        help="Bluetooth scanner supported / 支持扫码枪",
    )
    results = search_products(query, limit=50)

    if results.empty and query:
        st.info(t("no_results"))
        return
    if results.empty:
        return

    PAGE_SIZE = 10
    total = len(results)
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = st.session_state.get("_img_page", 1)
    last_q = st.session_state.get("_img_last_q")
    if last_q != query:
        page = 1
        st.session_state["_img_last_q"] = query
    page = min(max(1, page), total_pages)
    st.session_state["_img_page"] = page

    start = (page - 1) * PAGE_SIZE
    end = min(start + PAGE_SIZE, total)
    page_rows = results.iloc[start:end]

    st.caption(
        f"{t('showing')} {start + 1}–{end} {t('of')} {total} "
        f"{t('results_count')} · {t('page')} {page} {t('of')} {total_pages}"
    )

    # ----- Render each product with image preview + uploader ----------
    for idx, row in page_rows.iterrows():
        item_code = str(row.get("ItemCode", "") or "")
        barcode   = str(row.get("Barcode", "") or "")
        name      = str(row.get("Name", "") or "")
        with st.container(border=True):
            cl, cm, cr = st.columns([1, 3, 2])

            # ---- Current image preview (or placeholder) ----
            img_path = get_product_image_path(item_code, barcode)
            with cl:
                if img_path is not None:
                    st.image(str(img_path), width=120)
                else:
                    st.markdown(
                        "<div style='width:120px;height:120px;"
                        "background:#f0f0f0;border:1px dashed #bbb;"
                        "border-radius:6px;display:flex;align-items:center;"
                        "justify-content:center;color:#999;font-size:28px;'>"
                        "📷</div>",
                        unsafe_allow_html=True,
                    )

            # ---- Product info ----
            with cm:
                st.markdown(f"**{name}**")
                bits = []
                if item_code: bits.append(f"📋 {item_code}")
                if barcode:   bits.append(f"📊 {barcode}")
                if bits:
                    st.caption(" · ".join(bits))
                if img_path is not None:
                    st.caption(f"✅ {t('img_has_image')}: `{img_path.name}`")
                else:
                    st.caption(f"⚪ {t('img_no_image')}")

            # ---- Upload + delete ----
            with cr:
                # Stable widget keys per product so the input survives
                # rerenders without colliding with other rows.
                upload_key = f"img_up_{item_code}_{barcode}_{idx}"
                up = st.file_uploader(
                    t("img_upload"),
                    type=[e.lstrip(".") for e in IMAGE_EXTS],
                    key=upload_key,
                    label_visibility="collapsed",
                )
                if up is not None:
                    try:
                        save_product_image(
                            item_code, barcode, name,
                            up.getvalue(), up.name,
                        )
                        st.success(f"✅ {t('img_uploaded')}")
                        # Reset the uploader so the success message doesn't
                        # re-fire on the next rerun
                        st.session_state.pop(upload_key, None)
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

                if img_path is not None:
                    if st.button(t("img_delete"),
                                 key=f"img_del_{item_code}_{barcode}_{idx}",
                                 use_container_width=True):
                        if delete_product_image(item_code, barcode):
                            st.success(t("img_deleted"))
                            st.rerun()

    # ----- Pagination -------------------------------------------------
    if total_pages > 1:
        pc1, pc2, pc3 = st.columns([1, 2, 1])
        with pc1:
            if st.button(t("prev_page"), key="img_prev",
                         disabled=(page <= 1),
                         use_container_width=True):
                st.session_state["_img_page"] = page - 1
                st.rerun()
        with pc2:
            st.markdown(
                f"<div style='text-align:center;padding-top:6px;'>"
                f"{t('page')} <b>{page}</b> {t('of')} <b>{total_pages}</b>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with pc3:
            if st.button(t("next_page"), key="img_next",
                         disabled=(page >= total_pages),
                         use_container_width=True):
                st.session_state["_img_page"] = page + 1
                st.rerun()


def page_admin_backup() -> None:
    """Local DB backup & restore management for admins."""
    st.markdown(f"### 💾 {t('backup_title')}")
    st.caption(t("backup_subtitle"))

    # ----- Current DB info ---------------------------------------------
    if DB_PATH.exists():
        size_kb = round(DB_PATH.stat().st_size / 1024, 1)
        with db_conn() as conn:
            n_orders = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
            n_ships  = conn.execute("SELECT COUNT(*) FROM shipments").fetchone()[0]
            n_short  = conn.execute("SELECT COUNT(*) FROM shortages").fetchone()[0]
        st.markdown(
            f"**📂 {t('current_db_info')}:** `{DB_PATH}` &nbsp;·&nbsp; "
            f"**{size_kb} KB** &nbsp;·&nbsp; "
            f"**{t('backup_counts')}:** {n_orders} / {n_ships} / {n_short}"
        )

    # ----- Action buttons ----------------------------------------------
    a1, a2, a3 = st.columns(3)
    with a1:
        if st.button(t("backup_now"), type="primary",
                     use_container_width=True, key="bk_now"):
            try:
                p = backup_database("manual")
                if p is None:
                    st.warning(t("no_data"))
                else:
                    st.success(f"{t('backup_done')}: `{p.name}`")
                    st.rerun()
            except Exception as e:
                st.error(f"{t('backup_failed')}: {e}")

    with a2:
        # Inline export-SQL — runs on click then surfaces a download_button
        if st.button(t("export_sql"), use_container_width=True, key="exp_sql"):
            data = export_db_dump_sql()
            if not data:
                st.warning(t("no_data"))
            else:
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                st.session_state["_sql_dump"] = (data, f"orders_dump_{stamp}.sql")
        if "_sql_dump" in st.session_state:
            data, fname = st.session_state["_sql_dump"]
            st.download_button(
                f"⬇️ {fname}",
                data=data, file_name=fname, mime="application/sql",
                use_container_width=True, key="dl_sql_dump",
            )

    with a3:
        # Import from SQL dump (text). File upload reuses st.file_uploader.
        with st.expander(t("import_sql")):
            up = st.file_uploader(t("upload_sql"), type=["sql", "txt"],
                                  key="imp_sql_file")
            if up is not None:
                if st.button(f"♻️ {t('import_sql')}", key="do_imp_sql",
                             type="primary"):
                    try:
                        safety = restore_from_sql_dump(up.getvalue())
                        st.success(
                            f"{t('import_done')}`{safety.name if safety else 'n/a'}`"
                        )
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))

    st.divider()

    # ----- Backup list -------------------------------------------------
    st.markdown(f"#### {t('available_backups')}")
    backups = list_backups()
    if not backups:
        st.info(t("no_backups"))
        return

    # If a restore is awaiting confirmation, show the dialog inline
    pending = st.session_state.get("_pending_restore")
    if pending:
        st.warning(f"{t('confirm_restore_q')}\n\n📁 `{pending}`")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button(t("yes_restore"), type="primary",
                         use_container_width=True, key="confirm_restore"):
                try:
                    safety = restore_from_backup(Path(pending))
                    st.success(
                        f"{t('restore_done')}`{safety.name if safety else 'n/a'}`"
                    )
                    st.session_state.pop("_pending_restore", None)
                    st.rerun()
                except Exception as e:
                    st.error(str(e))
                    st.session_state.pop("_pending_restore", None)
        with cc2:
            if st.button(t("cancel"), use_container_width=True, key="cancel_restore"):
                st.session_state.pop("_pending_restore", None)
                st.rerun()
        st.divider()

    # Render each backup as a row with download / restore / delete buttons
    for b in backups:
        with st.container(border=True):
            r1, r2 = st.columns([3, 2])
            with r1:
                # Color the row name based on label so manual / pre_restore
                # snapshots stand out from the auto-snapshots
                icon = "💾"
                if "_manual_" in b["name"]:
                    icon = "📌"
                elif "_pre_restore_" in b["name"]:
                    icon = "🛟"
                st.markdown(f"{icon} **`{b['name']}`**")
                st.caption(
                    f"📅 {b['created']} &nbsp;·&nbsp; "
                    f"💽 {b['size_kb']} KB &nbsp;·&nbsp; "
                    f"📊 {b['counts']['orders']} / "
                    f"{b['counts']['shipments']} / "
                    f"{b['counts']['shortages']}"
                )
            with r2:
                bc1, bc2, bc3 = st.columns(3)
                # Download — read bytes then surface download_button
                with bc1:
                    dl_key = f"dl_bk_{b['name']}"
                    if st.button(t("download_backup"), key=dl_key,
                                 use_container_width=True):
                        data = b["path"].read_bytes()
                        st.session_state[f"_dl_bk_{b['name']}"] = data
                    if f"_dl_bk_{b['name']}" in st.session_state:
                        st.download_button(
                            "⬇️", data=st.session_state[f"_dl_bk_{b['name']}"],
                            file_name=b["name"],
                            mime="application/x-sqlite3",
                            use_container_width=True,
                            key=f"dl2_bk_{b['name']}",
                        )
                with bc2:
                    if st.button(t("restore_backup"),
                                 key=f"rst_bk_{b['name']}",
                                 use_container_width=True):
                        st.session_state["_pending_restore"] = str(b["path"])
                        st.rerun()
                with bc3:
                    if st.button(t("delete_backup"),
                                 key=f"del_bk_{b['name']}",
                                 use_container_width=True):
                        try:
                            b["path"].unlink()
                            # Drop any cached download bytes for this file
                            st.session_state.pop(f"_dl_bk_{b['name']}", None)
                            st.rerun()
                        except OSError as e:
                            st.error(str(e))


def page_admin_email() -> None:
    """Configure SMTP, per-event recipients, and test sending."""
    st.markdown(f"### 📧 {t('email_title')}")
    st.caption(t("email_subtitle"))

    cfg = load_email_config()

    # ----- Master toggle ------------------------------------------------
    enabled = st.checkbox(
        t("email_enabled"), value=bool(cfg.get("enabled", False)),
        key="em_enabled",
    )

    # ----- SMTP settings ------------------------------------------------
    st.markdown(f"#### 🔧 {t('smtp_settings')}")
    s1, s2 = st.columns(2)
    with s1:
        smtp_host = st.text_input(
            t("smtp_host"), value=cfg.get("smtp_host", ""),
            placeholder="smtp.gmail.com", key="em_host",
        )
        smtp_user = st.text_input(
            t("smtp_user"), value=cfg.get("smtp_user", ""),
            key="em_user",
        )
        from_addr = st.text_input(
            t("smtp_from"), value=cfg.get("from_addr", ""),
            placeholder="orders@example.com", key="em_from",
        )
    with s2:
        smtp_port = st.number_input(
            t("smtp_port"), min_value=1, max_value=65535,
            value=int(cfg.get("smtp_port", 587)), step=1, key="em_port",
        )
        smtp_password = st.text_input(
            t("smtp_password"), value=cfg.get("smtp_password", ""),
            type="password", key="em_pw",
        )
        use_tls = st.checkbox(
            t("smtp_use_tls"), value=bool(cfg.get("use_tls", True)),
            key="em_tls",
            help="Port 465 → implicit SSL (this option ignored). "
                 "Port 587 → STARTTLS (this option recommended).",
        )

    # ----- Per-event recipients -----------------------------------------
    st.markdown(f"#### 🔔 {t('event_settings')}")
    ev_cfg = cfg.get("events", {})
    new_events: dict[str, dict] = {}
    for key, label_key in [
        ("new_order",  "ev_new_order"),
        ("dispatched", "ev_dispatched"),
        ("shortage",   "ev_shortage"),
    ]:
        cur = ev_cfg.get(key, {"enabled": True, "to": []})
        with st.container(border=True):
            ec1, ec2 = st.columns([1, 3])
            with ec1:
                ev_on = st.checkbox(
                    t(label_key),
                    value=bool(cur.get("enabled", True)),
                    key=f"em_ev_{key}",
                )
            with ec2:
                to_text = st.text_area(
                    t("recipients"),
                    value="\n".join(cur.get("to", [])),
                    key=f"em_to_{key}",
                    height=80,
                )
            new_events[key] = {
                "enabled": bool(ev_on),
                "to": [
                    line.strip() for line in to_text.splitlines()
                    if line.strip()
                ],
            }

    # ----- Per-branch recipient (optional) -----------------------------
    st.markdown(f"#### 🏪 {t('branch_recipients')}")
    st.caption(t("branch_email_hint"))
    branch_emails = cfg.get("branch_emails", {}) or {}
    new_branch_emails: dict[str, str] = {}
    for branch in BRANCHES:
        new_branch_emails[branch] = st.text_input(
            branch, value=branch_emails.get(branch, ""),
            key=f"em_br_{branch}", placeholder="(optional)",
        )

    # ----- Save ---------------------------------------------------------
    save_col, _ = st.columns([1, 3])
    with save_col:
        if st.button(t("save_email_cfg"), type="primary",
                     use_container_width=True, key="em_save"):
            new_cfg = {
                "enabled":       bool(enabled),
                "smtp_host":     smtp_host.strip(),
                "smtp_port":     int(smtp_port),
                "smtp_user":     smtp_user.strip(),
                "smtp_password": smtp_password,
                "from_addr":     from_addr.strip(),
                "use_tls":       bool(use_tls),
                "events":        new_events,
                "branch_emails": {k: v.strip() for k, v in
                                  new_branch_emails.items() if v.strip()},
            }
            save_email_config(new_cfg)
            st.success(t("email_cfg_saved"))
            st.rerun()

    st.divider()

    # ----- Test send ---------------------------------------------------
    st.markdown(f"#### 🧪 {t('test_email')}")
    tc1, tc2 = st.columns([3, 1])
    with tc1:
        test_to = st.text_input(
            t("test_to"), value="", key="em_test_to",
            placeholder=cfg.get("from_addr", ""),
        )
    with tc2:
        st.markdown("&nbsp;", unsafe_allow_html=True)  # vertical spacer
        if st.button(t("test_email"), key="em_test_btn",
                     use_container_width=True):
            if not test_to.strip():
                st.warning("⚠️ recipient empty")
            else:
                # Use the *current form values* (not the saved ones), so
                # admin can test before saving.
                test_cfg = {
                    "smtp_host":     smtp_host.strip(),
                    "smtp_port":     int(smtp_port),
                    "smtp_user":     smtp_user.strip(),
                    "smtp_password": smtp_password,
                    "from_addr":     from_addr.strip(),
                    "use_tls":       bool(use_tls),
                }
                ok, err = send_test_email(test_cfg, test_to.strip())
                if ok:
                    st.success(t("test_sent_ok"))
                else:
                    st.error(f"{t('test_sent_fail')}: {err}")

    st.divider()

    # ----- Activity log ------------------------------------------------
    st.markdown(f"#### 📜 {t('email_log_title')}")
    log = list(reversed(load_email_log()))  # newest first
    if not log:
        st.info(t("email_log_empty"))
    else:
        # Show last 50 entries; deeper history kept on disk
        rows = []
        for entry in log[:50]:
            rows.append({
                t("log_time"):   entry.get("ts", ""),
                t("log_event"):  entry.get("event", ""),
                t("log_to"):     ", ".join(entry.get("to", [])),
                "Subject":       entry.get("subject", ""),
                t("log_status"): "✅ " + t("log_ok") if entry.get("ok")
                                 else "❌ " + t("log_fail"),
                "Error":         entry.get("error", ""),
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True,
                     hide_index=True)
        if st.button(t("log_clear"), key="em_log_clear"):
            try:
                if EMAIL_LOG_PATH.exists():
                    EMAIL_LOG_PATH.unlink()
                st.rerun()
            except Exception as e:
                st.error(str(e))


def page_admin_export() -> None:
    st.markdown(f"### 📥 {t('nav_export')}")
    today = datetime.now().date()
    c1, c2 = st.columns(2)
    with c1: d1 = st.date_input(t("from_date"), value=today, key="exp_d1")
    with c2: d2 = st.date_input(t("to_date"),   value=today, key="exp_d2")
    d1s = d1.strftime("%Y-%m-%d"); d2s = d2.strftime("%Y-%m-%d")

    for label_key, fn, fname in [
        ("exp_picking", export_picking_list,    "picking_list"),
        ("exp_recon",   export_reconciliation,  "reconciliation"),
        ("exp_short",   export_shortage_report, "shortage_report"),
    ]:
        st.divider()
        st.markdown(f"#### {t(label_key)}")
        if st.button(f"{t('generate')} — {t(label_key)}", key=f"gen_{fname}"):
            bio = fn(d1s, d2s)
            if bio is None:
                st.warning(t("no_data"))
            else:
                st.download_button(
                    f"⬇️ {t('download')} {fname}_{d1s}_{d2s}.xlsx",
                    data=bio,
                    file_name=f"{fname}_{d1s}_{d2s}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


# =========================================================================
# WORKSPACE ROUTERS
# =========================================================================
def _render_sidebar_nav(items: list[tuple[str, str]]) -> None:
    for page_key, label in items:
        is_active = st.session_state.page == page_key
        if st.sidebar.button(
            label, key=f"nav_{page_key}",
            use_container_width=True,
            type="primary" if is_active else "secondary",
        ):
            st.session_state.page = page_key
            # Leaving any page should drop the order-confirm modal state.
            st.session_state["confirming"] = False
            st.rerun()


def render_branch() -> None:
    render_header(f"🛒 {st.session_state.branch}")
    render_lang_switch()

    st.sidebar.markdown(f"**{t('current_branch')}**")
    st.sidebar.markdown(f"`{st.session_state.branch}`")
    st.sidebar.divider()

    _render_sidebar_nav([
        ("order",     t("nav_order")),
        ("my_orders", t("nav_my_orders")),
        ("my_short",  t("nav_my_short")),
    ])
    st.sidebar.divider()
    with st.sidebar:
        logout_button("logout_branch")

    pages = {
        "order":     page_branch_order,
        "my_orders": page_branch_my_orders,
        "my_short":  page_branch_shortages,
    }
    pages.get(st.session_state.page, page_branch_order)()


def render_warehouse() -> None:
    render_header(t("role_warehouse"))
    render_lang_switch()

    st.sidebar.markdown(f"**{t('role_warehouse')}**")
    st.sidebar.divider()
    _render_sidebar_nav([
        ("pending",  t("nav_pending")),
        ("short_in", t("nav_short_in")),
        ("history",  t("nav_dispatch_history")),
    ])
    st.sidebar.divider()
    with st.sidebar:
        logout_button("logout_warehouse")

    pages = {
        "pending":  page_warehouse_pending,
        "short_in": page_warehouse_shortages,
        "history":  page_warehouse_dispatch_history,
    }
    pages.get(st.session_state.page, page_warehouse_pending)()


def render_admin() -> None:
    render_header(t("role_admin"))
    render_lang_switch()

    st.sidebar.markdown(f"**{t('role_admin')}**")
    st.sidebar.divider()
    _render_sidebar_nav([
        ("dashboard",  t("nav_dashboard")),
        ("all_orders", t("nav_all_orders")),
        ("dispatch",   t("nav_dispatch")),
        ("short_mgmt", t("nav_short_mgmt")),
        ("export",     t("nav_export")),
        ("images",     t("nav_images")),
        ("backup",     t("nav_backup")),
        ("email",      t("nav_email")),
    ])
    st.sidebar.divider()
    with st.sidebar:
        logout_button("logout_admin")

    pages = {
        "dashboard":  page_admin_dashboard,
        "all_orders": page_admin_all_orders,
        "dispatch":   page_warehouse_pending,    # admin reuses warehouse view
        "short_mgmt": page_warehouse_shortages,  # ditto
        "export":     page_admin_export,
        "images":     page_admin_images,
        "backup":     page_admin_backup,
        "email":      page_admin_email,
    }
    pages.get(st.session_state.page, page_admin_dashboard)()


# =========================================================================
# ROUTER
# =========================================================================
def route() -> None:
    if st.session_state.lang is None:
        page_pick_language()
        return
    if not is_authenticated():
        page_login()
        return
    role = st.session_state.role
    if role == Role.BRANCH:
        render_branch()
    elif role == Role.WAREHOUSE:
        render_warehouse()
    elif role == Role.ADMIN:
        render_admin()
    else:
        st.error("Unknown role; please log out.")
        logout_button("logout_unknown")


# =========================================================================
# ENTRY
# =========================================================================
def main() -> None:
    st.set_page_config(
        page_title="SUNSHINE · 阳光集团 订货系统",
        page_icon="☀️",
        layout="wide",
        initial_sidebar_state="auto",
    )
    inject_css()
    init_db()
    init_session()

    # Take a one-time backup snapshot per Streamlit process so even if a
    # code change later corrupts the DB, we have a "fresh process start"
    # restore point. The throttled auto-snapshot logic kicks in beyond that.
    if not st.session_state.get("_startup_backup_done"):
        try:
            auto_snapshot_if_due()
        except Exception as e:
            log_exception("startup_auto_snapshot", e)
        st.session_state["_startup_backup_done"] = True

    # Auto-import any image files dropped into the root folder. Stash the
    # summary in session_state so the admin Images page can show what got
    # moved at startup. Done once per process (same pattern as backup).
    if not st.session_state.get("_startup_scan_done"):
        try:
            st.session_state["_startup_scan_summary"] = \
                scan_and_import_root_images()
        except Exception as e:
            st.session_state["_startup_scan_summary"] = {
                "moved": [], "skipped_unmatched": [],
                "errors": [f"scan failed: {e}"],
            }
        st.session_state["_startup_scan_done"] = True

    route()


if __name__ == "__main__":
    main()
